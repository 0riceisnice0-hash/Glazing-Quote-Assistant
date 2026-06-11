import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\zacpl\Desktop\Tender Documents")


def main():
    results = []
    for pack in sorted([p for p in ROOT.iterdir() if p.is_dir()]):
        quote_files = find_quote_files(pack)
        results.append({
            "pack": pack.name,
            "quoteFiles": [str(p) for p in quote_files],
            "candidates": [extract_candidates(p) for p in quote_files],
        })
    print(json.dumps(results, indent=2, ensure_ascii=False))


def find_quote_files(pack):
    files = []
    for p in pack.rglob("*"):
        if not p.is_file():
            continue
        lower_path = str(p).lower()
        lower_name = p.name.lower()
        if "supplier quote" in lower_path or "supplier quotes" in lower_path:
            continue
        if "do not send" in lower_name:
            continue
        if p.suffix.lower() not in [".pdf", ".xlsx", ".xlsm", ".xls"]:
            continue
        if "client quote" in lower_path or "quotation" in lower_name or "glazing quote" in lower_name:
            files.append(p)
    return sorted(files)


def extract_candidates(path):
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_candidates(path)
    if suffix in [".xlsx", ".xlsm", ".xls"]:
        return extract_workbook_candidates(path)
    return {"file": str(path), "type": suffix, "lines": []}


def extract_pdf_candidates(path):
    lines = []
    text = ""
    try:
        reader = PdfReader(str(path))
        for page in reader.pages:
            text += "\n" + (page.extract_text() or "")
    except Exception as exc:
        return {"file": str(path), "type": "pdf", "error": str(exc), "lines": []}

    for raw in text.splitlines():
        line = compact(raw)
        if not line:
            continue
        if "£" in line or re.search(r"\b(total|subtotal|vat|grand total|quote total|amount due)\b", line, re.I):
            amounts = parse_amounts(line)
            if amounts or re.search(r"\b(total|subtotal|vat|grand total|quote total|amount due)\b", line, re.I):
                lines.append({"line": line[:240], "amounts": amounts})
    return {
        "file": str(path),
        "type": "pdf",
        "bestTotal": choose_best_amount(lines),
        "lines": lines[-30:],
    }


def extract_workbook_candidates(path):
    candidates = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"file": str(path), "type": "workbook", "error": str(exc), "lines": []}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for r_idx, row in enumerate(rows, start=1):
            values = [v for v in row if v is not None and str(v).strip() != ""]
            if not values:
                continue
            row_text = " | ".join(str(v).strip() for v in values)
            has_keyword = re.search(r"\b(total|subtotal|vat|grand total|quote total|amount due)\b", row_text, re.I)
            nums = [float(v) for v in values if isinstance(v, (int, float)) and abs(float(v)) >= 1]
            if has_keyword or ("£" in row_text):
                candidates.append({
                    "sheet": ws.title,
                    "row": r_idx,
                    "line": row_text[:240],
                    "amounts": nums + parse_amounts(row_text),
                })
    return {
        "file": str(path),
        "type": "workbook",
        "bestTotal": choose_best_amount(candidates),
        "lines": candidates[-40:],
    }


def choose_best_amount(lines):
    scored = []
    for item in lines:
        line = item.get("line", "")
        amounts = item.get("amounts", [])
        if not amounts:
            continue
        amount = max(amounts)
        score = 0
        if re.search(r"\bgrand\s+total|amount\s+due|total\s+including|total\s+inc|total\s+incl|total\s+price\b", line, re.I):
            score += 5
        if re.search(r"\btotal\b", line, re.I):
            score += 3
        if re.search(r"\bvat\b", line, re.I):
            score -= 1
        if re.search(r"\bsubtotal|sub-total|net\b", line, re.I):
            score -= 1
        score += min(amount / 100000, 2)
        scored.append((score, amount, line))
    if not scored:
        return None
    scored.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return {"amount": round(scored[0][1], 2), "line": scored[0][2]}


def parse_amounts(text):
    amounts = []
    for m in re.finditer(r"£\s*([0-9][0-9,]*(?:\.[0-9]{2})?)", text):
        amounts.append(float(m.group(1).replace(",", "")))
    return amounts


def compact(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


if __name__ == "__main__":
    main()
