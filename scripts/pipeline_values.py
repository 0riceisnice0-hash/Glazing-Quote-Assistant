# -*- coding: utf-8 -*-
"""Stage 3 of the quote-pipeline truth exercise (Zac, dashmsg-16).

For every job that has a client quote on file, read the issued value out of the
house pricing workbook, and match the job against the Estimating Log to see
whether an outcome was ever recorded.

Reads  scratchpad/pipeline-stage1.json
Writes scratchpad/pipeline-stage3.json
"""
import datetime as dt
import json
import os
import re

import openpyxl

LOG = (r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\13. Estimating"
       r"\Leads\Estimating Log.xlsx")
IN = os.path.join("scratchpad", "pipeline-stage1.json")
OUT = os.path.join("scratchpad", "pipeline-stage3.json")
SINCE = dt.datetime(2025, 9, 1).timestamp()


def words(s):
    return set(re.findall(r"[a-z0-9]{3,}", (s or "").lower()))


def read_total(path):
    """House pricing docs label the bottom line TOTAL* - take the first number
    to the right of that label. Returns (value, doc_date_text) or (None, None)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None, None
    best, datetxt = None, None
    for ws in wb.worksheets:
        try:
            rows = list(ws.iter_rows())
        except Exception:
            continue
        for row in rows:
            for i, c in enumerate(row):
                v = c.value
                if isinstance(v, str):
                    s = v.strip().lower()
                    if s.startswith("date:") and not datetxt:
                        datetxt = v.strip()[5:].strip()
                    if s.startswith("total") and "vat" not in s:
                        for c2 in row[i + 1:]:
                            if isinstance(c2.value, (int, float)) and c2.value:
                                if best is None or c2.value > best:
                                    best = float(c2.value)
                                break
    try:
        wb.close()
    except Exception:
        pass
    return best, datetxt


def load_log():
    wb = openpyxl.load_workbook(LOG, data_only=True, read_only=True)
    ws = wb["Estimating Log"]
    out = []
    header = None
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if header is None:
            if vals and "Client/Prospect" in [str(v) for v in vals if v]:
                header = {str(v).strip(): i for i, v in enumerate(vals) if v}
            continue
        def get(name):
            i = header.get(name)
            return vals[i] if i is not None and i < len(vals) else None
        client, project = get("Client/Prospect"), get("Project name")
        if not client and not project:
            continue
        out.append({
            "client": str(client or "").strip(),
            "project": str(project or "").strip(),
            "enquiry": str(get("Date of Enquiry") or "")[:10],
            "deadline": str(get("Deadline") or "")[:10],
            "notes": str(get("Notes ") or get("Notes") or "").strip(),
            "issued_for_checking": str(get("Issued for checking") or "").strip(),
            "wl": str(get("W/L") or "").strip(),
        })
    return out


def main():
    rows = json.load(open(IN, encoding="utf-8"))
    rows = [r for r in rows if r["latest_quote_mtime"] >= SINCE]
    log = load_log()
    print("jobs quoted since 01/09/2025: %d   estimating-log rows: %d" % (len(rows), len(log)))

    logged_outcome = sum(1 for l in log if l["wl"])
    print("log rows carrying a W/L outcome: %d" % logged_outcome)

    for n, r in enumerate(rows, 1):
        # Prefer the sendable house pricing workbook in the client-quote folder.
        cands = [q for q in r["quotes"]
                 if q["file"].lower().endswith(".xlsx") and q["in_client_folder"]]
        cands += [q for q in r["quotes"] if q["file"].lower().endswith(".xlsx")]
        r["value"], r["value_from"], r["doc_date"] = None, None, None
        for q in cands[:4]:
            val, datetxt = read_total(os.path.join(r["path"], q["rel"]))
            if val:
                r["value"], r["value_from"], r["doc_date"] = val, q["file"], datetxt
                break

        rw, rp = words(r["client"]), words(r["project"])
        best, score = None, 0
        for l in log:
            s = len(rw & words(l["client"])) * 2 + len(rp & words(l["project"])) * 3
            if s > score:
                best, score = l, s
        r["log"] = best if score >= 3 else None
        r["log_score"] = score
        if n % 40 == 0:
            print("  ...%d/%d" % (n, len(rows)))

    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, indent=1, ensure_ascii=False)

    priced = [r for r in rows if r["value"]]
    print("\nvalue read for %d of %d jobs" % (len(priced), len(rows)))
    print("matched to an estimating-log row: %d" % sum(1 for r in rows if r["log"]))
    print("total value of quotes with a value: GBP %s" % format(sum(r["value"] for r in priced), ",.2f"))


if __name__ == "__main__":
    main()
