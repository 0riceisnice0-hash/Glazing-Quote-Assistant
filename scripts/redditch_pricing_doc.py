# -*- coding: utf-8 -*-
"""Redditch Library BLBS0956 - the client-facing pricing document and proposal.

Adam, 28/07 19:43: "Can you put this into our pricing document, but we need to
undercut Joedan. We also need to ensure we inform the client of our 10 year
warranty, which we can manually put in a proposal document. We will also need
to include the same inclusions/exclusions as Joedan."

Three instructions, and the third one costs money. Matching Joedan's scope is
not free: their clause 12 INCLUDES removal of the existing elements, clause 13
a 45mm cloaking profile, clause 4 trickle vents, clause 15 perimeter sealing at
the time of installation. Our benchmark of 28/07 carried none of the site-applied
ones. Costing them is most of what this script does.

  python scripts/redditch_pricing_doc.py

Writes:
  outputs/Redditch Library - Fenster Pricing Document.xlsx   (house template)
  outputs/Redditch Library - Fenster Proposal.html/.pdf      (10 year warranty)
  outputs/redditch-pricing.json                              (the build-up)
"""
import json
import math
import os
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import mary_pricing as engine
from redditch_takeoff import SCHEDULE, code_for

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "outputs")

# --------------------------------------------------------------- the evidence
# BSW QT250834, 15/06/2026 - Sheerline Prestige quoted TO US for Pride
# Developments' Severn Trent job. 6 lines, 27 units, 72.578 m2, reconciling to
# GBP 34,902.35. Nearest comparable buy Fenster holds: same supplier, same
# client, six weeks old, sizes bracketing Redditch's.
ST_A, ST_B = 721.4685, -0.40934          # rate = a * area^b, R2 0.9934
ST_REF = "BSW QT250834, 15/06/2026, Pride Developments - Severn Trent"

# Measured supplier factors from Fenster's own sent pricing documents, same
# code and band. BSW +5.7% (n=272), Aplus -1.6% (n=83), 4Ali -1.5% (n=82).
# The Severn Trent curve is a BSW curve, so this converts it to an Aplus one.
BSW_FACTOR, APLUS_FACTOR = 1.057, 0.984
SECOND_SUPPLIER = APLUS_FACTOR / BSW_FACTOR      # 0.9309

MASTIC_RATE = 5.0                # GBP/linear metre - js/pricing.js masticRate
MCD = 0.025                      # to mirror Joedan's commercial terms
STRIP_OUT_ALLOWANCE = 3000.00    # ALLOWANCE, NOT A RATE. No strip-out rate
                                 # exists anywhere in Fenster's records - 362
                                 # archive workbooks searched 28/07. Adam's to
                                 # confirm or move (REQ-24).

JOEDAN_GROSS = 90687.17          # Appendix 2, JCQ.9727, gross of 2.5% MCD
JOEDAN_NET = JOEDAN_GROSS * (1 - MCD)


def describe(ref, w, h, cfg, code):
    th, fl, fp, sd, dd = cfg["TH"], cfg["FL"], cfg["FP"], cfg["SD"], cfg["DD"]
    parts = []
    if th:
        parts.append("%d top hung opening light%s" % (th, "s" if th > 1 else ""))
    if fl:
        parts.append("%d fixed light%s" % (fl, "s" if fl > 1 else ""))
    if fp:
        parts.append("%d infill panel%s" % (fp, "s" if fp > 1 else ""))
    if sd:
        parts.append("single leaf commercial doorset")
    if dd:
        parts.append("double leaf commercial doorset")
    if not parts:
        parts.append("fixed light (configuration not stated on the tender schedule)")
    kind = "Aluminium doorset" if (sd or dd) else "Aluminium window"
    if (sd or dd) and (th or fl or fp):
        kind = "Aluminium window and doorset assembly"
    return "Ref %s - %s: %s" % (ref, kind, ", ".join(parts))


def build():
    with open(os.path.join(OUT, "redditch-takeoff.json"), encoding="utf-8") as fh:
        take = json.load(fh)
    cfg = {l["ref"]: l["config"] for l in take["lines"]}

    rows, lines = [], []
    perimeter_m = 0.0
    for ref, w, h, th, fl, fp, sd, dd in SCHEDULE:
        area = w / 1000.0 * h / 1000.0
        code = code_for(area, sd, dd)
        bsw_rate = ST_A * area ** ST_B                      # GBP/m2, BSW curve
        rate = bsw_rate * SECOND_SUPPLIER                   # second supplier
        line = engine.price_line(code, w, h, qty=1, supply_rate=rate)
        line["ref"] = ref
        line["rate_per_m2"] = round(rate, 2)
        lines.append(line)
        perimeter_m += 2 * (w + h) / 1000.0
        rows.append({
            "code": code,
            "desc": describe(ref, w, h, cfg[ref], code),
            "size": "%d x %d" % (w, h),
            "qty": 1,
            "unit": "nr",
            "frames": round(line["supply"], 2),
        })

    area_total = sum(w * h for _, w, h, *_r in SCHEDULE) / 1e6
    job = engine.price_job(lines)
    frames = sum(l["supply"] for l in lines)
    adders = sum(l["adder"] for l in lines)
    installation = job["installation"]

    # Solar-control glass. Spec cl.8 demands a 4mm bronze anti-sun toughened
    # outer pane; the curve is built from ordinary softcoat units. Register
    # glass-unit medians: 103.03 solar, 89.74 plain softcoat.
    solar = area_total * (103.03 - 89.74)

    # Perimeter sealing. Joedan cl.15 include it; pack p70 cl.11 REQUIRES it
    # ("Allow for external grade waterproof mastic sealant to all windows").
    # Fenster's house documents have always carried this as an OPTIONAL extra -
    # that is the Princess Beatrice REQ-6 error and it must not repeat here.
    mastic = perimeter_m * MASTIC_RATE

    rows.append({
        "code": "", "desc": "Perimeter sealing - external grade low modulus waterproof "
                            "mastic to all window and doorset joints with the structure, "
                            "applied at the time of installation. Includes 45mm white "
                            "cellular uPVC cloaking profile to the internal perimeter "
                            "where required.",
        "size": "%.0f lin m" % perimeter_m, "qty": 1, "unit": "item",
        "unitRateOverride": round(mastic, 2)})
    rows.append({
        "code": "", "desc": "Solar control glazing - upgrade of the outer pane to 4mm "
                            "bronze anti-sun toughened throughout, per specification cl.8.",
        "size": "%.2f m2" % area_total, "qty": 1, "unit": "item",
        "unitRateOverride": round(solar, 2)})
    rows.append({
        "code": "", "desc": "Strip out of the existing metal framed windows and doorsets, "
                            "set down on site for disposal by the main contractor.",
        "size": "43 openings", "qty": 1, "unit": "item",
        "unitRateOverride": round(STRIP_OUT_ALLOWANCE, 2)})

    net_subtotal = frames + adders + installation + solar + mastic + STRIP_OUT_ALLOWANCE
    mcd_uplift = net_subtotal * (1 / (1 - MCD) - 1)
    gross = net_subtotal + mcd_uplift

    # Joedan state their rates gross of 2.5% MCD and say so in one line rather
    # than showing the discount as an item. Mirror that: spread the uplift
    # across the rates pro rata, into the template's "Additional" column so the
    # "Frames" column keeps carrying the true frame cost. Installation is a
    # template formula and cannot be grossed, so its share rides on the rows.
    line_net = [r.get("frames", 0) + engine.CODE_VALUE.get(r["code"], 0) * engine.ADDER_FACTOR
                if r.get("code") else r["unitRateOverride"] for r in rows]
    base = sum(line_net)
    shares = [round(mcd_uplift * v / base, 2) for v in line_net]
    shares[-1] = round(shares[-1] + (mcd_uplift - sum(shares)), 2)  # penny residual
    for r, share in zip(rows, shares):
        if r.get("code"):
            r["additional"] = share
        else:
            r["unitRateOverride"] = round(r["unitRateOverride"] + share, 2)

    return {
        "rows": rows, "lines": lines, "area_m2": area_total,
        "perimeter_m": perimeter_m, "frames": frames, "adders": adders,
        "installation": installation, "solar": solar, "mastic": mastic,
        "strip_out": STRIP_OUT_ALLOWANCE, "net_subtotal": net_subtotal,
        "mcd_uplift": mcd_uplift, "gross": gross,
    }


# ------------------------------------------------------------------ Joedan's
# Verbatim from Appendix 2, pack pages 151-153. Adam: "include the same
# inclusions/exclusions as Joedan." The ONE deliberate divergence is the
# warranty - theirs is 12 months (their cl.21), ours is ten years.
INCLUSIONS = [
    "Polyester powder coated thermally broken aluminium windows, and commercial aluminium doorsets, in a single standard RAL colour internally and externally.",
    "Argon filled double glazed units with warm edge spacer bar - inner pane 4mm clear soft coat low-E toughened, outer pane 4mm bronze anti-sun toughened. Area weighted U-value no worse than 1.4 W/m2K to windows and 3.0 W/m2K to doors.",
    "Infill panels of solid phenolic core faced both sides with 2.0mm polyester powder coated aluminium sheets in a matt finish, U-value no greater than 1.2 W/m2K.",
    "All casements with shootbolt espagnolette locking, Securistyle Defender friction hinges, and separate concealed devices restricting initial opening to 100mm where required.",
    "Trickle ventilators throughout.",
    "Ironmongery to doors 32, 34, 37 and 41 - anti-finger-trap stiles, 100mm bottom rail, 100mm midrail, Axim 8800 concealed non hold-open transom closer, PR7100 exit panic device without external access, flush bolts to head and threshold of the slave leaf of all double doors, mill finished low threshold.",
    "Removal of the existing windows and doorsets, set down on site for disposal by the main contractor.",
    "45mm white cellular uPVC cloaking profile to the internal perimeter where required.",
    "Sealing of all joints between framing and structure with external grade low modulus silicone, applied at the time of installation.",
    "One site visit to undertake the manufacturing survey and take dimensions.",
    "FENSTER'S TEN YEAR GUARANTEE on frames and glazed units - in place of the twelve month product warranty customary in this sector.",
]

EXCLUSIONS = [
    "Building Control.",
    "Secure containers for our sole use (2no 20ft x 8ft).",
    "Skips adjacent to our works, and disposal of the removed windows and doors.",
    "Heras fencing and safety balustrades.",
    "Access scaffolding internally and externally, with board heights to suit the window installation and suitable access to distribute windows and remove debris from all board levels.",
    "Any access equipment including towers, podiums, scaffolding and MEWPs.",
    "Lifting and hoisting equipment suitable for window frames and glass.",
    "Removal and disposal of asbestos in the areas proposed for window replacement.",
    "Removal or reinstatement of utility based items, for example pipes, sinks, radiators.",
    "Protection of frames once fitted.",
    "Welfare facilities.",
    "Internal window boards.",
    "Manifestations and fire signage.",
    "Moving electrics or fire alarms.",
    "Removal and reinstatement of extractor fans.",
    "Disconnecting existing keypads and reconnecting them to the new doors.",
    "Removal and reinstatement of internal blinds.",
    "Magnetic locks.",
    "Door stops and door restraints.",
    "Suspended ceiling removal, reinstatement or repairs.",
    "Up-stand where windows sit onto the flat roof.",
    "Making good internally beyond the cloaking profile described above, including localised decoration to reveals - builder's work.",
    "Fire rated items and fire rated glass - no item is fire rated.",
    "Works outside normal working hours. One continuous visit is allowed; free and clear access to create a full day's work for our operatives is required, and return visits are chargeable.",
    "The defined provisional sum of GBP 5,000 for preparing out-of-square openings (specification cl.10) - all items are priced on the basis of square and plumb apertures.",
]

SUMMARY = [
    "Fenster Glazing are pleased to submit our proposal for the replacement of the external windows and doors at Redditch Library, 15 Market Place, Redditch, B98 8AR, forming part of the wider refurbishment of the building for Worcestershire County Council.",
    "Our offer covers 43 window and doorset positions across 41 references and 136.53 m2, taken from the pricing schedule issued with tender BLBS0956 and verified line by line against elevational drawings BLBS0956-GLE-RL-XX-DR-B-02 and -B-03. Every reference on the schedule appears on the elevations; none is missing and none is duplicated.",
    "The library remains open and in public use throughout. Our sequence is planned opening by opening so that no reveal is left unglazed at the end of a shift, with the Market Place frontage managed to keep the public highway clear and safe at all times.",
    "Fenster provide a TEN YEAR GUARANTEE on frames and glazed units. This is materially longer than the twelve month product warranty customary in this sector, and it is offered here without qualification or additional cost.",
]


def main():
    b = build()
    print("REDDITCH LIBRARY - PRICING DOCUMENT BUILD-UP")
    print("=" * 76)
    print("  frame supply (second supplier basis)   %12s" % "{:,.2f}".format(b["frames"]))
    print("  house code adders                      %12s" % "{:,.2f}".format(b["adders"]))
    print("  installation (fit only)                %12s" % "{:,.2f}".format(b["installation"]))
    print("  solar control glazing                  %12s" % "{:,.2f}".format(b["solar"]))
    print("  perimeter sealing (%.0f lin m x %.2f)  %12s"
          % (b["perimeter_m"], MASTIC_RATE, "{:,.2f}".format(b["mastic"])))
    print("  strip out ALLOWANCE (not a rate)       %12s   = GBP %.2f per opening"
          % ("{:,.2f}".format(b["strip_out"]), b["strip_out"] / 43))
    print("  " + "-" * 40)
    print("  net                                    %12s" % "{:,.2f}".format(b["net_subtotal"]))
    print("  add 2.5%% Main Contractor's Discount    %12s" % "{:,.2f}".format(b["mcd_uplift"]))
    print("  ======================================= %12s  TENDER SUM (gross of MCD)"
          % "{:,.2f}".format(b["gross"]))
    print()
    print("  Joedan JCQ.9727 gross %s / net %s"
          % ("{:,.2f}".format(JOEDAN_GROSS), "{:,.2f}".format(JOEDAN_NET)))
    print("  We undercut their gross by GBP %s (%.2f%%) and their net by GBP %s"
          % ("{:,.2f}".format(JOEDAN_GROSS - b["gross"]),
             (JOEDAN_GROSS - b["gross"]) / JOEDAN_GROSS * 100,
             "{:,.2f}".format(JOEDAN_NET - b["net_subtotal"])))
    print()
    print("  margin: adders GBP %s = %.1f%% of the net sum"
          % ("{:,.2f}".format(b["adders"]), b["adders"] / b["net_subtotal"] * 100))

    jobdef = {
        "client": "Pride Developments Group Ltd",
        "fao": "Leonard White, Senior Quantity Surveyor",
        "projectRef": "Redditch Library - BLBS0956 - window and door replacement",
        "siteAddress": "Redditch Library, 15 Market Place, Redditch, B98 8AR",
        "date": "28/07/2026",
        "supplier": "TBC - see clarifications",
        "note": "All figures exclusive of VAT and gross of 2.5% Main Contractor's Discount.",
        "rows": b["rows"],
        "mastic": 0, "epdm": 0,
        "summary": SUMMARY,
        "priceText": "GBP %s, gross of 2.5%% Main Contractor's Discount,"
                     % "{:,.2f}".format(b["gross"]),
        "inclusions": INCLUSIONS,
        "exclusions": EXCLUSIONS,
        "previousProjects": "Recent comparable projects include Headrow Court, Leeds (replacement "
                            "aluminium windows, doors and curtain wall, GBP630,000, occupied "
                            "city-centre phased installation) and Crestwood Park Primary School, "
                            "Dudley (replacement aluminium windows to an occupied school).",
        "outXlsx": os.path.join(OUT, "Redditch Library - Fenster Pricing Document.xlsx"),
        "outHtml": os.path.join(OUT, "Redditch Library - Fenster Proposal.html"),
    }
    defpath = os.path.join(OUT, "redditch-docdef.json")
    with open(defpath, "w", encoding="utf-8") as fh:
        json.dump(jobdef, fh, indent=1)

    b_out = {k: (round(v, 2) if isinstance(v, float) else v)
             for k, v in b.items() if k not in ("rows", "lines")}
    b_out["joedan_gross"] = JOEDAN_GROSS
    b_out["joedan_net"] = round(JOEDAN_NET, 2)
    b_out["undercut_on_gross"] = round(JOEDAN_GROSS - b["gross"], 2)
    b_out["undercut_on_net"] = round(JOEDAN_NET - b["net_subtotal"], 2)
    b_out["basis"] = ("%s, converted to a second-supplier position by the measured "
                      "factors BSW +5.7%% / Aplus -1.6%%. BENCHMARK - no supplier "
                      "quotation is held for this job." % ST_REF)
    with open(os.path.join(OUT, "redditch-pricing.json"), "w", encoding="utf-8") as fh:
        json.dump(b_out, fh, indent=1)

    subprocess.check_call([sys.executable,
                           os.path.join(REPO, "scripts", "generate-fenster-docs.py"),
                           defpath])
    print("\nwritten %s" % jobdef["outXlsx"])
    client = make_client_copy(jobdef["outXlsx"], b)
    print("written %s" % client)
    return b


def make_client_copy(src, b):
    """A print area protects a print, not the file.

    mary_checks caught this and it is right: the working file carries the
    product code in column B and our frame buy in J to O, all outside the
    printed range but all sitting in the .xlsx that would be attached to an
    email. Anybody who scrolls right reads our cost and our margin. So the
    client gets a separate file with the formulas resolved to values and the
    working columns emptied - the same shape as the RFQ/take-off separation
    that keeps Joedan's price and our buy out of anything Pride can open.
    """
    import openpyxl
    wb = openpyxl.load_workbook(src)
    ws = wb["Pricing Document "]

    r, last = 9, 9
    while ws["C%d" % r].value:
        code = ws["B%d" % r].value
        h = ws["H%d" % r].value
        if not isinstance(h, (int, float)):
            h = ((ws["J%d" % r].value or 0) + (ws["K%d" % r].value or 0)
                 + (ws["L%d" % r].value or 0)
                 + engine.CODE_VALUE.get(code, 0) * engine.ADDER_FACTOR)
        h = round(h, 2)
        ws["H%d" % r] = h
        ws["I%d" % r] = round(h * (ws["F%d" % r].value or 1), 2)
        for col in "BJKLMNO":
            ws["%s%d" % (col, r)] = None
        last, r = r, r + 1

    ws["I%d" % (last + 1)] = round(b["installation"], 2)      # INSTALLATION
    ws["I%d" % (last + 3)] = round(b["gross"], 2)             # TOTAL

    # Then sweep everything else outside the printed range - the "PRODUCT
    # CODES" caption, the "Supplier used:" pair, the stray spacer cells the
    # template carries in K and M. Naming our supplier to a main contractor
    # who is about to ask us for a better price is not a small leak.
    from openpyxl.cell.cell import MergedCell
    keep_cols = set("CDEFGHI")
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value in (None, ""):
                continue
            if cell.column_letter in keep_cols and cell.row <= 66:
                continue
            cell.value = None

    ws["C8"] = None            # the template's "SMA Smart Wall" code legend

    # And kill the OPTIONAL block. The template offers external mastic and EPDM
    # as extras; on this job perimeter sealing is REQUIRED by pack p70 cl.11,
    # it is inside the tender sum, and Joedan include it too. Leaving an
    # "OPTIONAL - EXTERNAL MASTIC" caption on the page is the Princess Beatrice
    # REQ-6 error exactly: it invites the QS to strike work we are obliged to do.
    for c in ("D60", "F61", "I61", "F62", "I62"):
        ws[c] = None
    out = src.replace(".xlsx", " (CLIENT COPY).xlsx")
    wb.save(out)

    # Strip the house template's inherited identity - Dan Parker of AG
    # Surveying in docProps, and two live external links into other people's
    # Outlook caches. REQ-27. Neither has been issued, so --in-place is right.
    subprocess.check_call([sys.executable,
                           os.path.join(REPO, "scripts", "mary_scrub_workbook.py"),
                           out, "--in-place"])
    subprocess.check_call([sys.executable,
                           os.path.join(REPO, "scripts", "mary_scrub_workbook.py"),
                           src, "--in-place"])
    return out


if __name__ == "__main__":
    main()
