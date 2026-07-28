"""Rebrand and de-trace an already-built client pack.

Written for Georgie's on 28/07/2026, when Adam spotted that the proposal issued to Pearce
Construction carried RRR Group's name AND logo on its cover - and the pricing workbook was
still carrying dan.parker@agsurveying.co.uk plus external links to two other people's
machines (REQ-27).

Nothing here touches a price. It edits three classes of thing only:

  1. TEXT      - literal client-name strings, in document.xml / sheet cells.
  2. IMAGES    - a logo that belongs to somebody else, replaced with a transparent PNG of
                 identical dimensions so the page layout does not move.
  3. METADATA  - dc:creator, and for workbooks the externalLinks parts, their relationships
                 and the <externalReferences> element that binds them.

Use it on a COPY. Where a file has already gone to a client, the issued file is the record of
what they received and must not be rewritten in place.

    python scripts\\clean_issued_pack.py --selftest
"""
from __future__ import annotations

import argparse
import io
import re
import shutil
import zipfile
from pathlib import Path

CREATOR_RE = re.compile(r"<dc:creator>.*?</dc:creator>", re.S)
LASTMOD_RE = re.compile(r"<cp:lastModifiedBy>.*?</cp:lastModifiedBy>", re.S)
EXTREFS_RE = re.compile(r"<externalReferences>.*?</externalReferences>", re.S)


def _rewrite_zip(src: Path, dst: Path, edit):
    """Copy every part of src to dst, passing each through edit(name, bytes).

    edit returns replacement bytes, or None to DROP the part entirely.
    """
    zin = zipfile.ZipFile(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            new = edit(item.filename, data)
            if new is None:
                continue
            zout.writestr(item, new)
    zin.close()


def _transparent_png(width: int, height: int) -> bytes:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGBA", (width, height), (0, 0, 0, 0)).save(buf, "PNG")
    return buf.getvalue()


def _set_creator(xml: bytes, creator: str) -> bytes:
    text = xml.decode("utf8")
    text = CREATOR_RE.sub("<dc:creator>%s</dc:creator>" % creator, text)
    text = LASTMOD_RE.sub("<cp:lastModifiedBy>%s</cp:lastModifiedBy>" % creator, text)
    return text.encode("utf8")


def clean_docx(src: Path, dst: Path, replacements: dict, drop_images: list, creator: str):
    """replacements: {'RRR GROUP': 'PEARCE...'}. drop_images: ['word/media/image4.png']."""
    zin = zipfile.ZipFile(src)
    sizes = {}
    for name in drop_images:
        from PIL import Image

        sizes[name] = Image.open(io.BytesIO(zin.read(name))).size
    zin.close()

    def edit(name: str, data: bytes):
        if name in sizes:
            return _transparent_png(*sizes[name])
        if name == "docProps/core.xml":
            return _set_creator(data, creator)
        if name.endswith(".xml") and (name.startswith("word/") or name == "docProps/app.xml"):
            text = data.decode("utf8")
            for old, new in replacements.items():
                text = text.replace(old, new)
            return text.encode("utf8")
        return data

    _rewrite_zip(src, dst, edit)


def clean_xlsx(src: Path, dst: Path, replacements: dict, creator: str, strip_external: bool = True):
    zin = zipfile.ZipFile(src)
    shared = zin.read("xl/sharedStrings.xml").decode("utf8") if "xl/sharedStrings.xml" in zin.namelist() else ""
    zin.close()

    def edit(name: str, data: bytes):
        if strip_external and name.startswith("xl/externalLinks/"):
            return None
        if name == "docProps/core.xml":
            return _set_creator(data, creator)
        if name == "xl/workbook.xml" and strip_external:
            return EXTREFS_RE.sub("", data.decode("utf8")).encode("utf8")
        if name == "xl/_rels/workbook.xml.rels" and strip_external:
            text = data.decode("utf8")
            text = re.sub(r'<Relationship[^>]*Type="[^"]*externalLink"[^>]*/>', "", text)
            return text.encode("utf8")
        if name in ("xl/sharedStrings.xml",) or (name.startswith("xl/worksheets/") and name.endswith(".xml")):
            text = data.decode("utf8")
            for old, new in replacements.items():
                text = text.replace(old, new)
            return text.encode("utf8")
        return data

    _rewrite_zip(src, dst, edit)
    _ = shared


TRACE_PATTERNS = [
    "agsurveying", "dan.parker", "Dan Parker", "Nicholas Baker",
    "LiamO'Donnell", "Users\\Parke", "Content.Outlook", "INetCache",
]


def audit(path: Path) -> list:
    """Return every third-party trace still readable inside path."""
    hits = []
    z = zipfile.ZipFile(path)
    for name in z.namelist():
        if name.endswith((".png", ".jpg", ".jpeg", ".emf", ".bin")):
            continue
        try:
            text = z.read(name).decode("utf8", "ignore")
        except Exception:
            continue
        for pat in TRACE_PATTERNS:
            if pat in text:
                hits.append((name, pat))
    z.close()
    return hits


def selftest():
    """Replays Georgie's: the issued pack must be dirty, the cleaned pack must be clean."""
    base = Path(__file__).resolve().parent.parent
    issued = base / "test-results/mary-inbox/processed/20260728T1301-zQFQAAAA-att"
    out = base / "outputs/georgies-reissue"
    src = issued / "Pearce Construction - Georgie's (formerly Rosebank) Pricing.xlsx"
    if not src.exists():
        print("SKIP - issued workbook not on disk at %s" % src)
        return
    before = audit(src)
    assert before, "expected the ISSUED workbook to carry third-party traces"
    tmp = out / "_selftest.xlsx"
    clean_xlsx(src, tmp, {"RRR Group": "Pearce Construction (Barnstaple) Ltd"}, "Fenster Glazing")
    after = audit(tmp)
    assert not after, "cleaned workbook still carries: %r" % after
    import openpyxl

    ws = openpyxl.load_workbook(tmp, data_only=True).active
    assert abs(ws["H25"].value - 89229.6125) < 0.001, "TOTAL moved: %r" % ws["H25"].value
    assert "RRR" not in str(ws["B3"].value), ws["B3"].value
    tmp.unlink()
    print("selftest OK - %d trace(s) before, 0 after, total unchanged at GBP 89,229.61" % len(before))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--audit", help="path to a file to audit for third-party traces")
    args = ap.parse_args()
    if args.selftest:
        selftest()
    elif args.audit:
        for name, pat in audit(Path(args.audit)) or [("", "")]:
            print("%-50s %s" % (name, pat) if name else "clean")
    else:
        ap.print_help()
