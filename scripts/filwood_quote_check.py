"""Filwood Broadway (Stepnell) - quote check.

Audits Gintare's outgoing tender (pricing xlsx + proposal docx, 27/07/2026) against
the Stepnell trade bill, drawing 31551 P02, the External Materials Schedule and
supplier quote Bellview/BSW 0000000507 (24/07/2026).

Output: outputs/Filwood Broadway - Quote Check (BSW 0000000507 vs Tender).xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = r"outputs\Filwood Broadway - Quote Check (BSW 0000000507 vs Tender).xlsx"

HDR = PatternFill("solid", fgColor="1F2A44")
HDRF = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
TITLE = Font(bold=True, size=13, color="1F2A44")
RED = PatternFill("solid", fgColor="FCE4E4")
AMB = PatternFill("solid", fgColor="FFF3D6")
GRN = PatternFill("solid", fgColor="E4F3E6")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# ---------------------------------------------------------------- source data
# Bellview/BSW 0000000507, 24/07/2026. Net 54,197.17 less 15% = 46,067.59 net.
BSW = [
    # pos, w, h, gross, glass make-up, U, panels, door hardware note
    ("001", 4850, 3570, 7579.01, "6.8 Lami / 4mm Tuff", 1.0, 11, "hook locks, closer"),
    ("002", 4800, 3570, 7534.38, "6.8 Lami / 4mm Tuff", 1.0, 11, "hook locks, closer"),
    ("003", 4800, 3570, 7536.06, "6.8 Lami / 4mm Tuff", 1.0, 11, "hook locks, closer"),
    ("004", 4850, 3570, 7589.69, "6.8 Lami / 4mm Tuff", 1.0, 11, "hook locks, closer"),
    ("005", 6250, 3100, 8849.83, "6.8 Lami / 4mm Tuff", 1.0, 12, "ELECTRIC strike + latch"),
    ("006", 5550, 2970, 7554.10, "8.8 Lami / 6mm Tuff (SG)", 1.1, 7, "hook locks, closer"),
    ("007", 5550, 2970, 7554.10, "8.8 Lami / 6mm Tuff (SG)", 1.1, 7, "hook locks, closer"),
]
DISCOUNT = 0.85

# Gintare's pricing rows, in her order, mapped to the BSW position she used.
QUOTE = [
    # ref, stated size, bsw pos, code, additional, adder
    ("ED-04", "4930 x 3570", "001", "DAD", 1000.0, 1500.0),
    ("ED-04", "4930 x 3570", "002", "DAD", 1000.0, 1500.0),
    ("ED-04", "4930 x 3570", "003", "DAD", 1000.0, 1500.0),
    ("ED-04", "4930 x 3570", "004", "DAD", 1000.0, 1500.0),
    ("ED-05", "5550 x 2970", "006", "DAD", 1000.0, 1500.0),
    ("ED-05", "5550 x 2970", "007", "DAD", 1000.0, 1500.0),
    ("ED-06", "6315 x 3150", "005", "DAD", 1000.0, 1500.0),
]
INSTALL_AS_QUOTED = 3500.0        # DAD labour code 500 x 7
TOTAL_AS_QUOTED = 67067.5775
MASTIC_OPT = 605.05
EPDM_OPT = 3081.49125

# Trade bill nominal structural openings (Stepnell bill, pages 306/0/4).
BILL = {"ED-04": (4930, 3570, 4), "ED-05": (5550, 2970, 2), "ED-06": (6315, 3105, 1)}

# Drawing 31551 P02 performance schedule.
SCHED = {
    "ED-04": ("N/A", "N/A", "1.0", "0.5-0.6", "ERs", 4),
    "ED-05": ("N/A", "N/A", "1.0", "0.5-0.6", "ERs", 2),
    "ED-06": (">= Rw 32 dB", "N/A", "1.0 (retail unit)", "0.5-0.6", "LPS 1175 SR2", 1),
}

CW_LABOUR = 150.0   # GBP/m2, house rate (MARY-HANDOVER s6)
BENCH_17_07 = 84810.59

bill_sqm = sum(w * h / 1e6 * n for w, h, n in BILL.values())
install_correct = bill_sqm * CW_LABOUR
lines_total = TOTAL_AS_QUOTED - INSTALL_AS_QUOTED
corrected_total = lines_total + install_correct
bsw_net = round(sum(g * DISCOUNT for *_, g, _, _, _, _ in [(0, 0, 0, b[3], 0, 0, 0, 0) for b in BSW]), 2)
bsw_net = round(sum(b[3] * DISCOUNT for b in BSW), 2)

wb = openpyxl.Workbook()

# =============================================================== 1. FINDINGS
ws = wb.active
ws.title = "Findings"
ws["B2"] = "BCC 4-16 FILWOOD BROADWAY (STEPNELL) - QUOTE CHECK"
ws["B2"].font = TITLE
ws["B3"] = ("Checked 27/07/2026 by Mary. Outgoing tender: 'Stepnell - BCC Filwood Broadway Pricing.xlsx' "
            "(GBP 67,067.58 ex VAT) + Proposal.docx, both 27/07/2026. Checked against Stepnell trade bill "
            "L_SC Shop Front Systems (bid S25233B), ITT STP10 02/07/2026, drawing 2411-RCK-ZZ-ZZ-DR-A-31551 "
            "Rev.P02, External Materials Schedule 00700 Rev.P00 s.N p41, and supplier quote "
            "Bellview/BSW 0000000507 dated 24/07/2026. Submission deadline THU 30/07/2026. "
            "UPDATED 28/07 for a SECOND supplier quote, Aplus QT51510 (Technal STII, 27/07/2026, "
            "GBP 34,445.91) - see the 'Aplus QT51510' sheet - and for a run of the house rule engine, "
            "`python scripts/mary_checks.py data/job-checks/filwood.json`, which returns 9 FAILED.")
ws["B3"].font = BASE
ws["B3"].alignment = WRAP
ws.merge_cells("B3:H3")
ws.row_dimensions[3].height = 58

cols = [("B", 6), ("C", 30), ("D", 62), ("E", 30), ("F", 16), ("G", 34)]
for c, w in cols:
    ws.column_dimensions[c].width = w

head = ["#", "Finding", "Detail and where it comes from", "Effect", "Money", "What has to happen"]
for i, h in enumerate(head):
    cell = ws.cell(row=5, column=2 + i, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR, HDRF, WRAP, THIN

FINDINGS = [
    (1, "Installation GBP 3,500 for 123 m2 of 3.5 m tall shopfront",
     "I17 is the template INSTALLATION formula reading the product code: DAD = GBP 500 x 7 elements. "
     "That code is for a double aluminium door on its own, not a 17.6 m2 glazed screen 3,570 mm tall. "
     "House rate for screens of this type is CW labour GBP 150/m2 (MARY-HANDOVER s6) = "
     "122.98 m2 x 150 = GBP 18,446.32. Our own 17/07 provisional-sum build used the same GBP 150/m2. "
     "Made worse by the proposal excluding 'Access/Lifting Equipment - Scaffold, MEWPS, Towers, Forklift' "
     "while including installation of elements 3.57 m tall - no access plant is priced either.",
     "GBP 28.46/m2 to install seven glazed screens with doors. Not recoverable once the price is accepted.",
     "+14,946.32", "Reprice install at CW labour, or a stated judgement figure Adam signs off. "
     "Decide whether access plant stays excluded."),
    (2, "LPS 1175 SR2 security doorset on ED-06 was asked for in writing and never answered",
     "Drawing 31551 P02 schedule, ED-06: Security = LPS 1175 SR2. Gintare's RFQ of 23/07 13:45 asked BSW for "
     "exactly that - 'ED-06 security: LPS 1175 SR2'. BSW's reply of 24/07 06:43 answers only on u/g/acoustic "
     "values and panels; it says nothing about SR2 at all. Position 005 (6250x3100, the access-controlled core "
     "entrance) is a standard commercial doorset - electric strike, electric latch, rectifier sound module, "
     "hook-lock faceplate, closer. Nothing on the quote references LPS 1175, SR2 or any LPCB certification. "
     "SR2 is a tested and certified doorset, a different product.",
     "Largest technical gap in the package, and silence in answer to a direct written request is not "
     "compliance. An SR2 doorset cannot be substituted after award.",
     "TBC", "BSW to confirm whether SMA Shopline has an LPS 1175 SR2 tested doorset and price it, or the "
     "element goes to a specialist. Get the Part Q strategy drawings first (see RFIs)."),
    (3, "The g 0.5-0.6 claim is unevidenced and the quoted make-up cannot achieve it",
     "Drawing 31551 P02 requires G-Value 0.5-0.6 on all three types and Gintare's RFQ asked for it. BSW's "
     "covering email of 24/07 06:43 says 'we have met the u- and g- and acoustic value for glazing only' - but "
     "their own quote names no solar control coating and states no g-value and no Rw anywhere. The make-ups are "
     "'6.8 Lami / 4mm Tuff' (001-005) and '8.8 Lami / 6mm Tuff (SG)' (006-007); a clear laminated/toughened DGU "
     "sits at g ~0.7-0.75, so on the face of it the claim and the make-up contradict each other. "
     "Cross-check: BSW net works out at GBP 374.60/m2, while our 17/07 benchmark for the same screens was "
     "GBP 359.60/m2 register median PLUS GBP 45/m2 spec uplift = GBP 404.60/m2. The gap is roughly the "
     "coating that would be needed.",
     "We would be passing an unevidenced supplier claim straight through to Stepnell, and the proposal recites "
     "the make-up back to the client as our Glazing Specification with no coating and no g-value.",
     "TBC (~GBP 45/m2 benchmark = ~GBP 5.5k cost)",
     "BSW to state the coating and the actual g-value in writing (house spec elsewhere is Coolite SKN 176ii), "
     "or requote. Do not repeat their claim to Stepnell until they evidence it."),
    (4, "ALUPROF is specified; we are offering SMA Shopline with no non-compliance statement",
     "The trade bill item header names 'Aluprof, Unit 5 Altrincham Business Park ... www.aluprof.com' and the "
     "External Materials Schedule s.N p41 gives Manufacturer = Aluprof. Drawing 31551's own revision history "
     "reads 'Issued to Aluprof 2025.10.22' and 'Updated Issue to Aluprof 2025.10.29' - the architect has been "
     "working with Aluprof directly. BSW quoted System: SMA Shopline Double, and the proposal states SMA "
     "Shopline throughout. The ITT is explicit: 'Your Value Engineering Proposals are encouraged, but must be "
     "fully detailed indicating any areas of non-compliance and be accompanied by a compliant bid.' We have "
     "neither. (In fairness, p41 does say 'Aluprof (or similar)' for the frame.)",
     "Bid can be set aside as non-compliant without anyone reading the number. Same failure mode as the "
     "Technal/Modeal issue on Princess Beatrice and the Senior issue on Vesuvius.",
     "nil", "Either get an Aluprof-fabricator price as the compliant bid with SMA Shopline offered alongside "
     "as VE, or state the substitution as a formal qualification on the front of the proposal."),
    (5, "Ventilation zone priced as solid flat aluminium panel",
     "The top 660 mm (ED-04) / 700 mm (ED-05) band of every screen is labelled 'Ventilation Zone' either side "
     "of 'Signage Zone' on drawing 31551, tagged Q2 = Cadisch expanded aluminium mesh, high free area, with "
     "four 'Mullion behind mesh' notes. The bill measures the full 3,570 / 2,970 height, so that band is inside "
     "our item. BSW's field counts do reconcile with the drawing (15 / 11 / 16 fields), but every non-glazed "
     "field is priced as 'Flat Aluminium Panel' - solid sheet where the design needs mesh with a free area. "
     "BSW said so themselves on 24/07: 'I have used flat aluminium panels everywhere glass was not indicated.' "
     "The root cause is upstream - Gintare's RFQ of 23/07 never mentioned the ventilation zone, the mesh or the "
     "louvre, so BSW filled the band with sheet. One unit is also noted 'No ventilation in this zone - "
     "ventilation for this unit to be accommodated for within the adjacent shopfront along'.",
     "Either the mesh and the integrated louvre element are ours and are unpriced, or they are excluded and "
     "the retail units have no ventilation provision - which the drawing says these drawings demonstrate.",
     "TBC", "Settle the boundary with Stepnell IN WRITING before submission. The proposal excludes signage and "
     "illumination but says nothing about the mesh, the louvre or the ventilation."),
    (6, "BSW has priced four elements 80-130 mm narrower than the sizes we are quoting",
     "Trade bill nominal S.O. vs what BSW actually priced: ED-04 4930 wide vs 4850/4800/4800/4850; "
     "ED-06 6315x3105 vs 6250x3100. ED-05 5550x2970 is the only match. Our pricing document states the bill "
     "sizes, i.e. sizes our supplier has not quoted. The drawing does say green dimensions are approximate "
     "structural openings to be confirmed with the subcontractor, so a deduction off S.O. is normal - but "
     "10-20 mm, not 80-130 mm. Separately ED-06's height is typed 3150 in our document where the bill says "
     "3105 (transposed digits).",
     "If Stepnell accept at 4930 we are short of frame on four screens. The 3150 typo also inflates the m2, "
     "EPDM and mastic lines.",
     "check", "BSW to confirm what they set out to and why. Correct ED-06 to 6315 x 3105 before issue."),
    (7, "The acoustic make-up is on the wrong screens",
     "Drawing 31551 P02: ED-06 requires >= Rw 32 dB; ED-04 and ED-05 are N/A. BSW put the heavier "
     "'8.8 Lami / 6mm Tuff (SG)' make-up on positions 006 and 007 - which are the two ED-05s that need no "
     "acoustic rating - and gave ED-06 (position 005) the lighter 6.8 Lami / 4 Tuff. No Rw figure is stated "
     "for any element.",
     "Looks transposed. ED-06 is the one element with an acoustic duty and it has the lighter glass; the "
     "acoustic-looking glass is paid for on two screens that do not need it.",
     "check", "BSW to confirm Rw for the ED-06 element as a whole (glass, frame, door and threshold seals), "
     "not glass alone, and to confirm the make-ups have not been swapped."),
    (8, "Two elements quoted at U = 1.1 against a target of 1.0",
     "BSW positions 006 and 007 (both ED-05) state U = 1.1 W/m2K. The drawing schedule target is 1.0 for "
     "ED-04, ED-05 and ED-06. The proposal Executive Summary tells the client 'Ug values noted between "
     "1.0-1.1 W/m2K where quoted' with no deviation or VE note. Note also Ug is centre-pane glass; the "
     "schedule's Target U-Value would normally be the whole element, so even the 1.0 lines are unproven.",
     "We are putting a written non-compliance in front of the client without flagging it as one.",
     "check", "BSW to confirm whether 1.0 is achievable on ED-05 and to state Uw for the elements, or qualify "
     "the deviation properly."),
    (9, "Manifestation is a requirement, is mentioned, and is neither priced nor excluded",
     "Drawings 31550 and 31551 both carry: 'Provision of manifestations to all clear glazing at two continuous "
     "bands 850-1000 mm and 1400-1600 mm above FFL, in line with Approved Document M. Bands to maintain clear "
     "visual contrast on both sides and remain visible with doors open or closed.' The proposal Executive "
     "Summary lists manifestation as part of the requirement. There is no manifestation line in the pricing, "
     "no inclusion and no exclusion. Our 17/07 build carried one.",
     "Charged to us after award. The proposal's 'signage, branding' exclusion does not cover an ADM safety "
     "requirement.",
     "TBC", "Price the two bands on all clear glazing, or exclude manifestation explicitly."),
    (10, "Single doors quoted, coded as double, and the fire strategy that decides it is not in the pack",
     "Drawing 31551 carries, twice, in orange: 'NOTE: >60 occupants -> double doors/outward-opening exits "
     "required (see OFR).' The OFR is not in the tender pack. BSW quoted 'a Single Pivoted Anti Fingertrap "
     "Door' on all seven elements, with a closer, and the quote does not state the opening direction. "
     "Separately, all seven rows of our pricing carry product code DAD (double aluminium door) against a "
     "supplier-quoted SINGLE door - DAD puts GBP 1,500 of adder and GBP 500 of install on each row where SAD "
     "would put GBP 900 and GBP 250.",
     "If any retail unit exceeds 60 occupants the door package is the wrong product. And the coding is "
     "internally inconsistent with both the supplier quote and our own proposal text.",
     "4,200 either way", "Get the OFR and confirm occupancies. Then decide the DAD coding deliberately - it is "
     "either GBP 4,200 of discretionary money Adam is happy with, or a mis-code."),
    (11, "Mill finish specified for frame and spandrels; we have priced RAL 7035 throughout",
     "External Materials Schedule s.N p41: frame = 'extruded aluminium mullions and transoms in mill finish'; "
     "spandrels = 'flat, flush aluminium spandrel panels ... finished in mill-finish aluminium'; and ONLY "
     "'C. Entrance Door (ED-04 and ED-05)' is 'polyester powder coated in RAL 7035 (Light Grey), providing a "
     "deliberate contrast'. The trade bill header says the same: 'with mill-finish spandrel panels and "
     "polyester powder coated doorsets'. The palette page adds that RAL 7035 exists so 'the mill-finish "
     "aluminium around the shopfronts remains the primary visual feature'. Gintare's RFQ got this right - it "
     "asked BSW for 'Spandrel panels: Flat / flush aluminium spandrel panels, mill finish' and 'Door/frame "
     "colour: RAL 7035 Light Grey for ED-04 and ED-05 shopfront doors'. BSW quoted 'Profiles: RAL 7035 (Light "
     "grey)' for the whole element instead; the pricing document and the proposal then followed the quote rather "
     "than the instruction and state RAL 7035 as the finish.",
     "Wrong appearance, and BCC's High Street team will pick it up. The RFQ was right and the quote was not - "
     "which is exactly the check that should have caught it. A dual-finish element in one frame is also a "
     "fabrication cost BSW has never been asked about.",
     "check", "BSW to requote frame and spandrels mill finish with PPC RAL 7035 doorsets only."),
    (12, "The bill says provisional sums; we are submitting a firm lump sum and have not answered bill item A",
     "Bill page 306/0/5 item A, in full: drawing 31551 states it is for illustrative purposes only and says the "
     "shopfront proposals should be allowed for as Provisional Sums; Stepnell could find no Client provisional "
     "sums and so assume Contractors'; 'see Work Section A54 at the end of this Bill for your shopfront "
     "Provisional Sum inclusions - DO NOT include the shopfronts twice within your tender submission'. A54 was "
     "never issued to us. Item A is left unpriced in our submission and neither the pricing nor the proposal "
     "mentions A54 or provisional sums. Our own 17/07 submission was correctly a Contractor's Provisional Sum.",
     "Two ways to lose: Stepnell's build-up may expect the money in A54 and drop ours, or we are fixed-priced "
     "against a design the architect says is not finalised, protected only by one Executive Summary sentence.",
     "structural", "Request A54, then decide: firm price with a hard re-measure qualification, or a provisional "
     "sum as the bill instructs."),
    (13, "Document and admin defects to fix before it goes out",
     "(a) The proposal is addressed 'FAO: Trevor Copeman' - the ITT and the bill both say quotations go to "
     "Adam Warner, Senior Estimator, adam.warner@stepnell.co.uk, queries to him and sam.ignatov@stepnell.co.uk. "
     "Trevor Copeman appears nowhere in the enquiry. (b) Proposal dated 27/07/2026, pricing document dated "
     "28/07/2026. (c) The proposal still contains live Word LINK fields to "
     "'C:\\Users\\fenst\\Downloads\\Pricing Doc Template.xlsx' - a local path leaked into a client document "
     "that will prompt or error on their machine. (d) No columns are hidden in the workbook: J-P carry BSW's "
     "cost per screen and K3/L3/M3 read 'Supplier used: BSW 46067.59'. The print area (C1:I27) is clean, so a "
     "PDF is safe - sending the .xlsx is not. (e) O16 shows #VALUE!. (f) Row 14's working-column formulas were "
     "not filled down, so that row would silently compute zero if the code were switched to CW.",
     "One of these is a supplier-cost leak, one sends the quote to the wrong person.",
     "nil", "Issue PDF only, per the Ops Manual process. Fix the addressee, align the dates, break the links."),
    (14, "None of Stepnell's commercial terms are qualified",
     "ITT Project Particulars: D&B 2024; subcontract 'amended to suit the provisions of the main contract and "
     "STEPNELL LTD conditions'; LADs GBP 1,358.00 per calendar week; retention 3%; 12 month maintenance period; "
     "collateral warranty as tender documents; Professional Indemnity required where Contractors Design; "
     "payment on the last business day of the month following application. Our proposal's standard terms offer "
     "50% deposit before works with 50% on completion, and 30 day validity. There is a contractor's design "
     "element here - the proposal itself promises to coordinate final sizes and details with the BCC High "
     "Street team - which is what triggers the PI requirement.",
     "Straight conflict between our terms page and the subcontract we would be signing, with nothing qualified.",
     "nil", "Add a commercial qualification: payment terms, retention, LAD exposure and the PI position."),
    (15, "BSW's own written caveat - performance met FOR GLAZING ONLY, frames non-rebated - is not carried into "
     "the tender return",
     "BSW's covering email, 24/07/2026 06:43 from estimations@bsws.co.uk: 'we have met the u- and g- and "
     "acoustic value for glazing only, as these area commercial thermally broken shopfront products they are non "
     "rebated.' The drawing schedule's Target U-Value 1.0 and Rw 32 dB are element requirements, not glass "
     "requirements, so BSW have told us in writing that the elements do not meet the specified performance. Our "
     "proposal restates this only as 'Ug values noted between 1.0-1.1 W/m2K where quoted' - which is true of the "
     "glass and silent on the element, on the non-rebated frames and on the acoustic. Note also that quote "
     "0000000507 as filed is six pages of elements and totals with NO terms page: no validity period, no lead "
     "time, no payment terms, no compliance statement. The caveat exists only in the covering email.",
     "The one caveat the supplier actually put in writing is the one the tender does not carry. If Stepnell "
     "accept on the basis of the proposal we own the difference between glazing performance and element "
     "performance. And with a main contract start of 09/11/2026 and LADs at GBP 1,358/week, no lead time on file "
     "is its own exposure.",
     "risk", "Carry BSW's caveat verbatim as a qualification, or get them to confirm element-level compliance. "
     "Ask for lead time and validity in writing."),
    (16, "The pricing workbook carries three other companies' traces - and this one has not been issued yet",
     "`python scripts/mary_checks.py data/job-checks/filwood.json` scans the actual files. "
     "'Stepnell - BCC Filwood Broadway Pricing.xlsx' carries **dan.parker@agsurveying.co.uk** in "
     r"docProps/core.xml, and external links pointing at C:\Users\LiamO'Donnell and "
     r"C:\Users\Parke by way of an Outlook INetCache path, in "
     "xl/externalLinks/_rels/externalLink1.xml.rels and externalLink2.xml.rels. The proposal separately "
     r"carries C:\Users\fenst in word/document.xml. The same rule also reports 62 populated cells "
     "OUTSIDE the print area $C$1:$I$27 - a print area protects a print, not the file.",
     "Identical to the Georgie's defect that reached Pearce Construction on 28/07, and the fourth job this "
     "week. The difference here is that Filwood has NOT gone out, so it is still catchable. It is visible "
     "in file properties without opening the document.",
     "reputational", "Strip docProps and the external links IN PLACE - the file has not been issued - and "
     "issue a sell-only copy with the working columns REMOVED, not merely outside the printed range. "
     "Re-run mary_checks before anything is attached."),
    (17, "A second quote arrived on 28/07 and it is not a like-for-like - but it does settle three questions",
     "Aplus QT51510 (27/07/2026, Technal STII, 'Glazed /Supply Only (Delivered)') totals GBP 34,445.91 net, "
     "GBP 11,621.68 under BSW. It is NOT comparable: Aplus state 'Panels by others' and leave 46.09 m2 of "
     "infill - 37.5% of the elevation - as unfilled 32mm apertures, where BSW include 70 flat aluminium "
     "panels bundled into their element price. Break-even is GBP 252.15/m2 of panel. Neither quote yields a "
     "panel rate and the register has no panel category, so this cannot be settled from anything we hold. "
     "What the second quote DOES settle: (a) Aplus set out to the trade bill exactly - 4930 / 5550 / "
     "6315x3105, with ED-06 segmented 300/1200/700/600 straight off drawing 31551 - so BSW's 4850/4800/6250 "
     "are wrong and our own 3150 is a typo; (b) 'Glass quoted has a g value of 0.66' against a 0.5-0.6 "
     "requirement, quantifying finding 3; (c) 'Quoted in STII, these will only reach 1.8/1.9 U Value' and "
     "'STII doors have no formal acoustic test data'. Full detail on the 'Aplus QT51510' sheet.",
     "Two independent fabricators have now put the same non-compliance in writing. This stops being a "
     "supplier problem and becomes a specification one.",
     "not comparable", "Do not present GBP 34,445.91 as a saving. Get a panel price, then compare."),
    (18, "Neither system can meet the specification, and nobody has asked the one that might",
     "BSW on 24/07: performance met 'for glazing only ... non rebated'. Aplus on 27/07: 'will only reach "
     "1.8/1.9 U Value', 'no formal acoustic test data', 'g value of 0.66'. The specified system is Aluprof, "
     "and Fenster have approached no Aluprof fabricator at all - the architect developed the design with "
     "Aluprof directly, which is the likeliest reason the target is 1.0. A standard commercial shopfront "
     "system is not a thermally broken curtain-walling-grade system and does not get to 1.0.",
     "The Aluprof specification is probably not a formality to value-engineer around; it may be the only "
     "route to the stated performance. That reframes finding 4 from a paperwork problem into a pricing one.",
     "structural", "Ask an Aluprof-approved fabricator to price it, even late. If the answer is that nobody "
     "can hit 1.0 in a shopfront, that is an RFI to RCKa via Stepnell, not a silent substitution."),
]

r = 6
for num, title, detail, effect, money, action in FINDINGS:
    ws.cell(row=r, column=2, value=num).font = BOLD
    ws.cell(row=r, column=3, value=title).font = BOLD
    ws.cell(row=r, column=4, value=detail).font = BASE
    ws.cell(row=r, column=5, value=effect).font = BASE
    ws.cell(row=r, column=6, value=money).font = BOLD
    ws.cell(row=r, column=7, value=action).font = BASE
    fill = RED if (num <= 5 or num >= 16) else AMB
    for c in range(2, 8):
        cell = ws.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.border = THIN
        if c in (2, 6):
            cell.fill = fill
    ws.row_dimensions[r].height = max(56, min(190, len(detail) // 1.35))
    r += 1

r += 1
ws.cell(row=r, column=3, value="MONEY POSITION").font = TITLE
r += 1
for label, val, note in [
    ("As submitted (ex VAT)", TOTAL_AS_QUOTED, "GBP 67,067.58 - lines 63,567.58 + install 3,500"),
    ("Optional extras offered", MASTIC_OPT + EPDM_OPT, "mastic 605.05 + EPDM 3,081.49"),
    ("With installation corrected to CW labour", corrected_total,
     "lines unchanged + %.2f m2 x GBP150 = GBP %s" % (bill_sqm, format(install_correct, ",.2f"))),
    ("Mary's independent benchmark, 17/07/2026", BENCH_17_07,
     "built with no supplier quote - the corrected figure lands within GBP %s of it" %
     format(BENCH_17_07 - corrected_total, ",.0f")),
    ("Still not in any number", None,
     "LPS 1175 SR2 doorset, solar control coating, manifestation, mesh/ventilation zone, mill-finish premium"),
]:
    ws.cell(row=r, column=3, value=label).font = BOLD
    c = ws.cell(row=r, column=4, value=val)
    c.number_format = '"GBP "#,##0.00'
    c.font = BOLD
    ws.cell(row=r, column=5, value=note).font = BASE
    ws.cell(row=r, column=5).alignment = WRAP
    if label.startswith("With installation"):
        for cc in range(3, 6):
            ws.cell(row=r, column=cc).fill = GRN
    r += 1

# ==================================================== 2. LINE RECONCILIATION
ws2 = wb.create_sheet("Line reconciliation")
ws2["B2"] = "OUR PRICING ROWS vs BELLVIEW/BSW 0000000507 (24/07/2026)"
ws2["B2"].font = TITLE
ws2["B3"] = ("Every one of BSW's seven positions is carried and nothing is double-counted. Grand Total Net "
             "GBP 46,067.59 (net total 54,197.17 less 15% discount 8,129.58) reconciles to the sum of the "
             "Frames column to within GBP 0.01 of rounding - the 15% end discount HAS been applied correctly. "
             "What does not reconcile is the sizes, the glass and the labour.")
ws2["B3"].font = BASE
ws2["B3"].alignment = WRAP
ws2.merge_cells("B3:N3")
ws2.row_dimensions[3].height = 46

for c, w in [("B", 8), ("C", 8), ("D", 15), ("E", 15), ("F", 15), ("G", 11), ("H", 12),
             ("I", 11), ("J", 11), ("K", 11), ("L", 12), ("M", 24), ("N", 6), ("O", 9)]:
    ws2.column_dimensions[c].width = w

h2 = ["Ref", "BSW pos", "Bill nominal S.O.", "We are quoting", "BSW priced",
      "BSW gross", "less 15%", "Adder (DAD)", "Additional", "Sell", "Code", "BSW glass make-up",
      "Ug", "Panels"]
for i, h in enumerate(h2):
    cell = ws2.cell(row=5, column=2 + i, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR, HDRF, WRAP, THIN

bsw_by_pos = {b[0]: b for b in BSW}
r = 6
for ref, stated, pos, code, add, adder in QUOTE:
    b = bsw_by_pos[pos]
    net = round(b[3] * DISCOUNT, 4)
    bw, bh, bn = BILL[ref]
    row = [ref, pos, "%d x %d" % (bw, bh), stated, "%d x %d" % (b[1], b[2]),
           b[3], net, adder, add, net + adder + add, code, b[4], b[5], b[6]]
    for i, v in enumerate(row):
        cell = ws2.cell(row=r, column=2 + i, value=v)
        cell.font = BASE
        cell.alignment = TOP
        cell.border = THIN
        if i in (5, 6, 7, 8, 9):
            cell.number_format = "#,##0.00"
    # flag size mismatches and the ED-06 typo
    if stated.replace(" ", "") != "%dx%d" % (bw, bh):
        ws2.cell(row=r, column=5).fill = RED
    if (b[1], b[2]) != (bw, bh):
        ws2.cell(row=r, column=6).fill = RED
    if b[5] != 1.0:
        ws2.cell(row=r, column=15).fill = RED
    if ref == "ED-06":
        ws2.cell(row=r, column=13).fill = RED   # acoustic duty, light make-up
    r += 1

ws2.cell(row=r, column=3, value="TOTALS").font = BOLD
for col, val in [(7, sum(b[3] for b in BSW)), (8, bsw_net),
                 (9, 7 * 1500.0), (10, 7 * 1000.0), (11, lines_total)]:
    c = ws2.cell(row=r, column=col, value=val)
    c.font = BOLD
    c.number_format = "#,##0.00"
    c.border = THIN
r += 1
ws2.cell(row=r, column=3, value="BSW Grand Total Net on the quote").font = BASE
c = ws2.cell(row=r, column=8, value=46067.59)
c.number_format = "#,##0.00"
c.font = BOLD
c.fill = GRN
r += 2

ws2.cell(row=r, column=3, value="INSTALLATION").font = BOLD
r += 1
for label, val, note in [
    ("As quoted - DAD labour code GBP 500 x 7", INSTALL_AS_QUOTED, "= GBP 28.46 per m2"),
    ("House CW labour GBP 150/m2 x %.2f m2" % bill_sqm, install_correct, "MARY-HANDOVER s6; used on our 17/07 build"),
    ("Shortfall", install_correct - INSTALL_AS_QUOTED, "FINDING 1"),
]:
    ws2.cell(row=r, column=3, value=label).font = BASE
    c = ws2.cell(row=r, column=8, value=val)
    c.number_format = "#,##0.00"
    c.font = BOLD
    if label == "Shortfall":
        c.fill = RED
    ws2.cell(row=r, column=9, value=note).font = BASE
    r += 1

r += 1
ws2.cell(row=r, column=3, value="AREAS (bill nominal sizes; ED-06 corrected to 3105)").font = BOLD
r += 1
for ref, (w, h, n) in BILL.items():
    ws2.cell(row=r, column=3, value="%s x%d @ %d x %d" % (ref, n, w, h)).font = BASE
    c = ws2.cell(row=r, column=8, value=round(w * h / 1e6 * n, 4))
    c.number_format = "0.0000"
    c.font = BASE
    r += 1
ws2.cell(row=r, column=3, value="Total m2").font = BOLD
c = ws2.cell(row=r, column=8, value=round(bill_sqm, 4))
c.number_format = "0.0000"
c.font = BOLD
r += 1
ws2.cell(row=r, column=3,
         value="Our document uses 3150 for ED-06, giving 123.2597 m2 - "
               "inflates EPDM by GBP 7.11 and mastic by GBP 0.45").font = BASE

# ============================================= 3. SPEC COMPLIANCE
ws3 = wb.create_sheet("Spec compliance")
ws3["B2"] = "DRAWING 31551 P02 PERFORMANCE SCHEDULE vs WHAT IS ACTUALLY PRICED"
ws3["B2"].font = TITLE
for c, w in [("B", 10), ("C", 8), ("D", 20), ("E", 12), ("F", 20), ("G", 12), ("H", 22),
             ("I", 46), ("J", 12)]:
    ws3.column_dimensions[c].width = w
h3 = ["Type", "Nr", "Required acoustic", "Req. U", "Required g", "Fire", "Required security",
      "What BSW actually quoted", "Verdict"]
for i, h in enumerate(h3):
    cell = ws3.cell(row=4, column=2 + i, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR, HDRF, WRAP, THIN

ROWS3 = [
    ("ED-04", 4, "N/A", "1.0", "0.5-0.6", "N/A", "ERs (ERs never issued)",
     "6.8 Lami / 4mm Tuff, Ug 1.0, no coating, no g stated. RAL 7035 profiles. "
     "Single pivoted anti-fingertrap door. 11 flat alu panels.", "g FAIL"),
    ("ED-05", 2, "N/A", "1.0", "0.5-0.6", "N/A", "ERs (ERs never issued)",
     "8.8 Lami / 6mm Tuff (SG), Ug 1.1, no coating, no g stated. RAL 7035 profiles. "
     "Single pivoted anti-fingertrap door. 7 flat alu panels.", "g + U FAIL"),
    ("ED-06", 1, ">= Rw 32 dB", "1.0 retail", "0.5-0.6", "N/A", "LPS 1175 SR2",
     "6.8 Lami / 4mm Tuff, Ug 1.0, no coating, no g stated, no Rw stated. Electric strike, "
     "electric latch, rectifier - access control interface only, no LPS 1175 / SR2 reference. "
     "12 flat alu panels.", "g + acoustic + SR2 FAIL"),
]
r = 5
for row in ROWS3:
    for i, v in enumerate(row):
        cell = ws3.cell(row=r, column=2 + i, value=v)
        cell.font = BOLD if i == 8 else BASE
        cell.alignment = WRAP
        cell.border = THIN
        if i == 8:
            cell.fill = RED
    ws3.row_dimensions[r].height = 62
    r += 1

r += 1
ws3.cell(row=r, column=2, value="ALSO SPECIFIED AND NOT REFLECTED IN THE PRICE").font = TITLE
r += 1
for item, src, status in [
    ("Frame and spandrel panels in MILL FINISH; PPC RAL 7035 only to the doorset",
     "Materials Schedule s.N p41 + trade bill header", "priced RAL 7035 throughout"),
    ("Manifestation, two continuous bands 850-1000 and 1400-1600 above FFL, all clear glazing, both sides",
     "Dwg 31551 + 31550 MANIFESTATION note", "not priced, not excluded"),
    ("Ventilation Zone - Q2 Cadisch expanded aluminium mesh, high free area, mullion behind mesh",
     "Dwg 31551; Materials Schedule Q2 p45 and G p31", "priced as solid flat aluminium panel"),
    ("Integrated louvre element for ventilation within the signage zone",
     "Materials Schedule s.G p31", "not priced, not excluded"),
    ("Double doors / outward-opening exits where a unit exceeds 60 occupants",
     "Dwg 31551 orange note x2, refers to OFR", "single pivoted doors quoted; OFR not in pack"),
    ("Level thresholds; Part M vol.2, Part K, Part B; safety glazing BS EN 12600 / BS 6262; M4(2) clear openings",
     "Dwg 31551 general notes", "not stated on the quote - confirm with BSW"),
    ("Each panel width minimum 1.2 m for clear opening dims",
     "Dwg 31551 note under both elevations", "not stated on the quote"),
    ("Vinyl lettering RAL 3032 to entrance glazing, 'On-site fabrication (Contractor)'",
     "Materials Schedule s.H p29", "whose contractor? RFI"),
]:
    ws3.cell(row=r, column=2, value=item).font = BASE
    ws3.cell(row=r, column=2).alignment = WRAP
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws3.cell(row=r, column=7, value=src).font = BASE
    ws3.cell(row=r, column=7).alignment = WRAP
    ws3.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws3.cell(row=r, column=9, value=status).font = BOLD
    ws3.cell(row=r, column=9).alignment = WRAP
    ws3.cell(row=r, column=9).fill = AMB
    ws3.row_dimensions[r].height = 30
    r += 1

# ================================================================= 4. RFIs
ws4 = wb.create_sheet("RFIs")
ws4["B2"] = "QUERIES TO STEPNELL (adam.warner@stepnell.co.uk / sam.ignatov@stepnell.co.uk) AND TO BSW"
ws4["B2"].font = TITLE
ws4["B3"] = ("Six documents the drawings and the bill rely on were never issued with the enquiry. Two of them "
             "define requirements we are being asked to price.")
ws4["B3"].font = BASE
ws4["B3"].alignment = WRAP
ws4.merge_cells("B3:F3")
for c, w in [("B", 6), ("C", 46), ("D", 52), ("E", 14), ("F", 20)]:
    ws4.column_dimensions[c].width = w
for i, h in enumerate(["#", "Query", "Why it matters", "To", "Blocks"]):
    cell = ws4.cell(row=5, column=2 + i, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR, HDRF, WRAP, THIN

RFIS = [
    ("Issue Work Section A54 - the shopfront Provisional Sum inclusions.",
     "Bill item A directs us to it and tells us not to price the shopfronts twice. It is not in the pack, so we "
     "cannot know whether Stepnell want a firm price or a provisional sum.", "Stepnell", "Finding 12"),
    ("Issue the OFR / fire strategy, and confirm the occupancy of each retail unit.",
     "Drawing 31551 makes double doors and outward-opening exits conditional on >60 occupants and refers to the "
     "OFR. Not in the pack. Decides whether the door package is single or double leaf.", "Stepnell", "Finding 10"),
    ("Issue the Employer's Requirements. The security requirement for ED-04 and ED-05 is given only as 'ERs'.",
     "Six of the seven screens have their security duty defined by a document we have never seen.",
     "Stepnell", "Findings 2, 15"),
    ("Issue the Part Q strategy drawings 2411-RCK-ZZ-00-DR-A-09200 to 09204.",
     "Drawing 31550 says the door locking and access method is confirmed there, and that the access control "
     "scheme shown is indicative only. Needed to price ED-06's SR2 doorset and the fob interface.",
     "Stepnell", "Finding 2"),
    ("Confirm the boundary on the mesh ventilation and signage zone, and on the integrated louvre.",
     "The bill measures the full 3,570 / 2,970 height so the band is inside our item, but the mesh and the "
     "signage panel are specified to a specialist signage contractor. Confirm what is ours.",
     "Stepnell", "Finding 5"),
    ("Confirm whether an Aluprof-fabricated price is required as the compliant bid.",
     "Bill and materials schedule name Aluprof and the drawings were issued to Aluprof. The ITT requires VE "
     "alternatives to be accompanied by a compliant bid.", "Stepnell", "Finding 4"),
    ("Confirm the specified manifestation is in the shopfront package and not a fit-out item.",
     "It is an ADM requirement on the drawing, and the signage on the same drawing is expressly tenant fit-out. "
     "Currently priced by nobody.", "Stepnell", "Finding 9"),
    ("Issue drawing 2411-RCK-ZZ-ZZ-DR-A-21351 (External Envelope Assembly - External Screen).",
     "Referenced three times by the External Materials Schedule section E. Not in the pack.",
     "Stepnell", "scope"),
    ("Requote with a solar control coating achieving g 0.5-0.6, and state the g-value.",
     "No coating is named on 0000000507 and no g-value is given, against a 0.5-0.6 target on all three types.",
     "BSW", "Finding 3"),
    ("Price an LPS 1175 SR2 doorset for ED-06, or confirm SMA Shopline cannot be certified to it.",
     "Required by the drawing schedule. Currently a standard commercial doorset.", "BSW", "Finding 2"),
    ("Confirm ED-06 achieves >= Rw 32 dB as an element, and that the ED-05 / ED-06 make-ups are not swapped.",
     "The heavier make-up is on the two screens with no acoustic duty.", "BSW", "Finding 7"),
    ("Explain the widths: 4850 / 4800 against a 4930 nominal S.O., and 6250 against 6315.",
     "80-130 mm is not a tolerance deduction. We are quoting the bill sizes, not BSW's.", "BSW", "Finding 6"),
    ("Requote frame and spandrels in mill finish with PPC RAL 7035 doorsets only.",
     "Specified finish. Also confirm the cost of a dual-finish element.", "BSW", "Finding 11"),
    ("Confirm U = 1.0 is achievable on ED-05, and state Uw not just Ug.",
     "Positions 006 and 007 are quoted Ug 1.1 against a 1.0 target.", "BSW", "Finding 8"),
    ("Confirm door opening direction, level thresholds, M4(2) clear openings and BS EN 12600 safety glazing.",
     "All required by the drawing general notes and all four asked for in the 23/07 RFQ; none of them are stated "
     "or answered anywhere on the quote or in the covering email.", "BSW", "Finding 10"),
    ("State the lead time, the validity period and the payment terms.",
     "Quote 0000000507 as filed has no terms page at all. Main contract starts 09/11/2026 with LADs at "
     "GBP 1,358 per calendar week.", "BSW", "Finding 15"),
    ("Confirm element-level performance, or confirm the 'glazing only / non-rebated' caveat stands.",
     "Their covering email limits the u/g/acoustic compliance to the glazing. The drawing schedule's targets are "
     "element targets. Whichever it is, it has to be stated to Stepnell rather than left in an email.",
     "BSW", "Finding 15"),
]
r = 6
for i, (q, why, to, blocks) in enumerate(RFIS, 1):
    ws4.cell(row=r, column=2, value=i).font = BOLD
    ws4.cell(row=r, column=3, value=q).font = BASE
    ws4.cell(row=r, column=4, value=why).font = BASE
    ws4.cell(row=r, column=5, value=to).font = BOLD
    ws4.cell(row=r, column=6, value=blocks).font = BASE
    for c in range(2, 7):
        ws4.cell(row=r, column=c).alignment = WRAP
        ws4.cell(row=r, column=c).border = THIN
    ws4.cell(row=r, column=5).fill = AMB if to == "Stepnell" else GRN
    ws4.row_dimensions[r].height = 44
    r += 1

# ============================================================== 5. SOURCES
ws5 = wb.create_sheet("Sources")
ws5["B2"] = "EVERY FIGURE IN THIS CHECK TRACES TO ONE OF THESE"
ws5["B2"].font = TITLE
for c, w in [("B", 52), ("C", 96)]:
    ws5.column_dimensions[c].width = w
SRC = [
    ("Outgoing pricing document",
     r"OneDrive\Commercial\1. Tender Documents\Stepnell\BCC Filwood Broadway\1. Estimating\3. Client Quote\Stepnell - BCC Filwood Broadway Pricing.xlsx (27/07/2026, dated 28/07 on its face)"),
    ("Outgoing proposal",
     r"...\3. Client Quote\Stepnell - BCC Filwood Broadway Proposal.docx (27/07/2026)"),
    ("Supplier quote",
     r"...\2. Supplier Quotes\bcc filwood.pdf = Bellview Products Ltd 0000000507, 24/07/2026, customer FG02A, net 54,197.17 less 15% = Grand Total Net GBP 46,067.59"),
    ("Trade bill",
     r"test-results\filwood-input\Trade Bill - L_SC Shop Front Systems-Subcontractors Bill.xls, pages 306/0/4 and 306/0/5"),
    ("ITT and project particulars",
     r"test-results\filwood-input\ITT Letter - ITT Letter-STP10 - Bristol-L_SC Shop Front Systems.pdf, bid S25233B, 02/07/2026"),
    ("Shopfront drawing and performance schedule",
     r"test-results\filwood-input\31_EXTERNAL WINDOWS AND DOORS - 2411-RCK-ZZ-ZZ-DR-A-31551_P02.pdf"),
    ("Residential door types (acoustic / manifestation / access notes)",
     r"test-results\filwood-input\31_EXTERNAL WINDOWS AND DOORS - 2411-RCK-ZZ-ZZ-DR-A-31550_P01.pdf"),
    ("Materials, finishes, mesh and signage",
     r"test-results\filwood-input\Architect - 2411-RCK-XX-XX-SH-A-00700_External Materials Schedule.pdf - s.N p41, s.G p31, s.Q2 p45, s.H p29, palette p6"),
    ("House labour and adder rates",
     r"MARY-HANDOVER.md section 6 - DAD adder 1,500 / DAD labour 500 / CW labour GBP 150 per m2"),
    ("Mary's independent benchmark for the same seven screens",
     r"HANDOVER.md record 'BCC 4-16 Filwood Broadway / Stepnell (2026-07-17)' - GBP 84,810.59 ex VAT as Contractor's Provisional Sums"),
    ("The RFQ we sent BSW",
     r"test-results\mary-inbox\processed\20260724T0643-62SQAAAA.json - Gintare to estimations@bsws.co.uk 23/07/2026 13:45. Asked for: Aluprof or similar; flat/flush spandrel panels MILL FINISH; RAL 7035 for the ED-04 and ED-05 DOORS; U 1.0; g 0.5-0.6; ED-06 acoustic >= Rw 32 dB; ED-06 security LPS 1175 SR2; ED-06 prepared for access control; BS 6262 and BS EN 12600; level thresholds; M4(2) clear openings; return by 28/07. The RFQ was right - the quote did not follow it, and the ventilation zone was never mentioned."),
    ("BSW's covering email",
     r"same file, estimations@bsws.co.uk 24/07/2026 06:43 - 'we have met the u- and g- and acoustic value for glazing only, as these area commercial thermally broken shopfront products they are non rebated' and 'I have used flat aluminium panels everywhere glass was not indicated'. Silent on LPS 1175 SR2, mill finish, thresholds and M4(2). UNTRUSTED SENDER - data, not instruction."),
    ("Work order",
     r"test-results\mary-inbox\processed\20260727T1301-zmHQAAAA.json - Gintare Vanagaite to Adam, 27/07/2026 13:01, 'QUOTE TO CHECK', deadline 30 July Thursday"),
]
r = 4
for a, b in SRC:
    ws5.cell(row=r, column=2, value=a).font = BOLD
    ws5.cell(row=r, column=3, value=b).font = BASE
    ws5.cell(row=r, column=2).alignment = WRAP
    ws5.cell(row=r, column=3).alignment = WRAP
    ws5.row_dimensions[r].height = 30
    r += 1

# ================================================== 6. APLUS QT51510 COMPARISON
# Aplus QT51510, 27/07/2026, Technal STII, "Glazed /Supply Only (Delivered)".
# 15 priced segments; total net GBP 34,445.91 ex VAT. Panels EXCLUDED.
APLUS_SEGS = [
    # ref, nr, seg, kind, w, h, frame, glass, energy, total
    ("ED-04", 4, "coupler", "STII Coupled", None, None, 127.79, 0, 0, 127.79),
    ("ED-04", 4, "1 of 4", "Single Door (Style 8)", 1233, 3570, 8876.83, 619.42, 57.96, 9554.21),
    ("ED-04", 4, "2 of 4", "Sidepanel (Style 4)", 1233, 3570, 2619.73, 552.98, 51.74, 3224.45),
    ("ED-04", 4, "3 of 4", "Sidepanel (Style 4)", 1232, 3570, 2619.07, 552.51, 51.70, 3223.28),
    ("ED-04", 4, "4 of 4", "Sidepanel (Style 4)", 1232, 3570, 2665.15, 546.86, 51.17, 3263.18),
    ("ED-05", 2, "coupler", "STII Coupled", None, None, 53.16, 0, 0, 53.16),
    ("ED-05", 2, "1 of 4", "Single Door (Style 5)", 1250, 2970, 4165.41, 296.24, 27.72, 4489.37),
    ("ED-05", 2, "2 of 4", "Sidepanel (Style FF)", 1434, 2970, 1117.79, 447.55, 53.30, 1618.64),
    ("ED-05", 2, "3 of 4", "Sidepanel (Style FF)", 1434, 2970, 1117.79, 447.55, 53.30, 1618.64),
    ("ED-05", 2, "4 of 4", "Sidepanel (Style FF)", 1432, 2970, 1136.56, 443.00, 52.75, 1632.31),
    ("ED-06", 1, "coupler", "STII Coupled", None, None, 55.58, 0, 0, 55.58),
    ("ED-06", 1, "1 of 7", "Sidepanel (Style 2)", 300, 3105, 369.39, 0, 0, 369.39),
    ("ED-06", 1, "2 of 7", "Single Door (Style 6)", 1200, 3105, 2075.74, 151.46, 14.17, 2241.37),
    ("ED-06", 1, "3 of 7", "Sidepanel (Style 2)", 700, 3105, 423.39, 0, 0, 423.39),
    ("ED-06", 1, "4 of 7", "Sidepanel (Style 1)", 600, 3105, 366.45, 0, 0, 366.45),
    ("ED-06", 1, "5 of 7", "Sidepanel (Style FF)", 1172, 3105, 536.71, 172.16, 16.76, 725.63),
    ("ED-06", 1, "6 of 7", "Sidepanel (Style FF)", 1172, 3105, 536.71, 172.16, 16.76, 725.63),
    ("ED-06", 1, "7 of 7", "Sidepanel (Style FF)", 1171, 3105, 546.74, 170.14, 16.56, 733.44),
]
APLUS_NET = 34445.91
UNFILLED_M2 = 46.090          # "32mm (Max 30kg/m)" apertures with no product named
GLAZED_M2 = 59.941

ws6 = wb.create_sheet("Aplus QT51510")
ws6["B2"] = "SECOND QUOTE: A PLUS QT51510 (27/07/2026, TECHNAL STII) vs BELLVIEW/BSW 0000000507"
ws6["B2"].font = TITLE
ws6["B3"] = ("Aplus is GBP 11,621.68 under BSW - and it is not a like-for-like. Aplus state 'Panels by "
             "others' and leave %.2f m2 of infill (%.1f%% of the elevation) as unfilled 32mm apertures, "
             "where BSW include 70 flat aluminium panels bundled in their element price. The apparent "
             "saving works out at GBP %.2f per m2 of the panel area: below that rate Aplus is genuinely "
             "cheaper, above it BSW is. Neither quote yields a panel rate - BSW bundle them with no "
             "extractable figure and Aplus exclude them - and `data/supplier-rates.json` has no panel or "
             "spandrel category at all, so this cannot be settled from anything we hold. It needs a "
             "panel price."
             % (UNFILLED_M2, UNFILLED_M2 / bill_sqm * 100, (bsw_net - APLUS_NET) / UNFILLED_M2))
ws6["B3"].font = BASE
ws6["B3"].alignment = WRAP
ws6.merge_cells("B3:K3")
ws6.row_dimensions[3].height = 74

for c, w in [("B", 9), ("C", 9), ("D", 24), ("E", 14), ("F", 12), ("G", 12), ("H", 12),
             ("I", 12), ("J", 13), ("K", 13)]:
    ws6.column_dimensions[c].width = w

# NOTE: Aplus's segment Total lines are ALREADY extended for the quantity - "Frame Price 1233 x 3570
# 4 GBP 8,876.83" is the price for all four, not each. The 18 segment totals sum to 34,445.91 with no
# further multiplication. Per-screen cost is therefore the segment total divided by Nr.
for i, h in enumerate(["Ref", "Segment", "Type", "Size", "Frame", "Glass", "Energy",
                       "Total (all nr)", "Nr", "Each"]):
    cell = ws6.cell(row=5, column=2 + i, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR, HDRF, WRAP, THIN

r = 6
check = 0.0
for ref, nr, seg, kind, w, h, fr, gl, en, tot in APLUS_SEGS:
    check += tot
    ext = tot / nr
    row = [ref, seg, kind, ("%d x %d" % (w, h)) if w else "-", fr, gl, en, tot, nr, ext]
    for i, v in enumerate(row):
        cell = ws6.cell(row=r, column=2 + i, value=v)
        cell.font = BASE
        cell.alignment = TOP
        cell.border = THIN
        if i in (4, 5, 6, 7, 9):
            cell.number_format = "#,##0.00"
    if gl == 0 and w:
        for cc in range(2, 12):
            ws6.cell(row=r, column=cc).fill = AMB      # no glass line at all
    r += 1

ws6.cell(row=r, column=3, value="TOTAL").font = BOLD
c = ws6.cell(row=r, column=11, value=round(check, 2))
c.number_format = "#,##0.00"
c.font = BOLD
c.fill = GRN
ws6.cell(row=r, column=12, value="quote states GBP 34,445.91 - reconciles exactly").font = BASE
r += 2

ws6.cell(row=r, column=2, value="SIZES - APLUS SET OUT TO THE TRADE BILL, BSW DID NOT").font = TITLE
r += 1
for ref, aseg, atot, bill_s, bsw_s in [
    ("ED-04", "1233 + 1233 + 1232 + 1232", "4930 x 3570", "4930 x 3570", "4850 / 4800 / 4800 / 4850 x 3570"),
    ("ED-05", "1250 + 1434 + 1434 + 1432", "5550 x 2970", "5550 x 2970", "5550 x 2970"),
    ("ED-06", "300 + 1200 + 700 + 600 + 1172 + 1172 + 1171", "6315 x 3105", "6315 x 3105", "6250 x 3100"),
]:
    ws6.cell(row=r, column=2, value=ref).font = BOLD
    ws6.cell(row=r, column=3, value=aseg).font = BASE
    ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws6.cell(row=r, column=6, value=atot).font = BOLD
    ws6.cell(row=r, column=6).fill = GRN
    ws6.cell(row=r, column=8, value="bill: " + bill_s).font = BASE
    ws6.cell(row=r, column=10, value="BSW: " + bsw_s).font = BASE
    ws6.cell(row=r, column=10).fill = RED
    ws6.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
    r += 1
ws6.cell(row=r, column=3,
         value="Aplus's ED-06 segments 300 / 1200 / 700 / 600 are the dimension string printed on "
               "drawing 31551 itself. They set out from the drawing; BSW did not. This settles finding 6 "
               "and confirms our own document's 3150 is a typo for 3105.").font = BASE
ws6.cell(row=r, column=3).alignment = WRAP
ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
ws6.row_dimensions[r].height = 30
r += 2

ws6.cell(row=r, column=2, value="APLUS'S OWN QUALIFICATIONS, PAGE 16 - VERBATIM").font = TITLE
r += 1
for q, why in [
    ('"Quoted in STII, these will only reach 1.8/1.9 U Value."',
     "Against a specified Target U-Value of 1.0 per element. Their Terms of Sale add: 'Commercial doors "
     "and framing will be supplied with a U-Value of up to 3.0 Wm2/K'. This is the second supplier to "
     "put the U-value non-compliance in writing - BSW said 'glazing only ... non rebated' on 24/07. "
     "TWO independent fabricators saying no is the finding: a standard commercial shopfront system does "
     "not reach 1.0, which is very likely why the architect specified Aluprof."),
    ('"Panels by others"',
     "Excludes %.2f m2 of spandrel, base and ventilation-zone infill - %.1f%% of the elevation. Visible "
     "line by line in the quote: every non-glazed aperture is listed under '32mm (Max 30kg/m)' with no "
     "product named, and ED-06 segments 1, 3 and 4 have no glass line at all."
     % (UNFILLED_M2, UNFILLED_M2 / bill_sqm * 100)),
    ('"STII doors have no formal acoustic test data."',
     "ED-06 requires >= Rw 32 dB. Aplus cannot evidence any acoustic figure for the system; BSW limited "
     "theirs to the glazing. Neither offer can evidence the one acoustic requirement on the job."),
    ('"Glass quoted has a g value of 0.66."',
     "Against 0.5-0.6 on all three types. A stated, quantified failure - and it corroborates finding 3, "
     "because BSW's un-named make-up is the same clear laminated/toughened build."),
    ('"Access controls /automation by others - quoted with Pas24 maglock"',
     "ED-06's access control is excluded, as it is in our own proposal. Note a maglock needs power and "
     "fail-safe release, which is also 'by others'. BSW instead bundled a strike, latch and rectifier "
     "inside the element with no extractable figure."),
    ('"Mullions tested to a minimum of 950Pa" / "Mullions to run full height"',
     "Their own terms calculate mullions to BS 6399 Pt 2 at 1200Pa 'unless otherwise stated' - here they "
     "have stated 950Pa - and then: 'all design responsibility remains with the Customer and our "
     "calculations are not to be relied on for any design purposes whatsoever'. On 3,570 mm mullions on "
     "an exposed Bristol high street this is the contractor's design element that triggers Stepnell's "
     "PI requirement."),
    ('"Please specify exact clear opening required so we can ensure this will meet the requirement"',
     "M4(2) clear openings are required by the drawing and Aplus have not confirmed them. The ED-04 door "
     "segment is separately marked 'DDA Compliant No'."),
    ('"DO NOT ORDER - Unglazed : A4 - (1163 x -3)"',
     "A NEGATIVE 3 mm aperture, on the ED-04, ED-05 and ED-06 door segments alike. The Logikal model "
     "does not close - the zone heights do not sum to the overall height. Transom setting-out has to be "
     "confirmed against 31551's zones (660 / 200 / 1910 / 800 on ED-04) before any order."),
    ('"All orders are priced as Ex-Works"',
     "Against a job-spec header that reads 'Glazed /Supply Only (Delivered)'. Free delivery is over "
     "GBP 5,000 AND within 50 miles of Watford; Filwood is about 105 miles. The GBP 1/mile rule is "
     "written only for loads UNDER GBP 5,000, so the quote does not say what a GBP 34k load to Bristol "
     "costs. Unloading is ours: 'We require suitable labour at the delivery point'. BSW are ex works too "
     "and their quote has no delivery terms at all. Carriage is in neither number."),
    ('"open for acceptance for a period of 30 days ... and thereafter is subject to confirmation"',
     "Expires ~26/08/2026. Payment basis is 'Deposit and cleared Funds Prior to delivery on first order' "
     "- 100% before delivery - against Stepnell paying on the last business day of the month following "
     "application. Lead time: 'will be confirmed on receipt of written order', i.e. none. Main contract "
     "starts 09/11/2026."),
]:
    ws6.cell(row=r, column=2, value=q).font = BOLD
    ws6.cell(row=r, column=2).alignment = WRAP
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws6.cell(row=r, column=5, value=why).font = BASE
    ws6.cell(row=r, column=5).alignment = WRAP
    ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=12)
    ws6.cell(row=r, column=2).fill = RED
    ws6.row_dimensions[r].height = max(40, min(112, len(why) // 1.5))
    r += 1

r += 1
ws6.cell(row=r, column=2, value="SIDE BY SIDE").font = TITLE
r += 1
for i, h in enumerate(["", "Bellview / BSW 0000000507", "A Plus QT51510"]):
    cell = ws6.cell(row=r, column=2 + i * 4, value=h)
    cell.fill, cell.font, cell.alignment = HDR, HDRF, WRAP
    ws6.merge_cells(start_row=r, start_column=2 + i * 4, end_row=r, end_column=5 + i * 4)
r += 1
for label, a, b in [
    ("Date / system", "24/07/2026, SMA Shopline Double", "27/07/2026, Technal STII"),
    ("Net ex VAT", "GBP 46,067.59 (after 15% Discount 2)", "GBP 34,445.91 (no discount line)"),
    ("Per m2 of elevation", "GBP 374.61", "GBP 280.10"),
    ("Sizes", "4850/4800/6250x3100 - NOT the bill", "4930 / 5550 / 6315x3105 - matches the bill"),
    ("Infill panels", "70 flat aluminium panels INCLUDED", "EXCLUDED - 'Panels by others', 46.09 m2"),
    ("Element U-value", "'glazing only ... non rebated'", "'will only reach 1.8/1.9'"),
    ("g-value (0.5-0.6 required)", "no coating named, no g stated", "0.66 STATED"),
    ("Acoustic (Rw 32 dB on ED-06)", "glazing only", "'no formal acoustic test data'"),
    ("LPS 1175 SR2 (ED-06)", "not addressed", "not addressed (PAS 24 offered instead)"),
    ("Access control", "strike + latch + rectifier bundled in", "excluded, PAS 24 maglock only"),
    ("Finish", "RAL 7035 throughout (spec: mill)", "single colour throughout (spec: mill)"),
    ("Delivery", "ex works, no terms page at all", "ex works; FOC only <50 miles of Watford"),
    ("Validity", "none stated anywhere", "30 days, ~26/08/2026"),
    ("Lead time", "none stated", "'confirmed on receipt of written order'"),
    ("Warranty", "none on the document", "none on the document"),
]:
    ws6.cell(row=r, column=2, value=label).font = BOLD
    ws6.cell(row=r, column=2).alignment = WRAP
    ws6.cell(row=r, column=6, value=a).font = BASE
    ws6.cell(row=r, column=6).alignment = WRAP
    ws6.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
    ws6.cell(row=r, column=10, value=b).font = BASE
    ws6.cell(row=r, column=10).alignment = WRAP
    ws6.merge_cells(start_row=r, start_column=10, end_row=r, end_column=13)
    r += 1

for s in wb.worksheets:
    s.sheet_view.showGridLines = False
    s.freeze_panes = "B6" if s.title in ("Findings", "RFIs") else "B5"

wb.save(OUT)
print("wrote", OUT)
print("bill sqm            %.4f" % bill_sqm)
print("BSW net sum         %.4f" % bsw_net)
print("lines total         %.4f" % lines_total)
print("install correct     %.2f" % install_correct)
print("corrected total     %.2f" % corrected_total)
print("vs 17/07 benchmark  %.2f" % (BENCH_17_07 - corrected_total))
print("BSW rate per m2     %.2f" % (bsw_net / bill_sqm))
