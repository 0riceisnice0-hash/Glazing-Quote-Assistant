# -*- coding: utf-8 -*-
"""Stage 4 - the deliverable for Zac's Jacob/BD question (dashmsg-16).

Joins the archive sweep (what we issued), the Estimating Log (what outcome was
recorded) and 9k+ estimating@ messages (who last spoke to whom) into one
workbook of quotes that are out with no recorded result.

Reads  scratchpad/pipeline-stage3.json, scratchpad/pipeline-mailbox.json
Writes outputs/Fenster Quote Pipeline - Issued Quotes Without an Outcome.xlsx
"""
import datetime as dt
import json
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STAGE3 = os.path.join("scratchpad", "pipeline-stage3.json")
MAILBOX = os.path.join("scratchpad", "pipeline-mailbox.json")
OUT = os.path.join("outputs", "Fenster Quote Pipeline - Issued Quotes Without an Outcome.xlsx")
TODAY = dt.datetime(2026, 7, 27, 23, 59)
INTERNAL = "fensterglazing.com"
# An external reply is not a CLIENT reply. Suppliers and portals chase us
# constantly - on Gordon Court the only recent inbound was AFS chasing their own
# quote, which would otherwise read as "the client came back".
SUPPLIERS = ("aluminiumfiresystems", "aplusaluminium", "bswbuilding", "bellview",
             "vetroseal", "strongdor", "cnglass", "ikon", "mercuryglazing",
             "mercuryspecialistframes", "in-tend", "intend", "procontract",
             "delta-esourcing", "sellerservices", "noreply", "no-reply",
             "donotreply", "do-not-reply")
STOP = {"the", "and", "ltd", "limited", "road", "street", "house", "school", "construction",
        "development", "developments", "project", "works", "new", "phase", "unit", "units",
        "quote", "quotation", "client", "level", "site"}


def toks(s):
    return {w for w in re.findall(r"[a-z0-9]{4,}", (s or "").lower()) if w not in STOP}


def parse_dt(s):
    try:
        return dt.datetime.strptime((s or "")[:19], "%Y-%m-%dT%H:%M:%S")
    except ValueError:
        return None


def main():
    jobs = json.load(open(STAGE3, encoding="utf-8"))
    msgs = json.load(open(MAILBOX, encoding="utf-8"))
    for m in msgs:
        m["_dt"] = parse_dt(m["received"])
        m["_tok"] = toks(m["subject"])
    msgs = [m for m in msgs if m["_dt"]]

    # Two different clients can run a job with the same name - the archive holds
    # a Gordon Court for Chigwell AND one for Target Maintenance. Subject matching
    # cannot separate them, so mark them rather than quietly merge their traffic.
    by_name = {}
    for j in jobs:
        by_name.setdefault(frozenset(toks(j["project"])) or frozenset([j["project"].lower()]),
                           []).append(j)
    for group in by_name.values():
        for j in group:
            j["name_clash"] = len(group) - 1

    for j in jobs:
        ptok, ctok = toks(j["project"]), toks(j["client"])
        issued = dt.datetime.fromtimestamp(j["latest_quote_mtime"])
        j["_issued"] = issued
        hits = []
        for m in msgs:
            # A subject that carries the project name is the strong signal;
            # client name alone is too loose on repeat customers.
            if ptok and len(ptok & m["_tok"]) >= min(2, len(ptok)):
                hits.append(m)
            elif ctok and len(ctok & m["_tok"]) >= 2:
                hits.append(m)
        j["_hits"] = len(hits)
        inb = [m for m in hits if m["inbound_external"]]
        out_ = [m for m in hits if not m["inbound_external"]]
        client_inb = [m for m in inb if not any(s in m["from"] for s in SUPPLIERS)]
        j["last_in"] = max((m["_dt"] for m in inb), default=None)
        j["last_out"] = max((m["_dt"] for m in out_), default=None)
        last = max([d for d in (j["last_in"], j["last_out"]) if d], default=None)
        j["last_contact"] = last
        last_client_in = max((m["_dt"] for m in client_inb), default=None)
        j["replied_after_issue"] = bool(
            last_client_in and last_client_in >= issued - dt.timedelta(days=1))
        j["last_inbound_from"] = ""
        if inb:
            j["last_inbound_from"] = max(inb, key=lambda m: m["_dt"])["from"]
        j["days_silent"] = max(0, (TODAY - last).days) if last else None
        # Who it went to: recipients of the most recent outbound on this job.
        j["issued_to"] = ""
        if out_:
            newest = max(out_, key=lambda m: m["_dt"])
            ext = [a for a in (newest["to"] + newest["cc"]) if a and INTERNAL not in a]
            j["issued_to"] = ", ".join(sorted(set(ext))[:3])

        wl = (j.get("log") or {}).get("wl", "")
        if j["in_projects"]:
            j["status"] = "WON (in 2. Projects)"
        elif wl:
            j["status"] = "RECORDED (%s)" % wl
        elif j["days_silent"] is None:
            j["status"] = "NO TRACE - not in mailbox window"
        elif j["days_silent"] <= 21:
            j["status"] = "LIVE"
        elif j["days_silent"] <= 60:
            j["status"] = "COOLING"
        else:
            j["status"] = "COLD"

    open_jobs = [j for j in jobs if not j["in_projects"] and not (j.get("log") or {}).get("wl")]
    closed = [j for j in jobs if j not in open_jobs]
    order = {"LIVE": 0, "COOLING": 1, "COLD": 2, "NO TRACE - not in mailbox window": 3}
    open_jobs.sort(key=lambda j: (order.get(j["status"], 9), -(j["value"] or 0)))

    os.makedirs("outputs", exist_ok=True)
    wb = openpyxl.Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="002D3A")
    money = '#,##0.00'

    def sheet(title, headers, widths):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = hdr, fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    wb.remove(wb.active)

    ws = sheet("Open quotes",
               ["Status", "Client", "Project", "Value ex VAT", "Value source", "Date issued",
                "Issued to (external)", "Last contact", "Direction", "Client replied since issue?",
                "Last inbound was from", "Days silent", "Emails seen", "Contact data reliable?",
                "Log row?", "Folder"],
               [26, 30, 34, 15, 34, 12, 38, 12, 11, 13, 34, 11, 12, 30, 30, 60])
    for j in open_jobs:
        direction = ""
        if j["last_contact"]:
            direction = "IN" if j["last_in"] == j["last_contact"] else "OUT"
        log = j.get("log") or {}
        ws.append([
            j["status"], j["client"], j["project"],
            j["value"] if j["value"] else None,
            j.get("value_from") or "not read - no pricing workbook",
            j["_issued"].strftime("%d/%m/%Y"),
            j["issued_to"] or "not seen in mailbox window",
            j["last_contact"].strftime("%d/%m/%Y") if j["last_contact"] else "",
            direction,
            "yes" if j["replied_after_issue"] else "no",
            j.get("last_inbound_from") or "",
            j["days_silent"] if j["days_silent"] is not None else "",
            j["_hits"],
            ("NO - project name shared with %d other job(s)" % j["name_clash"])
            if j.get("name_clash") else "yes",
            ("%s / %s" % (log.get("client", ""), log.get("project", ""))).strip(" /") if log else "NOT ON LOG",
            j["path"],
        ])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = money

    ws2 = sheet("Outcome known",
                ["Status", "Client", "Project", "Value ex VAT", "Date issued", "Log W/L", "Folder"],
                [24, 30, 34, 15, 12, 10, 60])
    for j in sorted(closed, key=lambda j: -(j["value"] or 0)):
        ws2.append([j["status"], j["client"], j["project"], j["value"] or None,
                    j["_issued"].strftime("%d/%m/%Y"),
                    (j.get("log") or {}).get("wl", ""), j["path"]])
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = money

    live = [j for j in open_jobs if j["status"] == "LIVE"]
    cooling = [j for j in open_jobs if j["status"] == "COOLING"]
    cold = [j for j in open_jobs if j["status"] == "COLD"]
    notrace = [j for j in open_jobs if j["status"].startswith("NO TRACE")]

    def tot(rows):
        return sum(r["value"] or 0 for r in rows)

    ws3 = sheet("Summary", ["Measure", "Jobs", "Value ex VAT where read", "Note"], [46, 10, 22, 78])
    for r in [
        ("Jobs with a client quote on file (since 01/09/2025)", len(jobs), tot(jobs),
         "The archive, not the log. A job counts once, at its most recent client-quote document."),
        ("- outcome known (won, or W/L on the log)", len(closed), tot(closed),
         "Won is inferred from the job also existing under Commercial\\2. Projects."),
        ("- NO RECORDED OUTCOME", len(open_jobs), tot(open_jobs), "This is the list Jacob needs."),
        ("   LIVE - contact within 21 days", len(live), tot(live), "Someone is still talking to them."),
        ("   COOLING - 22 to 60 days silent", len(cooling), tot(cooling), "Chase candidates."),
        ("   COLD - over 60 days silent", len(cold), tot(cold), "Chase or write off."),
        ("   NO TRACE in the mailbox window", len(notrace), tot(notrace),
         "Quote on file but no matching email since 01/09/2025 - likely issued from another mailbox."),
        ("Estimating Log rows", 349, 0, "3 won, 13 lost, the rest blank - which is why this was built from the archive."),
        ("estimating@ messages read", len(msgs), 0,
         "01/09/2025 to 27/07/2026, all folders including sent. Metadata only."),
        ("Largest single open quote", 1, max((j["value"] or 0) for j in open_jobs),
         "Elkins Construction, Brandon Estate EWI Remediation - checked by hand against cell H67 of "
         "'Pricing Document - Brandon Estate REV 2.xlsx'. It is a real quoted total, not a misread, "
         "but it is 28% of the open value on its own - read the rest of the numbers with that in mind."),
    ]:
        ws3.append([r[0], r[1], r[2] or None, r[3]])
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].number_format = money

    ws4 = sheet("What Mary cannot see", ["Gap", "Effect on this list", "What would close it"], [34, 62, 52])
    for row in [
        ("commercial@fensterglazing.com",
         "Paul Taylor and the project managers run award and delivery traffic through it. Mary only sees it when someone copies estimating@ - the Riverside site address arrived that way today, by luck. Awards and lost-bid notices addressed to commercial@ are invisible, so jobs here can look cold when they were actually won or lost weeks ago.",
         "Mail.Read on commercial@. This is the single biggest gap and it is already being arranged for Jacob."),
        ("info@fensterglazing.com",
         "First point of contact for inbound enquiries. Nothing that arrives there is visible, so a client replying to a quote via the website or a general address leaves no trace in this list.",
         "Mail.Read on info@. Already on Mary's own next-candidates list since the Hightown miss."),
        ("adam@ and individual staff mailboxes",
         "Adam is copied on most quotes but clients often reply to him directly. His replies only appear here when estimating@ is on the thread.",
         "Mail.Read on adam@, or a rule copying client replies to estimating@."),
        ("Contractor portals (In-Tend, Delta, ProContract, Ermine's own)",
         "Awards and rejections are frequently posted on the portal with only a bare notification emailed - the Ninn Lane message today is exactly that. The outcome exists but is unreadable from any mailbox.",
         "Portal logins, or a standing instruction that whoever opens a portal message records the outcome on the log."),
        ("Phone, WhatsApp and site conversations",
         "Adam said today that Paul would WhatsApp the Riverside address. Decisions taken this way never reach any mailbox and cannot be recovered by any amount of access.",
         "Only a discipline fix: the outcome has to be written down somewhere Mary or Jacob can read."),
        ("Anything issued before 01/09/2025",
         "The mailbox pull starts there. Older quotes still on file show as NO TRACE rather than genuinely cold.",
         "Extend the Graph pull further back - it is only a date filter, but it costs time and adds little for chasing."),
        ("The Estimating Log itself",
         "349 rows, 3 marked won and 13 lost. It cannot be used to confirm an outcome, only to find one when it happens to be filled in. It also does not carry the quoted value.",
         "Either the log gets filled in, or the archive plus mailbox stays the source of truth - which is what this workbook does."),
    ]:
        ws4.append(list(row))
    for r in ws4.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws5 = sheet("Method", ["Step", "How"], [30, 110])
    for row in [
        ("What counts as issued",
         "A job counts if it has a client-facing document in its '3. Client Quote' folder, or a file named quote/quotation/proposal/pricing elsewhere in the job. Files marked DO NOT SEND, review, take-off, master, example or template are excluded."),
        ("Date issued", "Last-modified date of the most recent such document. Not the date it was emailed."),
        ("Value", "The TOTAL* cell of the house pricing workbook, ex VAT. Blank where no workbook exists or the total could not be read - those are shown as 'not read', never guessed."),
        ("Won", "The same client and job also existing under Commercial\\2. Projects. That is a strong signal but not a guarantee, and it will miss a job won under a different folder name."),
        ("Last contact", "Newest estimating@ message whose SUBJECT carries the project name (or two or more distinct client words). Subject matching is imperfect - a renamed thread breaks it, which is why 'Emails seen' is shown so you can judge."),
        ("Client replied", "At least one external inbound message on that subject dated on or after the issue date, EXCLUDING known suppliers and portals (AFS, Aplus, BSW, Bellview, Vetroseal, Strongdor, CN Glass, IKON, Mercury, In-Tend and noreply addresses). Without that exclusion Gordon Court reads as 'client replied today' when the only inbound was Aluminium Fire Systems chasing their own quote. The 'Last inbound was from' column is there so you can see who actually wrote."),
        ("Live / cooling / cold", "Days since the last contact in either direction: 21 or fewer LIVE, 22-60 COOLING, over 60 COLD. A cut-off, not a judgement about the client."),
        ("Same name, two jobs", "Where two clients run a job with the same name the mailbox cannot tell them apart, and their contact data is merged. The archive holds a Gordon Court for Chigwell AND one for Target Maintenance; Gordon Court's row shows 'issued to' addresses at tm-gb.co.uk for exactly that reason. Every affected row is marked NO in 'Contact data reliable?' - treat the value and issue date as sound and the contact columns as unusable."),
        ("Known weakness", "Everything here depends on estimating@ being on the thread. See 'What Mary cannot see'."),
    ]:
        ws5.append(list(row))
    for r in ws5.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT)
    print("written: %s" % OUT)
    print("open (no outcome): %d jobs, GBP %s" % (len(open_jobs), format(tot(open_jobs), ",.2f")))
    print("  LIVE %d (GBP %s) | COOLING %d (GBP %s) | COLD %d (GBP %s) | NO TRACE %d" % (
        len(live), format(tot(live), ",.2f"), len(cooling), format(tot(cooling), ",.2f"),
        len(cold), format(tot(cold), ",.2f"), len(notrace)))
    print("\ntop 15 by value with no recorded outcome:")
    for j in sorted(open_jobs, key=lambda j: -(j["value"] or 0))[:15]:
        print("  %-9s %-28s %-34s %14s  silent %s" % (
            j["status"][:9], j["client"][:28], j["project"][:34],
            format(j["value"], ",.2f") if j["value"] else "-",
            j["days_silent"] if j["days_silent"] is not None else "n/a"))


if __name__ == "__main__":
    main()
