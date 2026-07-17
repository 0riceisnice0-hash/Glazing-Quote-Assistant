# -*- coding: utf-8 -*-
"""Generate Fenster house-format client documents.

1. Pricing document: clones templates/MASTER PRICING DOC.xlsx (the live house
   template - logo, formulas, code adders, installation SUMPRODUCT) and fills
   the client block + item rows. All sell arithmetic stays in the template's
   own formulas so the sheet behaves exactly like a hand-filled one.
2. Proposal: builds an HTML replica of MASTER COVER LETTER (exec summary,
   project approach, inclusions/exclusions, previous projects, delivery team,
   full T&Cs) with the extracted brand images; project-specific photos are
   placeholder slots unless supplied. Print to PDF with headless Chrome.

Usage: python scripts/generate-fenster-docs.py job.json
job.json schema: see build_demo() at the bottom of this file.
"""
import base64
import copy
import json
import os
import re
import sys

import openpyxl
from openpyxl.drawing.image import Image as XLImage

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_XLSX = os.path.join(REPO, "templates", "MASTER PRICING DOC.xlsx")
IMAGES_DIR = os.path.join(REPO, "templates", "proposal-images")
SHEET = "Pricing Document "  # note trailing space in the master
FIRST_ROW, LAST_ROW = 9, 20  # template item rows carrying formulas


def _clone_row_formulas(ws, src_row, dst_row):
    """Copy the template's per-row formulas/styles from src_row to dst_row."""
    for col in "BCDEFGHIJKLMNOP":
        src = ws["%s%d" % (col, src_row)]
        dst = ws["%s%d" % (col, dst_row)]
        v = src.value
        if isinstance(v, str) and v.startswith("="):
            dst.value = v.replace(str(src_row), str(dst_row))
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)


def make_pricing_doc(job, out_path):
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb[SHEET]

    ws["D3"] = job.get("client", "")
    ws["D4"] = job.get("projectRef", "")
    ws["D5"] = job.get("siteAddress", "")
    ws["D6"] = job.get("date", "")
    ws["L3"] = job.get("supplier", "")
    if job.get("note"):
        ws["H5"] = job["note"]

    rows = job["rows"]
    n = len(rows)
    last = FIRST_ROW + n - 1
    if last > LAST_ROW:
        ws.insert_rows(LAST_ROW, amount=last - LAST_ROW)
        for r in range(LAST_ROW, last + 1):
            _clone_row_formulas(ws, FIRST_ROW, r)

    # clear template example rows then fill
    for r in range(FIRST_ROW, max(last, LAST_ROW if last <= LAST_ROW else last) + 1):
        for col in ("B", "C", "E", "F", "G", "J", "K", "L"):
            ws["%s%d" % (col, r)] = None
    for i, row in enumerate(rows):
        r = FIRST_ROW + i
        ws["B%d" % r] = row.get("code", "")
        ws["C%d" % r] = row.get("desc", "")
        ws["E%d" % r] = row.get("size", "")
        ws["F%d" % r] = row.get("qty", 1)
        ws["G%d" % r] = row.get("unit", "nr")
        if row.get("frames") is not None:
            ws["J%d" % r] = row["frames"]
        if row.get("glass") is not None:
            ws["K%d" % r] = row["glass"]
        if row.get("additional") is not None:
            ws["L%d" % r] = row["additional"]
        if row.get("unitRateOverride") is not None:
            ws["H%d" % r] = row["unitRateOverride"]  # manual price, like Adam's revisions

    inst_row = last + 1 if last > LAST_ROW else 21
    total_row = inst_row + 2
    if last > LAST_ROW:
        # install/total formulas shifted by insert_rows; retarget their ranges
        shift = last - LAST_ROW
        inst = ws["I%d" % (21 + shift)]
        if isinstance(inst.value, str) or inst.value is None:
            pass
        inst_row = 21 + shift
        total_row = 23 + shift
        inst_cell = ws["I%d" % inst_row]
        if inst_cell.value and hasattr(inst_cell.value, "text"):
            txt = inst_cell.value.text
        else:
            txt = inst_cell.value
        if isinstance(txt, str):
            new = txt.replace("B9:B20", "B9:B%d" % last).replace("F9:F20", "F9:F%d" % last).replace("N9:N20", "N9:N%d" % last)
            from openpyxl.worksheet.formula import ArrayFormula
            inst_cell.value = ArrayFormula("I%d" % inst_row, new)
    ws["I%d" % total_row] = "=SUM(I%d:I%d)+I%d" % (FIRST_ROW, last, inst_row)

    # optional extras
    mastic_row, epdm_row = total_row + 4, total_row + 5
    ws["I%d" % mastic_row] = job.get("mastic", 0) or 0
    ws["I%d" % epdm_row] = job.get("epdm", 0) or 0

    wb.save(out_path)
    return out_path


def _img64(name):
    p = os.path.join(IMAGES_DIR, name)
    if not os.path.exists(p):
        return ""
    ext = "png" if name.endswith("png") else "jpeg"
    return "data:image/%s;base64,%s" % (ext, base64.b64encode(open(p, "rb").read()).decode())


def make_proposal_html(job, out_html):
    content = json.load(open(os.path.join(REPO, "templates", "proposal-content.json"), encoding="utf-8"))
    tandc = []
    in_tc = False
    for p in content["paragraphs"]:
        if p["text"].strip() == "TERMS AND CONDITIONS":
            in_tc = True
            continue
        if in_tc and p["text"].strip():
            tandc.append(p["text"].strip())
    team = content["tables"][4]

    def esc(s):
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    photos = job.get("photos", [])
    photo_html = ""
    for i in range(3):
        if i < len(photos) and os.path.exists(photos[i]):
            b64 = base64.b64encode(open(photos[i], "rb").read()).decode()
            photo_html += '<div class="photo"><img src="data:image/jpeg;base64,%s"></div>' % b64
        else:
            photo_html += '<div class="photo placeholder">Project image %d<br>(insert site/product photo)</div>' % (i + 1)

    approach = [
        ("Technical Review &amp; Survey", "Following appointment, Fenster Glazing will review all project documentation and undertake a detailed site survey to confirm dimensions, interfaces and access requirements prior to manufacture."),
        ("Procurement &amp; Manufacture", "Materials will be procured through our established supply chain following survey completion and approval of all fabrication drawings and sign-off documentation."),
        ("Project Management", "A dedicated Project Manager and Installation Manager will be assigned to oversee the works from award through to completion, coordinating closely with your site team."),
        ("Installation", "Installation works will be undertaken by experienced glazing operatives in accordance with project-specific RAMS, programme and site rules."),
        ("Quality Assurance &amp; Handover", "Quality inspections are undertaken throughout survey, manufacture and installation in accordance with our QA procedures, concluding with sign-off documentation, warranties and O&amp;M information at handover."),
    ]

    rows_incl = job.get("inclusions", [])
    rows_excl = job.get("exclusions", [])
    incl_html = "".join("<li>%s</li>" % esc(x) for x in rows_incl)
    excl_html = "".join("<li>%s</li>" % esc(x) for x in rows_excl)
    team_html = ""
    for row in team:
        for cell in row:
            lines = [l.strip() for l in cell.split("\n") if l.strip()]
            if lines:
                team_html += '<div class="tm"><strong>%s</strong><br>%s</div>' % (esc(lines[0]), "<br>".join(esc(l) for l in lines[1:]))

    tc_html = "".join("<li>%s</li>" % esc(t) for t in tandc)

    html = """<!DOCTYPE html><html><head><meta charset="utf-8"><title>%(ref)s - Fenster Glazing Proposal</title>
<style>
@page { size: A4; margin: 14mm 13mm; }
* { box-sizing: border-box; }
body { font-family: Arial, Helvetica, sans-serif; color: #222; font-size: 10px; margin: 0; }
.header { display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #002D3A; padding-bottom: 8px; }
.header img { height: 62px; }
h1 { color: #002D3A; font-size: 19px; margin: 14px 0 4px 0; }
h2 { color: #002D3A; font-size: 13px; margin: 16px 0 5px 0; border-bottom: 1px solid #d5dde0; padding-bottom: 2px; }
.meta td { padding: 2px 12px 2px 0; font-size: 10.5px; }
.meta td:first-child { font-weight: bold; color: #002D3A; }
.photos { display: flex; gap: 8px; margin: 8px 0; }
.photo { flex: 1; height: 150px; overflow: hidden; border-radius: 4px; }
.photo img { width: 100%%; height: 100%%; object-fit: cover; }
.placeholder { background: #eef2f4; border: 1px dashed #9ab; display: flex; align-items: center; justify-content: center; text-align: center; color: #667; font-size: 9px; }
.price { background: #eaf4ee; border: 1px solid #bfd9c8; padding: 10px 14px; font-size: 13px; margin: 10px 0; }
.price strong { font-size: 16px; }
ol.approach { margin: 4px 0 0 16px; padding: 0; }
ol.approach li { margin-bottom: 6px; }
.twocol { display: flex; gap: 14px; }
.twocol > div { flex: 1; }
ul { margin: 4px 0 0 16px; padding: 0; }
li { margin-bottom: 3px; }
.team { display: flex; flex-wrap: wrap; gap: 10px; }
.tm { border: 1px solid #d5dde0; border-radius: 4px; padding: 7px 10px; min-width: 200px; font-size: 9.5px; }
ol.tc { margin: 4px 0 0 16px; padding: 0; font-size: 8.3px; color: #444; }
ol.tc li { margin-bottom: 4px; }
.pagebreak { page-break-before: always; }
.accreds img { height: 44px; margin-right: 10px; }
.footer { font-size: 8px; color: #888; margin-top: 14px; border-top: 1px solid #ddd; padding-top: 4px; }
</style></head><body>
<div class="header"><img src="%(logo)s" alt="Fenster Glazing"><div style="text-align:right;font-size:9.5px">97-98 Alston Drive, Bradwell Abbey, MK13 9HF<br>01908 429200 | www.fensterglazing.com<br>VAT: 305818213</div></div>
<h1>Thank you for choosing Fenster Glazing to support you on this project.</h1>
<table class="meta">
<tr><td>Client:</td><td>%(client)s</td></tr>
<tr><td>FAO:</td><td>%(fao)s</td></tr>
<tr><td>Project Reference:</td><td>%(ref)s</td></tr>
<tr><td>Site Address:</td><td>%(site)s</td></tr>
<tr><td>Date:</td><td>%(date)s</td></tr>
</table>
<div class="photos">%(photos)s</div>
<h2>Executive Summary</h2>
%(summary)s
<div class="price">Our price for the works described: <strong>%(price)s</strong> excluding VAT.<br>For a full breakdown of costs please refer to our relevant pricing document.</div>
<h2>Project Approach</h2>
<ol class="approach">%(approach)s</ol>
<div class="twocol">
<div><h2>Inclusions</h2><ul>%(incl)s</ul></div>
<div><h2>Exclusions</h2><ul>%(excl)s</ul></div>
</div>
<h2>Previous Projects</h2>
<p>%(previous)s Further completed projects can be found on our website: fensterglazing.com/commercial</p>
<h2>Project Delivery Team</h2>
<div class="team">%(team)s</div>
<div class="pagebreak"></div>
<h2>Terms and Conditions</h2>
<ol class="tc">%(tc)s</ol>
<div class="footer">Fenster Glazing &amp; Locks Ltd - %(ref)s - generated by Mary (Glazing Quote Assistant) - quotation valid 30 days</div>
</body></html>""" % {
        # fenster-logo.png is extracted from the MASTER PRICING DOC letterhead;
        # the cover-letter docx media images include CLIENT logos (e.g. Elkins)
        # from past jobs, so never use those for the header.
        "logo": _img64("fenster-logo.png"),
        "client": esc(job.get("client", "")),
        "fao": esc(job.get("fao", "")),
        "ref": esc(job.get("projectRef", "")),
        "site": esc(job.get("siteAddress", "")),
        "date": esc(job.get("date", "")),
        "photos": photo_html,
        "summary": "".join("<p>%s</p>" % esc(p) for p in job.get("summary", [])),
        "price": esc(job.get("priceText", "")),
        "approach": "".join("<li><strong>%s</strong><br>%s</li>" % a for a in approach),
        "incl": incl_html,
        "excl": excl_html,
        "previous": esc(job.get("previousProjects", "Recent comparable projects include Headrow Court, Leeds (replacement aluminium windows, doors and curtain wall, GBP630,000, occupied city-centre phased installation).")),
        "team": team_html,
        "tc": tc_html,
    }
    with open(out_html, "w", encoding="utf-8") as fh:
        fh.write(html)
    return out_html


if __name__ == "__main__":
    job = json.load(open(sys.argv[1], encoding="utf-8"))
    out_x = job.get("outXlsx")
    out_h = job.get("outHtml")
    if out_x:
        make_pricing_doc(job, out_x)
        print("pricing doc:", out_x)
    if out_h:
        make_proposal_html(job, out_h)
        print("proposal html:", out_h)
