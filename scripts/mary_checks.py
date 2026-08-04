# -*- coding: utf-8 -*-
"""Every mistake Mary has caught, turned into a check that runs every time.

The money is not in the rates. Look at what she has actually caught: 46 panes
of glass missing at Stoke Park, six units short on the Vesuvius RFQ, chapel
folding doors absent from the Grange Hill scope, no panic bars on fire-exit
doors at SM5, Sheerline coupled to Smart Wall. A perfect rate table would have
caught NONE of those. They are scope, quantity, compliance and system errors -
and each one is a rule you can run.

The design point: a rule that cannot find its facts returns UNKNOWN, not PASS.
Silence is how these get missed in the first place, so an unanswered question
fails the run exactly like a broken rule does. That makes this a forcing
function - Mary has to state the facts before a quote can leave.

  python scripts/mary_checks.py --new <job>      # write a blank manifest
  python scripts/mary_checks.py <manifest.json>  # run the checks
  python scripts/mary_checks.py --list           # what it checks and why

Exit code 0 only when every check passes.
"""
import argparse
import datetime as dt
import json
import os
import re
import sys
import textwrap

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MANIFEST_DIR = os.path.join(REPO, "data", "job-checks")

PASS, FAIL, UNKNOWN, NA = "PASS", "FAIL", "UNKNOWN", "n/a"

# Riverside House, 27/07. A supplier quote expiring the same day our own price
# closes clears "held as long as ours" and is still useless - the client accepts
# on the last day and there is nothing left to place the order against. Days of
# headroom below this raise a question rather than a pass.
THIN_MARGIN_DAYS = 14

# Systems and their frame depth in mm. Frames can only be coupled within the
# same depth - Adam's ruling, 24/07/2026.
SYSTEM_DEPTH = {
    "smart wall": 100, "sma smart wall": 100, "mc600": 100, "smart alitherm 600": 100,
    "sheerline": 70, "sheerline prestige": 70, "smart alitherm 400": 70,
    "technal": 70, "senior sf52": 52, "senior pure": 75, "aluprof mb-78ei": 78,
    # Redditch Library BLBS0956, 28/07. Joedan's own systems, named in Gleeds'
    # spec 3.5.3 as the minimum standard. The depth is the system designation -
    # the same convention that makes Senior SF52 52mm and Smart Alitherm 600
    # 100mm. EL75 windows coupled to AC100 doors on refs 32 and 34 is the
    # Sheerline/Smart Wall problem again, this time written into the tender.
    "el75": 75, "el75mm squareline": 75, "ac100": 100, "ac100 commercial": 100,
}


def depth_of(system):
    s = str(system or "").strip().lower()
    for name, d in SYSTEM_DEPTH.items():
        if name in s:
            return d, name
    return None, ""


# Systems that have NO window product at all, so "quote the window in the door's
# system" - the remedy this file has carried since 24/07 - cannot be done.
#
# SM5 Wexham, 29/07/2026. BSW, in writing, asked directly whether the coupled
# windows could be requoted to match the doors: "There are no compatible windows
# that can coupler to smart wall even in smarts smartwall as this is a door and
# screen product only. we do not manufacture a standalone smarts window system."
#
# That matters beyond one job. Adam's coupling ruling says move the window into
# the door's system; for Smart Wall there is nothing to move it into, and a screen
# is fixed glazing, so an opening vent cannot live there either. Grange Hill hit
# the identical coupling the same morning and would have been sent down the same
# dead end. The real remedy is to move the WHOLE run to a system that makes both,
# or to decouple the window from the door.
NO_WINDOW_SYSTEM = {
    "smart wall": "BSW 29/07/2026: 'a door and screen product only... we do not "
                  "manufacture a standalone smarts window system'",
    "sma smart wall": "BSW 29/07/2026: 'a door and screen product only... we do not "
                      "manufacture a standalone smarts window system'",
}


def no_window_system(system):
    s = str(system or "").strip().lower()
    for name, why in NO_WINDOW_SYSTEM.items():
        if name in s:
            return name, why
    return None, ""


def result(rule, status, detail, catch="", remedy=""):
    """A finding, and separately what to do about it.

    `remedy` is its own field because of a measurement riverside prompted on
    28/07/2026. The remedy used to be the last sentence of `detail`, after the
    list of offending items - so it was displaced further the more items there
    were, while the truncation that hid it was triggered by that same length.
    Across 13 manifests: 9 of 9 details over 400 characters had the remedy past
    the cut, against 3 of 35 under it. The same rule, "delivery actually
    included", showed the remedy at 0% and visible on ten one-supplier jobs and
    at 78-89% and cut on the three multi-supplier ones.

    So the sentence telling you what to do vanished exactly on the jobs where
    most was wrong. Keeping it in its own field means no future abridgement -
    a summary, a dashboard excerpt, a --brief mode - can displace it again.
    """
    return {"rule": rule, "status": status, "detail": detail,
            "from_catch": catch, "remedy": remedy}


# ---------------------------------------------------------------- the rules

def check_system_coupling(m):
    """SM5 Wexham, 24/07. Sheerline (70mm) cannot be coupled to a Smart Wall
    door (100mm) - there is no way to join them. Any window in the same
    coupled run as a door must be quoted in the door's system."""
    runs = m.get("coupled_runs")
    if runs is None:
        return result("system-depth coupling", UNKNOWN,
                      "List every run where frames are coupled together, and the system of each "
                      "element ('coupled_runs'). If nothing is coupled, say so with an empty list.",
                      "SM5 Wexham")
    if not runs:
        return result("system-depth coupling", NA, "nothing coupled on this job", "SM5 Wexham")
    bad, blocked = [], {}
    for run in runs:
        depths = {}
        for el in run.get("elements", []):
            d, name = depth_of(el.get("system"))
            if d is None:
                return result("system-depth coupling", UNKNOWN,
                              "System %r on %s is not one I know a depth for - add it to "
                              "SYSTEM_DEPTH or state it." % (el.get("system"), el.get("ref", "?")),
                              "SM5 Wexham")
            depths.setdefault(d, []).append("%s (%s)" % (el.get("ref", "?"), name))
        if len(depths) > 1:
            bad.append("%s couples %s" % (run.get("name", "a run"),
                                          " to ".join("%dmm: %s" % (d, ", ".join(v))
                                                      for d, v in sorted(depths.items()))))
            for el in run.get("elements", []):
                name, why = no_window_system(el.get("system"))
                if name and "%s on %s" % (name, run.get("name", "a run")) not in blocked:
                    blocked["%s on %s" % (name, run.get("name", "a run"))] = why
    if bad:
        remedy = ("Requote every element in the run in the same system as the door, and reprice - "
                  "the coupled system is usually the dearer one.")
        if blocked:
            remedy = ("DO NOT simply requote the windows in the door's system - it has no window "
                      "product: " + "; ".join("%s (%s)" % (k, v) for k, v in sorted(blocked.items()))
                      + ". A screen is fixed glazing, so an opening vent cannot go there either. "
                      "Either move the WHOLE run to a system that makes both windows and doors at "
                      "one depth, or separate the window from the door so nothing is coupled. Both "
                      "change the price - get it requoted, do not adjust the existing one.")
        return result("system-depth coupling", FAIL,
                      "Frames of different depths are coupled - they cannot be joined. " + "; ".join(bad),
                      "SM5 Wexham", remedy=remedy)
    return result("system-depth coupling", PASS, "every coupled run is one system depth", "SM5 Wexham")


def check_panic_hardware(m):
    """SM5 Wexham. ED.02 fire-exit doors were quoted with no panic bar, though
    the drawing ironmongery schedule required one. Adam confirmed a real error."""
    doors = m.get("doors")
    if doors is None:
        return result("fire-exit panic hardware", UNKNOWN,
                      "List the doors with a 'fire_exit' flag and whether panic hardware is priced "
                      "('doors': [{ref, fire_exit, panic_hardware_priced}]).", "SM5 Wexham")
    missing = [d.get("ref", "?") for d in doors
               if d.get("fire_exit") and not d.get("panic_hardware_priced")]
    if missing:
        return result("fire-exit panic hardware", FAIL,
                      "Fire-exit doors with no panic hardware priced: %s. The ironmongery schedule "
                      "governs, not the supplier's default." % ", ".join(missing), "SM5 Wexham")
    fe = [d for d in doors if d.get("fire_exit")]
    return result("fire-exit panic hardware", PASS if fe else NA,
                  "%d fire-exit door(s), all with panic hardware priced" % len(fe) if fe
                  else "no fire-exit doors on this job", "SM5 Wexham")


def check_glass_ownership(m):
    """Stoke Park, 27/07. Aplus supplied frames UNGLAZED, so the glass was
    Fenster's to buy - and the glass order was 46 panes short of the final
    sizes list. An unglazed frame order always transfers the glass buy."""
    frames = m.get("frame_supply")
    if frames is None:
        return result("unglazed frames need a glass order", UNKNOWN,
                      "State whether the frames are supplied glazed or unglazed "
                      "('frame_supply': 'glazed'|'unglazed'), and if unglazed, "
                      "'glass_order': {placed, panes_ordered, panes_required}.", "Stoke Park")
    if str(frames).lower() != "unglazed":
        return result("unglazed frames need a glass order", NA, "frames supplied glazed", "Stoke Park")
    g = m.get("glass_order") or {}
    if not g:
        return result("unglazed frames need a glass order", FAIL,
                      "Frames are UNGLAZED - the glass is Fenster's to buy - and no glass order is "
                      "recorded. This is exactly the Stoke Park failure.", "Stoke Park")
    req, ordered = g.get("panes_required"), g.get("panes_ordered")
    if req is None or ordered is None:
        return result("unglazed frames need a glass order", UNKNOWN,
                      "Give panes_required (from the FINAL sizes list) and panes_ordered.", "Stoke Park")
    if ordered < req:
        return result("unglazed frames need a glass order", FAIL,
                      "Glass order is %d panes short: %d ordered against %d required."
                      % (req - ordered, ordered, req), "Stoke Park",
                      remedy="Check for a systematic gap - at Stoke Park one toplight per bay was "
                             "missing on every type.")
    return result("unglazed frames need a glass order", PASS,
                  "%d panes ordered against %d required" % (ordered, req), "Stoke Park")


def check_quantities(m):
    """Vesuvius Way, 27/07. The RFQ that went out to suppliers asked for six
    fewer units than the trade bill - drawings said Qty 1 where the bill said
    4no. Reconcile drawings against the bill before anything is issued."""
    q = m.get("quantities")
    if q is None:
        return result("drawing vs bill quantities", UNKNOWN,
                      "Reconcile every item: 'quantities': [{ref, bill_qty, drawing_qty}].",
                      "Vesuvius Way")
    bad = ["%s (bill %s, drawings %s)" % (i.get("ref", "?"), i.get("bill_qty"), i.get("drawing_qty"))
           for i in q if i.get("bill_qty") != i.get("drawing_qty")]
    if bad:
        return result("drawing vs bill quantities", FAIL,
                      "Quantities disagree: %s." % "; ".join(bad), "Vesuvius Way",
                      remedy="Do not issue an RFQ until the covering email states the quantities "
                             "explicitly.")
    return result("drawing vs bill quantities", PASS,
                  "%d item(s) reconcile between bill and drawings" % len(q), "Vesuvius Way")


def check_scope_gaps(m):
    """Grange Hill, 24/07. Spec clause 3.15 - chapel folding doors, about
    GBP 10k - was not in the supplier RFQ at all. Every priced item in the
    spec must be either quoted or explicitly excluded."""
    items = m.get("spec_items")
    if items is None:
        return result("spec covered or excluded", UNKNOWN,
                      "List every priceable item in the spec and its treatment: "
                      "'spec_items': [{ref, treatment: 'priced'|'excluded'|'provisional'}].",
                      "Grange Hill")
    loose = [i.get("ref", "?") for i in items if i.get("treatment") not in
             ("priced", "excluded", "provisional")]
    if loose:
        return result("spec covered or excluded", FAIL,
                      "Spec items neither priced nor explicitly excluded: %s. A silent gap reads as "
                      "included to the client." % ", ".join(loose), "Grange Hill")
    return result("spec covered or excluded", PASS,
                  "all %d spec item(s) priced, excluded or carried as provisional" % len(items),
                  "Grange Hill")


def check_supplier_quote_currency(m):
    """Princess Beatrice, 23/07. The pack mixed a current Aplus quote with
    LAST YEAR's letter bearing the same job name. Supplier quotes are valid
    about 30 days - date-check every one."""
    quotes = m.get("supplier_quotes")
    if quotes is None:
        return result("supplier quotes in date", UNKNOWN,
                      "List them: 'supplier_quotes': [{supplier, ref, date: 'YYYY-MM-DD'}].",
                      "Princess Beatrice")
    if not quotes:
        return result("supplier quotes in date", NA, "no supplier quotes held - benchmark pricing",
                      "Princess Beatrice")
    today = dt.date.today()
    stale = []
    for q in quotes:
        try:
            d = dt.date.fromisoformat(str(q.get("date")))
        except Exception:
            return result("supplier quotes in date", UNKNOWN,
                          "Quote %s has no usable date." % q.get("ref", "?"), "Princess Beatrice")
        age = (today - d).days
        if age > 30:
            stale.append("%s %s is %d days old" % (q.get("supplier", "?"), q.get("ref", "?"), age))
    if stale:
        return result("supplier quotes in date", FAIL,
                      "Expired supplier pricing: %s. Re-confirm before relying on it." % "; ".join(stale),
                      "Princess Beatrice")
    return result("supplier quotes in date", PASS, "%d supplier quote(s), all inside 30 days" % len(quotes),
                  "Princess Beatrice")


def check_net_pricing(m):
    """Standing rule. Bellview and BSW quotes carry end discounts - using the
    gross figure overprices the job. Always use Grand Total Net."""
    quotes = m.get("supplier_quotes")
    if quotes is None:
        return result("net of supplier discount", UNKNOWN, "see supplier_quotes above", "standing rule")
    if not quotes:
        return result("net of supplier discount", NA, "no supplier quotes held", "standing rule")
    unclear = [q.get("ref", "?") for q in quotes if q.get("used_net") is None]
    if unclear:
        return result("net of supplier discount", UNKNOWN,
                      "State used_net for: %s. Gross where an end discount exists overprices the job."
                      % ", ".join(unclear), "standing rule")
    gross = [q.get("ref", "?") for q in quotes if q.get("used_net") is False]
    if gross:
        return result("net of supplier discount", FAIL,
                      "Priced off GROSS supplier figures: %s. Use Grand Total Net." % ", ".join(gross),
                      "standing rule")
    return result("net of supplier discount", PASS, "all supplier figures taken net", "standing rule")


def check_full_height_screens(m):
    """Greenfields calibration, 22/07. Fenster's own sent quote priced
    full-height stair screens as curtain walling at 850/m2 + 150/m2 labour,
    NOT as windows. Mary coded them as windows and came out 6.3% high.

    WIDENED on Grange Hill, 29/07/2026, and in the other direction. The rule read
    'not curtain walling' as 'priced as windows', so it fired on a screen priced
    off a SUPPLIER QUOTATION - and on that job the convention was the error, not
    the fix: CW put the rate at ~GBP1,000/m2 sell where BSW quoted GBP598/m2 and
    build the thing as coupled casements over a door element. That was GBP13,000
    of a GBP27,560 benchmark, logged in data/calibration.json.

    The convention is what you use when nobody has quoted it. A supplier
    quotation supersedes it, exactly as the standing rule says - supplier-backed
    is always preferred, and a benchmark is evidence, never a firm price. So
    'supplier' is a PASS and says which basis it passed on; anything else is
    still the Greenfields error."""
    screens = m.get("full_height_screens")
    if screens is None:
        return result("full-height screens as curtain walling", UNKNOWN,
                      "Any full-height / floor-to-ceiling screens? 'full_height_screens': "
                      "[{ref, priced_as: 'curtain walling'|'supplier quotation'|'window'}].",
                      "Greenfields")
    if not screens:
        return result("full-height screens as curtain walling", NA, "no full-height screens", "Greenfields")
    cw, supplied, wrong = [], [], []
    for s in screens:
        basis = str(s.get("priced_as", "")).lower()
        if "curtain" in basis:
            cw.append(s.get("ref", "?"))
        elif "supplier" in basis:
            supplied.append(s.get("ref", "?"))
        else:
            wrong.append(s.get("ref", "?"))
    if wrong:
        return result("full-height screens as curtain walling", FAIL,
                      "Full-height screens priced as windows: %s. With no supplier quotation the "
                      "house convention is curtain walling - GBP850/m2 supply + GBP150/m2 labour."
                      % ", ".join(wrong), "Greenfields",
                      remedy="Price them on the curtain-walling convention, or get the screen "
                             "quoted and price it off the quotation.")
    bits = []
    if supplied:
        bits.append("%d off a supplier quotation, which supersedes the convention" % len(supplied))
    if cw:
        bits.append("%d on the curtain-walling convention" % len(cw))
    return result("full-height screens as curtain walling", PASS,
                  "%d screen(s): %s" % (len(screens), "; ".join(bits)), "Greenfields")


# A 'fabricator' that is actually a statement that nobody can make it. The rule
# below used to test the field for truthiness alone, so "NONE APPROACHED CAN MAKE
# IT" and "their own system, not available to Fenster" both counted as a
# fabricator and passed. Vesuvius Way and Redditch Library were both sitting on
# that when it was found (29/07).
_NO_FABRICATOR = re.compile(
    r"\bnone\b|\bnobody\b|\bno[- ]one\b|\bno fabricator\b|\bno approved\b"
    r"|\bcannot\b|\bcan't\b|\bunable\b|\bnot available\b|\bnot approached\b"
    r"|\bnever approached\b|\bdoes not fabricate\b|\bdo not fabricate\b",
    re.I)


def check_fabricator_can_make_it(m):
    """Vesuvius Way, 27/07. The whole pack was Senior, and none of BSW
    (Sheerline), Aplus (Technal) or Bellview (SMA) fabricate Senior. A tender
    priced on a system nobody can make is not a tender.

    Widened 29/07, on Vesuvius again. The rule only asked whether the
    'fabricator' field was non-empty, so the honest answer - writing "NONE
    APPROACHED CAN MAKE IT" into it - PASSED the very check that exists to catch
    that. Redditch Library was passing on "their own system, not available to
    Fenster" at the same time. Naming the problem in the field is not the same as
    having a fabricator, so a denial now fails, and an explicit
    'can_make_it': false fails whatever the prose says.
    """
    systems = m.get("systems_specified")
    if systems is None:
        return result("someone can actually fabricate it", UNKNOWN,
                      "'systems_specified': [{system, fabricator}] - who is making each system? "
                      "Set 'can_make_it': false where the answer is that nobody on our supply "
                      "chain can.",
                      "Vesuvius Way")
    orphans, denied = [], []
    for s in systems:
        name = s.get("system", "?")
        fab = s.get("fabricator")
        if not fab:
            orphans.append(name)
        elif s.get("can_make_it") is False or _NO_FABRICATOR.search(str(fab)):
            denied.append("%s (recorded as: %s)" % (name, str(fab)[:90]))
    if orphans or denied:
        parts = []
        if orphans:
            parts.append("No fabricator identified for: %s." % ", ".join(orphans))
        if denied:
            parts.append("The 'fabricator' field itself says nobody can make it: %s."
                         % "; ".join(denied))
        return result("someone can actually fabricate it", FAIL,
                      " ".join(parts), "Vesuvius Way",
                      remedy="Either find an approved one or qualify an alternative system formally "
                             "in the tender. Writing the problem into the manifest is not the same "
                             "as solving it - the client reads the price, not the manifest.")
    return result("someone can actually fabricate it", PASS,
                  "every specified system has a fabricator", "Vesuvius Way")


def check_uvalue_basis(m):
    """SM5 Wexham. Mary failed a door on a whole-installation U-value of 1.6.
    Adam corrected her: that figure is an AVERAGE across the package, not a
    per-element limit. Do not reject an element against an average."""
    u = m.get("u_value")
    if u is None:
        return result("U-value read as average not per-element", UNKNOWN,
                      "'u_value': {required, basis: 'whole installation'|'per element'}.", "SM5 Wexham")
    basis = str(u.get("basis", "")).lower()
    if not basis:
        return result("U-value read as average not per-element", UNKNOWN,
                      "State whether the required U-value is a whole-installation average or a "
                      "per-element limit - they are checked completely differently.", "SM5 Wexham")
    if "whole" in basis and u.get("rejected_elements"):
        return result("U-value read as average not per-element", FAIL,
                      "Elements rejected against a whole-installation average: %s. Cold elements pass "
                      "when averaged with better ones." % ", ".join(u["rejected_elements"]), "SM5 Wexham")
    return result("U-value read as average not per-element", PASS,
                  "U-value treated as %s" % basis, "SM5 Wexham")


def check_system_performance(m):
    """St Mary's, 27/07 - and the fifth instance of this shape in a month.

    A system can be fabricable and still be incapable of the performance the
    spec demands. SM5 Wexham: SMA Smart Wall Pocket doors could not meet a
    whole-installation U-value of 1.6 because the system is not thermally
    broken. Brocks Hill: Smart Wall is not manufactured in triple glazing at
    all. In both cases the quote looked fine and the shortfall was in the
    system, not the glass - so it cannot be closed by changing the make-up.

    'systems_specified' entries may carry a 'performance' block:
        {system, fabricator,
         performance: {required, capable: true|false|null, evidence}}

    capable=false FAILS. capable=null (unknown, nobody asked the supplier)
    returns ASK, because on both founding jobs the answer existed and nobody
    had gone and got it."""
    systems = m.get("systems_specified")
    if systems is None:
        return result("system can meet the specified performance", UNKNOWN,
                      "see systems_specified above", "St Mary's / SM5 Wexham")
    rated = [s for s in systems if isinstance(s.get("performance"), dict)]
    if not rated:
        return result("system can meet the specified performance", UNKNOWN,
                      "No performance requirement recorded against any system. State it: "
                      "'performance': {required, capable, evidence}. If the pack sets no thermal, "
                      "acoustic or security performance at all, say so with an empty requirement.",
                      "St Mary's / SM5 Wexham")
    incapable, unknown = [], []
    for s in rated:
        p = s["performance"]
        if not p.get("required"):
            continue
        if p.get("capable") is False:
            incapable.append("%s cannot meet %s (%s)"
                             % (s.get("system", "?"), p.get("required"),
                                p.get("evidence") or "no evidence cited"))
        elif p.get("capable") is None:
            unknown.append("%s against %s" % (s.get("system", "?"), p.get("required")))
    if incapable:
        return result("system can meet the specified performance", FAIL,
                      "System cannot meet the specification: %s. This is a system change or a formal "
                      "qualification - it cannot be closed by changing the glass."
                      % "; ".join(incapable), "St Mary's / SM5 Wexham")
    if unknown:
        return result("system can meet the specified performance", UNKNOWN,
                      "Nobody has asked the supplier whether these can meet the requirement: %s."
                      % "; ".join(unknown), "St Mary's / SM5 Wexham",
                      remedy="Get it in writing - on both founding jobs the answer existed and no "
                             "one had gone and got it.")
    return result("system can meet the specified performance", PASS,
                  "%d system(s) confirmed capable of the specified performance" % len(rated),
                  "St Mary's / SM5 Wexham")


def _ral(s):
    m = re.search(r"ral\s*(\d{4})", str(s or "").lower())
    return m.group(1) if m else None


_SUBSTRATE = re.compile(
    r"\s*\b(?:foil\s+)?on\s+(?:white|cream|grey|gray|black|brown|anthracite)\b.*$")


def _visible_face(desc):
    """Strip the substrate off a foiled finish description.

    Gordon Court, 27/07. BSW describe a dual-colour Liniar frame as 'Grey Foil
    On White (7016)' - grey foil laminated onto a white substrate, i.e. the
    visible external face is GREY. The substring match below saw the word
    'White' inside it, decided the external face matched the white internal
    face, and reported a correctly-priced dual-colour job as single colour.

    A false FAIL is as expensive as a missed one: it teaches people to click
    past the checker. So drop the '... on <substrate>' clause and compare the
    face you can actually see."""
    return _SUBSTRATE.sub("", str(desc or "").strip().lower()).strip()


_COLOURS = ("white", "grey", "anthracite", "black", "brown", "cream", "silver",
            "green", "blue", "red", "bronze", "ivory", "beige")


def _colours(desc):
    """The colour words in a finish description, with gray normalised to grey.

    Both sides of this comparison arrive wrapped in noise. The architect writes
    'PVC-U white internally' and 'dark grey to RAL XXX (TBC)'; the supplier
    writes '(9016) White' and '7016M Anthracite Grey - M'. Nothing is a
    substring of anything, so the old raw comparison called a correctly-quoted
    dual-colour job a substitution. Compare the colours, not the packaging."""
    d = _visible_face(desc).replace("gray", "grey")
    return {c for c in _COLOURS if c in d}


def _finish_matches(spec, quote):
    """Loose match - a RAL number on both sides is decisive, then a shared
    colour word, then one description containing the other. The point is to
    catch a substitution, not to argue about wording."""
    a, b = _visible_face(spec), _visible_face(quote)
    if not a or not b:
        return None
    ra, rb = _ral(a), _ral(b)
    if ra and rb:
        return ra == rb
    ca, cb = _colours(a), _colours(b)
    if ca and cb:
        return bool(ca & cb)
    return a in b or b in a


def check_finish_substitution(m):
    """Georgie's (formerly Rosebank), 27/07. Spec 2.28 required white aluminium
    internally and dark brown externally. Mercury QL004741 quoted every one of
    the 23 windows 'BROWN RAL TBC (SINGLE COLOUR ONLY)' - the specified white
    internal face was simply not in the price, and nobody had to say so out
    loud. A supplier's default finish is not the specified finish."""
    fins = m.get("finishes")
    if fins is None:
        return result("finish quoted is the finish specified", UNKNOWN,
                      "State the finish on both sides for every element: 'finishes': "
                      "[{ref, specified_internal, specified_external, quoted_internal, "
                      "quoted_external}]. Dual colour is a cost - it is never a default.",
                      "Georgie's")
    if not fins:
        return result("finish quoted is the finish specified", NA, "no finishes to check", "Georgie's")
    silent, single, wrong = [], [], []
    for f in fins:
        ref = f.get("ref", "?")
        si, se = f.get("specified_internal"), f.get("specified_external")
        qi, qe = f.get("quoted_internal"), f.get("quoted_external")
        if not qi or not qe:
            silent.append(ref)
            continue
        dual_specified = _finish_matches(si, se) is False
        if dual_specified and _finish_matches(qi, qe):
            single.append("%s (spec %s / %s, quoted %s both sides)" % (ref, si, se, qi))
            continue
        for side, s, q in (("internal", si, qi), ("external", se, qe)):
            if _finish_matches(s, q) is False:
                wrong.append("%s %s (spec %s, quoted %s)" % (ref, side, s, q))
    if single:
        return result("finish quoted is the finish specified", FAIL,
                      "Dual colour specified, single colour quoted: %s. The second colour is not in "
                      "the price." % "; ".join(single), "Georgie's")
    if wrong:
        return result("finish quoted is the finish specified", FAIL,
                      "Finish quoted does not match the finish specified: %s." % "; ".join(wrong),
                      "Georgie's")
    if silent:
        return result("finish quoted is the finish specified", UNKNOWN,
                      "Supplier states no finish for: %s. An unstated finish is the supplier's "
                      "standard, not yours." % ", ".join(silent), "Georgie's")
    return result("finish quoted is the finish specified", PASS,
                  "%d element(s) quoted in the specified finish" % len(fins), "Georgie's")


def check_supplier_covers_quantity(m):
    """Brocks Hill Phase 2, 27/07. The tender sold 2no Door Type E.04 at
    GBP 2,723.49 each. Bellview 0000000503 quoted ONE. The rate was simply
    applied twice, so the arithmetic looked perfect and GBP 2,723.49 of cost
    had no quote behind it. Reconciling a quote TOTAL is not the same as
    reconciling its QUANTITIES - the total ties either way.

    WHAT `qty_quoted` MEANS, because Gordon Court found on 28/07 that it had been
    carrying two different facts on two jobs. It is HOW MANY OF THAT QUOTATION'S
    UNITS THIS LINE USES - an allocation. It is NOT "how many the quotation
    contains for this reference"; that belongs in `qty_total` on the quote. Both
    jobs filled it with the wrong one, in opposite directions: Riverside credited
    every line with the quotation's whole quantity, and Gordon Court recorded
    what they SELL, which hid a surplus.

    The surplus arm below deliberately does not depend on which reading was used
    - it compares `qty_total` against the sum of `qty_sold`, which is the same
    question asked in a way that cannot be answered two ways.

    HOW TO COUNT `qty_total`, because Gordon Court filled it with the wrong fact
    within an hour of it being created and were right that the ambiguity moved
    up a level rather than going away. "What the quotation contains" is not one
    number: a door and its sidelight are ONE unit to a schedule, TWO to a
    factory and ONE to a delivery note, and all three answers are correct to
    different questions.

    COUNT SELLABLE UNITS - what you would sell the client - and the two traps
    are opposite ways round on the two quotations we hold:

      A Plus put a MULTIPLIER on one block: "Qty (2) O/A Sizes 1130mm x 1530mm".
      Counting blocks gives 1; the answer is 2. EXPAND the multiplier.

      BSW put one line per ELEMENT: "Qty: 1 Prestige Casement, Location D_E" and
      "Qty: 1 Prestige Open Out Door, Location D_E", joined by a "Std Coupler"
      line. Counting Qty: lines gives 14; the answer is 12. COLLAPSE coupled
      assemblies - the coupler line is the proof they are one unit.

    AND THE TEST IS STRUCTURAL, NOT LEXICAL. An earlier version of this note
    said "if a quotation shows a coupler, a screen, a sidelight or a mullion
    between two priced elements". Gordon Court found `screen` false-positive on
    "Outer: 80113 2 Rail Patio Screen" - a product name for a sliding leaf - and
    the same list run against QT51518 fires three more times, every one of them
    wrong: `screen` on a boilerplate note about curtain wall screens, `mullion`
    on a BS 6399 calculation note and on a cable-routing note, and `mull` on
    "Transom DF1421 Std Flat Tran/Mull", which is a PROFILE NAME. Three of the
    four keywords are unsafe and none of the hits on that quotation is a
    coupling. A keyword cannot establish a coupling.

    THE TEST: TWO OR MORE PRICED ELEMENTS CARRYING THE SAME LOCATION REFERENCE
    are candidates for one sellable unit. Gordon Court's real evidence was
    "Location: D_E" on two priced blocks; the coupler line only corroborated it.
    Confirm from the specification - a coupler line, a shared outerframe, one
    actuator - and never from a word alone. Where a location appears on several
    blocks at DIFFERENT SIZES, as their D_B does, they are separate positions."""
    cov = m.get("supplier_coverage")
    if cov is None:
        return result("supplier quote covers every unit sold", UNKNOWN,
                      "For every priced line, state what the supplier actually quoted: "
                      "'supplier_coverage': [{ref, qty_sold, qty_quoted, supplier_ref}].",
                      "Brocks Hill")
    if not cov:
        return result("supplier quote covers every unit sold", NA,
                      "no supplier-backed lines to check", "Brocks Hill")
    short, silent = [], []
    claimed = {}
    for c in cov:
        ref, sold, quoted = c.get("ref", "?"), c.get("qty_sold"), c.get("qty_quoted")
        if sold is None or quoted is None:
            silent.append(ref)
        elif quoted < sold:
            short.append("%s: selling %s, %s quoted %s"
                         % (ref, sold, c.get("supplier_ref", "the supplier"), quoted))
        else:
            key = c.get("supplier_ref")
            if key:
                claimed.setdefault(key, []).append((ref, quoted))
    # Riverside, 28/07. The Brocks Hill case is under-coverage - 2 sold, 1
    # quoted, GBP 2,723.49 with no quote behind it. This is the same money
    # problem from the other side: two lines each crediting the SAME quoted
    # units, so one of them is uncovered while the arithmetic still ties. It
    # was live here - both vents claimed qty_quoted 2 from a quotation whose
    # single position reads "Qty (2)", asserting four units against two sold -
    # and the rule passed, because it only ever asked whether quoted < sold.
    # Checked only where over-claim is possible: one supplier reference
    # credited on more than one line.
    # Match by whether the quotation's REFERENCE appears inside the coverage
    # entry's supplier_ref, rather than by reconstructing a composite key. The
    # first version built keys from ref, "supplier ref" and "firstword ref",
    # and matched none of them: coverage said "A Plus QT51518" and the quote
    # said supplier "A Plus Windows & Doors", ref "QT51518". So the rule
    # reported that nothing recorded the quantity when something did - a false
    # ASK, from assuming a string shape without printing the two strings. It
    # died the instant they were printed side by side, which is the whole of
    # Gordon Court's lesson.
    quote_totals = [(str(q.get("ref", "")).strip(), q.get("qty_total"))
                    for q in (m.get("supplier_quotes") or [])
                    if isinstance(q, dict) and q.get("qty_total") is not None
                    and str(q.get("ref", "")).strip()]

    def total_for(supplier_ref):
        s = str(supplier_ref).strip().lower()
        for ref, tot in quote_totals:
            if ref.lower() in s:
                return tot
        return None
    over, unbounded = [], []
    for key, lines in claimed.items():
        if len(lines) < 2:
            continue
        total = total_for(key)
        asked = sum(n for _, n in lines)
        if total is None:
            unbounded.append("%s is credited on %d lines (%s) with no qty_total recorded for it"
                             % (key, len(lines), ", ".join("%s x%s" % (r, n) for r, n in lines)))
        elif asked > total:
            try:
                over.append("%s: %d line(s) claim %s units between them but the quotation covers "
                            "%s" % (key, len(lines), asked, total))
            except Exception:
                pass
    if over:
        return result("supplier quote covers every unit sold", FAIL,
                      "The same quoted units are credited to more than one line, so at least one "
                      "line is not actually covered: " + "; ".join(over)
                      + ". The arithmetic ties either way - that is what makes it quiet.",
                      "Brocks Hill",
                      remedy="Split the quoted quantity across the lines it actually covers, or "
                             "get the missing units quoted.")
    # Gordon Court, 28/07 - the mirror of the over-claim above, and the third
    # state neither version of this rule reported. BSW quote two WE_14 and the
    # schedule has one, so GBP 921.29 of quoted cost had nothing sold against it
    # and sat inside the quotation total their workbook takes as cost. The
    # comparison below is deliberately independent of how `qty_quoted` was
    # read: what the quotation CONTAINS, against what is SOLD against it.
    sold_per_quote = {}
    for c in cov:
        key, sold = c.get("supplier_ref"), c.get("qty_sold")
        if key and isinstance(sold, (int, float)):
            sold_per_quote.setdefault(str(key).strip(), []).append(sold)
    surplus = []
    for key, sold_list in sold_per_quote.items():
        total = total_for(key)
        if total is None:
            continue
        try:
            gap = float(total) - float(sum(sold_list))
        except (TypeError, ValueError):
            continue
        if gap > 0:
            surplus.append("%s contains %s unit(s) and only %s are sold against it - %s quoted "
                           "unit(s) with nothing sold behind them"
                           % (key, total, sum(sold_list), int(gap) if gap == int(gap) else gap))
    if surplus and not short and not silent and not over:
        return result("supplier quote covers every unit sold", UNKNOWN,
                      "A supplier quotation contains more units than this job sells against it: "
                      + "; ".join(surplus)
                      + ". That is often right - a supplier prices the whole schedule, or scope "
                        "was cut after the enquiry. It becomes money only where the build-up "
                        "takes the quotation's TOTAL rather than its lines.",
                      "Brocks Hill",
                      remedy="First re-count qty_total in SELLABLE UNITS - two priced elements "
                             "carrying the same LOCATION reference are usually one unit, and a "
                             "Qty multiplier on a block is several. If the count holds, check how "
                             "the cost was taken: "
                             "where the build-up uses the quotation total, the surplus units are "
                             "in your cost with nothing sold against them, so ask the supplier "
                             "what they picked up that you did not.")
    if unbounded and not short and not silent:
        return result("supplier quote covers every unit sold", UNKNOWN,
                      "One supplier quotation is credited on several lines and nothing records "
                      "how many units it actually contains: " + "; ".join(unbounded)
                      + ". Without that, double-counting cannot be ruled out.",
                      "Brocks Hill",
                      remedy="Add 'qty_total' to that entry in 'supplier_quotes', counting "
                             "SELLABLE UNITS off the quotation - expand any Qty multiplier on a "
                             "position block, and collapse two or more priced elements carrying "
                             "the SAME LOCATION reference into the single unit they are, "
                             "confirming from the specification rather than from a keyword.")
    if short:
        return result("supplier quote covers every unit sold", FAIL,
                      "Units sold with no supplier quote behind them: %s. Extend the quote before "
                      "the order - the rate usually holds, but nobody has agreed it."
                      % "; ".join(short), "Brocks Hill")
    if silent:
        return result("supplier quote covers every unit sold", UNKNOWN,
                      "Quantities not stated for: %s." % ", ".join(silent), "Brocks Hill")
    return result("supplier quote covers every unit sold", PASS,
                  "%d line(s) fully covered by a supplier quote" % len(cov), "Brocks Hill")


def check_quote_validity_against_commitment(m):
    """Gordon Court, 27/07 - the third instance in one day, and the worst.

    check_supplier_quote_currency asks whether a quote is in date TODAY. That is
    the wrong horizon. What matters is how long OUR price has to stay open,
    because that is the period we are exposed for.

    Gordon Court: jLiving's Form of Tender says 'This tender remains open for
    consideration for a period of 180 days from the date of receipt of tenders'
    - receipt 22/07/2026, so our GBP 368,376.70 is committed to 18/01/2027. Both
    supplier quotes behind it run 30 days and reach the end of their stated
    validity in early August. GBP 201,304.36 of cost, 55% of the tender, is unfixed
    for 163 days against a firm lump sum executed as a deed under NEC3 Option A.

    WORDING FIXED 28/07/2026 after riverside found their own quotation says the
    price is "open for acceptance for 30 days AND THEREAFTER IS SUBJECT TO
    CONFIRMATION" - not that it lapses. Checked here too: BSW's four quotes say
    only "THIS QUOTATION IS ONLY VALID FOR THIRTY DAYS" with zero occurrences of
    lapse, expire, thereafter or withdraw; AFS say "Quotations are valid for 30
    days" and their five "expiry" references are all about expiry of the CONTRACT.
    So "lapses" was never any supplier's word - it was ours, and it is the harder
    word. This rule now reports the end of a STATED VALIDITY PERIOD, which is what
    the documents actually say, and leaves the consequence unasserted.

    Same shape twice more the same afternoon: John North Hall's ITT demands 90
    days because a Section 20 leasehold consultation takes months, and St Mary's
    reached it from the other side - quote validity against the CONTRACT START
    date, not the tender return date. Three jobs, one rule.

    Compare the end of each supplier quote's stated validity against the date our
    own price stops being open. A quote whose validity ends first is a repricing
    risk we own.

    SECOND WORDING FIX, 28/07/2026, same cause as the first. This rule said "a
    price we cannot withdraw". jLiving's Form of Tender says only "This tender
    remains open for consideration for a period of 180 days from the date of
    receipt of tenders" - and contains zero instances of withdraw, revoke,
    irrevocable, binding, cannot or may not. "Cannot withdraw" was ours, it is a
    stronger legal claim than the source makes, and our OWN terms carry a 30-day
    quotation validity that pulls the other way (Gordon Court 4N). The exposure the
    rule reports is real either way; what it must not do is settle a question our
    own two documents disagree about."""
    quotes = m.get("supplier_quotes")
    pc = m.get("price_commitment")
    if pc is None:
        return result("supplier price held as long as ours", UNKNOWN,
                      "How long must OUR price stay open? 'price_commitment': "
                      "{source, our_price_open_until: 'YYYY-MM-DD'}. Read the Form of Tender / ITT "
                      "validity clause and the contract start date - not the tender return date.",
                      "Gordon Court")
    if quotes is None:
        return result("supplier price held as long as ours", UNKNOWN,
                      "see supplier_quotes above", "Gordon Court")
    if not quotes:
        return result("supplier price held as long as ours", NA,
                      "no supplier quotes held - benchmark pricing", "Gordon Court")
    try:
        until = dt.date.fromisoformat(str(pc.get("our_price_open_until")))
    except Exception:
        return result("supplier price held as long as ours", UNKNOWN,
                      "price_commitment.our_price_open_until is not a usable date.", "Gordon Court")
    silent, gaps, exposed = [], [], 0.0
    for q in quotes:
        ref = "%s %s" % (q.get("supplier", "?"), q.get("ref", "?"))
        try:
            vu = dt.date.fromisoformat(str(q.get("valid_until")))
        except Exception:
            silent.append(ref)
            continue
        if vu < until:
            gap = (until - vu).days
            val = q.get("value")
            if isinstance(val, (int, float)):
                exposed += val
            gaps.append("%s validity ends %s, %d days before our price closes on %s%s"
                        % (ref, vu.isoformat(), gap, until.isoformat(),
                           "" if not isinstance(val, (int, float)) else " (GBP %s at risk)" % format(val, ",.2f")))
    if gaps:
        return result("supplier price held as long as ours", FAIL,
                      "Supplier validity ends inside our own commitment: %s. Total GBP %s of cost "
                      "unfixed against a price we have said stays open."
                      % ("; ".join(gaps), format(exposed, ",.2f")), "Gordon Court",
                      remedy="Get a written price hold to %s or carry a stated allowance for the gap."
                             % until.isoformat())
    if silent:
        return result("supplier price held as long as ours", UNKNOWN,
                      "No validity period stated for: %s. A quote with no stated validity is not a held "
                      "price." % ", ".join(silent), "Gordon Court")
    # Riverside House, 27/07. A quote that dies on the SAME DAY our price closes
    # technically passes the test above and is still no use: it leaves the client
    # no time to accept and us no time to order. Gordon Court failed by 163 days;
    # Riverside passes by zero, which is not the same as being covered.
    thin = []
    for q in quotes:
        try:
            vu = dt.date.fromisoformat(str(q.get("valid_until")))
        except Exception:
            continue
        margin = (vu - until).days
        if margin < THIN_MARGIN_DAYS:
            thin.append("%s %s validity ends %s, only %d day(s) after our price closes on %s"
                        % (q.get("supplier", "?"), q.get("ref", "?"), vu.isoformat(),
                           margin, until.isoformat()))
    if thin:
        return result("supplier price held as long as ours", UNKNOWN,
                      "Covered, but with no headroom: %s. Acceptance on the last day of our validity "
                      "leaves nothing to place the order against." % "; ".join(thin), "Riverside House",
                      remedy="Confirm the supplier price at the point of issue, or carry a stated "
                             "allowance.")
    return result("supplier price held as long as ours", PASS,
                  "%d supplier quote(s) held to at least %s" % (len(quotes), until.isoformat()),
                  "Gordon Court")


def check_free_delivery_threshold(m):
    """Riverside House, 27/07. A Plus QT51518 says 'Glazed /Supply Only
    (Delivered)' on its face. Their terms say something narrower: 'All orders
    are priced as Ex-Works', 'Loads over GBP 5000 + VAT will be delivered FOC
    within a 50-mile radius of Watford', and loads under GBP 5000 are batched
    or charged at GBP 1/mile each way. The Riverside order is GBP 4,845.22 -
    GBP 154.78 UNDER the threshold - so the word 'Delivered' on the quote does
    not mean delivery is in the price.

    Same shape as AFS on Gordon Court, whose Specifics page read 'Logistics:
    Delivered' while delivery sat in a priced extras block and T&C 8.1 put
    packaging, insurance and transport on the customer 'IN ADDITION'.

    'delivery_terms': [{supplier, ref, order_value, free_delivery_threshold,
                        charge_basis, delivery_priced}]"""
    terms = m.get("delivery_terms")
    if terms is None:
        return result("delivery actually included", UNKNOWN,
                      "State the delivery basis for every supplier: 'delivery_terms': "
                      "[{supplier, ref, order_value, free_delivery_threshold, charge_basis, "
                      "delivery_priced}]. A quote that says 'Delivered' on its face can still put "
                      "carriage on us in its terms. If a supplier genuinely carries delivery "
                      "unconditionally, say so with free_delivery_threshold 0.", "Riverside House")
    if not terms:
        return result("delivery actually included", NA, "no delivered supplier orders on this job",
                      "Riverside House")
    short, silent, prov = [], [], []
    for t in terms:
        ref = "%s %s" % (t.get("supplier", "?"), t.get("ref", "?"))
        val, thr = t.get("order_value"), t.get("free_delivery_threshold")
        # Gordon Court, 27/07 evening. Riverside's rule could say "always free"
        # (threshold 0) but had no way to say "NEVER free", which is the more
        # common case and the one on this job: AFS price delivery as a GBP 250
        # extra with no threshold at all, and every BSW quote is flatly "ex
        # works, additional delivery charges may apply". Leaving the threshold
        # null made both read as an unanswered question when AFS's omission is
        # a known, quantified GBP 250 hole. 'never' says so.
        never_free = str(thr).strip().lower() == "never"
        # Riverside, 28/07. Adversarially tested this rule against 12 variants
        # after Gordon Court's point that a detector validated on one positive
        # case has measured precision, not recall. Two real defects fell out.
        # (1) A numeric field written as a STRING - "5000" - crashed the whole
        # run on the >= comparison, aborting every later rule. That became more
        # likely the moment the field legitimately accepted "never", because a
        # reader who sees one string reasonably writes another. Coerce, and ASK
        # rather than crash on anything that is neither a number nor "never".
        if not never_free:
            try:
                val = None if val is None else float(val)
                thr = None if thr is None else float(thr)
            except (TypeError, ValueError):
                silent.append("%s (order_value %r / threshold %r is not a number)"
                              % (ref, t.get("order_value"), t.get("free_delivery_threshold")))
                continue
        else:
            try:
                val = None if val is None else float(val)
            except (TypeError, ValueError):
                silent.append("%s (order_value %r is not a number)" % (ref, t.get("order_value")))
                continue
        if val is None or (thr is None and not never_free):
            silent.append(ref)
            continue
        if not never_free and val >= thr:
            continue
        priced = t.get("delivery_priced")
        if never_free:
            gap = ("%s: the supplier never carries delivery free - no threshold exists, carriage is "
                   "chargeable on the whole GBP %s" % (ref, format(val, ",.2f")))
        else:
            gap = "%s: order GBP %s is GBP %s below the supplier's GBP %s free-delivery threshold" % (
                ref, format(val, ",.2f"), format(thr - val, ",.2f"), format(thr, ",.2f"))
        if priced is True:
            continue
        # A carriage cost the supplier makes CONTINGENT (A Plus batch sub-GBP5k
        # loads and only charge where batching fails) cannot be priced to the
        # penny by us. Identified-and-pending is a question, not a silent
        # omission - but it must never read as covered.
        if str(priced).lower() == "provisional" and t.get("charge_basis"):
            prov.append("%s; carried as provisional on the supplier's stated basis (%s)"
                        % (gap, t["charge_basis"]))
        # (2) An unrecognised value silently read as "not priced". delivery_priced
        # "yes" produced "Delivery is not in the price" - an assertion about the
        # world, from a value the rule simply did not understand. Misreading an
        # affirmative as a negative is the dangerous direction, so say so instead.
        elif priced not in (None, False) and str(priced).lower() not in ("false", "no", "0", "provisional"):
            silent.append("%s (delivery_priced is %r - use true, false, or \"provisional\")" % (ref, priced))
        else:
            short.append("%s and no carriage is priced (%s)"
                         % (gap, t.get("charge_basis") or "basis not stated"))
    if short:
        return result("delivery actually included", FAIL,
                      "Delivery is not in the price: %s." % "; ".join(short),
                      "Riverside House",
                      remedy="Either price the carriage or get the supplier to confirm the load is "
                             "being batched free.")
    if prov:
        return result("delivery actually included", UNKNOWN,
                      "Carriage identified but not yet fixed: %s." % "; ".join(prov),
                      "Riverside House",
                      remedy="Get the supplier to confirm the charge or that the load is batched "
                             "free before the price is issued.")
    if silent:
        return result("delivery actually included", UNKNOWN,
                      "Order value or free-delivery threshold not stated for: %s." % ", ".join(silent),
                      "Riverside House")
    return result("delivery actually included", PASS,
                  "%d supplier order(s) clear their delivery threshold or carry priced carriage" % len(terms),
                  "Riverside House")


# Widened 28/07 after riverside's sampling lesson. The first version was validated
# against ONE positive case - the founding one - so "0 false positives across 119
# spec items" measured precision and said nothing about recall. Tested against nine
# plausible ways of writing the same contradiction it caught five. These four were
# missed: "still to do", "never checked", "awaiting", "no answer yet".
_LABEL_STALE = re.compile(
    r"\b(NOT RUN|not run|outstanding|TBC|not yet done|still to do|to be (?:run|checked|done|confirmed)"
    r"|never (?:run|checked|done|asked)|awaiting|unanswered|not asked|no answer(?: yet)?|pending)\b",
    re.I)
_LABEL_DONE = re.compile(r"\b(RUN|DONE|resolved|CLEARED|answered|withdrawn|CORRECTED|confirmed|closed)\b")


# Riverside, 28/07, at Gordon Court's request and from their own near miss.
# They wrote "BSW terms and conditions of sale, available on request - no
# revision, no date, no title" into `document` - a careful, accurate,
# human-readable description of the fact that the document has no name - and the
# rule read it as a name. The field whose EMPTINESS was the signal had been
# filled with prose describing the emptiness. This is the most likely way a
# conscientious estimator defeats that branch, and it happened within an hour of
# the branch shipping.
ABSENCE_WORDS = re.compile(
    r"available on request|on request|not named|unnamed|no (?:revision|date|title|name|version)"
    r"|unnamed|unknown|not (?:stated|given|specified|provided)|\bn/a\b|none stated|no document"
    r"|tbc|unspecified|not held|unable to",
    re.I)


def _describes_absence(doc):
    """True when a 'document' value describes the absence of a name rather than
    being one. Deliberately narrow: it must contain one of a short list of
    phrases that only ever appear when someone is explaining that there is no
    title. A real document name - "Terms of Sale Revision V.01.2 - 08.01.2018" -
    matches none of them."""
    return bool(ABSENCE_WORDS.search(str(doc)))


# Riverside, 28/07, from Gordon Court's docProps finding. Kept deliberately
# narrow - an email address, a Windows or Mac user path, and the two folder
# names that only ever appear in an Outlook attachment cache. Anything looser
# fires on ordinary prose, which is the fault this week's noticeboard has been
# full of.
# Gordon Court, 28/07: this reported `ff@C.0` on their proposal PDF, which is
# bytes out of a compressed stream rather than an address. Riverside's printable
# guard does not cover it - every character in it is printable. So the address
# arm now requires a domain label of two or more characters and an ALPHABETIC
# TLD of two or more. Checked against every real address on both jobs:
# dan.parker@agsurveying.co.uk, hayley@hdplanning.co.uk, drawingoffice@aol.com,
# adam@fensterglazing.com, estimating@aplusaluminium.co.uk - all still match.
THIRD_PARTY_TRACE = re.compile(
    rb"[\w.+-]+@[\w-]{2,}(?:\.[\w-]{2,})*\.[A-Za-z]{2,}"
    rb"|C:\\+Users\\+[^\\\"<>]+"
    rb"|/Users/[^/\"<>]+"
    rb"|INetCache|Content\.Outlook", re.I)

# Parts of an OOXML package that are text stores rather than content. The
# founding case lived in the first of these, and last night's external-link
# clean missed it because it dropped xl/externalLinks/ and nothing else.
METADATA_PARTS = ("docProps/", "externalLink")


def scan_file_for_traces(path, allow=()):
    """Return the traces a file would carry to whoever opens it.

    Reads the raw bytes of every part of an OOXML package, and the whole file
    for anything else. `allow` is a tuple of substrings that are legitimately
    ours - our own domain, for instance.

    Returns (list of (part, trace), error string or None). A file that cannot
    be read returns an error rather than an empty list, because "no traces
    found" and "could not look" must never render the same.
    """
    import os
    import zipfile
    if not os.path.exists(path):
        return [], "file not found"
    hits = []
    try:
        if zipfile.is_zipfile(path):
            z = zipfile.ZipFile(path)
            for n in z.namelist():
                try:
                    raw = z.read(n)
                except Exception:
                    continue
                for m in set(THIRD_PARTY_TRACE.findall(raw)):
                    s = m.decode("utf-8", "ignore")
                    if any(a.lower() in s.lower() for a in allow):
                        continue
                    hits.append((n, s[:90]))
        else:
            with open(path, "rb") as fh:
                raw = fh.read()
            # A PDF is compression, not text. Tightening the pattern narrows the
            # odds of a false hit; reading the extracted text instead removes
            # the class of error. Only fall back to the bytes if the text cannot
            # be had - and if neither works, say so rather than return clean.
            if raw[:5] == b"%PDF-":
                try:
                    import pypdf
                    rd = pypdf.PdfReader(path)
                    raw = "".join((pg.extract_text() or "") for pg in rd.pages).encode(
                        "utf-8", "ignore")
                except Exception as exc:
                    return [], ("PDF text could not be extracted (%s: %s) - not scanned is "
                                "not the same as clean" % (type(exc).__name__, exc))
            # A compressed or binary file decoded as bytes throws up matches
            # that are not text at all - the Riverside drawings PDF produced
            # six "email addresses" out of 14 FlateDecode streams. Only accept
            # a trace that is printable.
            for m in set(THIRD_PARTY_TRACE.findall(raw)):
                s = m.decode("utf-8", "ignore")
                if not s or not all(32 <= ord(c) < 127 for c in s):
                    continue
                if any(a.lower() in s.lower() for a in allow):
                    continue
                hits.append((os.path.basename(path), s[:90]))
    except Exception as exc:
        return [], "%s: %s" % (type(exc).__name__, exc)
    return hits, None


def check_priced_document_view_is_intact(m):
    """Riverside House, 28/07, after two chats destroyed the same protection
    with the same line of code within an hour of each other.

    MASTER PRICING DOC.xlsx puts the supplier buy in columns J, K and L and sets
    a print area that stops at column I, so a printed or PDF'd quotation never
    carries it. That protection is stored as a defined name, `_xlnm.Print_Area`,
    in the same block as 50 foreign names inherited from an electrical
    template - and a regex that removes the block removes it too. So does
    `_xlnm.Print_Titles`, the repeating header rows, which Riverside missed on
    the first restore and Gordon Court found.

    Checks the document that actually goes to the client:
      - a print area exists;
      - NOTHING is populated outside it;
      - the repeating header rows survived.

    The middle one carries the weight. Gordon Court: a print area protects a
    print of one file and does nothing if the workbook is emailed; a second
    sell-only file protects the workbook and does nothing if somebody attaches
    the wrong one. Populated cells outside the print area mean you are covered
    against one failure mode of the two.

    Uses 'issued_documents': [{name, path, is_the_priced_document}]."""
    docs = m.get("issued_documents")
    if docs is None:
        return result("the client's view of the priced workbook", UNKNOWN,
                      "State the priced document and its path: 'issued_documents': "
                      "[{name, path, is_the_priced_document}].", "Riverside House",
                      remedy="Add a path to the priced document so it can be opened.")
    if isinstance(docs, dict):
        docs = [docs]
    if isinstance(docs, str) or not isinstance(docs, (list, tuple)):
        return result("the client's view of the priced workbook", UNKNOWN,
                      "'issued_documents' is %r - it must be a list." % (docs,),
                      "Riverside House", remedy="Rewrite the field as a list.")
    books = [d for d in docs if isinstance(d, dict) and d.get("is_the_priced_document")
             and client_facing(d)
             and str(d.get("path", "")).lower().endswith((".xlsx", ".xlsm"))]
    if not books:
        return result("the client's view of the priced workbook", NA,
                      "no priced workbook on this job - nothing to hide behind a print area",
                      "Riverside House")
    bad, unreadable = [], []
    for doc in books:
        name, path = doc.get("name", "?"), doc.get("path")
        try:
            import openpyxl
            ws = openpyxl.load_workbook(path).active
        except Exception as exc:
            unreadable.append("%s could not be opened (%s: %s)"
                              % (name, type(exc).__name__, exc))
            continue
        area = ws.print_area
        if not area:
            bad.append("%s has NO print area - the whole sheet prints, including any working "
                       "columns" % name)
            continue
        try:
            from openpyxl.utils import range_boundaries
            ref = area[0] if isinstance(area, (list, tuple)) else area
            ref = str(ref).split("!")[-1]
            c1, r1, c2, r2 = range_boundaries(ref)
        except Exception as exc:
            unreadable.append("%s has a print area (%r) this rule cannot parse (%s)"
                              % (name, area, type(exc).__name__))
            continue
        outside = []
        for row in ws.iter_rows():
            for c in row:
                if c.value in (None, ""):
                    continue
                if c.column < c1 or c.column > c2 or c.row < r1 or c.row > r2:
                    outside.append(c.coordinate)
        if outside:
            bad.append("%s has %d populated cell(s) OUTSIDE its print area %s (%s%s) - a print "
                       "area protects a print, not the file"
                       % (name, len(outside), ref, ", ".join(sorted(outside)[:6]),
                          "..." if len(outside) > 6 else ""))
        if not ws.print_title_rows:
            bad.append("%s has no repeating header rows (_xlnm.Print_Titles) - the header will "
                       "appear on page 1 only" % name)
    if unreadable:
        return result("the client's view of the priced workbook", UNKNOWN,
                      "; ".join(unreadable) + ". Not checked is not the same as intact.",
                      "Riverside House", remedy="Fix the path, then re-run.")
    if bad:
        return result("the client's view of the priced workbook", FAIL,
                      "; ".join(bad) + ".", "Riverside House",
                      remedy="Restore the print area and print titles - both are defined names "
                             "and a wholesale definedNames delete takes them - and issue a "
                             "sell-only copy with the working columns REMOVED rather than merely "
                             "outside the printed range.")
    return result("the client's view of the priced workbook", PASS,
                  "%d priced workbook(s): print area set, print titles intact, nothing populated "
                  "outside the printed range" % len(books), "Riverside House")


def check_no_third_party_traces_in_issued_files(m):
    """Riverside House, 28/07, from Gordon Court's finding on an already-issued
    document.

    `dc:creator` on their pricing document read "Dan Parker;
    dan.parker@agsurveying.co.uk" - a named person at another company, with his
    work email, recorded as the author of a quotation that went to a client. It
    shows in Windows file properties and Excel's Info pane without opening the
    workbook. Both jobs inherited it from MASTER PRICING DOC.xlsx, created
    2018-12-07.

    This rule opens the files rather than reading a manifest flag, because the
    entire point is that nobody knew the traces were there to declare. It also
    distinguishes "scanned and clean" from "could not be scanned" - a file that
    cannot be opened must never report the same as one that is clean.

    Uses 'issued_documents': [{name, path, is_the_priced_document}] - path
    relative to the repo root."""
    docs = m.get("issued_documents")
    if docs is None:
        return result("no third-party traces in issued files", UNKNOWN,
                      "State the documents that go to the client and where they are: "
                      "'issued_documents': [{name, path, is_the_priced_document}].",
                      "Riverside House",
                      remedy="Add a 'path' to each issued document so it can be opened and read.")
    if isinstance(docs, dict):
        docs = [docs]
    if isinstance(docs, str) or not isinstance(docs, (list, tuple)):
        return result("no third-party traces in issued files", UNKNOWN,
                      "'issued_documents' is %r - it must be a list." % (docs,),
                      "Riverside House", remedy="Rewrite the field as a list.")
    if not docs:
        return result("no third-party traces in issued files", NA,
                      "no issued documents on this job", "Riverside House")
    allow = tuple(m.get("own_domains") or ("fensterglazing.com",))
    dirty, unreadable, scanned = [], [], 0
    for doc in docs:
        if not isinstance(doc, dict):
            unreadable.append("%r is not a document entry" % (doc,))
            continue
        if not client_facing(doc):
            continue
        path = doc.get("path")
        name = doc.get("name", path or "?")
        if not path:
            unreadable.append("%s has no path, so it cannot be opened" % name)
            continue
        hits, err = scan_file_for_traces(path, allow)
        if err:
            unreadable.append("%s could not be read (%s)" % (name, err))
            continue
        scanned += 1
        for part, trace in hits:
            dirty.append("%s carries %r in %s" % (name, trace, part))
    if dirty:
        return result("no third-party traces in issued files", FAIL,
                      "A document that would go to the client carries somebody else's name, "
                      "email or file path: " + "; ".join(dirty)
                      + ". This is visible in file properties without opening the document.",
                      "Riverside House",
                      remedy="Rewrite docProps and strip external links on a COPY where the file "
                             "has already been issued - the issued file is the record of what the "
                             "client received - and in place where it has not.")
    if unreadable:
        return result("no third-party traces in issued files", UNKNOWN,
                      "Could not scan every issued document: " + "; ".join(unreadable)
                      + ". Not scanned is not the same as clean.",
                      "Riverside House",
                      remedy="Fix the path or the entry, then re-run.")
    return result("no third-party traces in issued files", PASS,
                  "%d issued document(s) scanned, no third-party name, email or path in any of "
                  "them" % scanned, "Riverside House")


def check_exposures_state_our_recourse(m):
    """Riverside House, 28/07, from Gordon Court's withdrawal that ran in their
    own favour.

    Every re-read this week was driven by suspicion that something was worse
    than recorded. Nothing drives a re-read in the other direction, because a
    pessimistic position feels prudent - so entitlement you already own goes
    unclaimed and an exposure gets reported as unbacked when it is partly
    backed. On Riverside the A Plus storage clock was written up as a cost that
    grows with the delay, with no mention that our own terms make client-caused
    delay costs recoverable.

    'exposures': [{item, lands_on, our_recourse}] - our_recourse is the term,
    clause or document that backs us, or the string "none" where genuinely
    nothing does. "none" is a good answer. Silence is not, and neither is
    prose that only restates the exposure."""
    exps = m.get("exposures")
    if exps is None:
        return result("exposures state our recourse", UNKNOWN,
                      "List what this job is exposed to and what backs us on each: "
                      "'exposures': [{item, lands_on, our_recourse}]. Where nothing backs us, "
                      "say so with our_recourse 'none' - that is an answer. A pessimistic "
                      "position feels prudent and is simply wrong in the other direction.",
                      "Riverside House",
                      remedy="For each exposure, read your own terms for the clause that bears "
                             "on it before recording it as unbacked.")
    if isinstance(exps, dict):
        exps = [exps]
    if isinstance(exps, str) or not isinstance(exps, (list, tuple)):
        return result("exposures state our recourse", UNKNOWN,
                      "'exposures' is %r - it must be a list of {item, lands_on, our_recourse} "
                      "entries." % (exps,), "Riverside House",
                      remedy="Rewrite the field as a list, one entry per exposure.")
    if not exps:
        return result("exposures state our recourse", NA,
                      "no exposures recorded on this job", "Riverside House")
    silent = []
    for e in exps:
        if not isinstance(e, dict):
            silent.append("%r is not a {item, lands_on, our_recourse} entry" % (e,))
            continue
        item = str(e.get("item", "?"))[:70]
        rec = e.get("our_recourse")
        if rec is None or not str(rec).strip():
            silent.append("%s - our_recourse is unstated" % item)
            continue
        flat = str(rec).strip().lower()
        # "unknown", "tbc", "not looked at" are NOT answers - they are the
        # silence this rule exists to catch, wearing a value.
        if flat in ("unknown", "tbc", "?", "not checked", "not looked at", "unclear", "n/a"):
            silent.append("%s - our_recourse is %r, which is the silence this rule catches "
                          "rather than an answer" % (item, rec))
    if silent:
        return result("exposures state our recourse", UNKNOWN,
                      "Exposures recorded with no statement of what backs us: " + "; ".join(silent)
                      + ". An exposure written up one-sidedly reads as unbacked whether or not "
                        "it is.", "Riverside House",
                      remedy="Read your own terms and conditions, inclusions and exclusions for "
                             "the clause that bears on each, then record it - or record 'none'.")
    return result("exposures state our recourse", PASS,
                  "all %d recorded exposure(s) state what backs us, or that nothing does"
                  % len(exps), "Riverside House")


# Riverside, 28/07, from Gordon Court's n/a finding. `issued_documents` was
# being used for two different things: what we produced, and what the client
# receives. Riverside's list held the WORKING pricing document - which must
# never be sent - and an internal covering note to Adam. Three rules iterate it.
# A document counts as client-facing unless it says otherwise, so an unset flag
# behaves as it did before.
def client_facing(doc):
    return doc.get("goes_to_client", True) is not False


def check_exclusions_reach_the_issued_document(m):
    """Riverside House, 28/07. An exclusion that is not in the document you
    issue is not an exclusion.

    This chat spent three turns writing "excluded by us" about the AOV control
    system, Part K anti-fall protection, structural alterations and the
    structural design of fixings. All four ARE in Fenster's standard
    INCLUSIONS/EXCLUSIONS schedule - twelve lines of it - which lives in
    templates/proposal-content.json, the proposal and cover-letter path.
    Riverside was generated from MASTER PRICING DOC.xlsx, which has no
    exclusions block at all, and the only exclusion on its face was the one
    sentence someone had typed into a spec note.

    So the company had an answer and the job did not carry it. The gap is not
    in the drafting, it is between the template that holds the exclusions and
    the template that gets issued.

    'issued_documents': [{name, is_the_priced_document, exclusions_stated}] -
    exclusions_stated is the count of exclusions written on the face of that
    document, or a list of them."""
    docs = m.get("issued_documents")
    relied = [i for i in (m.get("spec_items") or [])
              if str(i.get("treatment", "")).lower() == "excluded"]
    if docs is None:
        return result("exclusions reach the issued document", UNKNOWN,
                      "State what we would actually hand the client and how many exclusions are "
                      "written on its face: 'issued_documents': [{name, is_the_priced_document, "
                      "exclusions_stated}]. A standard exclusions schedule that lives in a "
                      "template this job was not generated from protects nobody.",
                      "Riverside House",
                      remedy="Open the document you would send and count the exclusions on it.")
    if not relied:
        return result("exclusions reach the issued document", NA,
                      "nothing on this job is being carried as excluded", "Riverside House")
    if not docs:
        return result("exclusions reach the issued document", FAIL,
                      "%d item(s) are being carried as EXCLUDED and no issued document is "
                      "recorded at all. The exclusions exist only in this manifest."
                      % len(relied), "Riverside House",
                      remedy="Name the document that goes to the client, then put the exclusions "
                             "on it.")
    # Gordon Court, 28/07, referred back as a design question: should "the
    # priced document" mean ANY issued priced document carrying the exclusions,
    # or ALL of them? Their job issues two priced documents - a proposal that
    # carries the exclusions and a spreadsheet that does not - and the original
    # ALL reading failed them for it. They left it failing rather than edit a
    # flag, which was right.
    #
    # The ruling is neither. NO client-facing document carrying them is a FAIL
    # and always was: that is the founding case, a covering letter holding the
    # exclusions while the priced document does not. SOME BUT NOT ALL is an ASK,
    # because partial coverage is a judgement about how a pack will be used and
    # by whom, and a manifest cannot adjudicate it. An ASK keeps it visible
    # without asserting a defect that may not be one.
    bare, carrying = [], []
    for d in docs:
        if not isinstance(d, dict):
            return result("exclusions reach the issued document", UNKNOWN,
                          "%r is not a {name, is_the_priced_document, exclusions_stated} entry"
                          % (d,), "Riverside House",
                          remedy="Rewrite the entry, then re-run.")
        if not client_facing(d):
            continue
        stated = d.get("exclusions_stated")
        if not d.get("is_the_priced_document"):
            # A non-priced document does NOT rescue a bare priced one, and the
            # first draft of this ruling let it - my own test caught it. The
            # founding case is a covering letter holding the exclusions while
            # the priced document does not, and the reason it fails is that a
            # covering letter is detachable and unpriced: it will not travel
            # with the figure. Gordon Court's proposal is different in kind
            # because it is ITSELF priced. That distinction is the whole ruling,
            # so only priced documents count as carriers.
            continue
        if stated is None:
            return result("exclusions reach the issued document", UNKNOWN,
                          "%s does not say how many exclusions are on its face."
                          % d.get("name", "the priced document"), "Riverside House",
                          remedy="Open it and count them.")
        n = len(stated) if isinstance(stated, (list, tuple)) else stated
        try:
            n = int(n)
        except (TypeError, ValueError):
            return result("exclusions reach the issued document", UNKNOWN,
                          "%s states exclusions_stated as %r, which is neither a count nor a list."
                          % (d.get("name", "the priced document"), stated), "Riverside House",
                          remedy="Give a number or a list.")
        (bare if n <= 0 else carrying).append(d.get("name", "the priced document"))
    if bare and not carrying:
        return result("exclusions reach the issued document", FAIL,
                      "%d item(s) are being carried as EXCLUDED, and NOTHING going to the client "
                      "states any of them: %s. An exclusion that is not in the document you issue "
                      "is not an exclusion - a silent gap reads as included."
                      % (len(relied), ", ".join(bare)), "Riverside House",
                      remedy="Put the exclusions on the face of the priced document before it "
                             "is issued. It costs nothing before and is a dispute afterwards.")
    if bare:
        return result("exclusions reach the issued document", UNKNOWN,
                      "%d item(s) are carried as EXCLUDED and the pack states them UNEVENLY: %s "
                      "carr%s them, %s state%s none. Whether that matters depends on whether the "
                      "bare document can be relied on alone - forwarded, filed or quoted from "
                      "without the rest of the pack."
                      % (len(relied), ", ".join(carrying), "ies" if len(carrying) == 1 else "y",
                         ", ".join(bare), "s" if len(bare) == 1 else ""),
                      "Riverside House",
                      remedy="Either put the exclusions on the face of every priced document, or "
                             "record why the one that carries them will always travel with the "
                             "one that does not.")
    return result("exclusions reach the issued document", PASS,
                  "every client-facing priced document carries an exclusions schedule covering "
                  "%d relied-on exclusion(s)" % len(relied), "Riverside House")


def check_incorporated_terms_held(m):
    """Riverside House, 28/07. A supplier quote that incorporates its terms BY
    REFERENCE to a document we do not hold.

    A Plus QT51518 says the "Terms of Sale Revision V.01.2 - 08.01.2018" apply
    to the quotation and to any subsequent Contract, and that the DEFINITIONS -
    including who the "Customer" is - come from "Revision V.01 - 03.11.2017".
    Neither document is attached. Six files across the whole Commercial archive
    have "Terms of Sale" in the name and all six are the same Advisory Notes
    PDF; the Terms of Sale itself is in none of them. So the contract we would
    be ordering GBP 4,845.22 under has never been read here, on any job, in
    seven years.

    This is not the same as a quote with no terms at all - that is a gap you can
    see. An incorporation by reference reads as though the terms are settled and
    hides that you cannot say what they are. Asking a supplier for their terms
    costs one line of an email before an order and is a variation after one.

    'incorporated_terms': [{supplier, ref, document, held}] - held true when the
    named document is actually in our hands, false when it is only cited."""
    terms = m.get("incorporated_terms")
    if terms is None:
        return result("incorporated terms are actually held", UNKNOWN,
                      "State every document a supplier quote incorporates by reference and "
                      "whether we hold it: 'incorporated_terms': [{supplier, ref, document, "
                      "held}]. If a quote incorporates nothing by reference, say so with an "
                      "empty list.",
                      "Riverside House",
                      remedy="Read each supplier quote for the words 'apply to this quotation', "
                             "'Terms of Sale', 'conditions of sale' or 'as amended', then record "
                             "what it names and whether the file is in the job folder.")
    if isinstance(terms, dict):
        terms = [terms]
    if isinstance(terms, str) or not isinstance(terms, (list, tuple)):
        return result("incorporated terms are actually held", UNKNOWN,
                      "'incorporated_terms' is %r - it must be a list of "
                      "{supplier, ref, document, held} entries." % (terms,),
                      "Riverside House",
                      remedy="Rewrite the field as a list, one entry per incorporated document.")
    if not terms:
        return result("incorporated terms are actually held", NA,
                      "no supplier quote on this job incorporates terms by reference",
                      "Riverside House")
    missing, unnamed, unclear = [], [], []
    for t in terms:
        if not isinstance(t, dict):
            unclear.append("%r is not a {supplier, ref, document, held} entry" % (t,))
            continue
        ref = "%s %s" % (t.get("supplier", "?"), t.get("ref", "?"))
        doc = t.get("document")
        held = t.get("held")
        named = bool(doc) and bool(str(doc).strip()) and not _describes_absence(doc)
        # Gordon Court, 28/07, the first time this rule saw data that was not
        # mine. All four BSW quotations read "Orders are subject to acceptance
        # and terms and conditions of sale, AVAILABLE ON REQUEST" - no title, no
        # revision, no date. That is a WORSE position than my named A Plus case,
        # not a manifest-filling problem, and the rule used to grade it as the
        # lesser one and then hand back a remedy the estimator cannot carry out
        # ("say WHICH terms are incorporated" - the quotation does not say).
        # An unnamed incorporation gets its own bucket and its own remedy: ask
        # the supplier for the title, revision and date.
        if not named:
            if held in (True, 1) or str(held).strip().lower() in (
                    "true", "yes", "y", "held", "attached", "1"):
                unclear.append("%s is marked held but names no document - you cannot hold a "
                               "document you cannot name" % ref)
            else:
                unnamed.append("%s incorporates terms it does not even name" % ref)
            continue
        where = "%s incorporates \"%s\"" % (ref, str(doc).strip())
        # The lesson from check_free_delivery_threshold the same night: an
        # else-branch that produces an ASSERTION rather than a question will
        # eventually assert something false from a value it did not understand.
        # Only a documented vocabulary decides; anything else is asked about.
        if held in (True, 1):
            continue
        flat = str(held).strip().lower()
        if flat in ("true", "yes", "y", "held", "attached", "1"):
            continue
        if held is None or flat in ("", "none", "null"):
            missing.append("%s - and 'held' is unstated, which is not the same as held" % where)
        elif held in (False, 0) or flat in ("false", "no", "n", "0", "not held", "missing"):
            missing.append("%s - we do not hold it" % where)
        else:
            unclear.append("%s but 'held' is %r - use true or false" % (where, held))
    if unclear:
        return result("incorporated terms are actually held", UNKNOWN,
                      "Cannot tell whether the incorporated terms are held: " + "; ".join(unclear),
                      "Riverside House",
                      remedy="Fill the entry properly, then re-run before issuing anything.")
    if unnamed or missing:
        # Unnamed first, deliberately. A quote that names its terms tells you
        # what to ask for; one that says "available on request" leaves you
        # unable to say which version you have not read.
        parts = []
        if unnamed:
            parts.append("A supplier quote incorporates terms IT DOES NOT NAME: "
                         + "; ".join(unnamed)
                         + ". No title, revision or date, so we cannot even say which document "
                           "we have not read")
        if missing:
            parts.append("A supplier quote incorporates terms we have never read: "
                         + "; ".join(missing))
        return result("incorporated terms are actually held", UNKNOWN,
                      ". ".join(parts)
                      + ". The price rests on a contract whose contents we cannot state.",
                      "Riverside House",
                      remedy="Ask the supplier for the document - by title, revision and date "
                             "where the quote names one, and for whatever their quotation refers "
                             "to where it does not - before placing an order. It is one line "
                             "pre-order and a negotiation afterwards.")
    return result("incorporated terms are actually held", PASS,
                  "%d incorporated terms document(s) are in our hands" % len(terms),
                  "Riverside House")


def check_warranty_is_back_to_back(m):
    """Riverside House / Gordon Court, 28/07. What we promise the client against
    what the supplier promises us - compared as FOUR things, not one.

    Gordon Court found a five-year glass gap on AFS by comparing PERIODS. The
    same check run here on A Plus returned a period, an outright component
    exclusion and a cycle cap - three findings from one check - and their
    conclusion was that the check itself had been a quarter of a check:

        the PERIOD          10 years against 12 months
        the START DATE      ours states none at all; theirs runs from delivery
                            to our own yard, so the client's cover is spent
                            before the building is occupied
        the EXCLUSION LIST  four of six of A Plus's have no counterpart in ours
        a USAGE CAP         "15,000 cycles or 12 months, whichever is sooner" -
                            a period stated in years and capped in cycles is not
                            a period in years

    Two things this rule refuses to accept, both learned here:

    A PERIOD WITH NO START DATE IS NOT A PERIOD. Our own clause offers ten years
    and never says ten years from what. Both jobs had this defect and neither
    noticed while comparing the number of years.

    AN EXCLUSION LIST CANNOT BE COMPLETE WHERE THE TERMS ARE NOT HELD. AFS wrote
    theirs as 6.4.1-6.4.6 and it could be diffed. A Plus never wrote a list at
    all - theirs are conditional clauses scattered through Finishes, Hardware
    and the AOV notes, and the rest are in a Terms of Sale nobody has requested.
    So this rule reads `incorporated_terms`: if a supplier's terms are not held,
    `exclusions_complete: true` is a contradiction and is reported as one.

    THE RULING SPLITS BY WHOSE PROBLEM IT IS, corrected before this shipped. FAIL
    is for our own document being defective and for the record contradicting
    itself - a period with no start date, a list called complete that lives in
    terms we do not hold. Both are ours to fix unilaterally. The GAP itself -
    shorter period, unmatched exclusions, a usage cap - is an ASK, because a
    ten-year client warranty backed by twelve-month supplier terms is what the
    whole trade offers, and a gate that fails on the normal case stops being
    read. Surfacing it by name and handing the decision to a human is the job;
    vetoing a commercial position is not.

    'warranty': {'ours': {period_months, scope, start_date, usage_cap,
    exclusions[]}, 'suppliers': [{supplier, ref, covers, period_months,
    start_date, usage_cap, exclusions: [{exclusion, counterpart_in_ours}],
    exclusions_complete}]}. `counterpart_in_ours` is null where ours has none -
    that null is the finding."""
    w = m.get("warranty")
    if w is None:
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "State what we warrant and what each supplier warrants us: 'warranty': "
                      "{'ours': {period_months, scope, start_date, usage_cap, exclusions}, "
                      "'suppliers': [{supplier, ref, covers, period_months, start_date, "
                      "usage_cap, exclusions, exclusions_complete}]}. Compare four things, not "
                      "one - the period, the start date, the exclusion list, and whether "
                      "anything is capped by cycles or usage rather than time.",
                      "Riverside House / Gordon Court",
                      remedy="Find our guarantee clause and the supplier's, and read each one "
                             "through rather than for its number of years. A period stated in "
                             "years and capped in cycles is not a period in years.")
    if not isinstance(w, dict):
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "'warranty' is %r - it must be {'ours': {...}, 'suppliers': [...]}." % (w,),
                      "Riverside House / Gordon Court",
                      remedy="Rewrite the field with an 'ours' object and a 'suppliers' list.")
    ours = w.get("ours")
    if not isinstance(ours, dict):
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "'warranty.ours' is missing - state what WE offer the client before "
                      "comparing it with anything.",
                      "Riverside House / Gordon Court",
                      remedy="Quote our own guarantee clause into 'ours', including its start "
                             "date. If it states no start date, record start_date as null - "
                             "that is a finding, not a blank.")
    sups = w.get("suppliers")
    if sups is None:
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "'warranty.suppliers' is missing. If no supplier on this job states a "
                      "warranty at all, say so with an empty list - a supplier who states "
                      "nothing is a worse answer than a short period, not a better one.",
                      "Riverside House / Gordon Court",
                      remedy="Read each supplier quotation for 'guarantee', 'warrant', 'year' "
                             "and 'defect'. Record what you find, or the empty list.")
    if isinstance(sups, dict):
        sups = [sups]
    if isinstance(sups, str) or not isinstance(sups, (list, tuple)):
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "'warranty.suppliers' is %r - it must be a list." % (sups,),
                      "Riverside House / Gordon Court",
                      remedy="Rewrite the field as a list, one entry per supplier warranty.")

    ours_months = ours.get("period_months")
    ours_cap = ours.get("usage_cap")
    unheld = set()
    for it in (m.get("incorporated_terms") or []):
        if isinstance(it, dict) and not it.get("held"):
            unheld.add(str(it.get("supplier", "")).strip().lower())

    fails, asks, notes = [], [], []

    if not ours.get("start_date"):
        fails.append("we offer the client %s and our clause states NO START DATE - a period "
                     "with no start date is not a period"
                     % (ours.get("period") or _months(ours_months) or "a warranty"))
    if not ours.get("scope"):
        asks.append("what our warranty COVERS is not recorded - a clause worded for one kind "
                    "of product may not reach a component of another kind")

    for s in sups:
        if not isinstance(s, dict):
            asks.append("%r is not a supplier warranty entry" % (s,))
            continue
        who = "%s %s" % (s.get("supplier", "?"), s.get("ref", ""))
        who = who.strip()
        covers = s.get("covers")
        label = "%s%s" % (who, " (%s)" % covers if covers else "")
        sm = s.get("period_months")
        if sm is None:
            asks.append("%s states no period - a supplier who says nothing has not given us "
                        "an unlimited warranty, they have given us their terms of sale" % label)
        elif isinstance(sm, (int, float)) and isinstance(ours_months, (int, float)):
            if sm < ours_months:
                asks.append("%s gives us %s against the %s we offer the client - WE CARRY THE "
                            "%s IN BETWEEN. Whether the client-facing period is offered as it "
                            "stands is a commercial decision and needs a human to take it"
                            % (label, _months(sm), _months(ours_months),
                               _months(ours_months - sm)))
        if s.get("usage_cap") and not ours_cap:
            asks.append("%s is capped by USE, not time - \"%s\" - and our own warranty has no "
                        "equivalent cap. Work out what the cap means in service before "
                        "reporting it: a limit that cannot be reached inside the period is not "
                        "a finding, and one that can be is the real period"
                        % (label, s.get("usage_cap")))
        if s.get("start_date") and not ours.get("start_date"):
            notes.append("%s runs from \"%s\"" % (label, s.get("start_date")))

        excl = s.get("exclusions")
        if excl is None:
            asks.append("%s: exclusions not recorded. Read the clause through rather than for "
                        "its period - a supplier who writes no exclusion LIST still has "
                        "exclusions, scattered as conditions inside other paragraphs" % label)
            continue
        if isinstance(excl, dict):
            excl = [excl]
        orphans = []
        for e in (excl if isinstance(excl, (list, tuple)) else []):
            if isinstance(e, dict):
                if not e.get("counterpart_in_ours"):
                    orphans.append(str(e.get("exclusion", e))[:90])
            elif e:
                asks.append("%s: exclusion %r is not a {exclusion, counterpart_in_ours} entry "
                            "- the null counterpart IS the finding, so it has to be stated"
                            % (label, e))
        if orphans:
            asks.append("%s excludes %d thing(s) our warranty does not: %s. Where they decline "
                        "on one of these we still owe the client - decide which are worth "
                        "asking about and which are worth carrying"
                        % (label, len(orphans), "; ".join(orphans)))
        complete = s.get("exclusions_complete")
        supplier_key = str(s.get("supplier", "")).strip().lower()
        if complete and supplier_key in unheld:
            fails.append("%s: exclusions are recorded as COMPLETE while incorporated_terms says "
                         "we do not hold their terms of sale. Both cannot be true - the list "
                         "you have is the part they printed on the quotation" % label)
        elif not complete:
            asks.append("%s: the exclusion list is not complete, so the gaps above are a floor "
                        "and not a count" % label)

    if fails:
        # Gordon Court, 28/07. Returning on `fails` alone hid SEVEN asks behind one line
        # on this job - the 5-year glass gap, five orphan AFS exclusions, six orphan BSW
        # ones and a supplier stating no period at all, none of them printed. The split is
        # right and stays; what was wrong is that the reader could not tell there was
        # anything behind it. So the fails still lead and the asks are counted and named.
        return result("warranty is back-to-back with the supplier", FAIL,
                      "Our own warranty document is defective, or the comparison contradicts "
                      "itself - neither of these needs anyone's permission to fix: "
                      + "; ".join(fails)
                      + ("." if not notes else ". Also: " + "; ".join(notes) + ".")
                      + ("" if not asks else
                         " AND %d GAP(S) ARE QUEUED BEHIND THIS AND ARE NOT SHOWN ABOVE - they "
                         "come through as ASK once the FAIL is cleared, and they are the "
                         "larger half: %s." % (len(asks), "; ".join(asks))),
                      "Riverside House / Gordon Court",
                      remedy="Put a start date in the clause, and stop describing a list as "
                             "complete while the document it lives in has never been read. The "
                             "SIZE of the gap is a separate question and is listed after the "
                             "defect rather than instead of it.")
    if asks:
        return result("warranty is back-to-back with the supplier", UNKNOWN,
                      "The warranty we offer runs past the warranties we are given, or the "
                      "comparison is not finished: " + "; ".join(asks) + ".",
                      "Riverside House / Gordon Court",
                      remedy="Finish the four-part comparison - period, start date, exclusion "
                             "list, usage cap - then put the gap to a human. A ten-year client "
                             "warranty backed by twelve-month supplier terms is normal and may "
                             "be perfectly deliberate; what is not acceptable is nobody knowing.")
    return result("warranty is back-to-back with the supplier", PASS,
                  "%d supplier warranty(ies) compared on period, start date, exclusions and "
                  "usage cap - nothing we offer runs past what we are given" % len(sups),
                  "Riverside House / Gordon Court")


def _months(n):
    """A period in words, so a 108-month gap does not read as a number."""
    if not isinstance(n, (int, float)):
        return None
    n = int(n)
    if n and n % 12 == 0:
        y = n // 12
        return "%d year%s" % (y, "" if y == 1 else "s")
    if n < 12:
        return "%d month%s" % (n, "" if n == 1 else "s")
    return "%d months" % n


def check_spec_label_matches_evidence(m):
    """Gordon Court, 28/07. A spec item whose LABEL says outstanding while its own
    EVIDENCE says it was done.

    The founding case: 'Untagged glazing on the elevations - riverside's check,
    NOT RUN', whose evidence field ended 'TENTH TURN - RUN. Rendered the
    elevations at 110-260 dpi.' Both were written by me, ten turns apart. Only
    the ref was ever visible, because report() truncated the evidence - so the
    contradiction sat in the file for eleven turns and then cost a turn of
    re-derived work when I trusted the label.

    This matters to a price because check_scope_gaps reads the treatment field.
    A label that says GAP on something already resolved produces a FAIL nobody
    can action; a label that says NOT RUN on something already run sends
    somebody to do it twice.

    Verified before shipping: 0 fires across 119 spec items in 13 manifests,
    and it fires on the one real case recovered from git.
    """
    bad = []
    for s in m.get("spec_items") or []:
        label = "%s || %s" % (s.get("ref", ""), s.get("treatment", ""))
        ev = str(s.get("evidence", ""))
        if _LABEL_STALE.search(label) and _LABEL_DONE.search(ev):
            hit = _LABEL_DONE.search(ev)
            bad.append("%s - but its evidence says '%s'" % (
                str(s.get("ref", ""))[:70], ev[max(0, hit.start() - 30):hit.end() + 40].strip()))
    if bad:
        return result("spec item labels match their evidence", FAIL,
                      "A spec item is labelled outstanding while its own evidence records it as done: "
                      + "; ".join(bad)
                      + ". Fix the label or the evidence - whichever is wrong, somebody will act on the "
                        "one that is visible.",
                      catch="gordon-court")
    return result("spec item labels match their evidence", PASS,
                  "%d spec item label(s) agree with their evidence" % len(m.get("spec_items") or []))


_QUALS_VOID = re.compile(
    r"shall not be applicable|will not be applicable|not be applicable to this tender"
    r"|no qualification|without qualification|unqualified tender|qualifications will not be"
    r"|any caveats?, ?assumptions?, ?reservations? or exclusions?"
    r"|shall be deemed (?:to be )?(?:excluded|of no effect|void)"
    r"|shall have no effect|deemed not to apply", re.I)


def check_our_qualifications_survive_signature(m):
    """John North Hall, 28/07. The client's Form of Tender DISAPPLIES our
    exclusions, and signing it is how we submit.

    Three rules already ask whether an exclusion is written down:
    check_scope_gaps asks whether we addressed the item, and
    check_exclusions_reach_the_issued_document (riverside) asks whether the
    exclusion is on the face of the document we hand over. Both take for granted
    that an exclusion, once written on the right piece of paper, does something.

    John North Hall's ITT section 5.0, clause 4.1 (the second of four clauses
    all numbered 4.1) reads: "It is agreed that any other terms and conditions of
    contract or any caveats, assumptions, reservations or exclusions that may be
    printed on correspondence emanating from the tender, or any contract
    resulting from this tender, shall not be applicable to this tender or
    agreement." The clause immediately above offers to complete the works "For
    the firm price contained within the pricing summary".

    So on this job every one of the twelve lines in Fenster's standard
    inclusions/exclusions schedule is disapplied by our own signature - access
    plant, waste removal, making good, access control, the lot - and the tender
    becomes a firm lump sum for the whole of the works as described. Writing the
    exclusions more clearly does not help. Riverside's rule would return a clean
    PASS on a document whose exclusions have no legal effect, which is exactly
    the reassurance worth removing.

    The only two answers that work are pricing the item or getting the
    qualification accepted IN WRITING BEFORE the return date, because after
    signature the clause has already bitten.

    'qualification_regime': {document, clause, qualifications_permitted,
    we_must_sign} - qualifications_permitted false when the tender documents
    disapply, prohibit or deem void any caveat, assumption, reservation or
    exclusion we attach."""
    reg = m.get("qualification_regime")
    relied = [i for i in (m.get("spec_items") or [])
              if "exclud" in str(i.get("treatment", "")).lower()]
    if reg is None:
        return result("our qualifications survive signature", UNKNOWN,
                      "State what the tender documents do to our exclusions: "
                      "'qualification_regime': {document, clause, qualifications_permitted, "
                      "we_must_sign}. An exclusion is only worth writing if the contract we sign "
                      "lets it stand - and a Form of Tender is where that is decided, not the "
                      "specification.",
                      "John North Hall",
                      remedy="Read the Form of Tender / tender acceptance page word for word and "
                             "record whether it disapplies caveats, assumptions, reservations or "
                             "exclusions.")
    if isinstance(reg, (list, tuple)):
        reg = reg[0] if len(reg) == 1 and isinstance(reg[0], dict) else None
    if not isinstance(reg, dict):
        return result("our qualifications survive signature", UNKNOWN,
                      "'qualification_regime' is %r - it must be a {document, clause, "
                      "qualifications_permitted, we_must_sign} entry."
                      % (m.get("qualification_regime"),),
                      "John North Hall",
                      remedy="Rewrite the field as one object describing the tender's "
                             "qualification regime.")
    permitted = reg.get("qualifications_permitted")
    clause = str(reg.get("clause") or "")
    doc = str(reg.get("document") or "the tender documents").strip() or "the tender documents"
    if permitted is None and clause and _QUALS_VOID.search(clause):
        permitted = False
    if permitted is None:
        return result("our qualifications survive signature", UNKNOWN,
                      "%s is recorded but 'qualifications_permitted' is unstated. That is the "
                      "field the whole rule turns on - a tender that voids qualifications and a "
                      "tender that invites them look identical until someone reads the clause."
                      % doc, "John North Hall",
                      remedy="Set qualifications_permitted true or false against the words of the "
                             "clause, and paste the clause into 'clause'.")
    if isinstance(permitted, str):
        v = permitted.strip().lower()
        if v in ("yes", "true", "y", "permitted", "allowed"):
            permitted = True
        elif v in ("no", "false", "n", "void", "disapplied", "prohibited"):
            permitted = False
        else:
            return result("our qualifications survive signature", UNKNOWN,
                          "'qualifications_permitted' is %r - it must read true or false."
                          % (permitted,), "John North Hall",
                          remedy="Answer the question the clause answers: do our qualifications "
                                 "survive, yes or no?")
    # A dict, a list or any other container is truthy, and a truthy value here
    # means "our exclusions stand" - the reassuring answer. Caught by this
    # rule's own variant suite before it shipped, on {"v": false}.
    if not isinstance(permitted, bool) and permitted not in (0, 1):
        return result("our qualifications survive signature", UNKNOWN,
                      "'qualifications_permitted' is %r - it must read true or false, and a "
                      "value of any other shape would otherwise be read as the reassuring answer."
                      % (permitted,), "John North Hall",
                      remedy="Answer the question the clause answers: do our qualifications "
                             "survive, yes or no?")
    if permitted:
        return result("our qualifications survive signature", PASS,
                      "%s permits qualifications, so the %d item(s) we carry as excluded stand "
                      "as written" % (doc, len(relied)))
    if not relied:
        return result("our qualifications survive signature", PASS,
                      "%s disapplies qualifications, and nothing on this job is being carried as "
                      "excluded - so there is nothing for the clause to strike out" % doc)
    signing = reg.get("we_must_sign")
    accepted = [i for i in relied
                if str(i.get("qualification_accepted_in_writing", "")).lower()
                in ("true", "yes", "1")]
    exposed = [i for i in relied if i not in accepted]
    if not exposed:
        return result("our qualifications survive signature", PASS,
                      "%s disapplies qualifications, but all %d excluded item(s) have written "
                      "acceptance predating the return date" % (doc, len(relied)))
    return result("our qualifications survive signature", FAIL,
                  "%s disapplies our qualifications%s, and %d item(s) are being carried as "
                  "EXCLUDED with no written acceptance: %s. Once the Form of Tender is signed "
                  "these are not exclusions, they are unpriced work inside a firm lump sum. "
                  "Writing them more clearly on the proposal changes nothing."
                  % (doc,
                     " and signing it is how we submit" if signing else "",
                     len(exposed),
                     "; ".join(str(i.get("ref", "?"))[:70] for i in exposed[:8])
                     + ("; ..." if len(exposed) > 8 else "")),
                  catch="John North Hall",
                  remedy="For each one: price it, or get the qualification accepted in writing "
                         "BEFORE the return date. After signature the clause has already bitten.")


def check_priced_scope_is_not_excluded(m):
    """Crestwood Park, 28/07 - and Princess Beatrice, 27/07, which is the same
    error on another job in the same week.

    We charged Reynolds GBP 17,779.06 for Teleflex, 24% of the tender, and the
    clarifications on page 3 of the proposal that went out with it exclude
    "Teleflex controls / wiring". Princess Beatrice charges GBP 5,356.22 for
    external mastic against a proposal reading "External mastic is charged as
    an optional extra". Two documents, issued together, one saying we are doing
    it and the other saying we are not.

    This is not a drafting nicety. It is the worst possible pair to hold: the
    client can read the exclusion and employ someone else for work we have
    already bought and marked up, or read the price and hold us to scope we
    have told them we do not own. Whichever way it resolves, it resolves
    against us, because we wrote both halves.

    It is also invisible to every other rule here. check_scope_gaps asks
    whether a spec item is priced OR excluded and is satisfied by either;
    nothing asked whether it was somehow BOTH. The arithmetic ties, the
    supplier quote reconciles, the exclusion is properly written on the face of
    the document - and the contradiction sails through.

    'priced_lines': [{ref, amount, covered_by_our_exclusions}] -
    covered_by_our_exclusions is the TEXT of the exclusion that eats this line,
    or false if no exclusion touches it. Read the exclusions list against the
    priced lines one at a time; do not answer it from memory."""
    lines = m.get("priced_lines")
    if lines is None:
        return result("nothing priced is also excluded", UNKNOWN,
                      "List what the client is being charged for and whether any of our own "
                      "exclusions covers it: 'priced_lines': [{ref, amount, "
                      "covered_by_our_exclusions: \"<exclusion text>\"|false}].",
                      "Crestwood Park",
                      remedy="Open the priced document and the exclusions list side by side and "
                             "read one against the other.")
    if not lines:
        return result("nothing priced is also excluded", NA, "no priced lines recorded",
                      "Crestwood Park")
    clash, silent = [], []
    for ln in lines:
        ref = str(ln.get("ref", "?"))
        if "covered_by_our_exclusions" not in ln or ln.get("covered_by_our_exclusions") is None:
            silent.append(ref)
            continue
        cov = ln.get("covered_by_our_exclusions")
        if cov is False or (isinstance(cov, str) and not cov.strip()):
            continue
        amt = ln.get("amount")
        # Quote the exclusion at length, and if it has to be cut, cut it around
        # the words that matter rather than at a fixed offset. Crestwood's
        # exclusion is a 300-character run-on list and "Teleflex controls /
        # wiring" is the ninth item in it - a 90-character cap stopped at
        # "asbestos removal" and printed a clash whose evidence was invisible.
        # Same lesson as `remedy` in result(): the proof must not be what the
        # truncation eats.
        text = " ".join(str(cov).split())
        if len(text) > 220:
            key = str(ref).split("(")[0].strip().lower()
            at = text.lower().find(key) if key else -1
            if at > 110:
                text = "..." + text[at - 60:at + 160].strip() + "..."
            else:
                text = text[:220] + "..."
        clash.append("%s%s is charged and excluded by our own words: %r"
                     % (ref,
                        "" if amt in (None, "") else " (GBP %s)" % amt,
                        text))
    if clash:
        return result("nothing priced is also excluded", FAIL,
                      "%d line(s) are charged for and disclaimed in the same pack: %s. The client "
                      "holds both documents, so this is decided against us either way - they take "
                      "the exclusion and pay someone else for work we have bought, or they take "
                      "the price and hold us to scope we said was not ours."
                      % (len(clash), "; ".join(clash)),
                      "Crestwood Park",
                      remedy="Decide which one is true, then change the OTHER document before the "
                             "client reads them together. Withdrawing the exclusion is free; "
                             "withdrawing the price is not.")
    if silent:
        return result("nothing priced is also excluded", UNKNOWN,
                      "Not stated whether our exclusions touch: %s." % ", ".join(silent),
                      "Crestwood Park",
                      remedy="Answer it per line from the documents, not from memory.")
    return result("nothing priced is also excluded", PASS,
                  "%d priced line(s) checked against our own exclusions, no overlap" % len(lines),
                  "Crestwood Park")


def check_bought_in_lump_has_a_quantity_basis(m):
    """Crestwood Park, 28/07. WCI's quote WCIL/FEN4215 is GBP 14,223.25 for
    "13no. Sets each to operate 2 top hung vent 2pp" and "9no. Sets each to
    operate 2 top hung vent 2pp" - 22 sets, two vents apiece. Drawing A007
    requires 2No. operators per light and "Opening lights to operate with 1No.
    new White Teleflex Midi control EACH". Per light, not per window.

    Two vents per window is right on W1-W8, which the elevations split into 2.
    The other thirteen windows are split into 3, 5 and 6 parts. So the supplier
    priced a different question from the one the drawing asks, and the gap goes
    the expensive way.

    Nothing caught it, because on our own pricing document Teleflex is ONE ROW
    with no quantity and no rate - a lump. check_supplier_covers_quantity
    compares qty_sold against qty_quoted and a lump has neither, so the rule
    that exists for exactly this had nothing to compare. A number with no
    quantity behind it cannot be wrong, which is precisely what makes it
    dangerous: GBP 17,779.06, 24% of the tender, and the only thing anyone
    verified was that it multiplied up from the supplier's total correctly.

    So: any bought-in lump must state BOTH quantities and both bases. Where the
    supplier's basis is not the specification's basis, that is a finding even
    when no number is yet known - "we cannot say" is the answer, and it has to
    be said out loud rather than left as a lump nobody can question.

    'bought_in_lines': [{ref, amount, supplier_ref, supplier_qty,
    supplier_qty_basis, spec_required_qty, spec_qty_basis}]. spec_required_qty
    may be null where the drawing does not let you count it - that ASKS rather
    than passing."""
    lines = m.get("bought_in_lines")
    if lines is None:
        return result("bought-in lumps have a quantity basis", UNKNOWN,
                      "For every bought-in item carried as a lump: 'bought_in_lines': "
                      "[{ref, amount, supplier_ref, supplier_qty, supplier_qty_basis, "
                      "spec_required_qty, spec_qty_basis}].",
                      "Crestwood Park",
                      remedy="Read the supplier's own wording for what a unit of theirs covers, "
                             "then read the specification for what it counts.")
    if not lines:
        return result("bought-in lumps have a quantity basis", NA,
                      "no bought-in lump lines on this job", "Crestwood Park")
    short, mismatch, silent = [], [], []
    for ln in lines:
        ref = str(ln.get("ref", "?"))
        sq, rq = ln.get("supplier_qty"), ln.get("spec_required_qty")
        sb, rb = ln.get("supplier_qty_basis"), ln.get("spec_qty_basis")
        if not sb or not rb:
            silent.append("%s (quantity basis not stated on %s)"
                          % (ref, "the supplier quote" if not sb else "the specification"))
            continue
        # The bases are prose, not codes. Equality is only meaningful when they
        # are written the same way; anything else is for a human to read.
        same_basis = str(sb).strip().lower() == str(rb).strip().lower()
        if sq is not None and rq is not None and sq < rq:
            short.append("%s: %s quoted %s (%s), the spec requires %s (%s)"
                         % (ref, ln.get("supplier_ref", "the supplier"), sq, sb, rq, rb))
        elif not same_basis:
            mismatch.append("%s: %s priced per %r, the specification counts per %r%s"
                            % (ref, ln.get("supplier_ref", "the supplier"), str(sb)[:60],
                               str(rb)[:60],
                               "" if rq is not None else
                               " - and the required quantity has not been established"))
    if short:
        return result("bought-in lumps have a quantity basis", FAIL,
                      "Bought-in quantity is short of the specification: %s. A lump sum hides "
                      "this: there is no quantity on our document for anyone to challenge."
                      % "; ".join(short), "Crestwood Park",
                      remedy="Go back to the supplier with the specification's own count before "
                             "the price is relied on.")
    if mismatch:
        return result("bought-in lumps have a quantity basis", FAIL,
                      "Supplier and specification are counting different things: %s. The totals "
                      "reconcile and the scope does not." % "; ".join(mismatch),
                      "Crestwood Park",
                      remedy="Re-ask the supplier on the specification's basis. Do not reconcile "
                             "a lump to its own total and call it checked.")
    if silent:
        return result("bought-in lumps have a quantity basis", UNKNOWN,
                      "Quantity basis missing for: %s." % "; ".join(silent), "Crestwood Park")
    return result("bought-in lumps have a quantity basis", PASS,
                  "%d bought-in lump(s) counted on the specification's own basis" % len(lines),
                  "Crestwood Park")


def check_rfq_answered(m):
    """Brocks Hill Phase 2, 29/07 - and the third time this month after Filwood
    and Georgie's. Gintare's RFQ to BSW asked for six things including SOLAR
    CONTROL GLAZING and OBSCURE GLAZING WHERE REQUIRED. BSW answered on colour,
    triple glazing and panic gear, quoted 'Clr' on every line, and said nothing
    about the other two either way. The tender was then built on the quote
    instead of the instruction, and the omission read as the estimator's.

    Silence is not compliance. Every line of the RFQ has to be ticked off
    against the return, and a supplier who does not mention an item has not
    priced it."""
    items = m.get("rfq_items")
    if items is None:
        return result("the RFQ was answered line by line", UNKNOWN,
                      "Tick every item the RFQ asked for against the quote that came back: "
                      "'rfq_items': [{item, requested, quoted_response}]. Use null for "
                      "quoted_response where the supplier said nothing.", "Brocks Hill")
    if not items:
        return result("the RFQ was answered line by line", NA,
                      "no RFQ to reconcile on this job", "Brocks Hill")
    silent, refused = [], []
    for i in items:
        name = i.get("item", "?")
        if not i.get("requested"):
            continue
        resp = i.get("quoted_response")
        if resp is None or str(resp).strip() == "":
            silent.append(name)
        elif str(resp).strip().lower() in ("no", "not available", "not quoted", "excluded"):
            refused.append("%s (%s)" % (name, resp))
    if silent:
        return result("the RFQ was answered line by line", FAIL,
                      "Asked for and never answered: %s. A supplier who does not mention an item "
                      "has not priced it - do not let the quote overwrite the instruction."
                      % ", ".join(silent), "Brocks Hill")
    if refused:
        return result("the RFQ was answered line by line", FAIL,
                      "Supplier declined and the tender must say so: %s." % "; ".join(refused),
                      "Brocks Hill")
    return result("the RFQ was answered line by line", PASS,
                  "all %d RFQ item(s) answered by the supplier" % len(items), "Brocks Hill")


def check_uplift_applied(m):
    """Brocks Hill Phase 2, 04/08/2026. Five Strongdor steel doors were issued to
    SMD at GBP 2,728.81 each, which is Strongdor's GBP 2,637.01 per door-set plus
    the GBP 459.00 delivery split five ways - to the penny. The house uplift, code
    value x 75%, was never applied. GBP 7,500 of margin left the building inside a
    document that looked perfect, and the matching installation labour (5 x DAD
    500) was missing too, for GBP 10,000 in total.

    WHY THE TWO RULES EITHER SIDE OF THIS ONE BOTH MISS IT.
    `check_net_pricing` reconciles the total, and the total was internally
    consistent - 2,728.81 x 5 = 13,644.05, correct arithmetic on a wrong rate.
    `check_supplier_covers_quantity` compares the line against the supplier quote,
    and the line MATCHED the supplier quote exactly. That match is the symptom.
    Every other quantity- or total-based check reads a row sold at cost as clean,
    because at cost is where the two documents agree.

    So the question has to be asked a third way: SELL MINUS SUPPLY, ROW BY ROW,
    AGAINST THE CODE TABLE. On Brocks Hill that took one pass over nine rows -
    eight landed on their adder to the penny (SAD 900, DAD 1500, ELAW 637.50,
    LAW 487.50) and one came out at 91.80. It needs no re-pricing and no supplier
    document, only the workbook that is about to be issued.

    THE TELL, worth knowing because it is where to look first: a late line dropped
    in under deadline pressure. The steel cost landed at 14:41 on 31/07 and the
    tender went to the client at 15:12. The placeholder it replaced - GBP 2,000,
    described in writing as "plus our markup" - had no uplift in it either, so the
    error was in the row from the moment the row existed and survived a check by
    the Commercial Director because the number moved and looked new.

    'priced_rows': [{ref, code, qty, supply_each, sell_each, additional_each}]
    `additional_each` is anything legitimately in the sell that is not the uplift -
    a delivery share, a carriage recharge. State it; do not fold it into supply,
    because folding it in is exactly how this row read as correct."""
    rows = m.get("priced_rows")
    if rows is None:
        return result("the house uplift is on every row", UNKNOWN,
                      "State every priced row against its code: 'priced_rows': [{ref, code, qty, "
                      "supply_each, sell_each, additional_each}]. A row sold at supply cost "
                      "reconciles against its own total AND against the supplier quote, so nothing "
                      "else in this file can see it.",
                      "Brocks Hill Phase 2",
                      remedy="Open the pricing workbook and read sell minus supply on each row "
                             "against the code table. It is one pass and needs no re-pricing.")
    if not rows:
        return result("the house uplift is on every row", NA,
                      "no priced rows on this job", "Brocks Hill Phase 2")
    try:
        import mary_pricing as _p
        code_value = _p.CODE_VALUE
    except Exception as exc:
        return result("the house uplift is on every row", UNKNOWN,
                      "the code table could not be read from mary_pricing (%s: %s), so no row "
                      "could be checked - and 'not checked' is not 'correct'"
                      % (type(exc).__name__, exc),
                      "Brocks Hill Phase 2")
    short, unknown = [], []
    for r in rows:
        ref = r.get("ref", "?")
        code = str(r.get("code", "")).strip().upper()
        if code not in code_value:
            unknown.append("%s (code %r is not in the code table)" % (ref, r.get("code")))
            continue
        try:
            supply = float(r["supply_each"])
            sell = float(r["sell_each"])
            extra = float(r.get("additional_each") or 0)
        except (TypeError, ValueError, KeyError):
            unknown.append("%s (supply_each %r / sell_each %r is not a number)"
                           % (ref, r.get("supply_each"), r.get("sell_each")))
            continue
        due = round(code_value[code] * 0.75, 2)
        got = round(sell - supply - extra, 2)
        if abs(got - due) <= 0.01:
            continue
        try:
            qty = int(r.get("qty") or 1)
        except (TypeError, ValueError):
            qty = 1
        short.append("%s (%s): uplift is GBP %s, code %s is worth GBP %s - GBP %s x %d = "
                     "GBP %s adrift"
                     % (ref, code, format(got, ",.2f"), code, format(due, ",.2f"),
                        format(due - got, ",.2f"), qty, format((due - got) * qty, ",.2f")))
    if short:
        return result("the house uplift is on every row", FAIL,
                      "Rows priced at or near supply cost: %s." % "; ".join(short),
                      "Brocks Hill Phase 2",
                      remedy="Apply the code adder, and check the installation line recomputes "
                             "from the labour codes WITH those rows in it - a row that lost its "
                             "uplift usually lost its labour too.")
    if unknown:
        return result("the house uplift is on every row", UNKNOWN,
                      "Rows that could not be checked: %s." % ", ".join(unknown),
                      "Brocks Hill Phase 2")
    return result("the house uplift is on every row", PASS,
                  "%d priced row(s) carry their code adder" % len(rows),
                  "Brocks Hill Phase 2")


def check_site_access_is_priced_or_excluded(m):
    """Luton Airport Departure Gates 1 & 2 (Ryebridge), issued 13/07/2026 at
    GBP 14,157.24. The site is AIRSIDE: every installer and supervisor needs an
    LLA induction for clearance and is escorted for the whole time they are
    landside of nothing and airside of everything. The priced document is two
    lines - 3 doors at 4,219.08, and INSTALLATION 1,500.00, which is exactly
    3 x the DAD labour code of 500. That 500 is the per-door-set FIT rate. It
    carries no mobilisation, no travel, no supervision and no site attendance.

    So the price contained ZERO airside content, and the proposal never said so.
    Page 3 reads "Airside working requirements, LLA inductions, security
    clearance and escorted access are to be coordinated prior to installation" -
    which names who arranges it and is silent on who pays, and sits in the
    narrative rather than in the EXCLUSIONS column where a client looks. On
    30/07 the client re-phased the works into TWO visits to keep the gates open
    and we agreed at no cost inside half an hour, which was the right
    relationship call on a job with 7,212.21 of headroom over the BSW buy - but
    it was made without anyone being able to say what a visit was worth, because
    nothing in the workbook answered that.

    WHY NO OTHER RULE SEES IT. Every check in this file reads the rows that ARE
    priced. This one is about a row that does not exist. The workbook reconciles
    perfectly, the uplift is on the line, the supplier quote covers the units -
    and the site is one where nobody can walk in unescorted.

    THE HABIT: whenever the site imposes a cost that is not glazing - escorted
    or inducted access, security clearance, permits, a live/occupied building,
    phasing into more than one visit, night or weekend working - it is either a
    PRICED ALLOWANCE or it is in the EXCLUSIONS of the issued document. Silence
    is neither, and silence reads to the client as included.

    'site_access': {'constraints': [...], 'allowance_gbp': n, 'excluded': [...],
                    'visits': n}
    `excluded` holds only constraints named in the issued document's exclusions -
    a sentence in the executive summary is not an exclusion, which is the whole
    lesson."""
    sa = m.get("site_access")
    if sa is None:
        return result("site access is priced or excluded", UNKNOWN,
                      "State it: 'site_access': {'constraints': [...], 'allowance_gbp': n, "
                      "'excluded': [...], 'visits': n}. Constraints are the things about THIS "
                      "site that cost money and are not glazing - escorted or inducted access, "
                      "security clearance, permits, live/occupied premises, phased visits, "
                      "night or weekend working. An unrestricted site is an empty list, not a "
                      "blank.",
                      "Luton Airport Departure Gates 1 & 2",
                      remedy="Read the enquiry and the proposal for what the site demands, then "
                             "read the workbook for the line that pays for it.")
    if not isinstance(sa, dict):
        return result("site access is priced or excluded", UNKNOWN,
                      "'site_access' should be an object, not %r" % type(sa).__name__,
                      "Luton Airport Departure Gates 1 & 2")
    constraints = [str(c).strip() for c in (sa.get("constraints") or []) if str(c).strip()]
    if not constraints:
        return result("site access is priced or excluded", NA,
                      "no access constraints stated on this site",
                      "Luton Airport Departure Gates 1 & 2")
    try:
        allowance = float(sa.get("allowance_gbp") or 0)
    except (TypeError, ValueError):
        return result("site access is priced or excluded", UNKNOWN,
                      "'allowance_gbp' is %r, which is not a number - and 'not checked' is not "
                      "'covered'" % sa.get("allowance_gbp"),
                      "Luton Airport Departure Gates 1 & 2")
    try:
        visits = int(sa.get("visits") or 1)
    except (TypeError, ValueError):
        visits = 1
    excluded = [str(e).strip().lower() for e in (sa.get("excluded") or []) if str(e).strip()]

    def is_excluded(c):
        c = c.lower()
        return any(e in c or c in e for e in excluded)

    naked = [c for c in constraints if not is_excluded(c)]
    if naked and allowance <= 0:
        detail = ("The site costs money in ways the price does not: %s. There is no allowance "
                  "in the workbook and none of it is in the issued exclusions."
                  % "; ".join(naked))
        if visits > 1:
            detail += (" And the programme is %d separate visits, so every one of them recurs "
                       "%d times." % (visits, visits))
        return result("site access is priced or excluded", FAIL, detail,
                      "Luton Airport Departure Gates 1 & 2",
                      remedy="Either put a mobilisation/attendance line in the workbook, or name "
                             "these in the EXCLUSIONS of the proposal. A sentence in the summary "
                             "saying they will be 'coordinated' is neither.")
    if naked:
        return result("site access is priced or excluded", PASS,
                      "GBP %s allowed against %d access constraint(s) over %d visit(s)"
                      % (format(allowance, ",.2f"), len(constraints), visits),
                      "Luton Airport Departure Gates 1 & 2")
    return result("site access is priced or excluded", PASS,
                  "%d access constraint(s), all named in the issued exclusions" % len(constraints),
                  "Luton Airport Departure Gates 1 & 2")


RULES = [
    check_rfq_answered, check_uplift_applied, check_site_access_is_priced_or_excluded,
    check_our_qualifications_survive_signature,
    check_priced_scope_is_not_excluded, check_bought_in_lump_has_a_quantity_basis,
    check_system_coupling, check_panic_hardware, check_glass_ownership, check_quantities,
    check_scope_gaps, check_supplier_quote_currency, check_net_pricing,
    check_full_height_screens, check_fabricator_can_make_it, check_uvalue_basis,
    check_finish_substitution, check_supplier_covers_quantity,
    check_system_performance, check_quote_validity_against_commitment,
    check_free_delivery_threshold, check_spec_label_matches_evidence,
    check_incorporated_terms_held, check_exclusions_reach_the_issued_document,
    check_exposures_state_our_recourse, check_no_third_party_traces_in_issued_files,
    check_priced_document_view_is_intact, check_warranty_is_back_to_back,
]


def blank_manifest(job):
    return {
        "job": job,
        "created": dt.datetime.now().strftime("%Y-%m-%d"),
        "_help": "Fill every field. A field left null fails the run - that is deliberate: "
                 "the errors this catches are all errors of silence.",
        "coupled_runs": None,
        "doors": None,
        "frame_supply": None,
        "glass_order": None,
        "quantities": None,
        "spec_items": None,
        "supplier_quotes": None,
        "full_height_screens": None,
        "systems_specified": None,
        "finishes": None,
        "u_value": None,
        "supplier_coverage": None,
        "priced_rows": None,
        "site_access": None,
        "rfq_items": None,
        "price_commitment": None,
        "delivery_terms": None,
        "incorporated_terms": None,
        "qualification_regime": None,
        "priced_lines": None,
        "bought_in_lines": None,
        "issued_documents": None,
        "exposures": None,
        "own_domains": None,
    }


def run(manifest):
    """Run every rule. A rule that raises loses itself, not the rest of the run.

    Hardened 28/07/2026 after riverside found that a `free_delivery_threshold`
    written as the string "5000" raised a TypeError inside
    check_free_delivery_threshold. This was a list comprehension, so that one
    exception aborted the WHOLE run - and because rules execute in list order,
    what you lost depended on where the crash sat. That rule is second from
    last, so the crash was silently taking check_spec_label_matches_evidence
    with it every single time.

    The specific TypeError was riverside's to fix and they have. This is the
    other half: no single rule should be able to decide how many of the other
    fifteen you get to see. A crash is now a FAIL on that rule alone, named,
    and the run continues - because a checker whose failure mode is "print
    nothing at all" is the worst version of every reporting bug found tonight.
    """
    out = []
    for rule in RULES:
        try:
            out.append(rule(manifest))
        except Exception as exc:
            out.append(result(getattr(rule, "__name__", "unknown rule"), FAIL,
                              "This rule CRASHED and checked nothing: %s: %s. The finding it "
                              "exists to catch has NOT been ruled out - treat it as unchecked, "
                              "not as passed." % (type(exc).__name__, exc),
                              remedy="Fix the rule or the manifest value that broke it, then "
                                     "re-run before issuing anything."))
    return out


def report(results, job=""):
    """Print every result, and every character of the ones that gate the price.

    Fixed 28/07/2026, Gordon Court, applying riverside's general form: a report
    that omits a category is worse than one that shows it wrongly, because the
    output looks clean and clean is not the same as complete.

    This function used to print detail[:96], then detail[96:200] for FAIL and
    ASK, and stop. On Gordon Court that threw away 1,877 of the 2,077 characters
    of the "spec covered or excluded" FAIL - 90%, cut mid-word, with no ellipsis
    and no count. That rule lists the items neither priced nor excluded: it named
    NINETEEN and three reached the screen. Among the sixteen nobody ever saw were
    curtain walling priced nowhere, the strip-out allocation, the demolition
    elevations, and the closing sentence "A silent gap reads as included to the
    client" - which was itself silently dropped.

    So: FAIL and ASK now wrap in FULL, because those are the lines that decide
    whether a price goes out. PASS and n/a stay on one line but say how much was
    cut, so nothing is lost without the reader being told.
    """
    fails = [r for r in results if r["status"] == FAIL]
    unknowns = [r for r in results if r["status"] == UNKNOWN]
    width = max(len(r["rule"]) for r in results)
    indent = " " * (width + 9)
    body = max(40, 118 - width)
    print("PRE-ISSUE CHECKS%s" % (" - " + job if job else ""))
    print("=" * (width + 60))
    for r in results:
        mark = {PASS: "PASS", FAIL: "FAIL", UNKNOWN: "ASK ", NA: "  - "}[r["status"]]
        detail = r["detail"]
        if r["status"] in (FAIL, UNKNOWN):
            lines = textwrap.wrap(detail, body) or [""]
            print("  [%s] %-*s  %s" % (mark, width, r["rule"], lines[0]))
            for extra in lines[1:]:
                print("%s%s" % (indent, extra))
            # Printed last but never abridged. It used to be the tail of `detail`,
            # which meant the longer the list of faults the further it slid past
            # the cut - see result().
            for i, extra in enumerate(textwrap.wrap(r.get("remedy") or "", body - 4)):
                print("%s%s %s" % (indent, "->" if i == 0 else "  ", extra))
        else:
            if len(detail) > body:
                detail = "%s... (+%d chars)" % (detail[:body], len(r["detail"]) - body)
            print("  [%s] %-*s  %s" % (mark, width, r["rule"], detail))
    print("=" * (width + 60))
    if fails:
        print("%d FAILED - do not issue this quote." % len(fails))
    if unknowns:
        print("%d question(s) unanswered. Unanswered is not the same as fine - every error this "
              "catches was an error of silence." % len(unknowns))
    if not fails and not unknowns:
        print("All checks pass.")
    return 0 if not fails and not unknowns else 1


# Riverside, 28/07/2026, written BEFORE check_incorporated_terms_held shipped
# rather than after. The whole point of the delivery exercise was that a rule
# validated on one positive case has measured precision and called it quality,
# so this one gets its variants first: eight that must FIRE and eight that must
# stay silent, including the three shapes that crash a rule rather than answer
# it - a dict where a list belongs, a bare string, and a non-dict entry.
TERMS_VARIANTS = [
    # (name, manifest value, expected status)
    ("field absent",           None,                                            UNKNOWN),
    ("empty list",             [],                                              NA),
    ("held true",              [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": True}],  PASS),
    ("held 'yes'",             [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": "yes"}], PASS),
    ("held 'attached'",        [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2",
                                 "held": "attached"}],                          PASS),
    ("held 1",                 [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": 1}], PASS),
    ("held false",             [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": False}], UNKNOWN),
    ("held 'no'",              [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": "no"}],  UNKNOWN),
    ("held 0",                 [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": 0}],     UNKNOWN),
    ("held unstated",          [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2"}],           UNKNOWN),
    ("held None",              [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": None}],  UNKNOWN),
    ("held 'maybe'",           [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2",
                                 "held": "maybe"}],                              UNKNOWN),
    ("no document named",      [{"supplier": "A Plus", "ref": "QT51518", "held": True}], UNKNOWN),
    ("one held, one not",      [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": True},
                                {"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Definitions V.01", "held": False}],  UNKNOWN),
    ("a dict, not a list",     {"supplier": "A Plus", "ref": "QT51518",
                                "document": "Terms of Sale V.01.2", "held": True},  PASS),
    ("a bare string",          "Terms of Sale V.01.2",                           UNKNOWN),
    ("entry is not a dict",    ["Terms of Sale V.01.2"],                         UNKNOWN),
    # Twelve more written AFTER the first seventeen passed, deliberately chosen
    # from shapes the implementation was not written against - because a suite
    # that passes first time may be testing the code's own assumptions back at
    # it rather than the behaviour. These all held; that is worth recording too.
    ("held 'TRUE' uppercase",  [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": "TRUE"}], PASS),
    ("held ' yes ' padded",    [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": " yes "}], PASS),
    ("document is whitespace", [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "   ", "held": True}],              UNKNOWN),
    ("held is an empty list",  [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": []}], UNKNOWN),
    ("held 2",                 [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": 2}],  UNKNOWN),
    ("held 'n/a'",             [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": "n/a"}], UNKNOWN),
    ("held is a dict",         [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2",
                                 "held": {"status": True}}],                     UNKNOWN),
    ("field is an int",        7,                                                UNKNOWN),
    ("a tuple of entries",     ({"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": False},), UNKNOWN),
    ("document is a number",   [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": 12345, "held": False}],             UNKNOWN),
    ("held 'Not Held'",        [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2",
                                 "held": "Not Held"}],                           UNKNOWN),
    ("entry is None",          [None],                                           UNKNOWN),
    # Gordon Court, 28/07, the first time this rule saw data that was not mine.
    # BSW's four quotations incorporate terms "available on request" - no title,
    # no revision, no date. The rule had no branch for it: an unnamed
    # incorporation fell into the manifest-filling bucket and got a remedy the
    # estimator could not carry out. These six fix that in both directions.
    ("unnamed, held false",    [{"supplier": "BSW", "ref": "Q1234", "held": False}],  UNKNOWN),
    ("unnamed, held unstated", [{"supplier": "BSW", "ref": "Q1234"}],                 UNKNOWN),
    ("unnamed, held 'no'",     [{"supplier": "BSW", "ref": "Q1234", "held": "no"}],   UNKNOWN),
    ("unnamed but marked held", [{"supplier": "BSW", "ref": "Q1234", "held": True}],  UNKNOWN),
    ("unnamed, document blank", [{"supplier": "BSW", "ref": "Q1234",
                                  "document": "", "held": False}],                    UNKNOWN),
    ("unnamed and named together",
                               [{"supplier": "BSW", "ref": "Q1234", "held": False},
                                {"supplier": "A Plus", "ref": "QT51518",
                                 "document": "Terms of Sale V.01.2", "held": False}],  UNKNOWN),
    # Gordon Court defeated the unnamed branch within an hour of it shipping, by
    # writing an accurate PROSE DESCRIPTION of the absence into the field whose
    # emptiness was the signal. These eleven test _describes_absence in both
    # directions, and deliberately use three drafting voices rather than one -
    # last turn's suite was 29 cases all written against A Plus's phrasing.
    ("document describes absence (Gordon Court's exact value)",
                               [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "BSW terms and conditions of sale, available on "
                                             "request - no revision, no date, no title",
                                 "held": False}],                                 UNKNOWN),
    ("document 'available on request'",
                               [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "Terms of Sale, available on request",
                                 "held": False}],                                 UNKNOWN),
    ("document 'TBC'",         [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "TBC", "held": False}],              UNKNOWN),
    ("document 'unnamed'",     [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "unnamed - the quote just says conditions apply",
                                 "held": False}],                                 UNKNOWN),
    ("document 'not stated'",  [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "revision not stated", "held": False}], UNKNOWN),
    ("document 'n/a'",         [{"supplier": "BSW", "ref": "Q1234",
                                 "document": "Conditions of Sale n/a", "held": False}], UNKNOWN),
    # ...and the negatives. A real document name must NOT be read as prose,
    # including names that contain risky-looking substrings.
    ("real name with a revision and date",
                               [{"supplier": "A Plus", "ref": "QT51518",
                                 "document": "A Plus Windows & Doors Limited Terms of Sale "
                                             "Revision V.01.2 - 08.01.2018",
                                 "held": True}],                                  PASS),
    ("real name, AFS voice",   [{"supplier": "AFS", "ref": "Q7585",
                                 "document": "AFS Conditions of Contract Q7585 "
                                             "(16pp, printed in full)", "held": True}],  PASS),
    ("real name containing 'NA/EU'",
                               [{"supplier": "X", "ref": "Q1",
                                 "document": "Terms and Conditions - NA/EU editions",
                                 "held": True}],                                  PASS),
    ("real name containing 'National'",
                               [{"supplier": "X", "ref": "Q1",
                                 "document": "Conditions of Sale - National Association of "
                                             "Glazing Contractors form", "held": True}],  PASS),
    ("real name, edition not revision",
                               [{"supplier": "BSW", "ref": "Q1",
                                 "document": "BSW Standard Conditions of Sale, edition 4, "
                                             "March 2024", "held": True}],        PASS),
]


DELIVERY_VARIANTS = [
    # Riverside, 28/07/2026. Gordon Court's point: a detector validated against one
    # positive case has measured PRECISION and called it quality. This rule shipped
    # on 27/07 with exactly one fixture - the one it was built from. Sixteen variants
    # of the same field found two real defects: a numeric field written as a string
    # crashed the whole run, and an unrecognised delivery_priced value was silently
    # read as "not priced", asserting something false about the world.
    # (name, term-dict overrides, expected status)
    ("baseline under threshold, not priced",   {},                                              FAIL),
    ("delivery_priced True",                   {"delivery_priced": True},                       PASS),
    ("provisional lowercase",                  {"delivery_priced": "provisional"},              UNKNOWN),
    ("provisional CAPITALISED",                {"delivery_priced": "PROVISIONAL"},              UNKNOWN),
    ("provisional, no charge_basis",           {"delivery_priced": "provisional",
                                                "charge_basis": None},                          FAIL),
    ("order value equals threshold",           {"order_value": 5000.0},                         PASS),
    ("threshold 0 - always free",              {"free_delivery_threshold": 0},                  PASS),
    ("threshold 'never'",                      {"free_delivery_threshold": "never"},            FAIL),
    ("order_value missing",                    {"order_value": None},                           UNKNOWN),
    ("threshold as string '5000'",             {"free_delivery_threshold": "5000"},             FAIL),
    ("delivery_priced 'yes'",                  {"delivery_priced": "yes"},                      UNKNOWN),
    ("delivery_priced None",                   {"delivery_priced": None},                       FAIL),
    ("order_value as string",                  {"order_value": "4845.22"},                      FAIL),
    ("threshold gibberish",                    {"free_delivery_threshold": "ask them"},         UNKNOWN),
    ("delivery_priced 'no'",                   {"delivery_priced": "no"},                       FAIL),
    ("'never' with string order_value",        {"free_delivery_threshold": "never",
                                                "order_value": "4845.22"},                      FAIL),
]


def selftest_coupling_remedy():
    """SM5 Wexham, 29/07/2026. The coupling rule used to imply one remedy - put
    the window in the door's system. BSW then ruled in writing that Smart Wall
    has no window in it at all, so on a Smart Wall run that remedy is a dead end.
    Assert both arms: a blocked system says so, an ordinary depth mismatch does
    not, and the mismatch itself still FAILs either way."""
    smartwall = {"coupled_runs": [{"name": "East run W.01 + ED.01", "elements": [
        {"ref": "W.01", "system": "Sheerline Prestige"},
        {"ref": "ED.01", "system": "SMA Smart Wall"}]}]}
    ordinary = {"coupled_runs": [{"name": "refs 32/34", "elements": [
        {"ref": "W32", "system": "EL75mm Squareline"},
        {"ref": "D34", "system": "AC100 Commercial"}]}]}
    a, b = check_system_coupling(smartwall), check_system_coupling(ordinary)
    checks = [
        ("smart wall run FAILs", a["status"] == FAIL),
        ("smart wall remedy warns off the dead end", "no window product" in a["remedy"]),
        ("smart wall remedy cites BSW", "BSW 29/07/2026" in a["remedy"]),
        ("ordinary run FAILs", b["status"] == FAIL),
        ("ordinary remedy does NOT warn", "no window product" not in b["remedy"]),
    ]
    bad = [n for n, got in checks if not got]
    print("  %-22s %d/%d coupling remedies behave as intended%s"
          % ("coupling remedy", len(checks) - len(bad), len(checks),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_delivery_variants():
    """Recall test for check_free_delivery_threshold - see DELIVERY_VARIANTS."""
    base = {"supplier": "A Plus", "ref": "QT51518", "order_value": 4845.22,
            "free_delivery_threshold": 5000.0, "charge_basis": "1/mile each way",
            "delivery_priced": False}
    bad = []
    for name, over, expect in DELIVERY_VARIANTS:
        t = dict(base)
        t.update(over)
        t = {k: v for k, v in t.items() if not (k == "charge_basis" and v is None)}
        try:
            got = check_free_delivery_threshold({"delivery_terms": [t]})["status"]
        except Exception as exc:
            got = "EXCEPTION %s" % type(exc).__name__
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d delivery variants behave as intended%s"
          % ("delivery recall", len(DELIVERY_VARIANTS) - len(bad), len(DELIVERY_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


# Riverside, 28/07. Written before check_exclusions_reach_the_issued_document
# shipped, and split evenly: eight that must FAIL or ASK, seven that must not.
ISSUED_VARIANTS = [
    # (name, spec_items, issued_documents, expected)
    ("field absent",            [{"ref": "x", "treatment": "excluded"}], None,          UNKNOWN),
    ("nothing excluded",        [{"ref": "x", "treatment": "priced"}],   [],            NA),
    ("no issued doc recorded",  [{"ref": "x", "treatment": "excluded"}], [],            FAIL),
    ("priced doc states none",  [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": 0}],                             FAIL),
    ("priced doc states some",  [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": 12}],                            PASS),
    ("stated as a list",        [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": ["testing", "scaffold"]}],       PASS),
    ("empty list is none",      [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": []}],                            FAIL),
    ("count unstated",          [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document",
                                  "is_the_priced_document": True}],                     UNKNOWN),
    ("count is prose",          [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": "a few"}],                       UNKNOWN),
    ("count as a numeric string", [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": "12"}],                          PASS),
    ("only a non-priced doc",   [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Covering letter", "is_the_priced_document": False,
                                  "exclusions_stated": 12}],                            PASS),
    ("covering letter carries them, priced doc does not",
                                [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Covering letter", "is_the_priced_document": False,
                                  "exclusions_stated": 12},
                                 {"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": 0}],                             FAIL),
    # Gordon Court, 28/07, referred back as a design question. Two PRICED
    # documents in one pack, one carrying the exclusions and one not. The
    # covering-letter case above still FAILS because a covering letter is
    # detachable and unpriced; this one ASKS, because partial coverage across
    # priced documents is a judgement about how the pack is used.
    ("two priced documents, one carries them, one does not",
                                [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Proposal.pdf", "is_the_priced_document": True,
                                  "exclusions_stated": 12},
                                 {"name": "Pricing.xlsx", "is_the_priced_document": True,
                                  "exclusions_stated": 0}],                             UNKNOWN),
    ("the bare priced document is not client-facing",
                                [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Proposal.pdf", "is_the_priced_document": True,
                                  "exclusions_stated": 12},
                                 {"name": "Working.xlsx", "is_the_priced_document": True,
                                  "exclusions_stated": 0, "goes_to_client": False}],    PASS),
    ("every priced document carries them",
                                [{"ref": "x", "treatment": "excluded"}],
                                [{"name": "Proposal.pdf", "is_the_priced_document": True,
                                  "exclusions_stated": 12},
                                 {"name": "Pricing.xlsx", "is_the_priced_document": True,
                                  "exclusions_stated": 13}],                            PASS),
    ("entry is not a dict",     [{"ref": "x", "treatment": "excluded"}], ["Pricing Document"],
                                                                                        UNKNOWN),
    ("no spec items at all",    None,
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": 0}],                             NA),
    ("provisional is not excluded", [{"ref": "x", "treatment": "provisional"}],
                                [{"name": "Pricing Document", "is_the_priced_document": True,
                                  "exclusions_stated": 0}],                             NA),
]


# Riverside, 28/07, written before the rule shipped. Seven that must fire,
# seven that must not - and the "unknown"/"tbc" cases matter most, because the
# obvious way to defeat this rule is to fill the field with a word that looks
# like a value and means silence. Gordon Court defeated the last new branch
# within an hour by writing prose into a field; this anticipates the same shape.
EXPOSURE_VARIANTS = [
    ("field absent",           None,                                              UNKNOWN),
    ("empty list",             [],                                                NA),
    ("recourse stated",        [{"item": "storage clock", "lands_on": "us",
                                 "our_recourse": "cl. Cancellation and Postponement"}], PASS),
    ("recourse 'none'",        [{"item": "free area basis", "lands_on": "us",
                                 "our_recourse": "none"}],                        PASS),
    ("recourse 'None' capitalised",
                               [{"item": "x", "lands_on": "us",
                                 "our_recourse": "None - nothing in our terms bears on it"}], PASS),
    ("recourse unstated",      [{"item": "storage clock", "lands_on": "us"}],     UNKNOWN),
    ("recourse null",          [{"item": "x", "lands_on": "us",
                                 "our_recourse": None}],                          UNKNOWN),
    ("recourse blank",         [{"item": "x", "lands_on": "us",
                                 "our_recourse": "   "}],                         UNKNOWN),
    ("recourse 'unknown'",     [{"item": "x", "lands_on": "us",
                                 "our_recourse": "unknown"}],                     UNKNOWN),
    ("recourse 'TBC'",         [{"item": "x", "lands_on": "us",
                                 "our_recourse": "TBC"}],                         UNKNOWN),
    ("recourse 'not checked'", [{"item": "x", "lands_on": "us",
                                 "our_recourse": "not checked"}],                 UNKNOWN),
    ("one stated one not",     [{"item": "a", "lands_on": "us", "our_recourse": "cl.9"},
                                {"item": "b", "lands_on": "us"}],                 UNKNOWN),
    ("a dict, not a list",     {"item": "a", "lands_on": "us", "our_recourse": "cl.9"}, PASS),
    ("a bare string",          "storage",                                         UNKNOWN),
    ("entry is not a dict",    ["storage"],                                       UNKNOWN),
]


# Riverside, 28/07, from Gordon Court's "print one real entry before comparing
# anything to anything". Printing supplier_coverage[0] showed both vents
# claiming qty_quoted 2 from a quotation whose single position reads "Qty (2)" -
# four units asserted against two sold - and the rule PASSED, because it only
# ever asked whether quoted < sold. The last two cases pin the reference
# matching, which failed on the real strings the first time it was written.
COVERAGE_VARIANTS = [
    ("Brocks Hill founding case - 2 sold, 1 quoted",
     [{"ref": "E.04", "qty_sold": 2, "qty_quoted": 1, "supplier_ref": "0000000503"}], [], FAIL),
    ("the Riverside bug - both lines claim the same 2 units",
     [{"ref": "AOV.01", "qty_sold": 1, "qty_quoted": 2, "supplier_ref": "A Plus QT51518"},
      {"ref": "AOV.02", "qty_sold": 1, "qty_quoted": 2, "supplier_ref": "A Plus QT51518"}],
     [{"supplier": "A Plus Windows & Doors", "ref": "QT51518", "qty_total": 2}], FAIL),
    ("corrected - one unit each",
     [{"ref": "AOV.01", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "A Plus QT51518"},
      {"ref": "AOV.02", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "A Plus QT51518"}],
     [{"supplier": "A Plus Windows & Doors", "ref": "QT51518", "qty_total": 2}], PASS),
    ("two lines, one quote, no qty_total recorded",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"},
      {"ref": "B", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"}], [], UNKNOWN),
    ("one line only - over-claim impossible, stays quiet",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"}], [], PASS),
    ("two quotes, one line each - stays quiet",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"},
      {"ref": "B", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT10"}], [], PASS),
    ("claim equals the quotation exactly",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"},
      {"ref": "B", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"}],
     [{"ref": "QT9", "qty_total": 2}], PASS),
    ("bare reference matches a composite one",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 2, "supplier_ref": "QT9"},
      {"ref": "B", "qty_sold": 1, "qty_quoted": 2, "supplier_ref": "QT9"}],
     [{"supplier": "Someone Ltd", "ref": "QT9", "qty_total": 2}], FAIL),
    ("a quote reference that matches nothing stays unbounded",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"},
      {"ref": "B", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"}],
     [{"ref": "QT99", "qty_total": 2}], UNKNOWN),
    # Gordon Court, 28/07 - the surplus arm, from their real numbers. BSW quote
    # two WE_14 and the schedule has one, so 118 units are on QT252247 and 117
    # are sold against it: GBP 921.29 of quoted cost with nothing sold behind
    # it, invisible to every earlier version of this rule.
    ("Gordon Court QT252247 - 118 contained, 117 sold",
     [{"ref": "WE_%d" % i, "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "BSW QT252247"}
      for i in range(117)],
     [{"supplier": "BSW", "ref": "QT252247", "qty_total": 118}], UNKNOWN),
    ("their reconciling quote - 44 contained, 44 sold",
     [{"ref": "P%d" % i, "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "BSW QT252248"}
      for i in range(44)],
     [{"supplier": "BSW", "ref": "QT252248", "qty_total": 44}], PASS),
    ("Riverside - 2 contained, 1 + 1 sold, balanced",
     [{"ref": "AOV.01", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "A Plus QT51518"},
      {"ref": "AOV.02", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "A Plus QT51518"}],
     [{"supplier": "A Plus Windows & Doors", "ref": "QT51518", "qty_total": 2}], PASS),
    ("a shortfall still beats a surplus to the answer",
     [{"ref": "E.04", "qty_sold": 2, "qty_quoted": 1, "supplier_ref": "0000000503"}],
     [{"ref": "0000000503", "qty_total": 1}], FAIL),
    ("surplus on one line only, no second line - still caught",
     [{"ref": "A", "qty_sold": 1, "qty_quoted": 1, "supplier_ref": "QT9"}],
     [{"ref": "QT9", "qty_total": 3}], UNKNOWN),
]


def selftest_coverage_variants():
    """Recall test for the over-claim arm of check_supplier_covers_quantity."""
    bad = []
    for name, cov, quotes, expect in COVERAGE_VARIANTS:
        try:
            got = check_supplier_covers_quantity(
                {"supplier_coverage": cov, "supplier_quotes": quotes})["status"]
        except Exception as exc:
            got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d coverage variants behave as intended%s"
          % ("coverage over-claim", len(COVERAGE_VARIANTS) - len(bad), len(COVERAGE_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_view_variants():
    """Recall test for check_priced_document_view_is_intact.

    Synthetic workbooks built and destroyed here. The founding cases are the
    two both chats actually committed - print area deleted, and print titles
    deleted while the print area was restored - plus the one that matters
    commercially: a cell populated outside the printed range.
    """
    import shutil
    import tempfile
    import openpyxl
    d = tempfile.mkdtemp(prefix="mary-view-")
    try:
        def book(name, area, titles, extra=None):
            p = os.path.join(d, name)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["C3"] = "Client:"
            ws["I23"] = 5990.22
            if extra:
                ws[extra] = 2331.075
            if area:
                ws.print_area = area
            if titles:
                ws.print_title_rows = titles
            wb.save(p)
            return p

        good = book("good.xlsx", "$C$1:$I$45", "$2:$7")
        no_area = book("noarea.xlsx", None, "$2:$7")
        no_titles = book("notitles.xlsx", "$C$1:$I$45", None)
        buy_outside = book("buy.xlsx", "$C$1:$I$45", "$2:$7", extra="J9")
        both_gone = book("bothgone.xlsx", None, None)

        VARIANTS = [
            ("field absent",             None,                                        UNKNOWN),
            ("empty list",               [],                                          NA),
            ("no priced workbook",       [{"name": "note", "path": "x.txt",
                                           "is_the_priced_document": True}],          NA),
            ("priced flag not set",      [{"name": "g", "path": good}],               NA),
            ("intact",                   [{"name": "g", "path": good,
                                           "is_the_priced_document": True}],          PASS),
            ("no print area",            [{"name": "n", "path": no_area,
                                           "is_the_priced_document": True}],          FAIL),
            ("print titles destroyed",   [{"name": "t", "path": no_titles,
                                           "is_the_priced_document": True}],          FAIL),
            ("buy price outside the printed range",
                                         [{"name": "b", "path": buy_outside,
                                           "is_the_priced_document": True}],          FAIL),
            ("both defined names gone",  [{"name": "x", "path": both_gone,
                                           "is_the_priced_document": True}],          FAIL),
            ("one intact one broken",    [{"name": "g", "path": good,
                                           "is_the_priced_document": True},
                                          {"name": "b", "path": buy_outside,
                                           "is_the_priced_document": True}],          FAIL),
            ("path does not exist",      [{"name": "x", "path": os.path.join(d, "no.xlsx"),
                                           "is_the_priced_document": True}],          UNKNOWN),
            ("a dict, not a list",       {"name": "g", "path": good,
                                          "is_the_priced_document": True},            PASS),
            ("a bare string",            "good.xlsx",                                 UNKNOWN),
            ("entry is not a dict",      ["good.xlsx"],                               NA),
        ]

        bad = []
        for name, value, expect in VARIANTS:
            m = {} if value is None else {"issued_documents": value}
            try:
                got = check_priced_document_view_is_intact(m)["status"]
            except Exception as exc:
                got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
            if got != expect:
                bad.append("%s: expected %s, got %s" % (name, expect, got))
        print("  %-22s %d/%d client-view variants behave as intended%s"
              % ("client view", len(VARIANTS) - len(bad), len(VARIANTS),
                 "" if not bad else "  MISSED: " + "; ".join(bad)))
        return not bad
    finally:
        shutil.rmtree(d, ignore_errors=True)


def selftest_trace_variants():
    """Recall test for check_no_third_party_traces_in_issued_files.

    Synthetic files, built and destroyed here, so the suite survives the
    template it was founded on being cleaned.
    """
    import shutil
    import tempfile
    import zipfile
    d = tempfile.mkdtemp(prefix="mary-trace-")
    try:
        def ooxml(name, core):
            p = os.path.join(d, name)
            with zipfile.ZipFile(p, "w") as z:
                z.writestr("docProps/core.xml", core)
                z.writestr("xl/worksheets/sheet1.xml", "<sheet><v>5990.22</v></sheet>")
            return p

        def flat(name, body):
            p = os.path.join(d, name)
            with open(p, "wb") as fh:
                fh.write(body)
            return p

        dirty = ooxml("dirty.xlsx", "<cp><dc:creator>Dan Parker;"
                                    "dan.parker@agsurveying.co.uk</dc:creator></cp>")
        clean = ooxml("clean.xlsx", "<cp><dc:creator>Fenster Glazing &amp; Locks Ltd"
                                    "</dc:creator></cp>")
        ours = ooxml("ours.xlsx", "<cp><dc:creator>adam@fensterglazing.com</dc:creator></cp>")
        path = ooxml("path.xlsx", "<cp><x>C:\\Users\\LiamO'Donnell\\AppData\\Local"
                                  "\\Microsoft\\Windows\\INetCache</x></cp>")
        plain = flat("plain.txt", b"Riverside House - nothing personal in here at all.")
        email = flat("email.txt", b"contact hayley@hdplanning.co.uk about the approval")
        # binary that is NOT text - the shape that produced six false "emails"
        # out of the drawings PDF before the printable guard went in
        binary = flat("binary.pdf", bytes(range(256)) * 40)

        # Gordon Court's exact false positive, and the shapes either side of it.
        ffat = flat("ffat.txt", b"noise ff@C.0 more noise")
        short_tld = flat("shorttld.txt", b"someone@example.c")
        numeric_tld = flat("numtld.txt", b"someone@example.11")
        real = flat("real.txt", b"write to dan.parker@agsurveying.co.uk about it")

        VARIANTS = [
            ("Gordon Court's 'ff@C.0' - not an address",
                                        [{"name": "f", "path": ffat}],              PASS),
            ("one-character TLD",       [{"name": "s", "path": short_tld}],         PASS),
            ("numeric TLD",             [{"name": "n", "path": numeric_tld}],       PASS),
            ("a real third-party address still fires",
                                        [{"name": "r", "path": real}],              FAIL),
            ("field absent",            None,                                       UNKNOWN),
            ("empty list",              [],                                         NA),
            ("clean ooxml",             [{"name": "c", "path": clean}],             PASS),
            ("our own domain allowed",  [{"name": "o", "path": ours}],              PASS),
            ("plain text, nothing",     [{"name": "p", "path": plain}],             PASS),
            ("binary, no real text",    [{"name": "b", "path": binary}],            PASS),
            ("third-party email in docProps",
                                        [{"name": "d", "path": dirty}],             FAIL),
            ("windows user path",       [{"name": "w", "path": path}],              FAIL),
            ("third-party email in a txt",
                                        [{"name": "e", "path": email}],             FAIL),
            ("one clean one dirty",     [{"name": "c", "path": clean},
                                         {"name": "d", "path": dirty}],             FAIL),
            ("no path given",           [{"name": "x"}],                            UNKNOWN),
            ("path does not exist",     [{"name": "x", "path": os.path.join(d, "nope.xlsx")}],
                                                                                    UNKNOWN),
            ("entry is not a dict",     ["clean.xlsx"],                             UNKNOWN),
            ("a dict, not a list",      {"name": "c", "path": clean},               PASS),
            ("a bare string",           "clean.xlsx",                               UNKNOWN),
        ]

        bad = []
        for name, value, expect in VARIANTS:
            m = {} if value is None else {"issued_documents": value}
            try:
                got = check_no_third_party_traces_in_issued_files(m)["status"]
            except Exception as exc:
                got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
            if got != expect:
                bad.append("%s: expected %s, got %s" % (name, expect, got))
        print("  %-22s %d/%d trace variants behave as intended%s"
              % ("third-party traces", len(VARIANTS) - len(bad), len(VARIANTS),
                 "" if not bad else "  MISSED: " + "; ".join(bad)))
        return not bad
    finally:
        shutil.rmtree(d, ignore_errors=True)


def selftest_exposure_variants():
    """Recall test for check_exposures_state_our_recourse."""
    bad = []
    for name, value, expect in EXPOSURE_VARIANTS:
        m = {} if value is None else {"exposures": value}
        try:
            got = check_exposures_state_our_recourse(m)["status"]
        except Exception as exc:
            got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d exposure variants behave as intended%s"
          % ("exposure recourse", len(EXPOSURE_VARIANTS) - len(bad), len(EXPOSURE_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_issued_variants():
    """Recall test for check_exclusions_reach_the_issued_document."""
    bad = []
    for name, items, docs, expect in ISSUED_VARIANTS:
        m = {"spec_items": items}
        if docs is not None:
            m["issued_documents"] = docs
        try:
            got = check_exclusions_reach_the_issued_document(m)["status"]
        except Exception as exc:
            got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d issued-document variants behave as intended%s"
          % ("exclusions issued", len(ISSUED_VARIANTS) - len(bad), len(ISSUED_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


WARRANTY_OURS = {
    "period": "10 years", "period_months": 120,
    "scope": "all glass and frame products supplied and installed",
    "start_date": "practical completion", "usage_cap": None,
    "exclusions": ["misuse", "vandalism"],
}


def _wsup(**kw):
    """A supplier warranty that is back-to-back, before the variant breaks it."""
    s = {"supplier": "A Plus", "ref": "Q1", "covers": "frames", "period_months": 120,
         "start_date": "date of delivery completion", "usage_cap": None,
         "exclusions": [{"exclusion": "misuse", "counterpart_in_ours": "misuse"}],
         "exclusions_complete": True}
    s.update(kw)
    return s


def _wours(**kw):
    o = dict(WARRANTY_OURS)
    o.update(kw)
    return o


HELD = [{"supplier": "A Plus", "ref": "Q1", "document": "Terms of Sale", "held": True}]
UNHELD = [{"supplier": "A Plus", "ref": "Q1", "document": "Terms of Sale", "held": False}]

# (name, warranty, incorporated_terms, expected)
WARRANTY_VARIANTS = [
    ("field absent",            None,                                       None, UNKNOWN),
    ("not a dict",              "10 years",                                 None, UNKNOWN),
    ("ours missing",            {"suppliers": [_wsup()]},                   None, UNKNOWN),
    ("suppliers missing",       {"ours": _wours()},                         None, UNKNOWN),
    ("suppliers not a list",    {"ours": _wours(), "suppliers": "A Plus"},  None, UNKNOWN),
    ("nobody to compare",       {"ours": _wours(), "suppliers": []},        None, PASS),
    ("fully back-to-back",      {"ours": _wours(), "suppliers": [_wsup()]}, HELD, PASS),
    ("a dict, not a list",      {"ours": _wours(), "suppliers": _wsup()},   HELD, PASS),

    # the founding case - both jobs offer a period and never say from when
    ("OURS HAS NO START DATE",  {"ours": _wours(start_date=None),
                                 "suppliers": [_wsup()]},                   HELD, FAIL),
    ("ours start date empty",   {"ours": _wours(start_date=""),
                                 "suppliers": [_wsup()]},                   HELD, FAIL),
    ("ours scope not stated",   {"ours": _wours(scope=None),
                                 "suppliers": [_wsup()]},                   HELD, UNKNOWN),

    # the period, which was the only part anyone was comparing
    ("supplier shorter",        {"ours": _wours(),
                                 "suppliers": [_wsup(period_months=12)]},   HELD, UNKNOWN),
    ("supplier longer",         {"ours": _wours(),
                                 "suppliers": [_wsup(period_months=180)]},  HELD, PASS),
    ("supplier equal",          {"ours": _wours(),
                                 "suppliers": [_wsup(period_months=120)]},  HELD, PASS),
    ("supplier states none",    {"ours": _wours(),
                                 "suppliers": [_wsup(period_months=None)]}, HELD, UNKNOWN),

    # a period in years capped in cycles is not a period in years
    ("capped by cycles",        {"ours": _wours(),
                                 "suppliers": [_wsup(usage_cap="15,000 cycles")]},
                                                                            HELD, UNKNOWN),
    ("capped, ours capped too", {"ours": _wours(usage_cap="15,000 cycles"),
                                 "suppliers": [_wsup(usage_cap="15,000 cycles")]},
                                                                            HELD, PASS),

    # the exclusion list, which is where the wider gap turned out to be
    ("exclusion no counterpart", {"ours": _wours(), "suppliers": [_wsup(
        exclusions=[{"exclusion": "powder coat adhesion", "counterpart_in_ours": None}])]},
                                                                            HELD, UNKNOWN),
    ("exclusion matched",       {"ours": _wours(), "suppliers": [_wsup(
        exclusions=[{"exclusion": "misuse", "counterpart_in_ours": "misuse"}])]},
                                                                            HELD, PASS),
    ("exclusions not recorded", {"ours": _wours(),
                                 "suppliers": [_wsup(exclusions=None)]},    HELD, UNKNOWN),
    ("exclusions empty list",   {"ours": _wours(),
                                 "suppliers": [_wsup(exclusions=[])]},      HELD, PASS),
    ("a bare string exclusion", {"ours": _wours(),
                                 "suppliers": [_wsup(exclusions=["misuse"])]},
                                                                            HELD, UNKNOWN),
    ("list not called complete", {"ours": _wours(),
                                  "suppliers": [_wsup(exclusions_complete=False)]},
                                                                            HELD, UNKNOWN),

    # the contradiction: you cannot have read a list you do not hold
    ("COMPLETE BUT TERMS UNHELD", {"ours": _wours(), "suppliers": [_wsup()]},
                                                                            UNHELD, FAIL),
    ("supplier entry is a string", {"ours": _wours(), "suppliers": ["A Plus"]},
                                                                            HELD, UNKNOWN),

    # the split: a document defect outranks a gap, and a gap alone never FAILs
    ("gap AND no start date",   {"ours": _wours(start_date=None),
                                 "suppliers": [_wsup(period_months=12, usage_cap="15,000 cycles",
                                 exclusions=[{"exclusion": "powder coat",
                                              "counterpart_in_ours": None}])]},
                                                                            HELD, FAIL),
    ("every gap, ours sound",   {"ours": _wours(),
                                 "suppliers": [_wsup(period_months=12, usage_cap="15,000 cycles",
                                 exclusions=[{"exclusion": "powder coat",
                                              "counterpart_in_ours": None}])]},
                                                                            HELD, UNKNOWN),
]


def selftest_warranty_variants():
    """Recall test for check_warranty_is_back_to_back."""
    bad = []
    for name, w, terms, expect in WARRANTY_VARIANTS:
        m = {} if w is None else {"warranty": w}
        if terms is not None:
            m["incorporated_terms"] = terms
        try:
            got = check_warranty_is_back_to_back(m)["status"]
        except Exception as exc:
            got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d warranty variants behave as intended%s"
          % ("warranty back-to-back", len(WARRANTY_VARIANTS) - len(bad), len(WARRANTY_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_terms_variants():
    """Recall test for check_incorporated_terms_held - see TERMS_VARIANTS.

    Written before the rule shipped, not after it fired. Eight of the seventeen
    are NEGATIVES: a rule that only ever says yes has not been tested.
    """
    bad = []
    for name, value, expect in TERMS_VARIANTS:
        m = {} if value is None else {"incorporated_terms": value}
        try:
            got = check_incorporated_terms_held(m)["status"]
        except Exception as exc:
            got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
        if got != expect:
            bad.append("%s: expected %s, got %s" % (name, expect, got))
    print("  %-22s %d/%d terms variants behave as intended%s"
          % ("incorporated terms", len(TERMS_VARIANTS) - len(bad), len(TERMS_VARIANTS),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


# John North Hall, 28/07/2026. Recall suite for
# check_our_qualifications_survive_signature. Two excluded items are held
# constant so every case turns only on the regime, and the awkward shapes are
# the ones the terms suite learned to expect: prose in place of a boolean,
# a dict where a list was meant, a field somebody answered in words.
_QR_RELIED = [{"ref": "Access plant and scaffolding", "treatment": "excluded"},
              {"ref": "Waste removal and disposal", "treatment": "Excluded by us"}]
_JNH_CLAUSE = ("It is agreed that any other terms and conditions of contract or any caveats, "
               "assumptions, reservations or exclusions that may be printed on correspondence "
               "emanating from the tender, or any contract resulting from this tender, shall "
               "not be applicable to this tender or agreement.")

QUALIFICATION_VARIANTS = [
    ("field absent",            None,                                              UNKNOWN),
    ("John North Hall as it is", {"document": "ITT section 5.0 Form of Tender", "clause": _JNH_CLAUSE,
                                  "qualifications_permitted": False, "we_must_sign": True}, FAIL),
    ("void, inferred from the clause alone",
                                {"document": "ITT s5.0", "clause": _JNH_CLAUSE},     FAIL),
    ("void, but nothing relied on", {"document": "ITT s5.0", "clause": _JNH_CLAUSE,
                                     "qualifications_permitted": False}, PASS),   # relied=[] below
    ("qualifications permitted", {"document": "JCT tender", "qualifications_permitted": True}, PASS),
    ("permitted as 'yes'",      {"document": "d", "qualifications_permitted": "yes"},  PASS),
    ("void as 'no'",            {"document": "d", "qualifications_permitted": "no"},   FAIL),
    ("void as 'disapplied'",    {"document": "d", "qualifications_permitted": "disapplied"}, FAIL),
    ("permitted unstated, clause silent",
                                {"document": "d", "clause": "Tenders must be returned by 9am."}, UNKNOWN),
    ("permitted unstated, no clause at all", {"document": "d"},                        UNKNOWN),
    ("permitted is prose",      {"document": "d", "qualifications_permitted": "probably not"}, UNKNOWN),
    ("permitted is a dict",     {"document": "d", "qualifications_permitted": {"v": False}}, UNKNOWN),
    ("a bare string",           "no qualifications allowed",                           UNKNOWN),
    ("a list of one",           [{"document": "d", "qualifications_permitted": False}], FAIL),
    ("a list of two",           [{"document": "d", "qualifications_permitted": False},
                                 {"document": "e", "qualifications_permitted": True}],  UNKNOWN),
    ("an int",                  7,                                                     UNKNOWN),
    ("empty dict",              {},                                                    UNKNOWN),
    # The escape hatch, and the shape that must NOT open it by accident.
    ("void but both accepted in writing",
                                {"document": "d", "qualifications_permitted": False}, PASS),
    ("void, one accepted one not",
                                {"document": "d", "qualifications_permitted": False},  FAIL),
    # Wordings from other tenders that mean the same thing.
    ("'shall have no effect'",  {"document": "d",
                                 "clause": "Any exclusions attached shall have no effect."}, FAIL),
    ("'unqualified tender'",    {"document": "d",
                                 "clause": "Only an unqualified tender will be considered."}, FAIL),
    ("'deemed not to apply'",   {"document": "d",
                                 "clause": "Contractor's standard terms are deemed not to apply."}, FAIL),
]


def selftest_qualification_variants():
    """Recall test for check_our_qualifications_survive_signature."""
    ok = True
    for name, value, want in QUALIFICATION_VARIANTS:
        relied = list(_QR_RELIED)
        if name == "void, but nothing relied on":
            relied = [{"ref": "Manifestation", "treatment": "priced"}]
        elif name == "void but both accepted in writing":
            relied = [dict(i, qualification_accepted_in_writing=True) for i in _QR_RELIED]
        elif name == "void, one accepted one not":
            relied = [dict(_QR_RELIED[0], qualification_accepted_in_writing=True), _QR_RELIED[1]]
        m = {"spec_items": relied}
        if value is not None or name == "field absent":
            m["qualification_regime"] = value
        try:
            got = check_our_qualifications_survive_signature(m)["status"]
        except Exception as exc:                      # noqa: BLE001 - a crash is a failure
            got = "CRASH: %s" % exc
        if got != want:
            print("  qualification variant %-42s wanted %s, got %s" % (name, want, got))
            ok = False
    print("  qualification regime   %d variant(s) checked%s"
          % (len(QUALIFICATION_VARIANTS), "" if ok else "  SOME FAILED"))
    return ok


FABRICATOR_VARIANTS = [
    # (name, systems_specified, want)
    ("field absent", None, UNKNOWN),
    ("named fabricator", [{"system": "Sheerline S1", "fabricator": "BSW"}], PASS),
    ("named, RFQ out, no return yet",
     [{"system": "Alu windows", "fabricator": "BSW (RFQ issued 24/07 15:14 and 15:29, no return)"}],
     PASS),
    ("named with quote ref",
     [{"system": "Smart Wall", "fabricator": "Bellview Products (0000000483 pos 007)"}], PASS),
    ("or-similar-approved basis",
     [{"system": "Senior PURe", "fabricator": "as above - BSW RFQ 28/07 15:22 'or similar approved'"}],
     PASS),
    ("empty field", [{"system": "Senior SF52", "fabricator": None}], FAIL),
    # The founding widening - all four were live text in real manifests on 29/07.
    ("prose says NONE APPROACHED",
     [{"system": "Senior SF52", "fabricator": "NONE APPROACHED CAN MAKE IT - BSW fabricate "
                                              "Sheerline, Aplus Technal, Bellview SMA Smart Wall."}],
     FAIL),
    ("prose says none - never asked",
     [{"system": "Senior PURe SLIDE", "fabricator": "none - dwg 001 was never attached to any RFQ"}],
     FAIL),
    ("prose says not available to us",
     [{"system": "Joedan casement", "fabricator": "Joedan Manufacturing (UK) Ltd - their own "
                                                  "system, not available to Fenster"}], FAIL),
    ("explicit can_make_it false beats the prose",
     [{"system": "Senior SF52", "fabricator": "BSW Window Solutions", "can_make_it": False}], FAIL),
    ("one good, one denied",
     [{"system": "Sheerline S1", "fabricator": "BSW"},
      {"system": "Senior SF52", "fabricator": "nobody on our supply chain"}], FAIL),
]


def selftest_fabricator_variants():
    """Recall test for check_fabricator_can_make_it.

    The rule was founded on Vesuvius and widened on Vesuvius: writing the honest
    answer into the 'fabricator' field used to satisfy the rule that exists to
    catch exactly that answer.
    """
    ok = True
    for name, value, want in FABRICATOR_VARIANTS:
        m = {}
        if value is not None:
            m["systems_specified"] = value
        try:
            got = check_fabricator_can_make_it(m)["status"]
        except Exception as exc:                      # noqa: BLE001 - a crash is a failure
            got = "CRASH: %s" % exc
        if got != want:
            print("  fabricator variant %-45s wanted %s, got %s" % (name, want, got))
            ok = False
    print("  %-22s %d variant(s) checked%s"
          % ("fabricator", len(FABRICATOR_VARIANTS), "" if ok else "  SOME FAILED"))
    return ok


SCREEN_VARIANTS = [
    ("field absent", None, UNKNOWN),
    ("no screens on the job", [], NA),
    # Greenfields, the founding error - a full-height screen coded as a window.
    ("priced as a window", [{"ref": "stair screen", "priced_as": "window"}], FAIL),
    ("priced blank", [{"ref": "stair screen", "priced_as": ""}], FAIL),
    ("curtain walling", [{"ref": "stair screen", "priced_as": "curtain walling"}], PASS),
    # Grange Hill, 29/07 - the convention was the error and the quotation was right.
    ("supplier quotation", [{"ref": "south screen", "priced_as": "supplier quotation, line by line"}], PASS),
    ("one of each", [{"ref": "a", "priced_as": "curtain walling"},
                     {"ref": "b", "priced_as": "supplier quotation"}], PASS),
    # A quotation for one and a guess for the other is still the founding error.
    ("supplier quote plus a window", [{"ref": "a", "priced_as": "supplier quotation"},
                                      {"ref": "b", "priced_as": "window"}], FAIL),
]


def selftest_screen_variants():
    """Recall test for check_full_height_screens.

    Greenfields must still fire, and the Grange Hill widening must not have
    turned 'anything that is not curtain walling' into 'anything at all'."""
    ok = True
    for name, value, want in SCREEN_VARIANTS:
        m = {}
        if value is not None:
            m["full_height_screens"] = value
        try:
            got = check_full_height_screens(m)["status"]
        except Exception as exc:                      # noqa: BLE001 - a crash is a failure
            got = "CRASH: %s" % exc
        if got != want:
            print("  screen variant %-45s wanted %s, got %s" % (name, want, got))
            ok = False
    print("  %-22s %d variant(s) checked%s"
          % ("full-height screens", len(SCREEN_VARIANTS), "" if ok else "  SOME FAILED"))
    return ok


def selftest_uplift_variants():
    """Brocks Hill Phase 2, 04/08/2026. The founding row is the fifth one below -
    a steel doorset sold at Strongdor's cost plus the delivery share and nothing
    else. Replay the real workbook: eight rows correct to the penny, one adrift by
    exactly GBP 1,500 a door.

    The other arms are the ways this rule could be wrong in the dangerous
    direction - reading a bad row as clean. A delivery share folded into the sell
    must NOT excuse a missing uplift (that is the founding case itself); a code
    the table does not know must ASK rather than pass; and an unparseable number
    must ASK rather than crash the run, which is the fault riverside found in the
    delivery rule on 28/07."""
    real = {"priced_rows": [
        {"ref": "E.02 door", "code": "SAD", "qty": 1, "supply_each": 2589.1085, "sell_each": 3489.1085},
        {"ref": "E.04 door", "code": "DAD", "qty": 2, "supply_each": 2878.661, "sell_each": 4378.661},
        {"ref": "E.01/E.03 window-door", "code": "SAD", "qty": 4, "supply_each": 3137.9195, "sell_each": 4037.9195},
        {"ref": "E.03 louvred door", "code": "DAD", "qty": 2, "supply_each": 2940.5155, "sell_each": 4440.5155},
        {"ref": "E.01 steel door", "code": "DAD", "qty": 5, "supply_each": 2637.01,
         "sell_each": 2728.81, "additional_each": 91.80},
        {"ref": "E.02 window", "code": "ELAW", "qty": 23, "supply_each": 1362.57, "sell_each": 2000.07},
        {"ref": "E.04 window", "code": "ELAW", "qty": 4, "supply_each": 1098.90, "sell_each": 1736.40},
        {"ref": "E.05 window", "code": "LAW", "qty": 2, "supply_each": 620.49, "sell_each": 1107.99},
        {"ref": "E.06 window", "code": "ELAW", "qty": 1, "supply_each": 984.64, "sell_each": 1622.14},
    ]}
    fixed = {"priced_rows": [dict(r, sell_each=4228.81) if r["ref"] == "E.01 steel door" else r
                             for r in real["priced_rows"]]}
    # the delivery share moved into the sell instead of being declared - the row
    # still has no uplift and must still fail
    hidden = {"priced_rows": [{"ref": "E.01 steel door", "code": "DAD", "qty": 5,
                               "supply_each": 2637.01, "sell_each": 2728.81}]}
    badcode = {"priced_rows": [{"ref": "mystery", "code": "ZZZ", "qty": 1,
                                "supply_each": 100.0, "sell_each": 200.0}]}
    badnum = {"priced_rows": [{"ref": "typo", "code": "DAD", "qty": 1,
                               "supply_each": "2,637.01", "sell_each": 4137.01}]}
    a = check_uplift_applied(real)
    b = check_uplift_applied(fixed)
    c = check_uplift_applied(hidden)
    d = check_uplift_applied(badcode)
    e = check_uplift_applied(badnum)
    f = check_uplift_applied({})
    g = check_uplift_applied({"priced_rows": []})
    checks = [
        ("the issued Brocks Hill workbook FAILs", a["status"] == FAIL),
        ("it names the steel door row", "E.01 steel door" in a["detail"]),
        ("it quantifies the hole at 7,500", "7,500.00 adrift" in a["detail"]),
        ("it does NOT accuse the eight correct rows", "E.02 window" not in a["detail"]),
        ("applying the 1,500 clears it", b["status"] == PASS),
        ("a delivery share folded into the sell still FAILs", c["status"] == FAIL),
        ("an unknown code ASKs, never passes", d["status"] == UNKNOWN),
        ("an unparseable number ASKs rather than crashing", e["status"] == UNKNOWN),
        ("an unfilled manifest ASKs", f["status"] == UNKNOWN),
        ("a job with no priced rows is N/A", g["status"] == NA),
    ]
    bad = [n for n, got in checks if not got]
    print("  %-22s %d/%d uplift variants behave as intended%s"
          % ("uplift", len(checks) - len(bad), len(checks),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_site_access_variants():
    """Luton Airport as issued, and the three ways out of it."""
    real = {"site_access": {
        "constraints": ["LLA induction and security clearance for every operative",
                        "escorted airside access for the duration of the works"],
        "allowance_gbp": 0, "excluded": [], "visits": 2}}
    priced = dict(real); priced["site_access"] = dict(real["site_access"], allowance_gbp=1800.0)
    qualified = dict(real); qualified["site_access"] = dict(
        real["site_access"], excluded=["LLA induction and security clearance for every operative",
                                       "escorted airside access for the duration of the works"])
    coordinated = dict(real); coordinated["site_access"] = dict(
        real["site_access"], excluded=["airside working to be coordinated prior to installation"])
    open_site = {"site_access": {"constraints": []}}
    a = check_site_access_is_priced_or_excluded(real)
    b = check_site_access_is_priced_or_excluded(priced)
    c = check_site_access_is_priced_or_excluded(qualified)
    d = check_site_access_is_priced_or_excluded(coordinated)
    e = check_site_access_is_priced_or_excluded(open_site)
    f = check_site_access_is_priced_or_excluded({})
    g = check_site_access_is_priced_or_excluded({"site_access": {
        "constraints": ["escorted airside access"], "allowance_gbp": "1,800"}})
    checks = [
        ("Luton as issued FAILs", a["status"] == FAIL),
        ("it names the escorted access", "escorted airside access" in a["detail"]),
        ("it says the constraint recurs on both visits", "2 separate visits" in a["detail"]),
        ("an allowance clears it", b["status"] == PASS),
        ("naming them in the exclusions clears it", c["status"] == PASS),
        ("a 'to be coordinated' sentence does NOT clear it", d["status"] == FAIL),
        ("an unrestricted site is N/A", e["status"] == NA),
        ("an unfilled manifest ASKs", f["status"] == UNKNOWN),
        ("an unparseable allowance ASKs rather than passing", g["status"] == UNKNOWN),
    ]
    bad = [n for n, got in checks if not got]
    print("  %-22s %d/%d site-access variants behave as intended%s"
          % ("site access", len(checks) - len(bad), len(checks),
             "" if not bad else "  MISSED: " + "; ".join(bad)))
    return not bad


def selftest_one_crash_costs_one_rule():
    """A rule that raises must lose itself, not the rest of the run.

    riverside's TypeError on 28/07 aborted the whole list comprehension, so a
    single bad manifest value silently took every later rule with it - and
    check_spec_label_matches_evidence is last, so it was being skipped every
    time. Persisted here rather than left in a transcript.
    """
    def exploding_rule(_manifest):
        raise TypeError("'<' not supported between instances of 'float' and 'str'")

    original = list(RULES)
    try:
        RULES.insert(4, exploding_rule)
        res = run({})
        got, want = len(res), len(original) + 1
        crashed = [r for r in res if "CRASHED" in r["detail"]]
        survived = any(r["rule"] == "spec item labels match their evidence" for r in res)
    finally:
        RULES[:] = original

    ok = got == want and len(crashed) == 1 and crashed[0]["status"] == FAIL and survived
    print("  %-22s one crash costs one rule: %d/%d results, crash reported as FAIL=%s, "
          "last rule survived=%s" % ("crash isolation", got, want,
                                     bool(crashed) and crashed[0]["status"] == FAIL, survived))
    return ok


def selftest():
    """Replay three jobs as they actually were and assert the rules still fire.

    These are not invented cases - they are the real state of SM5 Wexham,
    Stoke Park and Vesuvius Way at the moment the mistake was live. If a rule
    ever stops catching its own founding error, this fails loudly."""
    expected = {
        "_test-sm5.json": {"system-depth coupling", "fire-exit panic hardware",
                           "U-value read as average not per-element"},
        "_test-stoke.json": {"unglazed frames need a glass order"},
        "_test-vesuvius.json": {"drawing vs bill quantities", "spec covered or excluded",
                                "someone can actually fabricate it"},
        "_test-georgies.json": {"finish quoted is the finish specified"},
        "_test-brocks-hill.json": {"supplier quote covers every unit sold",
                                  "the RFQ was answered line by line"},
        "_test-st-marys.json": {"system can meet the specified performance"},
        "_test-gordon-court.json": {"supplier price held as long as ours"},
        "_test-riverside.json": {"delivery actually included"},
        "_test-john-north-hall.json": {"our qualifications survive signature"},
        "_test-crestwood.json": {"nothing priced is also excluded",
                                 "bought-in lumps have a quantity basis"},
    }
    # Rules whose founding error is a QUESTION rather than an outright error.
    # Asserted separately so that widening "fired" to include ASK cannot quietly
    # let a fixture that should FAIL degrade into merely asking.
    expected_ask = {
        "_test-riverside.json": {"supplier price held as long as ours"},
    }
    ok = True
    for name, must_fail in expected.items():
        path = os.path.join(MANIFEST_DIR, name)
        if not os.path.exists(path):
            print("  MISSING fixture %s" % name)
            ok = False
            continue
        with open(path, encoding="utf-8") as fh:
            results = run(json.load(fh))
        failed = {r["rule"] for r in results if r["status"] == FAIL}
        asked = {r["rule"] for r in results if r["status"] == UNKNOWN}
        missed = must_fail - failed
        missed_ask = expected_ask.get(name, set()) - asked
        note = ""
        if missed:
            note += "  MISSED: %s" % ", ".join(missed)
        if missed_ask:
            note += "  MISSED ASK: %s" % ", ".join(missed_ask)
        print("  %-22s %d rule(s) fired, %d asked%s" % (name, len(failed), len(asked), note))
        if missed or missed_ask:
            ok = False
    if not selftest_coupling_remedy():
        ok = False
    if not selftest_delivery_variants():
        ok = False
    if not selftest_terms_variants():
        ok = False
    if not selftest_warranty_variants():
        ok = False
    if not selftest_issued_variants():
        ok = False
    if not selftest_exposure_variants():
        ok = False
    if not selftest_trace_variants():
        ok = False
    if not selftest_view_variants():
        ok = False
    if not selftest_coverage_variants():
        ok = False
    if not selftest_qualification_variants():
        ok = False
    if not selftest_fabricator_variants():
        ok = False
    if not selftest_screen_variants():
        ok = False
    if not selftest_uplift_variants():
        ok = False
    if not selftest_site_access_variants():
        ok = False
    if not selftest_one_crash_costs_one_rule():
        ok = False
    print("selftest %s" % ("passed - every founding error is still caught" if ok else "FAILED"))
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("manifest", nargs="?")
    ap.add_argument("--new")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    if args.selftest:
        return selftest()

    if args.list:
        print("Checks, and the job that taught us each one:\n")
        for rule in RULES:
            doc = (rule.__doc__ or "").strip().split("\n")
            print("  %s" % doc[0])
            for line in doc[1:]:
                print("    %s" % line.strip())
            print()
        return 0

    if args.new:
        os.makedirs(MANIFEST_DIR, exist_ok=True)
        path = os.path.join(MANIFEST_DIR, "%s.json" % args.new.lower().replace(" ", "-"))
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(blank_manifest(args.new), fh, indent=1, ensure_ascii=False)
        print("wrote %s - fill it in, then run it" % path)
        return 0

    if not args.manifest:
        ap.print_help()
        return 2
    with open(args.manifest, encoding="utf-8") as fh:
        m = json.load(fh)
    return report(run(m), m.get("job", ""))


if __name__ == "__main__":
    sys.exit(main())
