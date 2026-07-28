# -*- coding: utf-8 -*-
"""JACOB - source: the BD Opportunity Log. What Fenster actually wins.

  python scripts/jacob_outcomes.py

Reads `4. Business Development\\Just in Case\\Opportunity Log 2025-2026.xlsx`
from the Commercial OneDrive, copies it into test-results first, and writes
`data/jacob/outcomes.json`.

READ ONLY. The workbook is opened from a copy under test-results and the
OneDrive original is never touched - Gintare, Adam and Steve are working in
that drive while this runs.

WHY THIS EXISTS
---------------
Until 28/07/2026 nothing on Jacob's board knew whether a lead was the kind
Fenster wins. It ranked by published contract value, which is exactly
backwards: 229 decided outcomes in this log say Fenster wins small work and
loses big work. The single hardest number in the file is that no job over
GBP 50,000 has been won in either year - 0 from 52.

Everything here is derived from the file at run time and carries its own
counts, because the log is hand-kept and the 2026 sheet is much thinner than
2025. Nothing is hard-coded from a previous read of it.
"""
import collections
import json
import os
import re
import shutil
import statistics
import sys
from datetime import date, datetime

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = (r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial"
       r"\4. Business Development\Just in Case\Opportunity Log 2025-2026.xlsx")
COPY_DIR = os.path.join(REPO, "test-results", "jacob-bd")
COPY = os.path.join(COPY_DIR, "Opportunity-Log-2025-2026.xlsx")
OUT = os.path.join(REPO, "data", "jacob", "outcomes.json")

COLS = ["client", "enquiry", "project", "adminbase", "deadline", "value",
        "returned", "notes", "wl", "lost", "commission", "chased"]

# The W/L column, as it is actually filled in. Only W and L are decided; the
# rest are states, and counting them as losses would halve a real win rate.
DECIDED = {"W": "won", "L": "lost"}
OPEN = {"O": "outstanding", "WIP": "work in progress", "DM": "deadline missed",
        "NA": "not applicable"}

# The bands a human thinks in, and the reason the board exists.
BANDS = [(0, 5_000), (5_000, 20_000), (20_000, 50_000),
         (50_000, 100_000), (100_000, 250_000), (250_000, None)]

# The Lost Reason legend is NOT recorded anywhere - not in the workbook, not
# in the BD notes, not in a comment or a validation list. I looked. So these
# are readings taken from the free text on the rows that carry each code, and
# they are labelled as readings until Adam confirms them (JAC-8).
LOST_READING = {
    "P": {"reading": "lost on price",
          "confidence": "high",
          "evidence": "44 of 50 rows say so in the notes, several verbatim "
                      "('Lost on price', 'we were most expensive', '80k out')"},
    "C": {"reading": "the main contractor lost it at their end - not a "
                     "Fenster loss at all",
          "confidence": "high",
          "evidence": "the dominant phrasing on these rows is 'Clegg lost this "
                      "to Bowmer and Kirkland', 'Borras lost this work. Our "
                      "price was competitive', 'lost at contractor stage'"},
    "V": {"reading": "unclear - the rows are a mixture (spec deviation, a "
                     "missing SBD accreditation, supply-only competitors, an "
                     "incumbent supplier, and three we admit we fumbled)",
          "confidence": "low",
          "evidence": "no single theme across 17 rows; do not build anything "
                      "on this code until the legend is confirmed"},
    "?": {"reading": "never found out - client went quiet",
          "confidence": "medium",
          "evidence": "'never heard back', 'not worth chasing', 'gone quiet' "
                      "recur across the 25 rows"},
    "NA": {"reading": "no reason recorded",
           "confidence": "medium",
           "evidence": "13 rows, mostly carrying only a generic chase note"},
}


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("\u00a3", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def iso(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    if isinstance(v, str):
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", v.strip())
        if m:
            return "%s-%02d-%02d" % (m.group(3), int(m.group(2)), int(m.group(1)))
    return None


def norm(v):
    return str(v).strip().upper() if v not in (None, "") else None


def band_label(lo, hi):
    if hi is None:
        return "over GBP %sk" % (lo // 1000)
    if lo == 0:
        return "under GBP %sk" % (hi // 1000)
    return "GBP %sk-%sk" % (lo // 1000, hi // 1000)


def read_rows():
    import openpyxl
    if not os.path.exists(SRC):
        sys.exit("Opportunity Log not found: %s" % SRC)
    os.makedirs(COPY_DIR, exist_ok=True)
    shutil.copy2(SRC, COPY)                # work on the copy, never the drive
    wb = openpyxl.load_workbook(COPY, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            rec = {"sheet": sheet, "row": r}
            blank = True
            for i, key in enumerate(COLS):
                v = ws.cell(r, i + 1).value
                if isinstance(v, str):
                    v = v.strip() or None
                if v is not None:
                    blank = False
                rec[key] = v
            if not blank:
                rows.append(rec)
    return rows, wb.sheetnames


def main():
    rows, sheets = read_rows()

    won = [r for r in rows if norm(r["wl"]) == "W"]
    lost = [r for r in rows if norm(r["wl"]) == "L"]
    decided = won + lost

    def vals(grp):
        return [v for v in (num(r["value"]) for r in grp) if v is not None]

    wv, lv = vals(won), vals(lost)

    bands = []
    for lo, hi in BANDS:
        def inband(r):
            v = num(r["value"])
            return v is not None and v >= lo and (hi is None or v < hi)
        w, l = len([r for r in won if inband(r)]), len([r for r in lost if inband(r)])
        bands.append({"label": band_label(lo, hi), "from": lo, "to": hi,
                      "won": w, "lost": l, "decided": w + l,
                      "winRate": round(100.0 * w / (w + l), 1) if w + l else None})

    per_client = collections.defaultdict(
        lambda: {"won": 0, "lost": 0, "open": 0, "wonValue": 0.0,
                 "lostValue": 0.0, "lastEnquiry": None, "projects": []})
    for r in rows:
        name = str(r["client"] or "").strip()
        if not name:
            continue
        c = per_client[name]
        st = norm(r["wl"])
        v = num(r["value"]) or 0
        if st == "W":
            c["won"] += 1
            c["wonValue"] += v
        elif st == "L":
            c["lost"] += 1
            c["lostValue"] += v
        elif st in OPEN:
            c["open"] += 1
        d = iso(r["enquiry"])
        if d and (not c["lastEnquiry"] or d > c["lastEnquiry"]):
            c["lastEnquiry"] = d
        if r["project"] and len(c["projects"]) < 6:
            c["projects"].append(str(r["project"])[:70])

    clients = []
    for name, c in per_client.items():
        dec = c["won"] + c["lost"]
        clients.append({
            "client": name, "won": c["won"], "lost": c["lost"], "open": c["open"],
            "decided": dec,
            "winRate": round(100.0 * c["won"] / dec, 1) if dec else None,
            "wonValue": round(c["wonValue"]),
            "lostValue": round(c["lostValue"]),
            "lastEnquiry": c["lastEnquiry"],
            "projects": c["projects"],
        })
    # A client Fenster converts, ordered by how much evidence there is for
    # saying so. One win from one enquiry is not a pattern.
    clients.sort(key=lambda c: (-(c["won"]), c["lost"], c["client"]))

    lost_reasons = []
    counts = collections.Counter(norm(r["lost"]) for r in rows if r["lost"])
    for code, n in counts.most_common():
        rd = LOST_READING.get(code, {})
        lost_reasons.append({"code": code, "count": n,
                             "shareOfLosses": round(100.0 * n / max(1, len(lost)), 1),
                             "reading": rd.get("reading", "unknown"),
                             "confidence": rd.get("confidence", "none"),
                             "evidence": rd.get("evidence", "")})

    chased = []
    for sheet in sheets:
        grp = [r for r in rows if r["sheet"] == sheet]
        filled = len([r for r in grp if r["chased"]])
        chased.append({"sheet": sheet, "rows": len(grp), "chased": filled,
                       "pct": round(100.0 * filled / max(1, len(grp)))})

    # Rows still open on the current sheet. These are not analysis - they are
    # a chase list Fenster already owns and is not working.
    this_year = sheets[-1]
    open_rows = []
    for r in rows:
        if r["sheet"] != this_year or norm(r["wl"]) not in ("O", "WIP"):
            continue
        open_rows.append({
            "client": str(r["client"] or "").strip(),
            "project": str(r["project"] or "").strip()[:90],
            "enquiry": iso(r["enquiry"]), "deadline": iso(r["deadline"]),
            "returned": iso(r["returned"]),
            "value": num(r["value"]),
            "state": OPEN.get(norm(r["wl"]), norm(r["wl"])),
            "chased": bool(r["chased"]),
            "notes": str(r["notes"] or "")[:200],
        })
    open_rows.sort(key=lambda r: (r["returned"] or r["enquiry"] or "", ), reverse=True)

    data = {
        "updated": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "Commercial OneDrive - 4. Business Development\\Just in Case"
                  "\\Opportunity Log 2025-2026.xlsx (read-only copy)",
        "sheets": sheets,
        "rows": len(rows),
        "summary": {
            "decided": len(decided), "won": len(won), "lost": len(lost),
            "winRate": round(100.0 * len(won) / max(1, len(decided)), 1),
            "wonMean": round(sum(wv) / max(1, len(wv))),
            "wonMedian": round(statistics.median(wv)) if wv else None,
            "wonMax": round(max(wv)) if wv else None,
            "wonTotal": round(sum(wv)),
            "lostMean": round(sum(lv) / max(1, len(lv))),
            "lostMedian": round(statistics.median(lv)) if lv else None,
            "lostMax": round(max(lv)) if lv else None,
            "valueFilled": len([r for r in rows if num(r["value"]) is not None]),
            # The ceiling. Stated as a fact about the record, not a rule.
            "biggestWon": round(max(wv)) if wv else None,
            "noWinAbove": next((b["from"] for b in bands
                                if b["decided"] and not b["won"]), None),
            "lostAboveThat": sum(b["lost"] for b in bands
                                 if b["decided"] and not b["won"]),
        },
        "bands": bands,
        "clients": clients,
        "lostReasons": lost_reasons,
        "lostLegend": {
            "status": "not recorded anywhere in the workbook or the BD folder",
            "note": "The readings below are taken from the free text on the "
                    "rows carrying each code. Adam has not confirmed them "
                    "(JAC-8). Nothing on the board is ranked by this column.",
        },
        "chased": chased,
        "openThisYear": open_rows,
        "states": collections.Counter(norm(r["wl"]) or "blank" for r in rows),
    }
    data["states"] = dict(data["states"])

    json.dump(data, open(OUT, "w", encoding="utf-8"), indent=1, ensure_ascii=False)
    s = data["summary"]
    print("%d rows over %s" % (len(rows), ", ".join(sheets)))
    print("  decided %d: %d won, %d lost -> %.1f%% win rate"
          % (s["decided"], s["won"], s["lost"], s["winRate"]))
    print("  won mean GBP %s, median GBP %s, biggest GBP %s"
          % (s["wonMean"], s["wonMedian"], s["biggestWon"]))
    print("  lost mean GBP %s, median GBP %s" % (s["lostMean"], s["lostMedian"]))
    for b in bands:
        print("    %-16s W%-4d L%-4d %s" % (b["label"], b["won"], b["lost"],
                                            "%.0f%%" % b["winRate"] if b["winRate"] is not None else "-"))
    print("  %d rows still open on the %s sheet" % (len(open_rows), sheets[-1]))
    print("-> %s" % OUT)


if __name__ == "__main__":
    main()
