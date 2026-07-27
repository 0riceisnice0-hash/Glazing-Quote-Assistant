"""Vesuvius Way (Worksop) - benchmark pricing review workbook.

Staniforth Construction LLP / BUSE Gas Solutions - Proposed New Gas Plant,
Plot 8 Vesuvius Way, Worksop S80 3NE. Trade bill L_SC Aluminium Doors & Windows.

No supplier quote exists yet (RFQ issued 27/07/2026), so every supply rate here is
either the house template curtain-wall rate or a size-banded median from
data/supplier-rates.json. Budget/benchmark only - NOT supplier backed.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REG = json.load(open('data/supplier-rates.json'))['register']


def median(supplier, category):
    for e in REG:
        if e['supplier'] == supplier and e['category'] == category:
            return e['median'], e['lineCount'], len(e.get('sources', []))
    raise KeyError(category)


# ---------------------------------------------------------------- rate basis
CW_SUPPLY = 850.00      # house MASTER PRICING DOC curtain-wall formula, GBP/m2
CW_LABOUR = 150.00      # house MASTER PRICING DOC curtain-wall labour, GBP/m2
SENIOR_UPLIFT = 0.15    # judgement: Senior PURe/SPD150 premium over Sheerline/SMA benchmark

WIN_SMALL_M, WIN_SMALL_N, WIN_SMALL_Q = median('bsw', 'aluminium casement window, glazed [<1.5m2]')
WIN_MED_M, WIN_MED_N, WIN_MED_Q = median('bsw', 'aluminium casement window, glazed [1.5-3m2]')
DOOR_M, DOOR_N, DOOR_Q = median('bsw', 'aluminium door, glazed [1.5-3m2]')

R_WIN_SMALL = round(WIN_SMALL_M * (1 + SENIOR_UPLIFT), 2)
R_WIN_MED = round(WIN_MED_M * (1 + SENIOR_UPLIFT), 2)
R_DOOR = round(DOOR_M * (1 + SENIOR_UPLIFT), 2)

# House template adders (product code value x 75%) and Adam's labour codes
ADDER = {'SAW': 337.50, 'MAW': 412.50, 'LAW': 487.50, 'SAD': 900.00, 'DAD': 1500.00}
LABOUR = {'SAW': 160.0, 'MAW': 160.0, 'LAW': 160.0, 'SAD': 250.0, 'DAD': 500.0}

# ---------------------------------------------------------------- take-off
# (ref, building, description, system, w, h, qty, area_each, kind, code, rate, rate_note)
CW = 'CW'
UNIT = 'UNIT'

items = [
    # ---- BUILDING 1 : WELFARE & SECURITY
    ('B1-B', 'Building 1 - Welfare', 'Curtain wall screen, Elevation 1 (raked head), 2 bays',
     'Senior SF52 (Zone Drained) Fully Capped 52mm', 2000, 2450, 1, 4.900, CW, None,
     CW_SUPPLY, 'House template CW rate'),
    ('B1-C', 'Building 1 - Welfare', 'Window, Elevation 1 (raked) - NO DRAWING IN PACK',
     'TBC - assumed Senior PURe', 1500, 1202, 1, 1.804, UNIT, 'MAW',
     R_WIN_MED, 'BSW alu casement glazed [1.5-3m2] median +15%'),
    ('B1-D', 'Building 1 - Welfare', 'Curtain wall screen w/ single AFT door + toplight, Elevation 2',
     'Senior SF52 + SPD150 door', 2950, 2450, 1, 7.2275, CW, 'SAD',
     CW_SUPPLY, 'House template CW rate; SAD adder+labour for door leaf'),
    ('B1-E', 'Building 1 - Welfare', 'Single high-usage door + 350 toplight, Elevation 2',
     'Senior SPD150 High Usage Door Standard', 1000, 2450, 2, 2.450, UNIT, 'SAD',
     R_DOOR, 'BSW alu door glazed [1.5-3m2] median +15%'),
    ('B1-F', 'Building 1 - Welfare', 'Window, Elevation 4',
     'Senior PURe No Profile Groove', 750, 950, 4, 0.7125, UNIT, 'SAW',
     R_WIN_SMALL, 'BSW alu casement glazed [<1.5m2] median +15%'),

    # ---- BUILDING 2 : OFFICE / SWITCH ROOM
    ('B2-A', 'Building 2 - Office', 'Curtain wall screen, Elevation 01 (raked, 5 bays @1380)',
     'Senior SF52 (Zone Drained) Fully Capped 52mm', 6900, 6000, 1, 35.78875, CW, None,
     CW_SUPPLY, 'House template CW rate; raked - 6.9x6.0 less 3.35x3.35 triangle/2'),
    ('B2-D', 'Building 2 - Office', 'Windows W07 + W08 set into cladding, Elevation 01',
     'Senior PURe No Profile Groove', 1350, 1450, 2, 1.9575, UNIT, 'MAW',
     R_WIN_MED, 'BSW alu casement glazed [1.5-3m2] median +15%'),
    ('B2-E', 'Building 2 - Office', 'Curtain wall screen w/ double AFT entrance door, Elevation 02',
     'Senior SF52 + double AFT door', 6500, 6000, 1, 39.000, CW, 'DAD',
     CW_SUPPLY, 'House template CW rate; DAD adder+labour for door pair'),
    ('B2-H', 'Building 2 - Office', 'Windows W09/W10/W11/W12 set into cladding, Elevation 03',
     'Senior PURe No Profile Groove', 1350, 1450, 4, 1.9575, UNIT, 'MAW',
     R_WIN_MED, 'BSW alu casement glazed [1.5-3m2] median +15%'),
    ('B2-I', 'Building 2 - Office', 'Windows W05 + W06 set into cladding, Elevation 04',
     'Senior PURe No Profile Groove', 1350, 1450, 2, 1.9575, UNIT, 'MAW',
     R_WIN_MED, 'BSW alu casement glazed [1.5-3m2] median +15%'),
]

# ---------------------------------------------------------------- TBC / not priced
tbc = [
    ('B1-A', 'Building 1 - Welfare', 'Access hatch, horizontal slider with EPS foam infill panel',
     'Senior PURe SLIDE Inline or Lift & Slide Square', '1450 x 1200', 1,
     'GBP 1,500 - 2,500', 'No register category for lift/slide hatches. Specialist unit - supplier quote required.'),
    ('B1-G', 'Building 1 - Welfare', 'Louvred double door (pair)',
     'PPC galvanised steel-core, louvred leaves', '1450 x 2110', 1,
     'GBP 3,500 - 5,000', 'STEEL not aluminium - Strongdor Steeldor Double median GBP 2,603 cost + louvre uplift + markup. Building attribution unclear (detail sits on the Building 02 door schedule).'),
    ('B2i-E', 'Building 2 - Internal', 'First floor partition window, fire/solar glass',
     'Senior PURe (internal) + Pilkington PyroStop Active Sun Cool 70/40', '2400 x 1500', 1,
     'see note', 'PyroStop is EI fire-resisting glass. Senior PURe is NOT a fire-rated system - a tested fire screen system is needed. Price and system both TBC.'),
    ('B2i-F', 'Building 2 - Internal', 'First floor partition window, fire/solar glass',
     'Senior PURe (internal) + Pilkington PyroStop Active Sun Cool 70/40', '3100 x 1500', 1,
     'GBP 6,000 - 9,000 for the pair', 'As above. Indicative only - specialist fire-screen quote required.'),
    ('B2-B/F', 'Building 2 - Office', 'Extra over for obscured spandrel panels, Elevations 01 & 02',
     '-', 'item', 2,
     'nil carried', 'Screen areas above are priced in FULL at the CW rate including spandrel zones, so no extra-over is added. Confirm with supplier.'),
    ('B2-C', 'Building 2 - Office', 'Extra over for panels obscured with reflective material',
     'Refer SAS Curtain Wall Design', 'item', 1,
     'nil carried', 'Reflective/mirrored finish may carry a genuine glass upcharge. RFI raised.'),
    ('B2-G', 'Building 2 - Office', 'Extra over for double curtain wall door',
     'Refer 2024-055-127 door schedule', 'item', 1,
     'included', 'Covered by the DAD adder + labour carried against B2-E.'),
]

# ---------------------------------------------------------------- compute
rows = []
tot_supply = tot_adder = tot_labour = 0.0
tot_area = 0.0
for (ref, bld, desc, system, w, h, qty, area_each, kind, code, rate, note) in items:
    area = area_each * qty
    supply = round(area * rate, 2) if kind == CW else round(area_each * rate, 2) * qty
    adder = round(ADDER[code] * qty, 2) if code and kind == UNIT else (ADDER[code] if code else 0.0)
    if kind == CW:
        labour = round(area * CW_LABOUR, 2) + (LABOUR[code] if code else 0.0)
    else:
        labour = round(LABOUR[code] * qty, 2) if code else 0.0
    sell = round(supply + adder + labour, 2)
    rows.append(dict(ref=ref, bld=bld, desc=desc, system=system, size='%d x %d' % (w, h),
                     qty=qty, area=round(area, 3), kind=kind, code=code or '-',
                     rate=rate, supply=supply, adder=adder, labour=labour, sell=sell, note=note))
    tot_supply += supply
    tot_adder += adder
    tot_labour += labour
    tot_area += area

grand = round(tot_supply + tot_adder + tot_labour, 2)

# ---------------------------------------------------------------- workbook
NAVY = 'FF1F2A44'
HDR = Font(bold=True, color='FFFFFFFF', size=10)
FILL = PatternFill('solid', fgColor=NAVY)
B = Font(bold=True)
thin = Side(style='thin', color='FFB0B7C3')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook()


def head(ws, cols, widths, row=1):
    for i, (c, wd) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HDR
        cell.fill = FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = wd


# --- Sheet 1: Summary
ws = wb.active
ws.title = 'Summary'
ws['A1'] = 'FENSTER GLAZING - PRICING REVIEW (BUDGET / BENCHMARK)'
ws['A1'].font = Font(bold=True, size=14)
meta = [
    ('Client', 'Staniforth Construction LLP (Joe Mayer)'),
    ('End client', 'BUSE Gas Solutions'),
    ('Project', 'Proposed New Gas Plant, Plot 8 Vesuvius Way, Worksop S80 3NE'),
    ('Trade bill', 'L_SC Aluminium Doors & Windows'),
    ('Architect', 'JHA Architecture Ltd, Doncaster (job ref 2024-055)'),
    ('Tender return', 'THURSDAY 30 JULY 2026'),
    ('Prepared', '27 July 2026'),
    ('Status', 'BUDGET / BENCHMARK ONLY - no supplier quote held. RFQ issued 27/07/2026.'),
]
r = 3
for k, v in meta:
    ws.cell(row=r, column=1, value=k).font = B
    ws.cell(row=r, column=2, value=v)
    r += 1

r += 1
ws.cell(row=r, column=1, value='COMMERCIAL POSITION').font = Font(bold=True, size=12)
r += 1
for k, v in [
    ('Priced scope - supply (benchmark)', tot_supply),
    ('Product code adders (house template, code value x 75%)', tot_adder),
    ('Installation (CW GBP150/m2 + Adam labour codes)', tot_labour),
]:
    ws.cell(row=r, column=1, value=k)
    c = ws.cell(row=r, column=2, value=round(v, 2))
    c.number_format = '#,##0.00'
    r += 1
ws.cell(row=r, column=1, value='BUDGET SELL TOTAL EX VAT').font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=grand)
c.font = Font(bold=True, size=12)
c.number_format = '#,##0.00'
r += 1
ws.cell(row=r, column=1, value='VAT @ 20%')
c = ws.cell(row=r, column=2, value=round(grand * 0.2, 2))
c.number_format = '#,##0.00'
r += 1
ws.cell(row=r, column=1, value='Budget total inc VAT').font = B
c = ws.cell(row=r, column=2, value=round(grand * 1.2, 2))
c.number_format = '#,##0.00'
c.font = B
r += 2
ws.cell(row=r, column=1, value='Glazed area priced (curtain wall + windows + doors)').font = B
ws.cell(row=r, column=2, value=round(tot_area, 2)).number_format = '#,##0.00'
r += 2
ws.cell(row=r, column=1, value='NOT INCLUDED IN THE ABOVE - see TBC & RFIs sheet').font = Font(bold=True, color='FFC00000')
r += 1
for t in ['Access hatch (Senior PURe SLIDE)', 'Louvred double door (steel)',
          '2no fire-rated first floor partition windows (Pilkington PyroStop)',
          'Extra-over obscured / reflective spandrel panels',
          'Roller shutter door 3500 x 5900 (not in this trade bill)']:
    ws.cell(row=r, column=1, value='  - ' + t)
    r += 1
ws.column_dimensions['A'].width = 58
ws.column_dimensions['B'].width = 62

# --- Sheet 2: Pricing lines
ws = wb.create_sheet('Pricing Lines')
cols = ['Bill ref', 'Building', 'Description', 'System', 'Size (mm)', 'Qty', 'Area m2',
        'Code', 'Supply rate', 'Supply cost', 'Code adder', 'Labour', 'Sell total', 'Rate basis']
head(ws, cols, [9, 20, 44, 34, 14, 6, 9, 7, 11, 12, 11, 11, 12, 46])
r = 2
for d in rows:
    vals = [d['ref'], d['bld'], d['desc'], d['system'], d['size'], d['qty'], d['area'],
            d['code'], d['rate'], d['supply'], d['adder'], d['labour'], d['sell'], d['note']]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BOX
        c.alignment = WRAP
        if i in (9, 10, 11, 12, 13):
            c.number_format = '#,##0.00'
    r += 1
ws.cell(row=r, column=3, value='TOTAL - priced scope').font = B
for col, val in ((7, tot_area), (10, tot_supply), (11, tot_adder), (12, tot_labour), (13, grand)):
    c = ws.cell(row=r, column=col, value=round(val, 2))
    c.font = B
    c.number_format = '#,##0.00'
ws.freeze_panes = 'A2'

# --- Sheet 3: TBC & RFIs
ws = wb.create_sheet('TBC & RFIs')
head(ws, ['Bill ref', 'Building', 'Description', 'System / spec', 'Size', 'Qty',
          'Indicative', 'Why it is not priced'], [9, 20, 44, 40, 14, 6, 26, 66])
r = 2
for t in tbc:
    for i, v in enumerate(t, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BOX
        c.alignment = WRAP
    r += 1

r += 2
ws.cell(row=r, column=1, value='RFIs / ACTIONS').font = Font(bold=True, size=12)
r += 1
rfis = [
    ('1', 'PROCUREMENT - the RFQ issued 27/07 carries drawings whose quantities are LOWER than the trade bill. '
          'Drawing 005 (Welfare Ele 4 window) says Quantity 1, bill item F says 4no. Drawing 004 (Welfare Ele 2 door) '
          'says Quantity 1, bill item E says 2no. Drawing 008 (Office window) says Quantity 6, window schedule 222P '
          'shows 8 marks (W05-W12) and the bill totals 8. Any supplier quoting off the drawings alone will return '
          'a price 2 doors, 3 windows and 2 windows short.'),
    ('2', 'Bill item B1-C - "Windows, Elevation 1, 1500 x 1255-1150" has NO drawing anywhere in the tender pack. '
          'Priced here on the bill dimensions as a Senior PURe unit. Confirm system, opening type and whether it exists.'),
    ('3', 'Bill item B1-G "Louvered Double Door" is listed under Building 1, but the only louvred double door detail '
          '(1450 x 2110) sits on the Building 02 door schedule 2024-055-221P. Confirm which building and how many.'),
    ('4', 'The external doors on door schedule 221P are specified as insulated STEEL-CORE, PPC galvanised double-skinned '
          'leaves in 90mm galvanised frames - not aluminium. Some are Rw 100 dB acoustic and some GRP fire-rated '
          '(drawing 127). Confirm what falls inside this aluminium trade bill.'),
    ('5', 'SYSTEM - the whole pack specifies Senior Architectural Systems (SF52 curtain wall, PURe windows, SPD150 doors, '
          'PURe SLIDE hatch). Fenster\'s regular fabricators quote Sheerline (BSW), Technal (Aplus) and SMA Smart Wall '
          '(Bellview) - none of them fabricate Senior. A Senior-approved fabricator is needed, or a formal alternative-system '
          'qualification. Every benchmark rate in this workbook comes from non-Senior quotes and is indicative only.'),
    ('6', 'Bill items B2i-E/F call for Pilkington PyroStop (fire-resisting) glass in Senior PURe frames on an internal '
          'first-floor partition. PURe is not a fire-rated system and the spec also asks for Active Sun Cool solar control '
          'on an internal screen. Confirm the fire rating required (EI30/EI60), and that solar control is intended.'),
    ('7', 'U-value 1.2 W/m2K is called up on every window, door and curtain wall panel. Confirm this is the whole-installation '
          'average, not a per-element requirement - per-element 1.2 on doors and curtain wall is demanding and drives the glass spec.'),
    ('8', 'Colour is TBC throughout ("Colour TBC by client prior to ordering"). Priced as a single PPC RAL colour, '
          'external and internal faces the same. Dual colour would be an extra.'),
    ('9', 'Trickle vents 4000mm2 are called for on all windows - allowed for within the benchmark rate. Confirm equivalent area.'),
    ('10', 'Roller shutter door 3500 x 5900 (SilentRoll 31, Rw 31dB, 3-phase electric) appears on door schedule 221P '
           'but is NOT in this trade bill. Excluded. Confirm it sits with another trade.'),
    ('11', 'Bill dimensions and drawing dimensions differ throughout (e.g. bill B1-B 2068 x 2540-2410 vs drawing 2000 x 2450; '
           'bill B2-I 8480 wide vs schedule 8890; bill window height 1980 vs schedule structural opening 1500). '
           'Bill figures read as structural openings, drawings as frame sizes. Final sizes by manufacturer survey.'),
    ('12', 'The reception screen quote in the pack (JS Office Environments ref MCJ/25204, 12/06/2025, GBP 4,595 + extras) '
           'is addressed to JHA Architecture for internal toughened reception screens. Not in this trade bill - excluded.'),
]
for n, t in rfis:
    ws.cell(row=r, column=1, value=n).font = B
    c = ws.cell(row=r, column=3, value=t)
    c.alignment = WRAP
    ws.row_dimensions[r].height = 46
    r += 1

# --- Sheet 4: Quantity check
ws = wb.create_sheet('Quantity Check')
head(ws, ['Item', 'Trade bill (Staniforth)', 'Logikal drawing issued with RFQ',
          'JHA schedule / elevation', 'Priced here', 'Comment'],
     [40, 24, 30, 30, 12, 54])
qc = [
    ('Welfare Ele 4 window 750x950 (bill F)', '4 nr', 'Pos 005 - Quantity: 1', 'not scheduled', '4',
     'DRAWING UNDER-STATES BY 3. Supplier quoting off drawing 005 returns 1 window.'),
    ('Welfare Ele 2 SPD150 door 1000x2450 (bill E)', '2 nr', 'Pos 004 - Quantity: 1', 'not scheduled', '2',
     'DRAWING UNDER-STATES BY 1. Supplier quoting off drawing 004 returns 1 door.'),
    ('Office windows 1350x1450 (bill D+H+I)', '8 (2+4+2)', 'Pos 008 - Quantity: 6',
     '222P shows W05-W12 = 8 marks', '8',
     'DRAWING UNDER-STATES BY 2. Bill and architect schedule agree at 8.'),
    ('Office window marks, bill item D', 'W06, W07', 'n/a', '222P Elevation 01 = W08, W07', '2',
     'Bill mislabels W08 as W06 (W06 then appears twice, in items D and I). Count is still 2.'),
    ('Welfare Ele 1 window 1500x1255-1150 (bill C)', '1 nr', 'NO DRAWING', 'not scheduled', '1',
     'No drawing exists in the tender pack for this item.'),
    ('Louvred double door (bill G)', '1 pr', 'NO DRAWING', '221P (Building 02 sheet) 1450x2110', '0 - TBC',
     'Steel-core louvred door, and listed against the wrong building.'),
    ('Access hatch (bill A)', '1 item', 'Pos 001 - Quantity: 1', 'not scheduled', '0 - TBC',
     'Drawing 001 was NOT attached to the RFQ that went out on 27/07.'),
    ('Curtain wall screens', '4 (B, D + Bldg2 A, E)', 'Pos 002/003/006/007 - Qty 1 each',
     '-', '4', 'Agree. 86.92 m2 total.'),
]
r = 2
for t in qc:
    for i, v in enumerate(t, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BOX
        c.alignment = WRAP
    ws.row_dimensions[r].height = 34
    r += 1

# --- Sheet 5: Source notes
ws = wb.create_sheet('Source Notes')
head(ws, ['Topic', 'Detail'], [30, 130])
notes = [
    ('Scope source', 'Trade bill "Aluminium Doors - Windows Bill.xls" (Staniforth, L_SC Aluminium Doors & Windows), '
                     'cross-read against Logikal drawings 001-008 and JHA drawings 2024-055-221P (door schedule), '
                     '222P (window schedule), 127 (door details) and 108D (welfare elevations).'),
    ('Pack location', 'OneDrive: Commercial\\1. Tender Documents\\Staniforth Construction LLP\\Worksop\\1. Estimating\\'
                      '1. Tender Documents\\Proposed Air Separation Unit, Vesuvius Way, Worksop Aluminium Doors Windows.zip '
                      '(55 files). The email of 27/07 08:20 carried only 10 of them.'),
    ('Working folder', 'test-results\\vesuvius-input (emailed subset) and test-results\\vesuvius-input\\full-pack (archive zip).'),
    ('Curtain wall rate', 'GBP %.2f/m2 supply + GBP %.2f/m2 labour - the MASTER PRICING DOC curtain-wall formula. '
                          'Same basis as the Grange Hill Methodist benchmark (24/07/2026).' % (CW_SUPPLY, CW_LABOUR)),
    ('Window rate <1.5m2', 'BSW "aluminium casement window, glazed [<1.5m2]" median GBP %.2f/m2 (%d lines, %d quotes) '
                           '+ %d%% Senior premium = GBP %.2f/m2.' % (WIN_SMALL_M, WIN_SMALL_N, WIN_SMALL_Q, SENIOR_UPLIFT * 100, R_WIN_SMALL)),
    ('Window rate 1.5-3m2', 'BSW "aluminium casement window, glazed [1.5-3m2]" median GBP %.2f/m2 (%d lines, %d quotes) '
                            '+ %d%% Senior premium = GBP %.2f/m2.' % (WIN_MED_M, WIN_MED_N, WIN_MED_Q, SENIOR_UPLIFT * 100, R_WIN_MED)),
    ('Door rate 1.5-3m2', 'BSW "aluminium door, glazed [1.5-3m2]" median GBP %.2f/m2 (%d lines, %d quotes) '
                          '+ %d%% Senior premium = GBP %.2f/m2.' % (DOOR_M, DOOR_N, DOOR_Q, SENIOR_UPLIFT * 100, R_DOOR)),
    ('Senior uplift', 'The +15%% is estimator judgement, not measured. Senior SF52/PURe/SPD150 is a premium specified system; '
                      'every register median available is from Sheerline (BSW), Technal (Aplus) or SMA (Bellview) quotes. '
                      'Replace with the real Senior fabricator quote as soon as it lands.'),
    ('Code adders', 'House MASTER PRICING DOC adders (product code value x 75%): SAW 337.50, MAW 412.50, SAD 900, DAD 1500. '
                    'Adam ruled 17/07/2026 that the template maths is the price.'),
    ('Labour', 'Adam\'s labour codes: SAW/MAW 160, SAD 250, DAD 500. Curtain wall at GBP150/m2.'),
    ('Not included', 'Mastic, EPDM, structural calculations, SAS curtain wall design, temporary works, scaffold/MEWP, '
                     'fire-stopping, cladding flashings/pressings (shown as Kingspan/Euroclad scope), roller shutter, '
                     'internal joinery doors (Howdens), reception screens, and the Building 03 package.'),
    ('Confidence', 'BUDGET. No supplier quote, a specified system Fenster has no live rate for, three quantity conflicts '
                   'in the tender documents and one item with no drawing. Do not issue as a fixed price.'),
]
r = 2
for k, v in notes:
    ws.cell(row=r, column=1, value=k).font = B
    ws.cell(row=r, column=1).alignment = WRAP
    c = ws.cell(row=r, column=2, value=v)
    c.alignment = WRAP
    ws.row_dimensions[r].height = 44
    r += 1

out = 'outputs/Vesuvius Way Worksop - Fenster Pricing Document and Review.xlsx'
wb.save(out)

print('Supply    %12s' % ('%.2f' % tot_supply))
print('Adders    %12s' % ('%.2f' % tot_adder))
print('Labour    %12s' % ('%.2f' % tot_labour))
print('GRAND     %12s ex VAT' % ('%.2f' % grand))
print('Area      %12s m2' % ('%.3f' % tot_area))
print()
for d in rows:
    print('%-6s %-52s qty=%-2s area=%8.3f supply=%10.2f adder=%8.2f lab=%9.2f sell=%10.2f'
          % (d['ref'], d['desc'][:52], d['qty'], d['area'], d['supply'], d['adder'], d['labour'], d['sell']))
print()
print('rates: win<1.5=%.2f win1.5-3=%.2f door=%.2f' % (R_WIN_SMALL, R_WIN_MED, R_DOOR))
print('saved ->', out)
