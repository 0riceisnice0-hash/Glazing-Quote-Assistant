# -*- coding: utf-8 -*-
"""Build the Redditch Library take-off workbook in the house take-off format.

Greenfields' three sheets - Project Information / Window & Door Schedule /
RFIs & Queries - plus two this job needs: the benchmark price, and the
inclusions and exclusions, because eleven items are being carried as excluded
and an exclusion that is not on the document is not an exclusion (Riverside).

  python scripts/redditch_takeoff_doc.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "outputs", "Redditch Library - Fenster Take-Off and Benchmark Price.xlsx")

HDR = PatternFill("solid", fgColor="1F2A44")
SUB = PatternFill("solid", fgColor="E8ECF3")
WARN = PatternFill("solid", fgColor="FCE9E6")
WHITE = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFC7D5")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


def sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    return ws


def head(ws, row, cells):
    for i, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill, c.font, c.border = HDR, WHITE, THIN
        c.alignment = WRAP
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, row, cells, fill=None, bold=False):
    for i, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border, c.alignment = THIN, WRAP
        if fill:
            c.fill = fill
        if bold:
            c.font = BOLD
    return row + 1


CONFIG = {
    "1": "1 fixed light + 1 fixed panel", "2": "1 fixed light + 1 fixed panel",
    "3": "2 fixed lights + 2 fixed panels", "4": "2 fixed lights + 2 fixed panels",
    "5": "2 fixed lights + 2 fixed panels", "6": "2 fixed lights + 2 fixed panels",
    "7": "2 fixed lights + 2 fixed panels", "8": "2 fixed lights + 2 fixed panels",
    "9": "2 top-hung opening lights", "10": "2 top-hung opening lights",
    "11": "2 top-hung opening lights", "12": "2 top-hung opening lights",
    "13": "2 top-hung opening lights", "14": "2 top-hung opening lights",
    "15": "1 top-hung opening light",
    "16": "2 fixed lights + 2 fixed panels", "17": "2 fixed lights + 2 fixed panels",
    "18": "2 fixed lights + 2 fixed panels",
    "19": "20 fixed lights", "20": "7 fixed lights + 1 fixed panel",
    "21": "1 top-hung + 1 fixed light",
    "22": "1 fixed light", "23": "1 fixed light",
    "24": "1 top-hung opening light", "25": "1 top-hung opening light",
    "26": "1 top-hung opening light", "27": "1 top-hung opening light",
    "28": "1 top-hung opening light",
    "29": "NOT STATED", "30": "NOT STATED", "31": "NOT STATED",
    "32.1": "1 top-hung + 1 fixed light + 1 fixed panel", "32.2": "single door",
    "33": "1 top-hung + 1 fixed light",
    "34.1": "1 top-hung + 2 fixed lights + 1 fixed panel", "34.2": "single door",
    "35": "1 top-hung + 1 fixed light", "36": "1 top-hung + 1 fixed light",
    "37": "single door", "38": "6 fixed lights", "39": "single door",
    "40": "2 fixed lights", "41": "double door",
}
ELEV = {}
for r in "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19".split():
    ELEV[r] = "A"
for r in ("35", "36", "37", "41"):
    ELEV[r] = "B"
for r in ("22 23 24 25 26 27 28 29 30 31 32.1 32.2 33 34.1 34.2 38 39 40").split():
    ELEV[r] = "C"
for r in ("20", "21"):
    ELEV[r] = "D"

NOTES = {
    "16": "RAKED - drawn as a parallelogram on the stair wall, sloping head and cill. 2250x2304 is a bounding box, not the frame. RFI-01.",
    "17": "RAKED - as ref 16. RFI-01.",
    "18": "RAKED - as ref 16. RFI-01.",
    "19": "17m continuous high-level ribbon over the whole of Elevation A. 20 bays counted on the drawing at 850mm centres - agrees with the schedule. Sits onto the flat roof: up-stand EXCLUDED.",
    "20": "7m high-level ribbon, Elevation D.",
    "29": "Configuration BLANK on Gleeds' own schedule and on Joedan's. Priced as a fixed light - stated assumption. RFI-02.",
    "30": "Configuration BLANK on both schedules. Priced as a fixed light. RFI-02.",
    "31": "Configuration BLANK on both schedules. Priced as a fixed light. RFI-02.",
    "32.1": "COUPLED to door 32.2 as ONE assembly on the elevation. Priced in a single thermally broken system throughout - see RFI-03.",
    "32.2": "COUPLED to window 32.1. Panic set per spec cl.5. Priced in the same system as 32.1.",
    "34.1": "COUPLED to door 34.2 as ONE assembly on the elevation. Priced in a single thermally broken system - RFI-03.",
    "34.2": "COUPLED to window 34.1. Panic set per spec cl.5.",
    "37": "Fire exit at roof level. Panic set per spec cl.5.",
    "38": "Drawn as a 5 x 3 = 15-PANE gridded screen; the schedule says 6 fixed lights. Framing content understated if the drawing is right. RFI-05.",
    "39": "The ONE door omitted from spec cl.5's ironmongery list (32, 34, 37, 41). No ironmongery specified anywhere. Priced as a plain single commercial doorset, NO panic set. RFI-06.",
    "41": "Double fire-exit door, Elevation B. Panic set + flush bolts to slave leaf per cl.5.",
}

RFIS = [
    ("RFI-01", "Refs 16, 17 and 18 are drawn on Elevation A as PARALLELOGRAMS following the stair soffit - sloping head and sloping cill - but scheduled as 2250 x 2304 rectangles, identically, all three. Please confirm the true frame geometry and the rake angle. A raked frame needs angled cuts to head, cill and both mullions and cannot be priced off a rectangle; the scheduled area also overstates the glass and understates the fabrication. GBP 8,509 of this take-off sits on these three."),
    ("RFI-02", "Refs 29, 30 and 31 (900 x 300) have EVERY configuration column blank on the tender's own pricing schedule as well as on Appendix 2. Please confirm what these three units are. They are priced here as fixed lights, which is the cheapest compliant reading; opening vents or louvres would change the price."),
    ("RFI-03", "Refs 32 and 34 are each drawn as ONE coupled assembly - a window frame joined directly to a door frame - but the specification puts the windows in EL75mm Squareline (thermally broken, 75mm) and the doors in AC100 Commercial (non-thermal, 100mm). Frames of different depths cannot be coupled. As written the specification cannot be built at these two positions. Please confirm which of the following is intended: (a) a thermally broken door so the whole run is one system and one depth; (b) a structural mullion or expressed movement joint between window and door so they are two separate frames; or (c) the window taken into the door system, which would put refs 32.1 and 34.1 outside the 1.4 W/m2K window requirement. This take-off assumes (a)."),
    ("RFI-04", "Specification 3.5.3 cl.6 requires 'a single standard RAL colours TBC'. Please confirm the RAL number. Priced on one standard RAL both faces; a non-standard, metallic or anodised finish is an extra."),
    ("RFI-05", "Ref 38 (2700 x 1700) is drawn on Elevation C as a 5-wide by 3-high grid - 15 panes - and scheduled as 6 fixed lights. Please confirm the mullion and transom arrangement. Four mullions and two transoms is not the same frame as 6 lights."),
    ("RFI-06", "Specification 3.5.3 cl.5 lists the ironmongery for doors 32, 34, 37 and 41. Ref 39 is a fifth door and appears nowhere in that clause. Please confirm its ironmongery, whether it is a fire escape, and whether it is glazed - it is drawn as a solid narrow leaf. It is priced here as a plain single commercial doorset with no panic hardware."),
    ("RFI-07", "Asbestos survey J449293 sample FM000125 records asbestos bitumen paint to metal cladding, chrysotile, 100 m2, 'present to left hand elevation', accessibility 2 (easily disturbed), with the recommendation 'REMOVE if affected by scheduled work'. Please confirm which elevation is 'left hand' and whether the window replacement disturbs that cladding. If it is the raking wall carrying refs 16, 17 and 18 then the window strip-out disturbs it. Note the four window-specific samples (FM000123, 124, 126, 127 - mastic and putty to the metal window frames, 75 linear metres) all returned NO ASBESTOS DETECTED. Asbestos removal is excluded from this take-off."),
    ("RFI-08", "The Form of Tender holds the tendered sum open for 10 WEEKS from submission; the JCT Minor Works preliminaries separately require the tender to be kept open 'for not less than 3 months'. Please confirm which period governs. Aluminium supplier quotations are firm for 30 days, so either period leaves the buy uncovered for 5 to 10 weeks after we commit."),
    ("RFI-09", "Specification 3.5.3 cl.3 places the design of any alternative window system on the tenderer under the Contractor's Designed Portion, and the preliminaries require GBP 5,000,000 professional indemnity for Contractor Designed Works - stating BOTH '12 years' and 'Six years' for the expiry of the required period. Please confirm which. Please also confirm the form and level of detail required for the compliance specification, which the specification requires to be submitted WITH the tender return rather than after award."),
    ("RFI-10", "The invitation allows the tendering contractor ONE site visit, arranged through Gleeds, with all sub-contractors in attendance. Specification 3.5.3 separately requires a measured survey of every window and external door before ordering and states that Gleeds' dimensions must not be relied on. Please confirm whether a further access visit is available to the successful window contractor before ordering, or whether the measured survey is a post-award obligation to be programmed within the 10-week contract period."),
]

INCLUSIONS = [
    "Supply and installation of 43no replacement aluminium windows and doors, refs 1-41 (32 and 34 each in two parts), to the sizes on the tender schedule.",
    "Polyester powder coated thermally broken aluminium windows, single standard RAL internally and externally.",
    "Commercial aluminium doorsets to refs 32.2, 34.2, 37, 39 and 41.",
    "Argon filled double glazed units, warm edge spacer, 4mm clear soft coat low-E toughened inner pane, 4mm bronze anti-sun toughened outer pane.",
    "Safety glass to all critical locations as defined by Part K.",
    "Solid phenolic core infill panels faced both sides in 2.0mm matt PPC aluminium, 1.2 W/m2K.",
    "Shootbolt espagnolette locking, Securistyle Defender friction hinges, concealed 100mm restrictors where required.",
    "Trickle ventilators to all casements.",
    "Panic ironmongery to doors 32, 34, 37 and 41 per specification cl.5 - Axim 8800 concealed transom closer, PR7100 exit panic device without external access, anti-finger-trap stiles, 100mm bottom and mid rails, mill finished low threshold, flush bolts to the slave leaf of ref 41.",
    "45mm white cellular uPVC cloaking profile to the internal perimeter where required.",
    "External grade waterproof mastic sealant to all windows and doors, applied at the time of installation.",
    "One site visit to undertake the manufacturing survey.",
    "Provisional sum of GBP 5,000 for preparing out-of-square openings, per specification cl.10.",
    "Fenster's standard ten year guarantee.",
]

EXCLUSIONS = [
    ("Strip out of the existing windows and doors, and carting debris off site", "NOT PRICED - NO RATE EXISTS. Gleeds ask for this as a separate line on p70 of the pack ('Cost for stripping out windows'). Fenster's installation rates are FIT ONLY - proven on Princess Beatrice and on Brocks Hill, a new build with nothing to remove, which recompute to the penny from the same per-unit codes. There is no strip-out rate in the rate register. THIS MUST BE PRICED BEFORE ANYTHING GOES TO PRIDE."),
    ("Making good internally, including localised decoration to all reveals", "Specification cl.11. Building trade, not glazing."),
    ("Removal, marking up, storage and re-fixing of internal blinds and curtains", "Strip Out cl.6."),
    ("Moving furniture and equipment, dust sheeting and protection of retained items", "Strip Out cl.5 and cl.7."),
    ("All access equipment - scaffolding, towers, podiums, MEWPs, hoisting and lifting equipment, and the temporary works design for it", "Specification 3.4. Joedan excluded the same. Note refs 19, 35, 36, 37 and 41 are at high level or roof level."),
    ("Site set-up, Heras fencing, compound, hoardings, and protection to the Elevation A public highway frontage", "Specification 3.2."),
    ("Flat roof covering replacement and all associated roofing work", "Appendix 1, Alumasc. Not our package - windows and doors only."),
    ("Edge protection guard rail to the flat roof", "Specification 3.5.2."),
    ("General repair and reinstatement - boxing in at roof level, rainwater goods, brickwork repairs", "Specification 3.5.4."),
    ("Up-stand where windows sit onto the flat roof", "Refs 19, 35, 36 and 37 sit at roof level. Joedan excluded the same item."),
    ("Building Control, skips, secure containers, welfare facilities, internal window boards, manifestations, fire signage, mag locks, door restraints, asbestos removal, and the moving of services, extractor fans, alarms or keypads", "Standard exclusions, and all matched by Joedan's own exclusion list."),
]


def main():
    with open(os.path.join(REPO, "outputs", "redditch-takeoff.json"), encoding="utf-8") as fh:
        d = json.load(fh)
    lines = {l["ref"]: l for l in d["lines"]}

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------------------------------------------------- sheet 1
    ws = sheet(wb, "Project Information", [30, 104])
    r = put(ws, 1, ["Redditch Library - Fenster take-off and benchmark price", ""], HDR)
    ws.cell(row=1, column=1).font = WHITE
    ws.cell(row=1, column=2).font = WHITE
    r += 1
    for k, v in [
        ("Project", "Redditch Library, 15 Market Place, Redditch, B98 8AR - flat roof refurbishment and external window and door replacement"),
        ("Tender reference", "BLBS0956, version 01, dated May 2026, 254 pages"),
        ("End client", "Worcestershire County Council"),
        ("Contract administrator", "Gleeds Cost Management Ltd, Birmingham - Shaun Wilkes BSc (Hons) MRICS, Director of Building Surveying"),
        ("We are invited by", "Pride Developments Group Ltd, Cwmbran - Leonard White, Senior Quantity Surveyor. Pride are bidding to Gleeds as main contractor; Fenster are their window and door sub-contractor."),
        ("Enquiry received", "22 July 2026 to info@, forwarded to Commercial@ and to Adam the same afternoon. Referred to estimating 28 July."),
        ("OUR DEADLINE", "NOT SET. Leonard White asked on 22/07 to acknowledge and 'submit your tender back asap' and gave no date. The 12 noon, Friday 26 June 2026 in the pack is GLEEDS' date to the main contractors and it has passed. Our date is whatever Pride need and nobody has asked them."),
        ("Contract", "JCT Minor Works with Contractor's Design Portion. 10 weeks to completion from commencement (TBC). Liquidated damages GBP 1,000 per calendar week. Retention 5%, releasing to 2.5% at practical completion. Rectification period six months. Public liability GBP 5m. CDP professional indemnity GBP 5m. No amendments to sections 3-7. The JCT form itself is not in the pack, and no sub-contract has been received from Pride."),
        ("Tender validity", "10 weeks from submission on the Form of Tender; 'not less than 3 months' in the preliminaries. Two figures, one pack - RFI-08."),
        ("Award criteria", "70% price, 30% quality. Lowest compliant tender scores 70."),
        ("Site constraints", "OPERATIONAL PUBLIC LIBRARY, occupied throughout. Working hours Mon-Fri 08:30-18:00 in the Schedule of Works and 08:30-17:30 in the preliminaries. Windows may not be left out over a shift - removal and replacement within the same working day. Elevation A fronts a public highway. Window-by-window programme required, rooms inspected three working days ahead, photographs before works. FORS Gold required for vehicles. ONE tender site visit only, arranged through Gleeds with all sub-contractors present."),
        ("Quantities", "43 items across 41 references, 136.54 m2. Taken from the tender's OWN blank pricing schedule (pack p77) and cross-checked line by line against Joedan's Appendix 2 schedule (p150) - the two are identical in every size and configuration - and against elevational drawings BLBS0956-GLE-RL-XX-DR-B-02 and -B-03. All 41 references appear on the elevations; none is missing and none is duplicated."),
        ("Systems specified", "Windows: EL75mm Squareline, polyester powder coated thermally broken aluminium. Doors: AC100 Commercial, polyester powder coated NON-THERMAL high traffic aluminium. BOTH ARE JOEDAN MANUFACTURING'S OWN SYSTEMS and Fenster cannot buy either. Specification 3.5.3 cl.3 expressly permits an alternative, subject to a compliance specification submitted WITH the tender return, written CA and Client approval, and the tenderer taking design responsibility under the Contractor's Designed Portion."),
        ("Thermal", "Windows 1.4 W/m2K area weighted; doors 3.0 W/m2K area weighted; infill panels 1.2 W/m2K. The window and door figures are PACKAGE AVERAGES, not per-element limits. 3.0 on the doors is a non-thermal door - for once the specification is easier than the house system, which is 1.8 W/m2K."),
        ("Glass", "Argon filled double glazed units with warm edge spacer. Inner pane 4mm clear soft coat low-E toughened. Outer pane 4mm BRONZE ANTI-SUN toughened. Safety glass to all Part K critical locations; transom and mullion heights to comply with Part K."),
        ("Warranty", "Fenster ten years. Gleeds specify no warranty period for the windows and doors - the only warranty requirement in the pack is 25 years insurance-backed on the roofing, which is not our package. Joedan, pricing the same building, offered TWELVE MONTHS."),
        ("Asbestos", "Bradley Environmental J449293, targeted R&D survey, 01/10/2025. All four window-related samples - mastic to metal framed windows, putty sealant, mastic to metal window frames, mastic to window joins, 75 linear metres - NO ASBESTOS DETECTED. But FM000125, asbestos bitumen paint to metal cladding, CHRYSOTILE, 100 m2, 'left hand elevation', easily disturbed, 'REMOVE if affected by scheduled work'. RFI-07. GBP 5,000 defined provisional sum in the pack covers the whole building including the roof."),
        ("Basis of this price", "BENCHMARK ONLY. No supplier has been asked and no quotation is held - the Commercial archive holds one file for this job and it is the tender pack. Rates are Fenster's own charged rates where enough history exists (learned) and supplier-quote medians otherwise (register). Every line states which. THIS IS NOT A QUOTABLE PRICE."),
    ]:
        r = put(ws, r, [k, v], bold=False)
        ws.cell(row=r - 1, column=1).font = BOLD

    # ---------------------------------------------------------- sheet 2
    ws = sheet(wb, "Window & Door Schedule", [8, 6, 9, 9, 8, 12, 34, 30])
    head(ws, 1, ["Ref", "Elev", "Width mm", "Height mm", "Area m2", "Type", "Configuration per the tender schedule", "Notes"])
    r = 2
    import redditch_takeoff as tk
    for ref in [l["ref"] for l in d["lines"]]:
        ln = lines[ref]
        cfg = ln["config"]
        typ = "door" if (cfg["SD"] or cfg["DD"]) else "window"
        r = put(ws, r, [ref, ELEV.get(ref, "?"), None, None, ln["area_m2"], typ,
                        CONFIG.get(ref, ""), NOTES.get(ref, "")],
                WARN if ref in ("16", "17", "18", "29", "30", "31", "38", "39", "32.1", "32.2", "34.1", "34.2") else None)
    # fill real sizes from the source schedule
    for i, (ref, wd, ht, *_rest) in enumerate(tk.SCHEDULE, start=2):
        ws.cell(row=i, column=3, value=wd)
        ws.cell(row=i, column=4, value=ht)
    r = put(ws, r, ["TOTAL", "", "", "", round(d["area_m2"], 2), "43 items", "41 references; 32 and 34 each split in two", ""], SUB, bold=True)

    # ---------------------------------------------------------- sheet 3
    ws = sheet(wb, "Benchmark Price", [8, 9, 9, 8, 7, 11, 13, 10, 12, 46])
    head(ws, 1, ["Ref", "Width mm", "Height mm", "Area m2", "Code", "Rate GBP/m2", "Frame + adder", "Fit labour", "Basis", "Where the rate comes from"])
    r = 2
    for (ref, wd, ht, *_rest) in tk.SCHEDULE:
        ln = lines[ref]
        r = put(ws, r, [ref, wd, ht, ln["area_m2"], ln["code"], ln["rate_per_m2"],
                        ln["line_total"], ln["labour"], ln["basis"], ln["provenance"]])
    r += 1
    r = put(ws, r, ["Frames and code adders", "", "", "", "", "", d["frames_and_adders"], "", "", "Supply at the benchmark rate, plus the house template's code adder at 75% of code value."], SUB, bold=True)
    r = put(ws, r, ["Installation (FIT ONLY)", "", "", "", "", "", d["installation"], "", "", "Adam's labour codes. These are fit rates and contain no strip-out - proven on Princess Beatrice and Brocks Hill."], SUB, bold=True)
    r = put(ws, r, ["Solar control glass premium", "", "", "", "", "", d["solar_glass_premium"], "", "", "136.54 m2 x GBP 13.29/m2. The specified outer pane is 4mm BRONZE ANTI-SUN; the frame rates above are built from ordinary softcoat units. Register glass medians: solar control toughened GBP 103.03/m2 against plain toughened softcoat GBP 89.74/m2."], SUB, bold=True)
    r = put(ws, r, ["BENCHMARK NET, EX VAT", "", "", "", "", "", d["benchmark_net_ex_vat"], "", "", "NOT A QUOTABLE PRICE - no supplier has been asked."], HDR, bold=True)
    for cc in range(1, 11):
        ws.cell(row=r - 1, column=cc).font = WHITE
    r += 1
    r = put(ws, r, ["NOT IN THE ABOVE", "", "", "", "", "", "", "", "", ""], SUB, bold=True)
    r = put(ws, r, ["Strip out existing windows and doors", "", "", "", "", "", "TBC", "", "", "Gleeds ask for this as a separate priced line (pack p70). NO RATE EXISTS ANYWHERE IN FENSTER'S RECORDS. The installation figure above is fit-only. This must be answered before a tender goes to Pride."], WARN, bold=True)
    r = put(ws, r, ["Provisional sum - preparing openings", "", "", "", "", "", 5000.00, "", "", "Gleeds specification 3.5.3 cl.10, defined provisional sum, to be used as directed by the CA."], WARN)
    r += 1
    r = put(ws, r, ["HOW THIS COMPARES", "", "", "", "", "", "", "", "", ""], SUB, bold=True)
    r = put(ws, r, ["Joedan JCQ.9727, 23/03/2026", "", "", "", "", "", 90687.17, "", "", "The competitor's fully priced schedule, left in the tender pack at Appendix 2. Gross of 2.5% main contractor's discount, so GBP 88,419.99 net to a main contractor. NOT LIKE FOR LIKE: Joedan's figure INCLUDES strip-out (their cl.12) and ours excludes it, so on the same scope we are already above them."])
    r = put(ws, r, ["Band-corrected sensitivity", "", "", "", "", "", 79047.55, "", "", "The same take-off with the register's known band errors applied to the supply component (St Mary's calibration: -35.5% under 1.5 m2, -1.2% at 1.5-3, +37.5% at 3-6, +35.2% over 6). 62% of this job's value sits in the 3-6 m2 band where the register runs high, so the headline above is more likely high than accurate."])

    # ---------------------------------------------------------- sheet 4
    ws = sheet(wb, "Inclusions & Exclusions", [60, 74])
    head(ws, 1, ["INCLUDED", ""])
    r = 2
    for t in INCLUSIONS:
        r = put(ws, r, [t, ""])
    r += 1
    head(ws, r, ["EXCLUDED", "Why / where it comes from"])
    r += 1
    for t, why in EXCLUSIONS:
        r = put(ws, r, [t, why], WARN if t.startswith("Strip out") else None)

    # ---------------------------------------------------------- sheet 5
    ws = sheet(wb, "RFIs & Queries", [10, 124])
    head(ws, 1, ["#", "Query for Pride Developments, to be passed to Gleeds"])
    r = 2
    for n, q in RFIS:
        r = put(ws, r, [n, q])
    r += 1
    r = put(ws, r, ["NOTE", "The invitation requires technical queries to be returned to Gleeds and assessed under the JCT Tendering Practice Note, 3rd Edition 2017. As sub-contractor our queries go through Pride."], SUB)

    # ---------------------------------------------------------- sheet 6
    mp = os.path.join(REPO, "outputs", "redditch-margin.json")
    if os.path.exists(mp):
        with open(mp, encoding="utf-8") as fh:
            m = json.load(fh)
        ws = sheet(wb, "Margin & Confidence", [44, 16, 74])
        head(ws, 1, ["Redditch Library - what we keep, and how sure I am of it", "GBP", "Note"])
        r = 2
        r = put(ws, r, ["WHAT THE BENCHMARK IS MADE OF", "", ""], SUB, bold=True)
        r = put(ws, r, ["Frame supply - COST", m["frame_supply"], "What a supplier would charge us. Benchmark, not a quotation - nobody has been asked."])
        r = put(ws, r, ["Solar-control glass premium - COST", m["solar_glass"], "The specified outer pane is 4mm bronze anti-sun; the frame rates are built from ordinary softcoat units."])
        r = put(ws, r, ["= Material we buy", m["material_cost"], ""], SUB, bold=True)
        r = put(ws, r, ["House code adders - MARGIN", m["adders"], "The template's per-unit adder at 75% of code value. This is the money we keep on materials."])
        r = put(ws, r, ["Installation - revenue", m["installation"], "Adam's per-unit labour codes. FIT ONLY. What fitting actually costs Fenster is recorded nowhere, so this is revenue and not margin."])
        r = put(ws, r, ["SELL", m["sell"], ""], HDR, bold=True)
        for cc in range(1, 4):
            ws.cell(row=r - 1, column=cc).font = WHITE
        r = put(ws, r, ["Gross margin if fitting breaks even", m["gross_margin"],
                        "%.1f%% of sell, %.1f%% mark-up on the material buy. NOT net profit - no prelims, supervision, survey, main contractor's discount or strip-out are in it."
                        % (m["margin_pct_of_sell"], m["markup_pct_on_material"])])
        r += 1
        r = put(ws, r, ["WHY THE MARGIN IS THIN HERE", "", "The house adder is a FIXED SUM PER UNIT, so it thins as units get bigger."], SUB, bold=True)
        for k in ("<1.5m2", "1.5-3m2", "3-6m2", ">6m2"):
            import mary_pricing as _e
            ls = [l for l in d["lines"] if _e.band_of(l["area_m2"]) == k]
            if not ls:
                continue
            c_ = sum(l["supply"] for l in ls)
            a_ = sum(l["adder"] for l in ls)
            r = put(ws, r, ["  %s - %d items" % (k, len(ls)), round(a_, 2),
                            "adder is %.1f%% of the frame line" % (a_ / (c_ + a_) * 100)])
        r = put(ws, r, ["  Redditch averages 3.18 m2 a unit", "", "adder = 24.7% of the frame line. Crestwood Park averaged 1.29 m2 and the adder was 42.9% (GBP 20,550 on a GBP 27,329.60 BSW buy). Big units earn less on this template."])
        r += 1
        r = put(ws, r, ["CONFIDENCE - AGAINST A REAL COMPARABLE BUY", "", m["severn_trent"]["ref"] + ". Sheerline Prestige aluminium casements, factory glazed, trickle vents, shootbolt locking - the nearest thing Fenster holds to this job. 6 lines, 27 units, 72.578 m2, reconciling exactly to their stated Total Nett Ex VAT of GBP 34,902.35."], SUB, bold=True)
        r = put(ws, r, ["  Fitted rate curve", "", "rate = %.2f x area^%.4f, R2 = %.4f. Two caveats, both pushing it HIGH for our purposes: that job is 3005 Wine Red metallic, and its outer pane is 6.8 laminated rather than 4mm toughened."
                        % (m["severn_trent"]["a"], m["severn_trent"]["b"], m["severn_trent"]["r2"])])
        r += 1
        r = put(ws, r, ["THREE ESTIMATES OF THE SAME JOB", "", ""], SUB, bold=True)
        r = put(ws, r, ["  Engine benchmark (as issued)", m["sell"], "Register and learned medians through mary_pricing."])
        r = put(ws, r, ["  Re-priced on the Severn Trent curve", m["severn_trent"]["sell"], "Same supplier, same client, six weeks old, same product family. The best single comparator we have."])
        r = put(ws, r, ["  St Mary's band correction", m["band_corrected_sell"], "The engine's measured band errors from the St Mary's calibration applied to the supply component."])
        r = put(ws, r, ["  Spread", round(m["sell"] - m["band_corrected_sell"], 2), "Almost entirely the 3-6 m2 band, which carries 62% of this job."])
        r += 1
        r = put(ws, r, ["CAN WE UNDERCUT JOEDAN?", "", ""], SUB, bold=True)
        r = put(ws, r, ["  Joedan gross of 2.5% MCD", m["joedan_gross"], "Their quotation JCQ.9727 as it sits in the tender pack."])
        r = put(ws, r, ["  Joedan NET to a main contractor", m["joedan_net"], "THE NUMBER TO BEAT - and it INCLUDES their strip-out (their cl.12). Ours does not."], HDR, bold=True)
        for cc in range(1, 4):
            ws.cell(row=r - 1, column=cc).font = WHITE
        for label, key in (("  Headroom at the engine benchmark", "engine benchmark (as issued)"),
                           ("  Headroom on the Severn Trent curve", "re-priced on the Severn Trent curve"),
                           ("  Headroom on the band correction", "St Mary's band correction")):
            v = m["headroom"][key]
            r = put(ws, r, [label, v, "= GBP %.2f per opening towards stripping 43 openings out of an occupied library." % (v / 43)],
                    WARN if v < 0 else None)

    wb.properties.creator = "Fenster Glazing & Locks Ltd"
    wb.properties.lastModifiedBy = "Fenster Glazing & Locks Ltd"
    wb.properties.title = "Redditch Library BLBS0956 - Take-Off and Benchmark Price"
    wb.properties.company = "Fenster Glazing & Locks Ltd"
    wb.save(OUT)
    print("written %s" % OUT)


if __name__ == "__main__":
    main()
