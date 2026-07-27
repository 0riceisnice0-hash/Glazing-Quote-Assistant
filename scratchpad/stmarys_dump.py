import openpyxl, os, sys

BASE = r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\1. Tender Documents\E T & S Construction"

targets = []
for root, dirs, files in os.walk(BASE):
    for f in files:
        if f.endswith(".xlsx") and "Refurbishment Pricing" in f:
            targets.append(os.path.join(root, f))

for p in targets:
    print("#" * 100)
    print("FILE:", p)
    wb = openpyxl.load_workbook(p, data_only=True)
    print("SHEETS:", wb.sheetnames)
    for ws in wb:
        print("=" * 90)
        print("SHEET:", ws.title, ws.dimensions)
        for r in ws.iter_rows():
            vals = [c.value for c in r]
            if any(v not in (None, "") for v in vals):
                cells = []
                for c in r:
                    v = c.value
                    if v in (None, ""):
                        continue
                    cells.append("%s=%s" % (c.coordinate, str(v)[:60]))
                print("  " + " | ".join(cells))
