# -*- coding: utf-8 -*-
"""Restore the print area I deleted last night, and extend it over the exclusions.

Gordon Court found their client holds 51 of our buy prices inside two files
called "Elevations". Running their filename check here meant opening the pricing
workbook properly for the first time - and columns J, K and L of rows 9 and 10
hold A Plus's buy, split three ways:

    J9 2331.075 frames   K9 85.655 glass   L9 5.88 surcharge     x2 = 4,845.22

against a sell of 5,990.22. Plus K3/L3 "Supplier used: A Plus (QT51518)".

Six turns of auditing this workbook and every dump I printed stopped at column
I, because that is where the part I was interested in stopped.

THE HOUSE FORMAT ALREADY SOLVES THIS AND I BROKE IT. The template's print area
is 'Pricing Document '!$C$1:$I$31 - deliberately set to stop at column I so the
buy columns never reach a printed or PDF'd copy. The Riverside file's print area
is EMPTY, and it is empty because last night's external-link strip did

    re.sub(r'<definedNames>.*?</definedNames>', '', s)

to remove 50 foreign defined names - and the print area is stored as a defined
name, `_xlnm.Print_Area`. I removed the fifty that were somebody else's and the
one that was ours in the same line, having verified only that no FORMULA used
any of them.

And restoring $C$1:$I$31 verbatim would repeat the fault in a quieter form: the
exclusions block added the night before lives at rows 33-45, outside it.
"""
import re
import shutil
import zipfile

import openpyxl

P = 'outputs/Riverside House - Fenster Pricing Document (house format).xlsx'
SHEET = 'Pricing Document '
AREA = "'%s'!$C$1:$I$45" % SHEET

b = openpyxl.load_workbook(P)
bws = b.active
before = (bws['I23'].value, type(bws['I21'].value).__name__, bws.print_area,
          sum(1 for row in bws.iter_rows() for c in row if c.value not in (None, '')),
          sum(1 for r in range(33, 46) if bws.cell(r, 3).value))

wb = openpyxl.load_workbook(P)
wb.active.print_area = AREA
wb.save(P)

a = openpyxl.load_workbook(P)
aws = a.active
after = (aws['I23'].value, type(aws['I21'].value).__name__, aws.print_area,
         sum(1 for row in aws.iter_rows() for c in row if c.value not in (None, '')),
         sum(1 for r in range(33, 46) if aws.cell(r, 3).value))

z = zipfile.ZipFile(P)
names = re.findall(r'<definedName[^>]*name="([^"]+)"', z.read('xl/workbook.xml').decode('utf-8'))
leak = [n for n in z.namelist()
        if re.search(rb"agsurveying|Dan Parker|LiamO|Content\.Outlook", z.read(n))]

print("  total formula   %-22s -> %s" % (before[0], after[0]))
print("  I21 type        %-22s -> %s" % (before[1], after[1]))
print("  print area      %-22s -> %s" % (before[2] or 'NONE', after[2]))
print("  populated cells %-22s -> %s" % (before[3], after[3]))
print("  exclusion rows  %-22s -> %s" % (before[4], after[4]))
print("  defined names now: %s" % (names or 'none'))
print("  third-party traces: %s" % (leak or 'none'))
print()
print("  What the print area now covers: C1:I45 - the priced items, the total, the")
print("  optional mastic, the footnote and all 13 exclusion rows, and NOT columns")
print("  J-L, which hold the supplier buy.")
