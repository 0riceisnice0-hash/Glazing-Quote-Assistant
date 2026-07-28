# -*- coding: utf-8 -*-
"""Open every file on this job and check it is the thing its filename claims.

Gordon Court's rule-20 side effect: enumerating issued documents to feed it made
them notice two client-facing PDFs had never been recorded as issued at all, and
opening those is how they found that "Window & Door Elevations.pdf" is in fact
all four BSW quotations - 51 of our buy prices, in the client's hands.

Their framing: my stale-draft lesson was wrong about WHEN. This is wrong about
WHAT, which is worse, because no amount of care about dates would catch it.

Two minutes each way. Run on everything Riverside holds and everything it would
send.
"""
import glob
import os
import re

import openpyxl
import pypdf

SUPPLIER = re.compile(r"a\s*plus|aplusaluminium|QT515\d\d|alston drive|4,845|4845\.22"
                      r"|2,422\.61|2422\.61", re.I)


def peek(path, n=600):
    e = os.path.splitext(path)[1].lower()
    try:
        if e == '.pdf':
            r = pypdf.PdfReader(path)
            txt = ''.join((p.extract_text() or '') for p in r.pages)
            return "PDF %dpp | %s" % (len(r.pages), re.sub(r'\s+', ' ', txt)[:n])
        if e == '.xlsx':
            wb = openpyxl.load_workbook(path)
            bits = []
            for ws in wb:
                for row in ws.iter_rows():
                    for c in row:
                        if isinstance(c.value, str) and c.value.strip():
                            bits.append(c.value.strip())
            return "XLSX %s | %s" % (wb.sheetnames, re.sub(r'\s+', ' ', ' / '.join(bits))[:n])
        with open(path, encoding='utf-8', errors='ignore') as fh:
            return "TEXT | %s" % re.sub(r'\s+', ' ', fh.read())[:n]
    except Exception as exc:
        return "COULD NOT OPEN: %s: %s" % (type(exc).__name__, exc)


print("=" * 100)
print("WHAT WE WOULD SEND - does each file contain what its name claims?")
print("=" * 100)
for f in sorted(glob.glob('outputs/Riverside*')):
    print("\n  %s" % os.path.basename(f))
    print("     %s" % peek(f, 420))

print()
print("=" * 100)
print("WHAT CAME IN - the 27/07 pack and the supplier attachments")
print("=" * 100)
for d in sorted(glob.glob('test-results/mary-inbox/processed/2026072*-att')):
    files = sorted(glob.glob(os.path.join(d, '*')))
    if not files:
        continue
    print("\n  %s" % d)
    for f in files:
        print("     %-52s %s" % (os.path.basename(f)[:52], peek(f, 190)))

print()
print("=" * 100)
print("DOES ANYTHING CLIENT-FACING NAME OUR SUPPLIER OR OUR BUY?")
print("=" * 100)
for f in sorted(glob.glob('outputs/Riverside*')):
    body = peek(f, 100000)
    hits = sorted(set(m.group(0) for m in SUPPLIER.finditer(body)))
    print("  %-64s %s" % (os.path.basename(f)[:64], hits or '-'))
