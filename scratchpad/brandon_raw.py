# -*- coding: utf-8 -*-
"""Raw cell dump of the top of the three Brandon revisions - the supplier block
lives in the first ten rows and the earlier revision's reader output
(_sup_vals=[(9, '', 362678.4)]) says it is reading a PRICED row's cell."""
import os, sys
import openpyxl

ROOT = os.path.join(os.path.expanduser("~"), "OneDrive - Fenster Glazing (1)",
                    "Commercial", "1. Tender Documents", "Elkins Construction",
                    "Brandon Estate EWI Remediation works", "1. Estimating",
                    "3. Client Quote")
FILES = [
    ("EARLIER (4.094)", os.path.join(ROOT, "SS", "DO NOT SEND Pricing Document - Brandon Estate.xlsx")),
    ("ELKINS  (1.006)", os.path.join(ROOT, "SS", "DO NOT SEND Elkins - Brandon Estate Pricing Document.xlsx")),
]
for label, path in FILES:
    print("=" * 110)
    print(label)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14), 1):
        cells = []
        for j, c in enumerate(row):
            if c.value is None:
                continue
            v = c.value
            if isinstance(v, float):
                v = "%.2f" % v
            cells.append("c%d=%r" % (j, v))
        print("  r%-3d %s" % (i, "  ".join(cells))[:400])
    wb.close()
