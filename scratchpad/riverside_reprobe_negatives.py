# -*- coding: utf-8 -*-
"""Re-test every ABSENCE this chat has claimed, with the assumptions removed.

Gordon Court probed their own proposal for two recourse clauses, got NOT PRESENT
on both, and both were there - one because the pattern required a trailing full
stop in a two-column table with no sentence terminators, one because of
apostrophe encoding. Their framing: the pattern encoded assumptions about the
DOCUMENT that the document does not honour.

Every negative finding this chat has published rests on a probe of the same kind.
Three are re-tested here with the assumptions stripped out:

  1. "MASTER PRICING DOC.xlsx has no exclusions section" - the original probe
     walked ws.iter_rows() only. An xlsx also carries text in headers, footers,
     comments, chart titles and drawing shapes, and none of that is a cell. That
     is the same fault one layer over: the pattern assumed all text lives where I
     looked.
  2. "Zero hits for the 'available on request' family on QT51518" - probed
     against pdfplumber output that is visibly full of U+FFFD where the PDF used
     bullets, curly quotes and en dashes. Any probe containing one of those
     characters would have failed silently.
  3. "Zero precedence statements across the Riverside outputs" - .txt files this
     chat wrote itself, so encoding is safe, but the xlsx half of that grep had
     fault 1.
"""
import io
import re
import unicodedata
import zipfile

import openpyxl


def norm(s):
    """Strip the assumptions a naive pattern makes about a document."""
    s = unicodedata.normalize('NFKD', str(s))
    for a, b in [(u'’', "'"), (u'‘', "'"), (u'“', '"'), (u'”', '"'),
                 (u'–', '-'), (u'—', '-'), (u'−', '-'), (u' ', ' '),
                 (u'�', ' '), (u'•', ' '), (u'▪', ' ')]:
        s = s.replace(a, b)
    return re.sub(r'\s+', ' ', s).lower()


PROBES = ['exclu', 'scaffold', 'welfare', 'site storage', 'waste removal', 'fire stopping',
          'internal finishing', 'final clean', 'testing', 'structural alteration',
          'design responsib', 'traffic management', 'dimensions provided', 'additional limitations',
          'builder', 'making good', 'not included', 'by others']

PRECEDENCE = ['govern', 'precede', 'conjunction', 'supersede', 'prevail', 'refer to the',
              'takes priority', 'in the event of', 'read with', 'accompan']

print("=" * 92)
print("1. EVERY PART OF THE XLSX, NOT JUST THE CELLS")
print("=" * 92)
for path in ['templates/MASTER PRICING DOC.xlsx',
             'outputs/Riverside House - Fenster Pricing Document (house format).xlsx']:
    print("\n--- %s" % path)
    wb = openpyxl.load_workbook(path)
    cell_hits = 0
    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and any(p in norm(c.value) for p in PROBES):
                    cell_hits += 1
        # the places iter_rows never reaches
        for label, val in [('header L', ws.oddHeader.left.text), ('header C', ws.oddHeader.center.text),
                           ('header R', ws.oddHeader.right.text), ('footer L', ws.oddFooter.left.text),
                           ('footer C', ws.oddFooter.center.text), ('footer R', ws.oddFooter.right.text)]:
            if val and any(p in norm(val) for p in PROBES + PRECEDENCE):
                print("   %-10s %s :: %s" % (ws.title, label, val[:110]))
        for cm in getattr(ws, 'legacy_drawing', None) and [] or []:
            pass
    print("   cell hits (normalised): %d" % cell_hits)
    # raw XML - catches shapes, text boxes, comments, drawings, defined names
    z = zipfile.ZipFile(path)
    parts = [n for n in z.namelist()
             if n.endswith(('.xml', '.vml')) and ('drawing' in n or 'comment' in n or 'note' in n
                                                  or 'header' in n or 'sharedStrings' in n)]
    print("   text-bearing parts: %s" % (parts or 'none beyond the sheet'))
    outside = []
    for n in z.namelist():
        if not n.endswith(('.xml', '.vml', '.rels')):
            continue
        if n.endswith('sheet1.xml') or 'sharedStrings' in n:
            continue
        raw = norm(z.read(n).decode('utf-8', 'ignore'))
        for p in PROBES + PRECEDENCE:
            if p in raw:
                outside.append((n, p))
    print("   probe hits in NON-cell parts: %s" % (outside or 'none'))

print()
print("=" * 92)
print("2. QT51518 - THE 'AVAILABLE ON REQUEST' FAMILY, NORMALISED, AND WITH TERMINATORS DROPPED")
print("=" * 92)
q = norm(io.open('scratchpad/qt51518_full.txt', encoding='utf-8').read())
fam = ['available on request', 'on request', 'subject to our standard', 'conditions of sale',
       'standard terms', 'as amended', 'terms of sale', 'apply to this', 'shall apply',
       'incorporat', 'from time to time', 'current at the date', 'copy available',
       'supplied on request', 'obtainable']
for p in fam:
    n = q.count(p)
    print("   %-26s %d" % (p, n))

print()
print("=" * 92)
print("3. PRECEDENCE, NORMALISED, ACROSS EVERY RIVERSIDE OUTPUT")
print("=" * 92)
import glob
for f in sorted(glob.glob('outputs/Riverside*')):
    if f.endswith('.xlsx') or f.endswith('.pdf'):
        continue
    body = norm(io.open(f, encoding='utf-8', errors='ignore').read())
    hits = [p for p in PRECEDENCE if p in body]
    print("   %-72s %s" % (f.split('/')[-1][:72], hits or '-'))
