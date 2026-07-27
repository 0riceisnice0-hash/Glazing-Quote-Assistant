import openpyxl, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

P = (r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\1. Tender Documents"
     r"\E T & S Construction\St Mary's Refurbishment\1. Estimating\3. Client Quote\SS"
     r"\ET & S Construction - St Mary's Refurbishment Pricing - DO NOT SEND.xlsx")
ws = openpyxl.load_workbook(P, data_only=True)["Pricing Document "]

# clause 2.24 applies to "glazed entrance doors and glazed screens"
DOORS_SCREENS = {"Type G", "Type I", "Type L", "Type O", "Type U", "Type AF", "Type AK"}
TALL_SCREENS = {"Type F", "Type H"}          # 3620mm windows with sills - arguable

def size(s):
    m = re.match(r"\s*(\d+)\s*x\s*(\d+)", str(s).replace(" ", ""))
    if not m:
        m = re.match(r"\s*(\d+)\s*x\s*(\d+)", str(s))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)

print("MANIFESTATION EXTENT - schedule 2376-09 clause 2.24")
print("two bands, 850-1000mm and 1400-1600mm above FFL, contrasting, both faces")
print("=" * 96)
print("%-9s %-13s %4s %8s %10s %12s" % ("type", "size", "qty", "width m", "per band", "2 bands"))

core = extra = 0.0
for r in range(9, 48):
    desc = str(ws.cell(r, 3).value or "").strip()
    if desc not in DOORS_SCREENS | TALL_SCREENS:
        continue
    w, h = size(ws.cell(r, 5).value)
    qty = ws.cell(r, 6).value or 0
    if w is None:
        print("  %-9s  UNPARSED SIZE %r" % (desc, ws.cell(r, 5).value)); continue
    band = w / 1000.0 * qty
    grp = "core" if desc in DOORS_SCREENS else "arguable"
    if desc in DOORS_SCREENS:
        core += band * 2
    else:
        extra += band * 2
    print("%-9s %-13s %4d %8.3f %10.3f %12.3f   %s"
          % (desc, "%dx%d" % (w, h), qty, w / 1000.0, band, band * 2, grp))

print("-" * 96)
print("CORE - glazed doors + screens (9 units, Types G/I/L/O/U/AF/AK): %.2f linear m of band" % core)
print("ARGUABLE - Types F and H, 3620mm tall but silled windows        : %.2f linear m" % extra)
print("IF BOTH                                                         : %.2f linear m" % (core + extra))
print()
print("STRIP-OUT QUANTITY (SOW 1.09, measured in m2 by MTCBC)")
print("=" * 96)
units = area = 0
for r in range(9, 48):
    if not ws.cell(r, 2).value:
        continue
    q = ws.cell(r, 6).value or 0
    units += q
    area += (ws.cell(r, 15).value or 0) * q
print("  existing openings to strip out and dispose of : %d units" % units)
print("  measured area (SOW 1.09 unit is m2)           : %.2f m2" % area)
print("  no strip-out or disposal rate exists in data/supplier-rates.json (0 of 80 categories)")
