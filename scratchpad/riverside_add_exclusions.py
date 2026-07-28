# -*- coding: utf-8 -*-
"""Put Fenster's standard exclusions onto the Riverside pricing document.

They were never on it. The twelve-line INCLUSIONS/EXCLUSIONS schedule lives in
templates/proposal-content.json, which is the proposal / cover-letter path;
Riverside was generated from MASTER PRICING DOC.xlsx, which has no exclusions
block at all. So every exclusion this chat has relied on - structural
alterations to the main contractor, design and structural calculations to
others, testing, site storage, scaffold, waste - exists only in a document this
job has never produced.

An exclusion that is not in the document you issue is not an exclusion.

Also rewrites the C31 footnote. It read "This pricing document should be read in
conjunction with the Terms and Conditions" - an incorporation by reference with
no title, no revision and no date, which is exactly the shape Gordon Court found
in BSW's quotations and described as worse than A Plus's named one. We were
doing to RRR what BSW do to us.
"""
import openpyxl
from openpyxl.styles import Font, Alignment

P = 'outputs/Riverside House - Fenster Pricing Document (house format).xlsx'
wb = openpyxl.load_workbook(P)
ws = wb.active

before_total = ws['I23'].value
before_i21 = type(ws['I21'].value).__name__

ws['C31'] = ("** This pricing document should be read with Fenster Glazing & Locks Ltd's "
             "Standard Terms and Conditions (issue 31.05.2026), a copy of which accompanies "
             "this document.")

BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')

rows = [
    ("EXCLUSIONS - the following are NOT included in the price above", True),
    ("Design responsibility - design calculations, structural calculations and engineer "
     "approvals, including any wind loading check and the structural design of fixings, "
     "unless specifically included within our scope", False),
    ("Structural alterations - forming, enlarging or making good any structural opening, "
     "to be completed by the main contractor", False),
    ("AOV control system - smoke control panel, mains and battery-backed supply, cabling, "
     "containment, fire-brigade override at ground floor access level, and commissioning", False),
    ("Testing - on or off site testing of the completed smoke ventilation system", False),
    ("Anti-fall protection to Part K, where the vent cill sits below 1100mm from finished "
     "floor level, and any preventative measures for trap hazard below 2.5m", False),
    ("Access and lifting equipment - scaffold, MEWPs, towers, forklift and the like", False),
    ("Site storage - materials will be delivered to site", False),
    ("Site welfare - welfare facilities, power, water, lighting", False),
    ("Waste removal, internal finishing, fire stopping and final clean", False),
    ("Traffic management - road closures, street licences, parking suspensions", False),
    ("Dimensions provided by others are assumed to be accurate. Any additional costs arising "
     "from incorrect dimensions will be treated as a variation and charged accordingly", False),
    ("The free area stated is GEOMETRIC. No aerodynamic figure has been provided by the "
     "manufacturer and none is warranted by this quotation", False),
]

r = 33
for text, is_head in rows:
    c = ws.cell(r, 3)
    c.value = ("" if is_head else "- ") + text
    c.alignment = WRAP
    if is_head:
        c.font = BOLD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 15 if is_head else 26
    r += 1

wb.save(P)

wb2 = openpyxl.load_workbook(P)
ws2 = wb2.active
print('I23 formula before/after: %r / %r' % (before_total, ws2['I23'].value))
print('I21 type before/after:    %s / %s' % (before_i21, type(ws2['I21'].value).__name__))
print('exclusion rows written:   33-%d' % (r - 1))
