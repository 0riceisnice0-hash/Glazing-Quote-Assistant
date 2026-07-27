import openpyxl, sys

P = ('C:/Users/zacpl/OneDrive - Fenster Glazing (1)/Commercial/1. Tender Documents/'
     'Chigwell (London) PLC/Gordon Court/1. Estimating/3. Client Quote/'
     'Chigwell Group - Gordon Court Pricing DO NOT SEND.xlsx')

CLIENT = P.replace(' DO NOT SEND', '')


def clean(s):
    return str(s).encode('ascii', 'replace').decode('ascii')


def dump(path, formulas=False):
    wb = openpyxl.load_workbook(path, data_only=not formulas)
    ws = wb.worksheets[0]
    print('#' * 20, 'FORMULAS' if formulas else 'VALUES', path.split('/')[-1])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = [(c.coordinate, c.value) for c in row if c.value not in (None, '')]
        if cells:
            print(' | '.join('%s=%s' % (k, clean(v)[:60]) for k, v in cells))


if __name__ == '__main__':
    which = sys.argv[1] if len(sys.argv) > 1 else 'v'
    if which == 'v':
        dump(P)
    elif which == 'f':
        dump(P, formulas=True)
    elif which == 'c':
        dump(CLIENT)
