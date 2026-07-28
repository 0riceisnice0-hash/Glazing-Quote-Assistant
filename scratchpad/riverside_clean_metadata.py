# -*- coding: utf-8 -*-
"""Clean the document properties on both Riverside deliverables.

Gordon Court found dc:creator = "Dan Parker;dan.parker@agsurveying.co.uk" in
docProps/core.xml of their issued pricing document. It replicates here exactly -
same person, same email, on the Riverside file - because both are clones of
MASTER PRICING DOC.xlsx, whose dcterms:created is 2018-12-07. That template has
been carrying a named third party's work email address as its author for seven
and a half years.

It survived last night's clean because that script dropped xl/externalLinks/ and
nothing else. My own lesson was "state where you looked", and I looked in cells,
then in external links, and stopped. docProps is a third store.

Two differences from their situation, both worth stating:

  - Riverside is UNISSUED, so the file is overwritten rather than copied. Their
    file went to Chigwell on 09/07 and they were right to leave it alone: it is
    the record of what the client actually received, and cleaning it would
    destroy the evidence of what was sent.
  - The decision about the personal data itself is not mine either way. Nothing
    of ours has been sent, so there is nothing to disclose on this job - but the
    template is everyone's problem and that goes to the board, not to a client.
"""
import re
import shutil
import zipfile

import openpyxl
import pypdf

# ------------------------------------------------------------------ xlsx
P = 'outputs/Riverside House - Fenster Pricing Document (house format).xlsx'
TMP = P + '.tmp'

b = openpyxl.load_workbook(P)
bws = b.active
before = (bws['I23'].value, type(bws['I21'].value).__name__, len(bws['H5'].value or ''),
          sum(1 for r in range(33, 46) if bws.cell(r, 3).value),
          sum(1 for row in bws.iter_rows() for c in row if c.value not in (None, '')))

CORE = (u'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        u'<cp:coreProperties '
        u'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        u'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        u'xmlns:dcterms="http://purl.org/dc/terms/" '
        u'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        u'<dc:creator>Fenster Glazing &amp; Locks Ltd</dc:creator>'
        u'<cp:lastModifiedBy>Fenster Glazing &amp; Locks Ltd</cp:lastModifiedBy>'
        u'<dc:title>Riverside House - AOV Smoke Vents - Pricing Document</dc:title>'
        u'<cp:category>Quotation</cp:category>'
        u'</cp:coreProperties>')

zin = zipfile.ZipFile(P)
with zipfile.ZipFile(TMP, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        n = item.filename
        if n == 'docProps/core.xml':
            zout.writestr(item, CORE.encode('utf-8'))
            continue
        data = zin.read(n)
        if n == 'docProps/app.xml':
            s = data.decode('utf-8', 'ignore')
            s = re.sub(r'<Company>[^<]*</Company>', '<Company>Fenster Glazing &amp; Locks Ltd</Company>', s)
            s = re.sub(r'<Manager>[^<]*</Manager>', '', s)
            data = s.encode('utf-8')
        zout.writestr(item, data)
zin.close()
shutil.move(TMP, P)

a = openpyxl.load_workbook(P)
aws = a.active
after = (aws['I23'].value, type(aws['I21'].value).__name__, len(aws['H5'].value or ''),
         sum(1 for r in range(33, 46) if aws.cell(r, 3).value),
         sum(1 for row in aws.iter_rows() for c in row if c.value not in (None, '')))
z = zipfile.ZipFile(P)
leak = [n for n in z.namelist()
        if re.search(rb"agsurveying|Dan Parker|LiamO|Content\.Outlook", z.read(n))]

print("XLSX      total formula  %s -> %s" % (before[0], after[0]))
print("          I21 type       %s -> %s" % (before[1], after[1]))
print("          H5 chars       %s -> %s" % (before[2], after[2]))
print("          exclusion rows %s -> %s" % (before[3], after[3]))
print("          populated cells %s -> %s" % (before[4], after[4]))
print("          parts holding a third-party name/path: %s" % (leak or 'none'))

# ------------------------------------------------------------------- pdf
D = 'outputs/Riverside House - AOV Smoke Vent Drawings.pdf'
r = pypdf.PdfReader(D)
print("\nPDF       before: %s" % {k: str(v)[:52] for k, v in (r.metadata or {}).items()})
w = pypdf.PdfWriter()
for pg in r.pages:
    w.add_page(pg)
w.add_metadata({
    '/Title': 'Riverside House - AOV Smoke Vent Drawings - Rev A',
    '/Author': 'Fenster Glazing & Locks Ltd',
    '/Subject': 'AOV.01 and AOV.02, stairwell smoke vents, 1130 x 1530',
    '/Creator': 'Fenster Glazing & Locks Ltd',
    '/Producer': 'Fenster Glazing & Locks Ltd',
})
with open(D, 'wb') as f:
    w.write(f)
r2 = pypdf.PdfReader(D)
print("PDF       after:  %s" % {k: str(v)[:52] for k, v in (r2.metadata or {}).items()})
print("PDF       pages %d -> %d, page 1 text intact: %s"
      % (len(r.pages), len(r2.pages),
         'RIVERSIDE HOUSE' in (r2.pages[0].extract_text() or '')))
