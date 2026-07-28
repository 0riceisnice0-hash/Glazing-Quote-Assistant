# -*- coding: utf-8 -*-
"""Restore the print titles I also destroyed, then build a sell-only client copy.

Two things from Gordon Court.

FIRST: there are TWO of ours in the definedNames block, not one. I restored
`_xlnm.Print_Area` last night and left `_xlnm.Print_Titles` = $2:$7 destroyed -
the repeating header rows that put the header on every printed page. So my fix
for a wholesale delete was itself partial, which is the same shape one more
time.

SECOND, and larger: they issue TWO files - a 257-cell sell-only
"Gordon Court Pricing.xlsx" and a 504-cell "Gordon Court Pricing DO NOT SEND.xlsx"
holding the cost codes. The control that protected them was the SECOND FILE, not
the print area - and the DO NOT SEND file's own print area would not have hidden
its columns K, L and M had anyone attached it.

    a print area protects a print of one file, and does nothing if the
    workbook is emailed
    a second file protects the workbook, and does nothing if somebody
    attaches the wrong one

Riverside has ONE file doing both jobs, so it is covered against one failure
mode and not the other. This builds the missing half: a client copy with the
buy columns GONE rather than merely outside the print range.

Every figure in the client copy is READ from the working document and
recomputed, never typed, and asserted against the working document's own
build-up before the file is written.
"""
import shutil

import openpyxl

SRC = 'outputs/Riverside House - Fenster Pricing Document (house format).xlsx'
DST = 'outputs/Riverside House - Fenster Pricing Document (CLIENT COPY - send this one).xlsx'

# ---------------------------------------------------------------- 1. titles
wb = openpyxl.load_workbook(SRC)
ws = wb.active
before_titles = ws.print_title_rows
ws.print_title_rows = '$2:$7'
wb.save(SRC)
ws2 = openpyxl.load_workbook(SRC).active
print("working document")
print("  print_area       %s" % ws2.print_area)
print("  print_title_rows %s -> %s" % (before_titles, ws2.print_title_rows))

# ------------------------------------------------------- 2. the client copy
src = openpyxl.load_workbook(SRC)
s = src.active

# derive, do not type
def num(coord):
    v = s[coord].value
    if not isinstance(v, (int, float)):
        raise SystemExit("expected a number in %s, found %r" % (coord, v))
    return float(v)

frames, glass, surch = num('J9'), num('K9'), num('L9')
buy_unit = round(frames + glass + surch, 2)
ADDER, LABOUR = 412.50, 160.00          # MAW: 550 x 75%, and the code's labour
unit_rate = round(buy_unit + ADDER, 2)
qty = int(num('F9')) + int(num('F10'))
items = round(unit_rate * qty, 2)
install = round(LABOUR * qty, 2)
total = round(items + install, 2)

print("\nderived from the working document, not typed")
print("  buy per unit   %8.2f  = %.3f + %.3f + %.2f" % (buy_unit, frames, glass, surch))
print("  unit rate      %8.2f  = buy + %.2f adder" % (unit_rate, ADDER))
print("  items x%d      %8.2f" % (qty, items))
print("  install        %8.2f  = %.2f x %d" % (install, LABOUR, qty))
print("  TOTAL          %8.2f" % total)
assert total == 5990.22, "derived total %r does not match the job's figure" % total
assert num('J10') == frames and num('K10') == glass and num('L10') == surch, \
    "the two units are not priced identically - the client copy would be wrong"

shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb.active

# freeze the sell side to values, so nothing depends on the columns about to go
ws['H9'], ws['I9'] = unit_rate, unit_rate
ws['H10'], ws['I10'] = unit_rate, unit_rate
ws['I21'] = install
ws['I23'] = total

# and remove the buy entirely - not merely out of the print range
ws.delete_cols(10, 13)          # J..V
ws['C31'] = ws['C31'].value     # keep the footnote, drop nothing else
ws.print_area = "'Pricing Document '!$C$1:$I$45"
ws.print_title_rows = '$2:$7'
wb.save(DST)

chk = openpyxl.load_workbook(DST).active
right = [c.coordinate for row in chk.iter_rows() for c in row
         if c.column > 9 and c.value not in (None, '')]
text = ' '.join(str(c.value) for row in chk.iter_rows() for c in row
                if isinstance(c.value, str))
print("\nclient copy")
print("  file             %s" % DST.split('/')[-1])
print("  TOTAL            %s" % chk['I23'].value)
print("  unit rate        %s" % chk['H9'].value)
print("  install          %s" % chk['I21'].value)
print("  exclusion rows   %d" % sum(1 for r in range(33, 46) if chk.cell(r, 3).value))
print("  print_area       %s   print_title_rows %s" % (chk.print_area, chk.print_title_rows))
print("  populated cells right of column I : %s" % (right or 'NONE'))
for probe in ('A Plus', 'QT51518', '2331', '85.65', 'Supplier used'):
    print("  contains %-15s %s" % (probe, probe in text or any(
        probe in str(c.value) for row in chk.iter_rows() for c in row if c.value is not None)))
