import openpyxl

BASE = ('C:/Users/zacpl/OneDrive - Fenster Glazing (1)/Commercial/1. Tender Documents/'
        'Chigwell (London) PLC/Gordon Court/1. Estimating/3. Client Quote/')
INT = BASE + 'Chigwell Group - Gordon Court Pricing DO NOT SEND.xlsx'
CLI = BASE + 'Chigwell Group - Gordon Court Pricing.xlsx'


def clean(s):
    return str(s).encode('ascii', 'replace').decode('ascii')


for label, path in (('INTERNAL', INT), ('CLIENT', CLI)):
    wbf = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    wsf, wsv = wbf.worksheets[0], wbv.worksheets[0]
    print('#' * 25, label, wsf.title, '| dims', wsf.dimensions)
    print('  print_area:', wsf.print_area)
    hidden = [k for k, d in wsf.column_dimensions.items() if d.hidden]
    print('  HIDDEN COLUMNS:', hidden or 'NONE')
    print('  formulas of interest:')
    for coord in ['M3', 'M4', 'M5', 'I61', 'I63', 'I57', 'I58', 'I59',
                  'H57', 'H58', 'H59', 'J57', 'J58', 'J59', 'I67', 'I68']:
        f, v = wsf[coord].value, wsv[coord].value
        print('    %-5s formula=%-46s value=%s' % (coord, clean(f)[:46], clean(v)[:28]))
    # anything in J..V on the client copy = leaked cost
    leak = []
    for row in wsv.iter_rows(min_col=10, max_col=22):
        for c in row:
            if c.value not in (None, ''):
                leak.append(c.coordinate)
    print('  CELLS POPULATED IN COLS J-V:', len(leak), leak[:18])
    print()
