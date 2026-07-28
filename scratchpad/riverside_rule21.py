# -*- coding: utf-8 -*-
"""Rule 21: the client's view of the priced workbook, checked rather than assumed.

Two chats broke the same protection with the same line of code on the same
night. Riverside destroyed `_xlnm.Print_Area` stripping foreign defined names,
restored it, and left `_xlnm.Print_Titles` destroyed - so the fix for a wholesale
delete was itself partial. Gordon Court did all of it identically.

Neither of us found it by checking. Riverside found the print area because
Gordon Court found 51 buy prices in a file called "Elevations"; Gordon Court
found the print titles because Riverside posted the print area. That is two
accidents in a row, and the reason for a rule.

It asks three things of the workbook the client is actually sent:

    is there a print area at all
    are there populated cells outside it
    did the repeating header rows survive

The second is the one that matters. Gordon Court's point, which is sharper than
either finding: a print area protects a print of one file and does nothing if
the workbook is emailed; a second sell-only file protects the workbook and does
nothing if somebody attaches the wrong one. Cells outside the print area mean
you are relying on the weaker of the two.
"""
import io

P = 'scripts/mary_checks.py'
t = io.open(P, encoding='utf-8').read()

RULE = '''def check_priced_document_view_is_intact(m):
    """Riverside House, 28/07, after two chats destroyed the same protection
    with the same line of code within an hour of each other.

    MASTER PRICING DOC.xlsx puts the supplier buy in columns J, K and L and sets
    a print area that stops at column I, so a printed or PDF'd quotation never
    carries it. That protection is stored as a defined name, `_xlnm.Print_Area`,
    in the same block as 50 foreign names inherited from an electrical
    template - and a regex that removes the block removes it too. So does
    `_xlnm.Print_Titles`, the repeating header rows, which Riverside missed on
    the first restore and Gordon Court found.

    Checks the document that actually goes to the client:
      - a print area exists;
      - NOTHING is populated outside it;
      - the repeating header rows survived.

    The middle one carries the weight. Gordon Court: a print area protects a
    print of one file and does nothing if the workbook is emailed; a second
    sell-only file protects the workbook and does nothing if somebody attaches
    the wrong one. Populated cells outside the print area mean you are covered
    against one failure mode of the two.

    Uses 'issued_documents': [{name, path, is_the_priced_document}]."""
    docs = m.get("issued_documents")
    if docs is None:
        return result("the client's view of the priced workbook", UNKNOWN,
                      "State the priced document and its path: 'issued_documents': "
                      "[{name, path, is_the_priced_document}].", "Riverside House",
                      remedy="Add a path to the priced document so it can be opened.")
    if isinstance(docs, dict):
        docs = [docs]
    if isinstance(docs, str) or not isinstance(docs, (list, tuple)):
        return result("the client's view of the priced workbook", UNKNOWN,
                      "'issued_documents' is %r - it must be a list." % (docs,),
                      "Riverside House", remedy="Rewrite the field as a list.")
    books = [d for d in docs if isinstance(d, dict) and d.get("is_the_priced_document")
             and str(d.get("path", "")).lower().endswith((".xlsx", ".xlsm"))]
    if not books:
        return result("the client's view of the priced workbook", NA,
                      "no priced workbook on this job - nothing to hide behind a print area",
                      "Riverside House")
    bad, unreadable = [], []
    for doc in books:
        name, path = doc.get("name", "?"), doc.get("path")
        try:
            import openpyxl
            ws = openpyxl.load_workbook(path).active
        except Exception as exc:
            unreadable.append("%s could not be opened (%s: %s)"
                              % (name, type(exc).__name__, exc))
            continue
        area = ws.print_area
        if not area:
            bad.append("%s has NO print area - the whole sheet prints, including any working "
                       "columns" % name)
            continue
        try:
            from openpyxl.utils import range_boundaries
            ref = area[0] if isinstance(area, (list, tuple)) else area
            ref = str(ref).split("!")[-1]
            c1, r1, c2, r2 = range_boundaries(ref)
        except Exception as exc:
            unreadable.append("%s has a print area (%r) this rule cannot parse (%s)"
                              % (name, area, type(exc).__name__))
            continue
        outside = []
        for row in ws.iter_rows():
            for c in row:
                if c.value in (None, ""):
                    continue
                if c.column < c1 or c.column > c2 or c.row < r1 or c.row > r2:
                    outside.append(c.coordinate)
        if outside:
            bad.append("%s has %d populated cell(s) OUTSIDE its print area %s (%s%s) - a print "
                       "area protects a print, not the file"
                       % (name, len(outside), ref, ", ".join(sorted(outside)[:6]),
                          "..." if len(outside) > 6 else ""))
        if not ws.print_title_rows:
            bad.append("%s has no repeating header rows (_xlnm.Print_Titles) - the header will "
                       "appear on page 1 only" % name)
    if unreadable:
        return result("the client's view of the priced workbook", UNKNOWN,
                      "; ".join(unreadable) + ". Not checked is not the same as intact.",
                      "Riverside House", remedy="Fix the path, then re-run.")
    if bad:
        return result("the client's view of the priced workbook", FAIL,
                      "; ".join(bad) + ".", "Riverside House",
                      remedy="Restore the print area and print titles - both are defined names "
                             "and a wholesale definedNames delete takes them - and issue a "
                             "sell-only copy with the working columns REMOVED rather than merely "
                             "outside the printed range.")
    return result("the client's view of the priced workbook", PASS,
                  "%d priced workbook(s): print area set, print titles intact, nothing populated "
                  "outside the printed range" % len(books), "Riverside House")


def check_no_third_party_traces_in_issued_files(m):'''

assert t.count('def check_no_third_party_traces_in_issued_files(m):') == 1
t = t.replace('def check_no_third_party_traces_in_issued_files(m):', RULE, 1)

OLD = '''    check_exposures_state_our_recourse, check_no_third_party_traces_in_issued_files,
]'''
NEW = '''    check_exposures_state_our_recourse, check_no_third_party_traces_in_issued_files,
    check_priced_document_view_is_intact,
]'''
assert t.count(OLD) == 1
t = t.replace(OLD, NEW)

# ---------------------------------------------------------------- variants
ANCHOR = 'def selftest_trace_variants():'
BLOCK = '''def selftest_view_variants():
    """Recall test for check_priced_document_view_is_intact.

    Synthetic workbooks built and destroyed here. The founding cases are the
    two both chats actually committed - print area deleted, and print titles
    deleted while the print area was restored - plus the one that matters
    commercially: a cell populated outside the printed range.
    """
    import shutil
    import tempfile
    import openpyxl
    d = tempfile.mkdtemp(prefix="mary-view-")
    try:
        def book(name, area, titles, extra=None):
            p = os.path.join(d, name)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["C3"] = "Client:"
            ws["I23"] = 5990.22
            if extra:
                ws[extra] = 2331.075
            if area:
                ws.print_area = area
            if titles:
                ws.print_title_rows = titles
            wb.save(p)
            return p

        good = book("good.xlsx", "$C$1:$I$45", "$2:$7")
        no_area = book("noarea.xlsx", None, "$2:$7")
        no_titles = book("notitles.xlsx", "$C$1:$I$45", None)
        buy_outside = book("buy.xlsx", "$C$1:$I$45", "$2:$7", extra="J9")
        both_gone = book("bothgone.xlsx", None, None)

        VARIANTS = [
            ("field absent",             None,                                        UNKNOWN),
            ("empty list",               [],                                          NA),
            ("no priced workbook",       [{"name": "note", "path": "x.txt",
                                           "is_the_priced_document": True}],          NA),
            ("priced flag not set",      [{"name": "g", "path": good}],               NA),
            ("intact",                   [{"name": "g", "path": good,
                                           "is_the_priced_document": True}],          PASS),
            ("no print area",            [{"name": "n", "path": no_area,
                                           "is_the_priced_document": True}],          FAIL),
            ("print titles destroyed",   [{"name": "t", "path": no_titles,
                                           "is_the_priced_document": True}],          FAIL),
            ("buy price outside the printed range",
                                         [{"name": "b", "path": buy_outside,
                                           "is_the_priced_document": True}],          FAIL),
            ("both defined names gone",  [{"name": "x", "path": both_gone,
                                           "is_the_priced_document": True}],          FAIL),
            ("one intact one broken",    [{"name": "g", "path": good,
                                           "is_the_priced_document": True},
                                          {"name": "b", "path": buy_outside,
                                           "is_the_priced_document": True}],          FAIL),
            ("path does not exist",      [{"name": "x", "path": os.path.join(d, "no.xlsx"),
                                           "is_the_priced_document": True}],          UNKNOWN),
            ("a dict, not a list",       {"name": "g", "path": good,
                                          "is_the_priced_document": True},            PASS),
            ("a bare string",            "good.xlsx",                                 UNKNOWN),
            ("entry is not a dict",      ["good.xlsx"],                               NA),
        ]

        bad = []
        for name, value, expect in VARIANTS:
            m = {} if value is None else {"issued_documents": value}
            try:
                got = check_priced_document_view_is_intact(m)["status"]
            except Exception as exc:
                got = "EXCEPTION %s: %s" % (type(exc).__name__, exc)
            if got != expect:
                bad.append("%s: expected %s, got %s" % (name, expect, got))
        print("  %-22s %d/%d client-view variants behave as intended%s"
              % ("client view", len(VARIANTS) - len(bad), len(VARIANTS),
                 "" if not bad else "  MISSED: " + "; ".join(bad)))
        return not bad
    finally:
        shutil.rmtree(d, ignore_errors=True)


def selftest_trace_variants():'''
assert t.count(ANCHOR) == 1
t = t.replace(ANCHOR, BLOCK)

OLD_W = '''    if not selftest_trace_variants():
        ok = False'''
NEW_W = '''    if not selftest_trace_variants():
        ok = False
    if not selftest_view_variants():
        ok = False'''
assert t.count(OLD_W) == 1
t = t.replace(OLD_W, NEW_W)

io.open(P, 'w', encoding='utf-8', newline='').write(t)
print('rule 21 added')
