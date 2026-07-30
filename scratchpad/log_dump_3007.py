# -*- coding: utf-8 -*-
"""READ-ONLY dump of the Estimating Log for the morning cross-check."""
import os
import datetime as dt
import openpyxl

SRC = r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\13. Estimating\Leads\Estimating Log.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log-dump-3007.txt")

wb = openpyxl.load_workbook(SRC, data_only=True)


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%d/%m/%Y")
    return str(v).replace("\n", " ").strip()


with open(OUT, "w", encoding="utf-8") as fh:
    fh.write("SHEETS: %s\n" % wb.sheetnames)
    ws = wb["Estimating Log"]
    fh.write("\n=== Estimating Log: %d rows x %d cols ===\n" % (ws.max_row, ws.max_column))
    rows = list(ws.iter_rows(values_only=True))
    # find header row
    hdr_i = 0
    for i, r in enumerate(rows[:10]):
        if sum(1 for c in r if c not in (None, "")) >= 4:
            hdr_i = i
            break
    hdr = [cell(c) for c in rows[hdr_i]]
    fh.write("HEADER(row %d): %s\n" % (hdr_i + 1, hdr))
    for i, r in enumerate(rows[hdr_i + 1:], start=hdr_i + 2):
        vals = [cell(c) for c in r]
        if not any(vals):
            continue
        fh.write("\n-- row %d --\n" % i)
        for h, v in zip(hdr, vals):
            if v:
                fh.write("   %-22s %s\n" % ((h or "?")[:22], v[:220]))
print("written", OUT)
