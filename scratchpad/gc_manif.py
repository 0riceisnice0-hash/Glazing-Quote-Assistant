# -*- coding: utf-8 -*-
"""Quantify Gordon Court manifestation using st-marys' method: width x 2 bands.

Manifestation is two bands (typically 850-1000mm and 1400-1600mm above floor)
running the full width of the glazed element, so linear metres = width x 2.
"""
import openpyxl

WB = ('C:/Users/zacpl/OneDrive - Fenster Glazing (1)/Commercial/1. Tender Documents/'
      'Chigwell (London) PLC/Gordon Court/1. Estimating/3. Client Quote/'
      'Chigwell Group - Gordon Court Pricing DO NOT SEND.xlsx')

ws = openpyxl.load_workbook(WB, data_only=True).worksheets[0]

# rows 49-59 are the external / communal doors, plus 10/12 the AOV+louvre
ROOMS = {
 'D_B': 'balcony doors to individual flats (private)',
 'D_C': 'plant room (no public route)',
 'D_D': 'Corridor 5-0 / Corridor 7-0 (COMMUNAL, single leaf)',
 'D_E': 'Flat 10 (private)',
 'D_U': 'Stair 2 (COMMUNAL, single leaf)',
 'D_A': 'external FD30S entrance doors, one at GR316 Entrance (COMMUNAL, DOUBLE)',
 'D_T': 'GR425 Store - scope disputed, RFI-1',
}

rows = []
for r in range(49, 60):
    code, ref, size, qty = (ws['B%d' % r].value, ws['C%d' % r].value,
                            ws['E%d' % r].value, ws['F%d' % r].value)
    if not (code and ref and size):
        continue
    w = float(str(size).split('x')[0].strip())
    rows.append((r, str(ref), w, int(qty or 0)))

def lin(sel):
    return sum(w / 1000.0 * 2 * q for _, ref, w, q in rows if ref in sel), \
           sum(q for _, ref, w, q in rows if ref in sel)

print('%-4s %-7s %-9s %-4s %-9s  %s' % ('row', 'ref', 'width mm', 'qty', 'lin m', 'location'))
for r, ref, w, q in rows:
    print('%-4d %-7s %-9.0f %-4d %-9.3f  %s' % (r, ref, w, q, w / 1000.0 * 2 * q,
                                                ROOMS.get(ref, '')))

print()
NARROW = {'D_A'}
MEDIUM = {'D_A', 'D_D', 'D_U'}
WIDE = {'D_A', 'D_D', 'D_U', 'D_T', 'D_E', 'D_B', 'D_C'}
for name, sel in (('NARROW  (D_A only - the doors that are unambiguously communal entrances)', NARROW),
                  ('MEDIUM  (+ D_D, D_U - the single-leaf communal doors cl.280 actually describes)', MEDIUM),
                  ('WIDE    (every glazed external door incl. private and plant)', WIDE)):
    m, n = lin(sel)
    print('%-74s %5.3f lin m over %2d units' % (name, m, n))

# the AOV / smoke-shaft units, as a separate question
aov = []
for r in (10, 12):
    ref, size, qty = ws['C%d' % r].value, ws['E%d' % r].value, ws['F%d' % r].value
    w = float(str(size).split('x')[0].strip())
    aov.append((str(ref), w, int(qty)))
    print()
print('SEPARATE QUESTION - full-height glazed elements in communal corridors:')
tot = 0
for ref, w, q in aov:
    tot += w / 1000.0 * 2 * q
    print('  %-6s %.0f mm wide x %d = %.3f lin m' % (ref, w, q, w / 1000.0 * 2 * q))
print('  subtotal %.3f lin m (only if Approved Doc K critical-location glazing is read to catch them)' % tot)
