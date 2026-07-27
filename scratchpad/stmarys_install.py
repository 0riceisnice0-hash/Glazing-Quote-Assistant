import openpyxl, os

LABOUR = {
    "SUPD": 250, "SAD": 250, "DUPD": 500, "DAD": 500, "ELAW": 250,
    "LAW": 160, "MAW": 160, "SAW": 160, "LPVC": 160, "MPVC": 160, "SPVC": 160,
    "SADLAW": 410, "SADMAW": 410, "SADSAW": 410,
}
CW_LABOUR_M2 = 150.0

P = (r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\1. Tender Documents"
     r"\E T & S Construction\St Mary's Refurbishment\1. Estimating\3. Client Quote\SS"
     r"\ET & S Construction - St Mary's Refurbishment Pricing - DO NOT SEND.xlsx")

wb = openpyxl.load_workbook(P, data_only=True)
ws = wb["Pricing Document "]

rows = []
for r in range(9, 48):
    code = ws.cell(r, 2).value          # B product code
    desc = ws.cell(r, 3).value          # C
    size = ws.cell(r, 5).value          # E
    qty = ws.cell(r, 6).value or 0      # F
    sqm = ws.cell(r, 15).value          # O CW SQM (unit area)
    cwlab = ws.cell(r, 14).value        # N CW LABOUR
    if not code:
        continue
    rows.append((r, code, desc, size, qty, sqm, cwlab))

tot_code, tot_area, tot_units = 0.0, 0.0, 0
print("%-4s %-8s %-9s %-13s %3s %8s %10s %10s" %
      ("row", "code", "desc", "size", "qty", "m2/ea", "codeLabour", "CW@150"))
for (r, code, desc, size, qty, sqm, cwlab) in rows:
    a = (sqm or 0) * qty
    if code == "CW":
        lab = a * CW_LABOUR_M2
    else:
        lab = LABOUR.get(code, 0) * qty
    cw_equiv = a * CW_LABOUR_M2
    tot_code += lab
    tot_area += a
    tot_units += qty
    print("%-4d %-8s %-9s %-13s %3d %8.3f %10.2f %10.2f" %
          (r, code, str(desc), str(size), qty, sqm or 0, lab, cw_equiv))

print()
print("total units          : %d" % tot_units)
print("total area m2        : %.3f" % tot_area)
print("labour by house codes: GBP %.2f" % tot_code)
print("quoted INSTALLATION  : GBP %.2f" % ws.cell(49, 9).value)
print("difference           : GBP %.2f" % (ws.cell(49, 9).value - tot_code))
print("all-lines-at-CW/150  : GBP %.2f" % (tot_area * CW_LABOUR_M2))
