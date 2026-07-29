# Build the WD001 tender return pack for Grange Hill Methodist Church.
#
# The document that exists today is the 24/07 BENCHMARK at GBP 27,560.07 - a number that
# died at 10:48 on 29/07 when BSW returned. It is also the only priced document on file, so
# if anyone reaches for it today they send the wrong figure AND our buy prices, to the QS
# who already derived our markup from the Gordon Court "Elevations" files (REQ-28).
#
# This replaces it: supplier-backed, grouped by specification clause, sell-only.
import json
import os
import subprocess
import sys

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "outputs")

# Grouped by spec clause, not by unit. Fine enough that Luke Baker can see the scope and
# strike a line if a group is not ours; coarse enough that it does not hand him our markup
# unit by unit - he already holds our Gordon Court buy prices from the same fabricator.
SOUTH = 14215.97      # 5900x2300 door element + 2No shaped gable units over
WEST = 11569.26       # 4588x2100 door element + 1200x2100 single door
WINDOWS = 11493.36    # the 11 plain rectangles nobody has located on a drawing
OPERATOR = 3000.00    # spec 3.13.1 - house allowance, no specialist approached
MANIFEST = 250.00     # spec 3.11.2 - house allowance
TOTAL = round(SOUTH + WEST + WINDOWS + OPERATOR + MANIFEST, 2)
assert abs(TOTAL - 40528.59) < 0.005, TOTAL

ROWS = [
    {"code": "", "size": "5900 x 2300 + 2/no 2900 x 2400", "qty": 1, "unit": "nr",
     "unitRateOverride": SOUTH,
     "desc": ("SOUTH ELEVATION - spec 3.14.1. Supply and install full storey height glazed entrance "
              "screen comprising 1/no 5900 x 2300mm door element with two 1200mm pivoted anti "
              "finger-trap door leaves and two fixed fields, level threshold, concealed panic bar, "
              "lever and cylinder, gate bolts to the secured leaf; with 2/no shaped units over "
              "(2900 x 2400 and 2900 x 2375) raking to the underside of the pitched roof. White "
              "polyester powder coated, polyamide thermal breaks, double glazed.")},
    {"code": "", "size": "4588 x 2100 + 1200 x 2100", "qty": 1, "unit": "nr",
     "unitRateOverride": WEST,
     "desc": ("WEST ELEVATION - spec 3.12.1. Supply and install full storey height glazed entrance "
              "screen comprising 1/no 4588 x 2100mm door element with a 1200mm pivoted anti "
              "finger-trap door leaf and two fixed fields, and 1/no separate 1200 x 2100mm single "
              "door element. Level threshold, concealed panic bar, lever and cylinder. White "
              "polyester powder coated, polyamide thermal breaks, double glazed.  ZERO RATED VAT "
              "ELEMENT - spec 3.12.1 (V).")},
    {"code": "", "size": "various - see qualification 1", "qty": 1, "unit": "item",
     "unitRateOverride": WINDOWS,
     "desc": ("WINDOWS - spec 3.11.1. Supply and install 11/no aluminium casement windows: 2/no "
              "1200 x 3000, 8/no 1200 x 1183 and 1/no 2000 x 2100. White polyester powder coated, "
              "polyamide thermal breaks, double glazed.  QUANTITIES AND LOCATIONS TO BE CONFIRMED "
              "- see qualification 1.")},
    {"code": "", "size": "-", "qty": 1, "unit": "item",
     "unitRateOverride": OPERATOR,
     "desc": ("ALLOWANCE - spec 3.13.1. Automatic operator package to the west disabled access door: "
              "operator sized to the door, safety sensor, additional door and frame strengthening, "
              "mounting plates, electrical supplies and equipment, internal and external push pads, "
              "internal emergency release and keyed isolating switch.  ALLOWANCE ONLY - no "
              "specialist quotation obtained, see qualification 8.  ZERO RATED VAT ELEMENT - spec "
              "3.13.1 (V).")},
    {"code": "", "size": "-", "qty": 1, "unit": "item",
     "unitRateOverride": MANIFEST,
     "desc": ("ALLOWANCE - spec 3.11.2. Apply fish symbol manifestations to the inner surface of all "
              "new glazing to the south and west elevations, approx 150 x 75mm, full width of glass, "
              "client to supply the pdf artwork.")},
]

QUALIFICATIONS = [
    ("1. WINDOW QUANTITIES ARE NOT AGREED. The 11/no windows above are taken from our fabricator's "
     "own re-scale of the elevations. There is no window schedule in the tender documents and no "
     "setting-out drawing has been issued, and specification clauses 3.11 to 3.14 describe the two "
     "screens and their doors only. Please confirm the quantity, size and location of these units. "
     "If they do not form part of WD001, deduct GBP 11,493.36."),
    ("2. CHAPEL FOLDING DOORS - spec 3.15.1 - EXCLUDED. No elevation of the chapel has been issued, "
     "so neither the door height nor the fixed glazed section above it has a dimension anywhere in "
     "the pack. Separately, the recessed bottom rail and the fixed glazed section over cannot both "
     "be provided within the folding door systems available to us. Please issue a chapel elevation "
     "and we will quote this element separately."),
    ("3. PRIVACY FILM - spec 3.15.2 - EXCLUDED, following 3.15.1."),
    ("4. INTERNAL FIRE CHECK DOORS - spec 3.16.1 and 3.16.2 - EXCLUDED. These are internal timber "
     "FD60 doorsets with timber infill framing, two layers of 12mm plasterboard both sides taped "
     "and skimmed, and magnetic hold-opens wired to the fire detection system. They appear on no "
     "drawing and in no door schedule. Please confirm whether they form part of WD001 and we will "
     "price them."),
    ("5. GLASS. Clause 3.11.1 names Pilkington Optitherm S1 with solar control, Arctic Blue, to the "
     "outer pane and Optitherm S1 plus to the inner. We have priced our fabricator's equivalent: "
     "6mm SKN 176 toughened / 16 / 6mm HP Neutral Low E argon filled to the windows, and 4mm "
     "Coolite SKN176II with 8.8 laminate / 6mm toughened Anti Sun Grey to the door elements. These "
     "are two different tints and we would recommend a single tint across both elevations. Samples "
     "will be provided before order in accordance with 3.11.1. If the named Pilkington products are "
     "required without substitution we will requote."),
    ("6. IRONMONGERY. Clauses 3.12.1 and 3.14.1 name a Briton 1438 fire escape push pad with Briton "
     "1413 access locking knob, and Yale Platinum 3-star euro cylinders with thumbturn, keyed alike. "
     "We have priced the fabricator's concealed panic bar with lever and screw-in cylinder, without "
     "internal thumbturns, offered under the 'or similar approved' provision. Please confirm "
     "acceptance or we will price the named products."),
    ("7. SUPPORT TO THE GLAZING ABOVE THE DOORS. The door elements are 100mm frames and the glazing "
     "above them is 70mm; the two cannot be mechanically coupled to one another. Our price assumes "
     "the glazing above each door element is independently supported - on the south elevation by the "
     "150 x 90 x 24 PFC facade header support beam shown on Silver Structural Surveys drawing "
     "S1323/02 Rev C01. Please confirm the supporting structure at the head of the west elevation "
     "door element. Where no independent support exists a change of system is required and the "
     "price will change."),
    ("8. THE AUTOMATIC OPERATOR IS AN ALLOWANCE, not a quotation. No specialist has been approached. "
     "It will be confirmed against a specialist price."),
    ("9. VAT. Priced exclusive of VAT. The specification identifies 3.12.1 (V) and 3.13.1 (V) as "
     "zero rated elements: GBP 14,569.26 of the sum above sits against those two clauses. The "
     "apportionment within 3.12.1 between the disabled access door and the surrounding screen is "
     "for your confirmation, and we will invoice in accordance with your direction and the "
     "appropriate customer certificate."),
    ("10. VALIDITY. The material costs behind this price are dated 29 July 2026 and are held for "
     "thirty days, expiring 28 August 2026. The works are programmed November 2026 to July 2027, so "
     "material costs are to be re-confirmed at the point of order."),
    ("11. All dimensions are our fabricator's, scaled from the 1:100 elevations. All sizes are "
     "subject to site survey before manufacture."),
    ("12. External perimeter mastic and EPDM membrane to the reveals are excluded. They can be "
     "included on confirmation of the perimeter to be sealed."),
]

INCLUSIONS = [
    "Supply and installation of the aluminium windows, doors and screens scheduled above",
    "White polyester powder coated finish, polyamide thermal breaks, double glazed units",
    "Level thresholds to the south and west entrance doors",
    "Panic hardware, lever furniture and cylinders to the entrance doors",
    "Allowance for the automatic operator package to the west disabled access door (see qualification 8)",
    "Allowance for the fish symbol manifestations, client to supply artwork",
    "Site survey before manufacture",
]

EXCLUSIONS_SHORT = [
    "Chapel folding doors and the glazed section over - spec 3.15.1 (qualification 2)",
    "Privacy film to the folding door glass - spec 3.15.2 (qualification 3)",
    "Removal of the existing internal fire check door - spec 3.16.1 (qualification 4)",
    "2/no FD60 internal fire check doorsets and associated builders work - spec 3.16.2 (qualification 4)",
    "Delivery to site - see the covering letter",
    "Structural support to the glazing above the door elements (qualification 7)",
    "External perimeter mastic and EPDM membrane to reveals (qualification 12)",
    "Making good, builders work, scaffolding, temporary works and out of hours working",
    "VAT",
]

JOB = {
    "client": "Chigwell (London) PLC",
    "fao": "Luke Baker, Senior Quantity Surveyor",
    "projectRef": "Grange Hill Methodist Church Ext - WD001 Windows and Doors",
    "siteAddress": "Grange Hill Methodist Church, Burrow Road, Chigwell, Essex, IG7 4HQ",
    "date": "29/07/2026",
    "supplier": "",
    "note": "",
    "rows": ROWS,
    # The 24/07 optional lines (mastic GBP 143.00, EPDM GBP 587.00) were sized on the dead
    # 23.49 m2 take-off. BSW measure 62.33 m2, so both are stale. Dropped rather than carried
    # - a stale number on a client document is the Redditch failure, not a rounding error.
    "mastic": 0,
    "epdm": 0,
    "priceText": "GBP %s" % "{:,.2f}".format(TOTAL),
    "summary": [
        "Fenster Glazing is pleased to submit its quotation for package WD001, Windows and Doors, "
        "at Grange Hill Methodist Church, Burrow Road, Chigwell.",
        "Our price covers specification clauses 3.11.1, 3.11.2, 3.12.1, 3.13.1 and 3.14.1. It is "
        "based on fabricator's quotations dated 29 July 2026 and is offered subject to the "
        "qualifications set out in our covering letter, which form part of this offer.",
    ],
    "inclusions": INCLUSIONS,
    "exclusions": EXCLUSIONS_SHORT,
    "outXlsx": os.path.join(OUT, "Grange Hill Methodist Church - WD001 Pricing Document (29-07-2026).xlsx"),
    "outHtml": os.path.join(OUT, "Grange Hill Methodist Church - WD001 Proposal (29-07-2026).html"),
}

job_path = os.path.join(REPO, "scratchpad", "grange-hill-wd001-job.json")
with open(job_path, "w", encoding="utf-8") as fh:
    json.dump(JOB, fh, indent=1)

r = subprocess.run([sys.executable, os.path.join(REPO, "scripts", "generate-fenster-docs.py"), job_path],
                   capture_output=True, text=True)
print(r.stdout.strip() or r.stderr.strip())
if r.returncode:
    sys.exit(r.returncode)

# --- make it sell-only ------------------------------------------------------------------
# The 28/07 workbook failed 'the client's view of the priced workbook' with 61 populated
# cells outside the print area. A print area protects a print, not the file: the client
# opens the xlsx and reads the working. Delete the working columns outright.
path = JOB["outXlsx"]
wb = openpyxl.load_workbook(path)
ws = wb["Pricing Document "]
ws.delete_cols(10, 8)                    # J..Q - Frames, Glass, Additional, CW, CW LABOUR, CW SQM
for r_ in range(1, ws.max_row + 1):      # B - the internal product codes
    ws.cell(row=r_, column=2).value = None
ws["C8"] = None                          # the template's own 'SMA Smart Wall' heading
# The template carries twelve example rows. Five are used, so seven still hold formulas that
# reference the working columns just deleted - they would open as #REF! on the client's screen.
last = 9 + len(ROWS) - 1
from openpyxl.cell.cell import MergedCell
for r_ in range(last + 1, 21):
    for c in ws[r_]:
        if not isinstance(c, MergedCell):
            c.value = None
# INSTALLATION is a SUMPRODUCT over the deleted columns, and labour is already inside the
# grouped rates above. Make it a plain zero so the TOTAL formula still adds.
ws["I21"] = 0
# Strip the OPTIONAL block: its two figures were sized on the superseded take-off.
for coord in ("D26", "F27", "I27", "F28", "I28"):
    ws[coord] = None

# Put the qualifications on the face of the PRICED document, not only in the covering
# letter. Riverside House is the founding case for that rule and SM5 Wexham is the recent
# one: a covering note and its attachments get separated, and what survives is the sheet
# with the number on it. Every exclusion that carries money has to be readable there.
from openpyxl.styles import Alignment, Font
qrow = 33   # below the template's own footnotes at C30 and C31 - do not overwrite those
ws["C%d" % qrow] = ("QUALIFICATIONS AND EXCLUSIONS - this quotation is offered subject to the "
                    "following, which form part of our offer.")
ws["C%d" % qrow].font = Font(bold=True)
for i, q in enumerate(QUALIFICATIONS):
    r_ = qrow + 1 + i
    rng = "C%d:I%d" % (r_, r_)
    if rng not in {str(x) for x in ws.merged_cells.ranges}:
        ws.merge_cells(rng)
    ws["C%d" % r_] = q
    ws["C%d" % r_].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r_].height = max(15, 11.0 * (int(len(q) / 105) + 1))
ws.print_area = "$C$1:$I$%d" % (ws.max_row)
# No ampersand: clean_issued_pack writes the creator straight into core.xml, and a bare '&'
# makes the part unparseable - it broke both files the first time this ran.
wb.properties.creator = "Fenster Glazing and Locks Ltd"
wb.properties.lastModifiedBy = "Fenster Glazing and Locks Ltd"
wb.save(path)

wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["Pricing Document "]
print("\nCLIENT COPY - %s" % os.path.basename(path))
for r_ in range(7, ws.max_row + 1):
    vals = [(c.coordinate, str(c.value)[:64]) for c in ws[r_] if c.value not in (None, "")]
    if vals:
        print("  ", vals)

qpath = os.path.join(REPO, "scratchpad", "grange-hill-qualifications.txt")
with open(qpath, "w", encoding="utf-8") as fh:
    fh.write("\n\n".join(QUALIFICATIONS))
print("\nqualifications: %s (%d)" % (qpath, len(QUALIFICATIONS)))
print("TOTAL GBP %s  (south %s / west %s / windows %s / operator %s / manifestations %s)"
      % tuple("{:,.2f}".format(x) for x in (TOTAL, SOUTH, WEST, WINDOWS, OPERATOR, MANIFEST)))
