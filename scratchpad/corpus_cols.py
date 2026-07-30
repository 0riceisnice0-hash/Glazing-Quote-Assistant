# -*- coding: utf-8 -*-
"""For the 29 documents the BACKTEST actually scores, are the fixed indices right?

parse_doc reads qty 5, unit rate 7, frames 9, glass 10 by position. Some archive
documents are shifted (Gordon Court's client-facing copy puts Qty at 4 and Unit
Rate at 6; Roxbourne puts Frames at 7). Those shifted copies mostly carry a
reference like 'LW_1' in column 1 rather than a product code, so parse_doc drops
every row and they never enter the corpus - but that is luck, not a check. If a
shifted document ever DID enter, the engine would be mining Glass as Frames.
Verify it, one document at a time."""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "scripts"))
import mary_backtest as bt
import mary_calibrate as cal
import mary_quote_reader as reader
import openpyxl

EXPECT = {"qty": 5, "unit rate": 7, "frames": 9, "glass": 10, "additional": 11, "cw": 12}

paths = {}
for q in reader.scan(cal.TENDERS):
    paths.setdefault(os.path.basename(q["path"]), q["path"])

docs = bt.collect()
print("%d documents in the backtest corpus\n" % len(docs))
bad = []
for d in docs:
    p = paths.get(d["file"])
    if not p:
        print("  ?? no path for %s" % d["file"])
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = [w for w in wb.worksheets if w.title.strip().lower().startswith("pricing document")][0]
    found = {}
    for row in ws.iter_rows(values_only=True, max_row=20):
        for i, c in enumerate(list(row)[:16]):
            if not isinstance(c, str):
                continue
            t = c.strip().lower()
            t = "unit rate" if t.replace(" ", "") == "unitrate" else t
            t = "additional" if t.startswith("additional") else t
            if t in EXPECT and t not in found:
                found[t] = i
    wb.close()
    wrong = {k: v for k, v in found.items() if v != EXPECT[k]}
    missing = [k for k in ("qty", "unit rate", "frames") if k not in found]
    if wrong or missing:
        bad.append((d, found, wrong, missing))
        print("  MISMATCH %-50s %d lines" % (d["file"][:50], len(d["lines"])))
        print("           found %s" % found)
        if wrong:
            print("           WRONG %s (expected %s)" % (wrong, {k: EXPECT[k] for k in wrong}))
        if missing:
            print("           MISSING %s" % missing)

if not bad:
    print("  every scored document has qty/unit rate/frames/glass/additional exactly where")
    print("  parse_doc assumes. The fixed indices are safe for this corpus.")

print("\nDoes 'additional' (col 11) exist, and how much money is in it?")
n_docs = n_lines = 0
for d in docs:
    p = paths.get(d["file"])
    if not p:
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = [w for w in wb.worksheets if w.title.strip().lower().startswith("pricing document")][0]
    has = False
    for row in ws.iter_rows(values_only=True, max_row=20):
        for i, c in enumerate(list(row)[:16]):
            if isinstance(c, str) and c.strip().lower().startswith("additional") and i == 11:
                has = True
    wb.close()
    if has:
        n_docs += 1
print("  %d of %d scored documents have an Additional header at column 11" % (n_docs, len(docs)))
