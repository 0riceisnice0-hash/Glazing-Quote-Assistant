# -*- coding: utf-8 -*-
"""Strip the foreign external link and defined names from the Riverside output.

MASTER PRICING DOC.xlsx carries a live external link to

    file:///C:\\Users\\LiamO'Donnell\\AppData\\Local\\Microsoft\\Windows\\
    INetCache\\Content.Outlook\\GM4B1OQ8\\Electrical Template - Draft - REV010.xlsx

- an Outlook attachment cache path on a named individual's machine, pointing at
  a third party's draft ELECTRICAL template - plus 48 defined names from two
  unrelated trades (electrical: FIRE_ALARM, CONTAINMENT, SMALL_POWER, PRELIMS;
  and structural steel: Beam, Column, RSJ, PFC, RHS, SHS).

Verified before touching anything: 74 formulas, NONE of them reference the
external workbook and NONE reference any of the 48 names. The GBP 5,990.22 does
not depend on any of it.

What it does do is put a named individual's local file path on a document we
would hand a client, and make Excel open it with "this workbook contains links
to one or more external sources that could be unsafe".

Only the Riverside OUTPUT is changed here. The template is shared and other
chats are quoting from it this week; breaking it mid-flight would be worse than
the fault. Flagged to the board instead.
"""
import os
import re
import shutil
import zipfile

import openpyxl

P = 'outputs/Riverside House - Fenster Pricing Document (house format).xlsx'
TMP = P + '.tmp'

before = openpyxl.load_workbook(P)
b_ws = before.active
b_total = b_ws['I23'].value
b_i21 = type(b_ws['I21'].value).__name__
b_h5 = b_ws['H5'].value
b_excl = sum(1 for r in range(33, 46) if b_ws.cell(r, 3).value)
b_names = len(before.defined_names)

DROP_PARTS = ('xl/externalLinks/',)
zin = zipfile.ZipFile(P)
with zipfile.ZipFile(TMP, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        n = item.filename
        if n.startswith(DROP_PARTS):
            continue
        data = zin.read(n)
        if n == 'xl/workbook.xml':
            s = data.decode('utf-8')
            s = re.sub(r'<definedNames>.*?</definedNames>', '', s, flags=re.S)
            s = re.sub(r'<externalReferences>.*?</externalReferences>', '', s, flags=re.S)
            data = s.encode('utf-8')
        if n == 'xl/_rels/workbook.xml.rels':
            s = data.decode('utf-8')
            s = re.sub(r'<Relationship[^>]*externalLink[^>]*/>', '', s)
            data = s.encode('utf-8')
        if n == '[Content_Types].xml':
            s = data.decode('utf-8')
            s = re.sub(r'<Override[^>]*externalLink[^>]*/>', '', s)
            data = s.encode('utf-8')
        zout.writestr(item, data)
zin.close()
shutil.move(TMP, P)

after = openpyxl.load_workbook(P)
a_ws = after.active
z = zipfile.ZipFile(P)
print("                       BEFORE                AFTER")
print("  I23 formula          %-20s %s" % (b_total, a_ws['I23'].value))
print("  I21 type             %-20s %s" % (b_i21, type(a_ws['I21'].value).__name__))
print("  H5 spec note         %-20s %s" % (len(b_h5 or ''), len(a_ws['H5'].value or '')))
print("  exclusion rows 33-45 %-20s %s"
      % (b_excl, sum(1 for r in range(33, 46) if a_ws.cell(r, 3).value)))
print("  defined names        %-20s %s" % (b_names, len(after.defined_names)))
print("  externalLink parts   %-20s %s"
      % (1, len([n for n in z.namelist() if 'externalLink' in n])))
print("  LiamO'Donnell in zip %-20s %s"
      % ('yes', any(b"LiamO" in z.read(n) for n in z.namelist()
                    if n.endswith(('.xml', '.rels')))))
