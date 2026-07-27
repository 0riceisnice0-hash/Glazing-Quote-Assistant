import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

P = (r"C:\Users\zacpl\Desktop\Glazing-Quote-Assistant\test-results\st-marys-input"
     r"\original-08-07\-\2, 3, 4 - SOW St. Marys.xlsx")

wb = openpyxl.load_workbook(P, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb:
    print("=" * 92)
    print("SHEET:", ws.title, ws.dimensions)
    n = 0
    for r in ws.iter_rows():
        cells = [(c.coordinate, str(c.value).strip()) for c in r
                 if c.value not in (None, "") and str(c.value).strip()]
        if not cells:
            continue
        n += 1
        if n > 160:
            print("   ... truncated")
            break
        print("  " + " | ".join("%s=%s" % (k, v[:70]) for k, v in cells))
