"""St Mary's Refurbishment, Merthyr Tydfil (E T & S Construction) - quote check and RFI schedule.

Consolidates four turns of audit on the tender ALREADY SUBMITTED 17/07/2026 at
GBP 174,546.37 ex VAT. The arithmetic is clean and is shown as such; everything
open is a question about what that price covers, not a change to it.

Sources, all read at source:
  - Internal workbook  "ET & S Construction - St Mary's Refurbishment Pricing - DO NOT SEND.xlsx"
  - Client proposal    "ET & S Construction - St Mary's Refurbishment Proposal.pdf"
  - Window schedule    2376-09 (priced) and 2376-09 rev A, 13.07.26
  - Door schedule      2376-08 (original 08/07 issue, empty revision block)
  - MTCBC ITT sections 2, 3, 4 - "2, 3, 4 - SOW St. Marys.xlsx" (prelims + schedule of works)
  - EDG02 Energy and Carbon Design Guidelines - Building Fabric
  - BSW Window Solutions QT252799, 15/07/2026  (Sheerline Prestige, windows)
  - Bellview Products 0000000483, 16/07/2026   (SMA Smart Wall Pocket + MC600 Plus)
  - Aplus QP70172, 22/07/2026 - an UNUSED alternative, not the backing
  - SMA "Smart Wall Profile.pdf" - manufacturer's published U-values

Output: outputs/St Marys Refurbishment - Quote Check and RFI Schedule.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"outputs\St Marys Refurbishment - Quote Check and RFI Schedule.xlsx"

HDR = PatternFill("solid", fgColor="1F2A44")
HDRF = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
TITLE = Font(bold=True, size=13, color="1F2A44")
RED = PatternFill("solid", fgColor="FCE4E4")
AMB = PatternFill("solid", fgColor="FFF3D6")
GRN = PatternFill("solid", fgColor="E4F3E6")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

TOTAL = 174546.365
INSTALL = 21915.05
BSW_NET = 61056.80
BELLVIEW_NET = 30352.38

# ---------------------------------------------------------------- helpers
def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = TITLE
    r = 2
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(size=9, italic=True, color="555555")
        r = 3
    return r + 1


def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c)
        cell.fill, cell.font, cell.border = HDR, HDRF, THIN
        cell.alignment = WRAP
    return row + 1


def widths(ws, *w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def put(ws, row, vals, fill=None, wrap_from=0, money=()):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font, c.border = BASE, THIN
        c.alignment = WRAP if (wrap_from and i >= wrap_from) else TOP
        if i in money and isinstance(v, (int, float)):
            c.number_format = '#,##0.00'
        if fill:
            c.fill = fill
    return row + 1


wb = openpyxl.Workbook()

# =============================================================== 1. Summary
ws = wb.active
ws.title = "Summary"
widths(ws, 34, 78)
r = title(ws, "ST MARY'S REFURBISHMENT - QUOTE CHECK AND RFI SCHEDULE",
          "Tender submitted 17/07/2026 (documents dated 16/07). Audit 27/07/2026. "
          "The price is NOT challenged - the arithmetic is exact. Everything below is a question "
          "about what that price covers.")

ROWS = [
    ("Job", "St Mary's Refurbishment to accommodate a special needs school"),
    ("Real name on the ITT", "Blessed Carlo Acutis Catholic School (St Mary's Campus), for Greenfield Special School"),
    ("Site", "Caedraw Rd, Merthyr Tydfil CF47 8HA  (our own documents say CF77 8HA - not a Merthyr postcode)"),
    ("Client", "E T & S Construction, FAO Tom Godfrey - Fenster is their glazing sub-contractor"),
    ("End client / architect", "Merthyr Tydfil CBC (Chris Evans, ref 2026-024) / cfw architects, drawing series 2376"),
    ("", ""),
    ("TENDER SUBMITTED", "GBP 174,546.37 ex VAT - 107 units, 202.80 m2"),
    ("  windows", "GBP 121,712.33 - 31 Sheerline Prestige type lines, 98 units"),
    ("  doors / screens / CW", "GBP 30,919.00 - 7 SMA lines, 9 units"),
    ("  installation", "GBP 21,915.05 - single global line"),
    ("  optional", "external mastic GBP 2,808.10 / EPDM GBP 5,028.61"),
    ("", ""),
    ("Supplier cost", "GBP 91,409.18 = BSW QT252799 GBP 61,056.80 + Bellview 0000000483 GBP 30,352.38"),
    ("Gross margin", "GBP 83,137.19 before install cost (47.6% of sell)"),
    ("NOT the backing", "Aplus QP70172 is dated 22/07 - AFTER we submitted - is Technal not Sheerline/SMA, "
                        "and is quoted UNGLAZED. It is an unused alternative."),
    ("", ""),
    ("ARITHMETIC", "CLEAN. All 31 window types tie to BSW exactly on quantity AND line total. All 7 SMA "
                   "lines tie to Bellview at the 15% discounted figure to the penny. Unit rates follow the "
                   "house template on every code. Install reconciles to the penny as the sum of the labour "
                   "codes. Every unit sold has a supplier quote behind it."),
    ("Filwood labour-code trap", "DID NOT BITE on the biggest line - Type AK is correctly coded CW and carries "
                                 "GBP 3,055.05 of curtain-wall labour at GBP 150/m2."),
    ("", ""),
    ("CONTRACT", "JCT MW 2016. Start 14/09/2026, completion 11/12/2026. Delay damages GBP 500/calendar day. "
                 "Retention 3% / 1.5%. Rectification 12 months."),
    ("mary_checks.py", "6 FAILED + 1 unanswered - every one of them an item below, not a new problem."),
]
for k, v in ROWS:
    if not k and not v:
        r += 1
        continue
    c = ws.cell(r, 1, k); c.font = BOLD; c.alignment = TOP; c.border = THIN
    c2 = ws.cell(r, 2, v); c2.font = BASE; c2.alignment = WRAP; c2.border = THIN
    if k in ("TENDER SUBMITTED", "ARITHMETIC", "CONTRACT"):
        c.fill = c2.fill = GRN
    r += 1

# ======================================================= 2. Commercial exposure
ws = wb.create_sheet("Commercial exposure")
widths(ws, 30, 13, 14, 62)
r = title(ws, "WHAT IS PROMISED OR REQUIRED BUT NOT IN THE GBP 174,546.37",
          "Nothing here is a benchmark price. Where no rate exists that is stated, because inventing one "
          "would be worse than saying so.")
r = header(ws, r, ["Item", "Quantity", "In the price?", "Position"])

EXPOSURE = [
    ("Strip-out and disposal of the existing windows",
     "107 openings / 202.80 m2", "NO",
     "Adam ruled 27/07 'we would include it for a job of this size'. MTCBC's SOW item 1.09 measures it in "
     "m2 and cross-refers it INTO our item 6.01, so their own document reads as though it is ours. The "
     "install line cannot absorb it - GBP 21,915.05 reconciles exactly to per-unit fit labour. NO strip-out, "
     "removal, disposal or waste category exists in data/supplier-rates.json (0 of 80 checked), so this needs "
     "a real price, not a benchmark.", RED),
    ("Manifestation to glazed doors and screens",
     "24.10 linear m (39.90 m if Types F and H count)", "NO",
     "Schedule 2376-09 clause 2.24 - two bands at 850-1000mm and 1400-1600mm, contrasting, both faces. Adam "
     "ruled 27/07 to allow it AND state it in the inclusions. Extent now measured: element width x 2 across "
     "the 9 glazed door and screen units. Whether the two silled 3,620mm screens count is the only judgement "
     "left. No manifestation rate exists in the register either.", RED),
    ("Delivery from BSW",
     "GBP 61,056.80 order", "NO",
     "'All estimates are ex works, additional delivery charges may apply' - no rate, no threshold, no distance "
     "rule. Delivery Address on the quote is 98 Alston Drive, Bradwell Abbey MK13 9HF - FENSTER'S OWN YARD.", RED),
    ("Delivery from Bellview",
     "GBP 30,352.38 order", "NO",
     "The quote is SILENT on delivery entirely - zero occurrences of deliver, ex works, carriage or free.", RED),
    ("Carriage from our yard to site",
     "107 units / 202.80 m2, ~150 miles", "NO",
     "Milton Keynes to Merthyr Tydfil. There is no carriage line anywhere in the pricing workbook.", RED),
    ("Access / lifting equipment",
     "55.97 m2 of glazing 3.62 m or taller; elements to 5,580mm", "EXCLUDED",
     "Correctly excluded in our proposal, and Adam has confirmed the proposal should say we allowed none. BUT "
     "the ruling settles what our document SAYS, not who PAYS: Prelims F and B require the Contractor to "
     "provide all scaffolding 'for himself and any Sub-Contractor'. An unqualified exclusion is a negotiating "
     "position, not an agreement.", AMB),
    ("Fobbed reader preparation",
     "D.01 and D.14", "NO",
     "Door schedule 2376-08 states 'fobbed reader' on both. We exclude access control and our own clarification "
     "says compatibility 'requires further review' - never done. Bellview's 7 units carry no electric strike, "
     "no rectifier and no transfer hinge. MTCBC carry the system separately as SOW 15.10, but the leaf and "
     "frame preparation is a separate question.", AMB),
    ("Anti-ligature ironmongery",
     "external doors", "NO",
     "2376-08 specifies 'Anti-Ligature Infilled Door Pull Handle on Plate, 300 x 75mm stainless', hinges to "
     "BS 7352 and 200mm kicking plates. Bellview list only concealed panic bars and closers. On a special "
     "needs school this is safeguarding, not finish.", RED),
]
for item, qty, inprice, pos, fill in EXPOSURE:
    r = put(ws, r, [item, qty, inprice, pos], fill=fill, wrap_from=1, money=())

r += 1
c = ws.cell(r, 1, "PRICE-HOLD EXPOSURE (separate from the above - this is risk, not omission)")
c.font = BOLD; c.fill = AMB
r += 1
r = header(ws, r, ["", "Date", "Value at risk", "Position"])
for a, b, cval, dd in [
    ("Our price must hold until", "11/12/2026", "", "JCT MW completion. There is NO tender validity clause in "
     "this pack - prelims and SOW both checked. We hold ITT sections 2, 3 and 4 only, because the ITT is the "
     "main contract tender between MTCBC and ET&S."),
    ("BSW QT252799 lapses", "14/08/2026", BSW_NET, "119 days before our price closes, and a month before the "
     "job starts on 14/09."),
    ("Bellview 0000000483 lapses", "15/08/2026", BELLVIEW_NET, "118 days early."),
    ("TOTAL UNFIXED", "", BSW_NET + BELLVIEW_NET, "52.4% of the sold price. Each 1% of supplier inflation is "
     "GBP 914.09 off the bottom line; 5% is GBP 4,570.46."),
]:
    r = put(ws, r, [a, b, cval, dd], fill=(AMB if a == "TOTAL UNFIXED" else None), wrap_from=4, money=(3,))

# ============================================================= 3. Findings
ws = wb.create_sheet("Findings")
widths(ws, 5, 10, 46, 34, 58)
r = title(ws, "FINDINGS", "Severity is about consequence if unresolved, not about whether the price is wrong.")
r = header(ws, r, ["#", "Severity", "Finding", "Source", "What it costs / what to do"])

FINDINGS = [
    (1, "HIGH", "THE DOOR SYSTEM CANNOT MEET THE SPECIFIED U-VALUE, UNDER ANY READING",
     "SMA 'Smart Wall Profile.pdf'; window schedule 2376-09; door schedule 2376-08; EDG02; our own proposal",
     "SMA publish 1.8 W/m2K for Smart Wall doors. Our proposal promises 1.4; the window schedule says 1.4; "
     "door schedule 2376-08 says 1.4 per door on D.01/D.17/D.22/D.26; EDG02 says 1.2. At 1.8 the 7 Smart Wall "
     "Pocket units - GBP 31,360.15 of sell - miss all of them. The 1.2 is expressly an AREA WEIGHTED AVERAGE, "
     "so it was computed rather than asserted: MC600 would have to achieve 0.55 W/m2K for the package to "
     "average 1.2. It is a system change or a formal qualification, not a glass swap. REQ-15/REQ-16.", RED),
    (2, "HIGH", "SEVEN DOOR UNITS SIT ON A SYSTEM OUR OWN RECORDS ALREADY FAILED",
     "SM5 Wexham record; Bellview 0000000483 positions 001-006",
     "SM5 Wexham found Smart Wall Pocket 'cannot meet the drawing's whole-installation U-value 1.6'. Fifth "
     "instance of this shape in a month. NOTE: position 007 is MC600 Plus - thermally broken, named in the "
     "SM5 Wexham record as part of the FIX. Do not lump the Type AK curtain walling in with the doors.", RED),
    (3, "HIGH", "TYPE G PUTS A SHEERLINE CASEMENT INSIDE A SMART WALL FRAME",
     "Schedule 2376-09 Type G/W.24; Bellview pos 001; BSW QT252799 'TYPE G INSERT'",
     "The schedule needs '1 top hung + 1 fixed + 1 external door'. Bellview quoted a door and TWO FIXED "
     "FIELDS with '1 x prepared for a thickness of 28mm'. BSW fill it with an 854x900 Sheerline casement at "
     "GBP 697.58. Sheerline is 70mm, Smart Wall 100mm, and BSW ruled on SM5 Wexham that the two cannot be "
     "coupled. GBP 8,499.66 of sell. Get the fabrication confirmed in writing before order. REQ-16.", RED),
    (4, "HIGH", "STRIP-OUT AND MANIFESTATION ARE PROMISED BUT NOT PRICED",
     "MTCBC SOW 1.09 into item 6.01; schedule 2376-09 cl.2.24; Adam's ruling 27/07",
     "See the Commercial exposure sheet. Neither can be benchmarked - no category exists in the register for "
     "either. REQ-22.", RED),
    (5, "HIGH", "DELIVERY AND CARRIAGE ARE IN NOBODY'S PRICE",
     "BSW QT252799 terms; Bellview 0000000483; the pricing workbook",
     "BSW is ex works to Fenster's own MK yard, 150 miles from site. Bellview is silent. No carriage line in "
     "the workbook. Two unpriced transport legs inside a fixed lump sum.", RED),
    (6, "MEDIUM", "OUR PRICE IS COMMITTED 119 DAYS LONGER THAN OUR COSTS ARE HELD",
     "BSW and Bellview 30-day validity; JCT MW completion 11/12/2026",
     "GBP 91,409.18 unfixed - 52.4% of the sold price. Both quotes die a month before the job starts. Either "
     "get the prices held in writing or accept the risk knowingly.", AMB),
    (7, "MEDIUM", "THE DOOR SCHEDULE REQUIRES HARDWARE THAT IS IN NEITHER QUOTE",
     "Door schedule 2376-08",
     "Fobbed readers on D.01/D.14; anti-ligature pull handles, BS 7352 hinges and 200mm kicking plates on the "
     "external doors; and 'No locking mechanism or latch' / 'Non-lockable device' on D.01/D.17/D.22/D.26 "
     "against the concealed panic bars we priced on all 7 units. Aplus flagged that same ambiguity in writing "
     "('It is unclear what a Non-Lockable Device is'); Bellview defaulted silently. REQ-19.", AMB),
    (8, "MEDIUM", "THE TWO ARCHITECT'S SCHEDULES DISAGREE ON EXTERNAL DOOR SIZES",
     "2376-08 vs 2376-09 rev A",
     "D.17 is 955x2570 where Type L is 955x2410; D.22 is 1530x2270 with no priced equivalent; D.26 is 930x2100 "
     "where Type U is 929x2370. Only D.01 maps cleanly (1530x2410 = Types I/O) - and D.01 is the '4 panes "
     "bi-folding' door our proposal substituted for commercial French doors. Reconcile before manufacture.", AMB),
    (9, "MEDIUM", "THE BIFOLD / FRENCH DOOR SUBSTITUTION HAS NEVER BEEN ACCEPTED IN WRITING",
     "Proposal p3 clarifications; door schedule D.01",
     "'Communal escape doors have been priced as panic bar doors. Bifold door locations have been priced as "
     "commercial French doors, as advised by the supplier.' That is a substitution the client has not agreed.", AMB),
    (10, "LOW", "NO SOLAR CONTROL ANYWHERE, AGAINST AN EDG02 g-VALUE OF 0.4-0.3",
     "BSW QT252799; Bellview 0000000483; EDG02",
     "Zero hits for solar, g-value, Suncool, SKN, Coolite or Planitherm across both quotes. 'Clr' means clear. "
     "For scale only, the register's matched solar-control categories imply GBP 8,796-16,489 of supply cost - "
     "benchmark, and it buys the g-value alone, not a compliant door.", AMB),
    (11, "LOW", "NEITHER SUPPLIER STATES AN INTERNAL FINISH, AND NO RAL WAS EVER FIXED",
     "BSW QT252799 'Ext Colour'; Bellview 'Profiles: Anthracite Grey'; schedule 2376-09",
     "The architect says only 'grey powder coated aluminium'. 7016 Anthracite is the supplier's choice, not "
     "cfw's. Confirm the RAL and whether the internal face matches.", AMB),
    (12, "LOW", "REV A CONTRADICTS ITSELF ON THE INTEGRAL BLIND",
     "2376-09 rev A note 6, Type AK W.92/W.93",
     "The revision note says the integral blind is omitted but one blind note survived - on our single most "
     "expensive line. Costs nothing today because we exclude blinds outright, but get it corrected in writing "
     "rather than argue it at manufacture.", GRN),
    (13, "LOW", "WE PRICED A SUPERSEDED DRAWING WITHOUT KNOWING",
     "2376-09 rev A dated 13.07.26; site plan rev E 08.07.26; issued 24/07",
     "Both were revised BEFORE our 17/07 quote and only issued on 24/07. Harmless this time - the change was "
     "an omission we had already excluded. Press ET&S to issue revisions when they make them.", GRN),
    (14, "LOW", "THE SITE POSTCODE ON OUR OWN DOCUMENTS IS WRONG",
     "Pricing document cell B5 and the proposal vs the MTCBC ITT",
     "Ours say CF77 8HA; the ITT, prelims and SOW all say CF47 8HA. CF77 is not a Merthyr postcode.", GRN),
]
for n, sev, f, src, cost, fill in FINDINGS:
    r = put(ws, r, [n, sev, f, src, cost], fill=fill, wrap_from=3)

# ================================================================= 4. RFIs
ws = wb.create_sheet("RFIs")
widths(ws, 5, 22, 100)
r = title(ws, "RFIs - QUESTIONS TO PUT, GROUPED BY WHO CAN ANSWER THEM",
          "Mary cannot email ET&S, the architect or the suppliers. These need Gintare or Adam to send.")
r = header(ws, r, ["#", "To", "Question"])

RFIS = [
    (1, "ET&S / cfw architects", "Which document governs the U-value - EDG02's 1.3 windows / 1.2 external doors "
        "/ g-value 0.4-0.3, or window schedule 2376-09's 1.4? Note this decides the WINDOWS; the doors miss "
        "under either reading."),
    (2, "BSW / Bellview / SMA", "Please issue an SMA U-value calculation for the 7 Smart Wall Pocket units as "
        "quoted. SMA's published figure is 1.8 W/m2K for Smart Wall doors - does that figure apply to Smart "
        "Wall POCKET, and what is the figure for the MC600 Plus curtain walling?"),
    (3, "BSW / Bellview", "How is Type G actually built? Bellview's element leaves an aperture 'prepared for a "
        "thickness of 28mm' and BSW supply an 854x900 Sheerline casement to fill it. Sheerline is 70mm and "
        "Smart Wall 100mm. Please confirm the detail in writing before order."),
    (4, "ET&S", "Does strip-out and disposal of the existing windows sit with us? MTCBC's SOW item 1.09 reads "
        "'Remove doors and windows; load into skip... (allowed in 6.01)' and 6.01 is our supply-and-fit item."),
    (5, "ET&S", "Who provides access? Our proposal excludes scaffold, MEWPs and towers; Prelims F and B require "
        "the Contractor to provide all scaffolding 'for himself and any Sub-Contractor'. We install to 5,580mm "
        "and 55.97 m2 of the glazing is 3.62 m or taller."),
    (6, "cfw architects", "What extent of manifestation is required to clause 2.24, and does it apply to the "
        "two 3,620mm silled screens (Types F and H) as well as the glazed doors and screens? We have measured "
        "24.10 linear m on the doors/screens, 39.90 m if F and H are included."),
    (7, "cfw architects", "What is a 'non-lockable device' on an escape door? D.01, D.17, D.22 and D.26 say 'No "
        "locking mechanism or latch' / 'Non-lockable device', and both suppliers have defaulted to concealed "
        "panic bars, which latch."),
    (8, "cfw architects", "Please confirm the anti-ligature ironmongery required on the external aluminium doors "
        "- pull handle on plate 300x75mm, hinges to BS 7352, 200mm kicking plates - none of which is in the "
        "supplier quotations."),
    (9, "ET&S", "Is fobbed reader preparation (cabling route, transfer hinge, electric strike or keep) in our "
        "package or MTCBC's under SOW item 15.10? 2376-08 requires fobbed readers on D.01 and D.14."),
    (10, "cfw architects", "Door schedule 2376-08 and window schedule 2376-09 rev A give different structural "
         "openings for the external doors (D.17, D.22, D.26). Which governs?"),
    (11, "ET&S", "Rev A of 2376-09 states the integral blind is omitted, but note 6 survives on Type AK "
         "(W.92/W.93). Please confirm in writing that the blind is omitted there too."),
    (12, "ET&S", "Please confirm acceptance of the substitutions stated in our proposal: communal escape doors "
         "priced as panic bar doors, and bifold locations priced as commercial French doors."),
    (13, "BSW / Bellview", "Please quote carriage to site at Caedraw Rd, Merthyr Tydfil CF47 8HA - both quotes "
         "are ex works or silent, and BSW's delivery address is our own Milton Keynes yard."),
    (14, "BSW / Bellview", "Will you hold these prices to the contract completion date of 11/12/2026? Both "
         "quotes run 30 days and lapse mid-August, a month before the job starts."),
]
for n, who, q in RFIS:
    r = put(ws, r, [n, who, q], wrap_from=3)

# =========================================================== 5. Reconciliation
ws = wb.create_sheet("Reconciliation")
widths(ws, 46, 18, 18, 46)
r = title(ws, "WHAT WAS CHECKED AND FOUND CORRECT",
          "Recorded because it matters as much as the findings - the price is sound and should not be reopened.")
r = header(ws, r, ["Check", "Result", "Scope", "Evidence"])

for a, b, c_, d_ in [
    ("Window lines vs BSW QT252799", "EXACT", "31 of 31 types",
     "Quantity AND line total match on every type, zero variance, 98 units."),
    ("SMA lines vs Bellview 0000000483", "EXACT", "7 of 7 lines",
     "All carried at the 15% end-discounted figure to the penny (e.g. Type AK 5,199.89 x 0.85 = 4,419.9065)."),
    ("Unit rate formula", "EXACT", "every code",
     "supply + (code value x 75%), verified on MAW, ELAW, LAW, SAW, SAD, DAD, SADMAW and CW."),
    ("Installation line", "EXACT", "GBP 21,915.05",
     "Reconciles to the penny as the sum of the house labour codes across all 39 lines."),
    ("Supplier cost total", "EXACT", "GBP 91,409.18",
     "BSW 61,056.80 + Bellview 30,352.38. Includes the GBP 697.58 Type G insert."),
    ("Supplier quantity coverage", "PASS", "39 lines",
     "Every unit sold has a supplier quote behind it (Brocks Hill rule)."),
    ("Filwood labour-code trap", "DID NOT BITE", "Type AK",
     "Correctly coded CW - 10.1835 m2 x 2 x GBP 150/m2 = GBP 3,055.05 of curtain-wall labour."),
    ("Obscure glazing to WCs", "PRICED", "9 panes",
     "BSW carry '6.8 Lam/18/4mm ObsTuff EcoPlus 1.0 Stippolyte 4mm'. Written 'ObsTuff', so a search for "
     "'obscure' returns nothing - do not re-raise it."),
    ("Trickle vents / restrictors / hinge protectors", "PRICED", "62 / 58 / 32 refs", "All in BSW QT252799."),
    ("Panic hardware on fire-exit doors", "PRICED", "all 6 door types", "ACIM453 concealed panic bars."),
    ("REQ-5 - the 24/07 addendum", "NO CHANGE", "2376-09 vs rev A",
     "209 window refs, 38 types, 28 s/o sizes, 38 opening patterns, 24 restrictor, 6 obscure, 33 U-value and "
     "38 SBD notes all identical. Only the blind note fell from 29 to 1, and blinds were already excluded."),
    ("'Making good'", "NOT OURS", "SOW section 8",
     "Only strip-out cross-refers into item 6.01. Making good and decoration are MTCBC's own section 8, so our "
     "Internal Finishing exclusion is safe. Do not widen the strip-out finding."),
    ("Tender validity clause", "NONE EXISTS", "prelims + SOW",
     "Zero hits for validity, remaining open or a price hold. The 90-day Section 20 trap seen elsewhere does "
     "not apply here."),
]:
    r = put(ws, r, [a, b, c_, d_], fill=GRN, wrap_from=4)

for s in wb:
    s.freeze_panes = "A5"
wb.save(OUT)
print("wrote", OUT)
print("sheets:", wb.sheetnames)
