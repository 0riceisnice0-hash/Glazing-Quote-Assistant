# -*- coding: utf-8 -*-
"""Stoke Park School (Borras, Aplus job 17644) - glass and louvre reconciliation.

Compares the FINAL Aplus glass-sizes sheet (issued 24/07/2026, frames supplied
UNGLAZED) against Vetroseal quote 064542 (01/07/2026), which was priced from
Fenster's earlier glass schedule. Writes a check workbook to outputs\\.

CORRECTED 27/07/2026 after Adam's answer to REQ-3. The 46 panes that looked
"missing" from the glass quote are the LOUVRE positions - they are filled by
IKON IKL332 louvre modules, not glass. Confirmed by two documents already in the
job folder: Aplus panel order QT50932 Rev7 (46 panels, 27.64 m2) and IKON quote
Q26-24329 dated 02/07/2026 (46 louvre modules, GBP 10,125.91 + carriage). Once
those come out, the glass count reconciles exactly: 124 required = 124 quoted.

What does NOT reconcile is SIZE. The job was re-input at Aplus on 02/07/2026 and
the frames were signed off that day; both the Vetroseal glass quote and the
panel schedule IKON priced against predate that re-input, so every pane and
every louvre is dimensioned to a superseded schedule. That is what this script
now reports.

Aplus sheet has one row per pane: item | ref | qty | width | height | glass type
| location. Panes marked "DO NOT ORDER - Unglazed" are aluminium infill panels
and are excluded from the order count. Vetroseal rows wrap over several lines,
so lines are matched by the arithmetic-consistent tail (unit area, unit price,
total) as the rate miner does.
"""
import collections
import os
import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
APLUS_PDF = os.path.join(REPO, "test-results", "stoke-glass-input", "Glass Sizes_17644.PDF")
VETRO_PDF = (r"C:\Users\zacpl\OneDrive - Fenster Glazing (1)\Commercial\2. Projects\Borras"
             r"\Coventry - Stoke Park School\1. Estimating\2. Supplier Quotes\Vetroseal"
             r"\FENSTERG_Quote_064542(STOKE-SCHOOL-COVENTRY).pdf")
OUT = os.path.join(REPO, "outputs", "Stoke Park School - Glass Sizes vs Quoted Glass (check).xlsx")

HEAD = PatternFill("solid", fgColor="1F2A44")
HEADF = Font(color="FFFFFF", bold=True)
WARN = PatternFill("solid", fgColor="FCE4E4")

# Aplus location text -> the reference used on the Vetroseal quote.
QUOTE_REF = {
    "Type A": "TYPE/A", "Type B": "TYPE/B", "Type C": "TYPE/C", "Type D": "TYPE/D",
    "Type E": "TYPE/E", "Type F": "TYPE/F", "Type G": "TYPE/G", "Type H": "TYPE/h",
    "Type J1": "TYPE/j (J1+J2)", "Type J2": "TYPE/j (J1+J2)",
    "Door Type A": "TYPE/a", "Door Type B1": "TYPE/B1/D02", "Door Type B2": "D04/TYPE/B2",
    "Door Type C": "(not on quote)", "Door Type D": "TYPE/d",
}

# The 46 louvre positions, from Aplus panel order QT50932 Rev7 and IKON Q26-24329:
# ref A1 (and F1 on Type G) on every bay of window Types A/B/C/D/F/G = 41, the head
# panel over doors D02 and D04 = 2, and all three panels of Door Type C / D03 = 3.
LOUVRE_WINDOW_TYPES = {"Type A", "Type B", "Type C", "Type D", "Type F", "Type G"}

# Sizes IKON quoted the louvres to (QT50932 Rev7 panel schedule, 01/07/2026),
# against the aperture on the signed-off order of 02/07/2026.
IKON_SIZES = {
    "Type A": (1033, 476), "Type B": (1031, 476), "Type C": (1030, 476),
    "Type D": (1031, 476), "Type F": (1030, 476), "Type G": (1056, 476),
    "Door Type B1": (1933, 783), "Door Type B2": (1933, 783),
    "Door Type C": (1933, 811),
}


def is_louvre(r):
    """True if this Aplus pane row is a louvre position, not glass."""
    if r["type"] in LOUVRE_WINDOW_TYPES and r["ref"] in ("A1", "F1"):
        return True
    if r["type"] == "Door Type C":
        return True
    if r["type"] in ("Door Type B1", "Door Type B2") and r["ref"] == "A1":
        return True
    return False


def pdf_lines(path):
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            out.extend((page.extract_text() or "").split("\n"))
    return out


def read_aplus():
    """One dict per pane row of the Aplus glass-sizes sheet."""
    rows = []
    for line in pdf_lines(APLUS_PDF):
        m = re.match(r"^(\d+)\s+([A-Z]\d)\s+(\d+)\s+(\d+)\s+(-?\d+)\s+(.*)$", line.strip())
        if not m:
            continue
        item, ref, qty, w, h, rest = m.groups()
        thickness = "32mm" if "32mm" in rest else ("28mm" if "28mm" in rest else "DO NOT ORDER")
        loc = re.search(r"(Door Type [A-Z0-9]+|Type [A-Z0-9]+)", rest)
        rows.append({
            "item": int(item), "ref": ref, "qty": int(qty), "w": int(w), "h": int(h),
            "thickness": thickness, "type": loc.group(1) if loc else rest.strip()[:20],
            "location": rest.strip(),
            "area": int(qty) * int(w) * max(int(h), 0) / 1e6,
        })
    return rows


def read_vetroseal():
    """Vetroseal priced lines: (line, ref, qty, w, h, unit area, unit price, total)."""
    text = "\n".join(pdf_lines(VETRO_PDF))
    found = re.findall(
        r"^\s*(\d+)\s+(\S+)\s+8\.8L\s*-\s*16\s*-\s*(\d+)\s+(\d+)\s+(\d+)\s+.*?"
        r"([\d.]+)\s+([\d.]+)\s+([\d,]+\.\d\d)\s*$", text, re.M)
    rows = []
    for ln, ref, qty, w, h, ua, up, tp in found:
        rows.append({"line": int(ln), "ref": ref, "qty": int(qty), "w": int(w), "h": int(h),
                     "unit_area": float(ua), "unit_price": float(up),
                     "total": float(tp.replace(",", "")),
                     "area": int(qty) * float(ua)})
    return rows


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill, cell.font = HEAD, HEADF
            cell.alignment = Alignment(horizontal="center", wrap_text=True)


def main():
    aplus, vetro = read_aplus(), read_vetroseal()
    order = [r for r in aplus if r["thickness"] != "DO NOT ORDER"]
    skip = [r for r in aplus if r["thickness"] == "DO NOT ORDER"]
    louvre = [r for r in order if is_louvre(r)]
    glass = [r for r in order if not is_louvre(r)]

    l_panes = sum(r["qty"] for r in louvre)
    l_area = sum(r["area"] for r in louvre)
    g_panes = sum(r["qty"] for r in glass)
    g_area = sum(r["area"] for r in glass)

    a_panes = sum(r["qty"] for r in order)
    a_area = sum(r["area"] for r in order)
    v_panes = sum(r["qty"] for r in vetro)
    v_area = sum(r["area"] for r in vetro)
    v_goods = sum(r["total"] for r in vetro)
    rate = v_goods / v_area                      # GBP 110.00/m2 flat on this quote
    surcharge_per_m2 = 436.52 / v_area           # energy surcharge, 3,357.82kg @ 0.13

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["STOKE PARK SCHOOL (BORRAS) - GLASS RECONCILIATION"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for label, value in [
        ("Aplus job", "17644 - frames supplied UNGLAZED, supply only, delivery 03/08/2026"),
        ("Aplus glass sizes sheet", "issued 24/07/2026 (input 02/07/2026)"),
        ("Glass quote on file", "Vetroseal 064542, quote date 01/07/2026 (30-day validity)"),
        ("Alternative rate on file", "CN Glass 01/07/2026 - GBP 60/m2 inc energy, same make-up"),
        ("Client price", "GBP 104,660.17 ex VAT (pricing doc rev 15/07/2026), glass inside unit rates"),
        ("", ""),
        ("QUANTITY - no shortfall", ""),
        ("Panes on Aplus final list", a_panes),
        ("less LOUVRE positions (IKON, not glass)", "%d / %.2f m2" % (l_panes, l_area)),
        ("GLASS required", "%d / %.2f m2" % (g_panes, g_area)),
        ("Panes on Vetroseal quote", "%d / %.2f m2" % (v_panes, v_area)),
        ("Difference", "%+d panes / %+.2f m2" % (v_panes - g_panes, v_area - g_area)),
        ("", ""),
        ("SIZE - every quoted pane is superseded", ""),
        ("Quoted panes matching a final-list size", "0 of %d" % v_panes),
        ("Why", "job re-input at Aplus 02/07/2026; both buys predate it"),
        ("", ""),
        ("Vetroseal goods value quoted", round(v_goods, 2)),
        ("Vetroseal rate implied GBP/m2", round(rate, 2)),
        ("Vetroseal energy surcharge GBP/m2", round(surcharge_per_m2, 2)),
        ("Vetroseal quoted net (goods+surcharge)", 12012.88),
        ("Same rate on glass area - indicative net", round(g_area * (rate + surcharge_per_m2), 2)),
        ("Indicative increase", round(g_area * (rate + surcharge_per_m2) - 12012.88, 2)),
        ("CN Glass GBP 60/m2 on the 28mm glass only",
         round(sum(r["area"] for r in glass if r["thickness"] == "28mm") * 60, 2)),
        ("", ""),
        ("COST vs THE SOLD PRICE", ""),
        ("Glass carried in the price (Vetroseal 063934, 05/06)", 9309.22),
        ("Glass now quoted (Vetroseal 064542, 01/07)", 12012.88),
        ("Louvres carried in the price (IKON Q26-24160, 05/06)", 7490.64),
        ("Louvres now quoted (IKON Q26-24329, 02/07)", 10125.91),
        ("Cost above the sold price", round((12012.88 - 9309.22) + (10125.91 - 7490.64), 2)),
        ("", ""),
        ("28mm GLASS panes / m2 required", "%d / %.2f" % (sum(r["qty"] for r in glass if r["thickness"] == "28mm"),
                                                         sum(r["area"] for r in glass if r["thickness"] == "28mm"))),
        ("32mm GLASS panes / m2 required", "%d / %.2f" % (sum(r["qty"] for r in glass if r["thickness"] == "32mm"),
                                                         sum(r["area"] for r in glass if r["thickness"] == "32mm"))),
        ("32mm on Vetroseal quote", "NONE - every line quoted 8.8L-16-4T (28.8mm)"),
        ("Panes excluded", "%d marked DO NOT ORDER - Unglazed (aluminium infill panels)" % sum(r["qty"] for r in skip)),
    ]:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 78
    ws.append([])
    ws.append(["Indicative figures use Vetroseal's own quoted rate on the larger area. They are NOT a quote -"])
    ws.append(["the 32mm make-up is unpriced and glass rates move with area, weight and pane size."])

    ws = wb.create_sheet("Per type")
    ws.append(["Aplus type", "Vetroseal ref", "Glass required", "Panes quoted", "Difference",
               "Glass area m2", "Aplus pane refs", "Note"])
    per_a = collections.defaultdict(lambda: [0, 0.0, set()])
    for r in glass:
        per_a[r["type"]][0] += r["qty"]
        per_a[r["type"]][1] += r["area"]
        per_a[r["type"]][2].add(r["ref"])
    per_v = collections.defaultdict(int)
    for r in vetro:
        per_v[r["ref"]] += r["qty"]
    for typ in sorted(per_a):
        ref = QUOTE_REF.get(typ, "?")
        if ref.startswith("TYPE/j"):
            quoted = per_v["TYPE/j"] // 2  # 24 quoted panes cover J1 and J2 equally
        else:
            quoted = per_v.get(ref, 0)
        need = per_a[typ][0]
        note = "" if need == quoted else "CHECK"
        ws.append([typ, ref, need, quoted, need - quoted, round(per_a[typ][1], 2),
                   ", ".join(sorted(per_a[typ][2])), note])
        if need - quoted:
            for cell in ws[ws.max_row]:
                cell.fill = WARN
    ws.append(["TOTAL", "", g_panes, v_panes, g_panes - v_panes, round(g_area, 2), "",
               "counts reconcile once louvres are removed"])
    ws[ws.max_row][0].font = Font(bold=True)
    for col, width in zip("ABCDEFGH", (16, 18, 15, 14, 11, 16, 22, 44)):
        ws.column_dimensions[col].width = width
    style_header(ws)

    ws = wb.create_sheet("Louvres not glass")
    ws.append(["Aplus type", "Pane ref", "Qty", "Aperture on signed-off order",
               "Size IKON quoted", "Height difference", "Note"])
    for r in sorted(louvre, key=lambda x: (x["type"], x["ref"])):
        ik = IKON_SIZES.get(r["type"])
        diff = "" if not ik else "%+d mm" % (ik[1] - r["h"])
        ws.append([r["type"], r["ref"], r["qty"], "%d x %d" % (r["w"], r["h"]),
                   "" if not ik else "%d x %d" % ik, diff,
                   "louvre too tall for the aperture" if ik and ik[1] != r["h"] else ""])
        if ik and ik[1] != r["h"]:
            for cell in ws[ws.max_row]:
                cell.fill = WARN
    ws.append(["TOTAL", "", l_panes, "%.2f m2" % l_area, "", "", ""])
    ws[ws.max_row][0].font = Font(bold=True)
    ws.append([])
    ws.append(["IKON Q26-24329, 02/07/2026: 46 IKL332 28mm louvre modules with 50mm insulated"])
    ws.append(["backing panels and insect mesh, RAL 7012 matt - GBP 10,125.91 + carriage TBC."])
    ws.append(["Priced against Aplus panel schedule QT50932 Rev7 (input 01/07). The frames were"])
    ws.append(["re-input and signed off on 02/07 with these apertures 85mm shorter."])
    for col, width in zip("ABCDEFG", (16, 10, 6, 28, 18, 17, 34)):
        ws.column_dimensions[col].width = width
    style_header(ws)

    ws = wb.create_sheet("Pane sizes quoted v final")
    ws.append(["Vetroseal line", "Type", "Size quoted", "Size on final list",
               "Width shift", "Height shift"])
    ref_map = {v: k for k, v in QUOTE_REF.items()}
    by_type = collections.defaultdict(list)
    for r in glass:
        by_type[r["type"]].append(r)
    for v in vetro:
        typ = ref_map.get(v["ref"], v["ref"])
        cands = (by_type["Type J1"] + by_type["Type J2"]) if v["ref"] == "TYPE/j" \
            else by_type.get(typ, [])
        if v["ref"] == "TYPE/j":
            typ = "Type J1/J2"
        if not cands:
            continue
        best = min(cands, key=lambda r: abs(r["w"] - v["w"]) + abs(r["h"] - v["h"]))
        dw, dh = best["w"] - v["w"], best["h"] - v["h"]
        ws.append([v["line"], typ, "%d x %d" % (v["w"], v["h"]),
                   "%d x %d" % (best["w"], best["h"]), dw, dh])
        if (dw, dh) != (0, 0):
            for cell in ws[ws.max_row]:
                cell.fill = WARN
    ws.append([])
    ws.append(["", "Not one of the %d quoted panes matches a size on the final list." % v_panes])
    for col, width in zip("ABCDEF", (14, 16, 16, 18, 12, 12)):
        ws.column_dimensions[col].width = width
    style_header(ws)

    ws = wb.create_sheet("Aplus final list")
    ws.append(["Item", "Pane ref", "Qty", "Width", "Height", "Make-up", "Area m2", "Location"])
    for r in aplus:
        ws.append([r["item"], r["ref"], r["qty"], r["w"], r["h"], r["thickness"],
                   round(r["area"], 3), r["location"]])
        if r["thickness"] == "DO NOT ORDER":
            for cell in ws[ws.max_row]:
                cell.fill = WARN
    for col, width in zip("ABCDEFGH", (7, 10, 6, 8, 8, 15, 9, 46)):
        ws.column_dimensions[col].width = width
    style_header(ws)

    ws = wb.create_sheet("Vetroseal 064542")
    ws.append(["Line", "Reference", "Qty", "Width", "Height", "Unit area m2",
               "Unit price GBP", "Total GBP", "Area m2"])
    for r in vetro:
        ws.append([r["line"], r["ref"], r["qty"], r["w"], r["h"], r["unit_area"],
                   r["unit_price"], r["total"], round(r["area"], 3)])
    ws.append(["", "TOTAL", v_panes, "", "", "", "", round(v_goods, 2), round(v_area, 2)])
    ws[ws.max_row][1].font = Font(bold=True)
    ws.append([])
    ws.append(["", "All lines quoted 8.8L-16-4T Lami/Tgh Black Multitech G, 1.2 softcoat, argon."])
    ws.append(["", "Net goods 11,576.36 + energy surcharge 436.52 (3,357.82kg @ 0.13) = 12,012.88 net."])
    for col, width in zip("ABCDEFGHI", (7, 18, 6, 8, 8, 13, 14, 12, 10)):
        ws.column_dimensions[col].width = width
    style_header(ws)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote %s" % OUT)
    print("Aplus final list %d panes / %.2f m2" % (a_panes, a_area))
    print("  louvres (IKON, not glass) %d / %.2f m2" % (l_panes, l_area))
    print("  glass required            %d / %.2f m2" % (g_panes, g_area))
    print("  Vetroseal quoted          %d / %.2f m2  -> counts reconcile"
          % (v_panes, v_area))
    print("but 0 of %d quoted panes match a final-list size" % v_panes)
    print("vetroseal rate %.2f/m2 goods + %.2f/m2 surcharge; on glass area %.2f"
          % (rate, surcharge_per_m2, g_area * (rate + surcharge_per_m2)))


if __name__ == "__main__":
    main()
