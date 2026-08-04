# Princess Beatrice House - take-off + tender audit workbook
# Sources: BPG schedules 5100/5101/5103 T02, client bill C0234, RBKC ERs Nov 2025,
# Fenster pricing doc 23/07/2026, Aplus QP70171 (21-Jul-2026) + Logikal offer (22/07/2026)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GREEN = "1F6B3B"
LIGHT = "E8F2EC"
AMBER = "FFF2CC"
RED = "F8CBCB"

hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=GREEN)
title_font = Font(bold=True, size=14, color=GREEN)
sub_font = Font(bold=True, size=11, color=GREEN)
thin = Side(style="thin", color="BBBBBB")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()

def sheet(name):
    if wb.sheetnames == ["Sheet"]:
        ws = wb.active; ws.title = name
    else:
        ws = wb.create_sheet(name)
    return ws

def put(ws, r, vals, header=False, fill=None):
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.border = box
        cell.alignment = wrap
        if header:
            cell.font = hdr_font; cell.fill = hdr_fill
        elif fill:
            cell.fill = PatternFill("solid", fgColor=fill)
    return r + 1

# ---------------- Sheet 1: Project Info ----------------
ws = sheet("Project Info")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95
ws["A1"] = "PRINCESS BEATRICE HOUSE - TAKE-OFF & TENDER AUDIT"; ws["A1"].font = title_font
ws["A2"] = "Prepared by Mary (Fenster estimating AI) - internal review of the submitted tender"; ws["A2"].font = Font(italic=True, size=9)
rows = [
    ("Client / Main Contractor", "Guildmore Planned Works Ltd (FAO Jason Mount, QS) - ref C0234-012"),
    ("Employer", "Royal Borough of Kensington and Chelsea (RBKC) - supported housing redevelopment"),
    ("Site", "Princess Beatrice House, 188-190 Finborough Road, Earls Court, London SW10 9BA"),
    ("Package", "External windows and doors - supply and install (Trade: External Windows and Doors)"),
    ("Architect / drawings", "BPG Architects - 4539-BPG series, schedules 5100/5101 (windows), 5103 (doors), rev T02 (03/07/2026), status PRELIMINARY / WORK IN PROGRESS"),
    ("Governing documents", "Client bill C0234 (Block 1); RBKC Employers Requirements Nov 2025; Guildmore Standard Subcontract; JCT D&B 2016 with amendments (main contract)"),
    ("Tender return", "17 July 2026 'at the absolute latest' (Fenster quote dated 23/07/2026 - late; confirm Guildmore accepted)"),
    ("Basis of tender", "Lump sum, firm. Supply & Fix tender to be INCLUSIVE of 2.5% main contractor discount. Retention 5%. Defects liability 12 months."),
    ("Fenster quote", "TOTAL GBP 272,771.68 + VAT (items 233,091.68 + installation 39,680.00). Optional: external mastic 5,356.22, EPDM 8,276.91."),
    ("Systems quoted", "Modeal (Aplus) aluminium casement windows + Modeal complex coupled doors (Types 1-5); Technal STII commercial door screens (Types 6-8). RAL 7016M matt."),
    ("Supplier costs", "Aplus QP70171 Crystal sheet 21-Jul-2026, GBP 111,185.75 nett glazed supply-only delivered (windows + coupled doors); Aplus Logikal offer 22/07/2026, GBP 17,499.74 (Technal STII door screens). Both current."),
    ("Glazing", "6.8mm laminated / 18 / 4mm toughened soft-coat DGUs, black warm edge. Doors: SCA LAM / TGH HS (heat-soaked). PAS24 multipoint locks, standard closers."),
    ("Key spec demands", "Bill item A: all windows and doors 'fully in accordance with the Technal UK specification'. ERs: Part Q + Secured by Design Silver, heat-soaked Kitemarked toughened glass, fully-fitted window sample, durability assessment, NHBC/Defects Insurer standards."),
    ("Programme (ITT template)", "Commencement/completion dates in the ITT are stale (Oct 2025 - Mar 2026) - to be confirmed with Guildmore."),
]
r = 4
for k, v in rows:
    ws.cell(r, 1, k).font = Font(bold=True); ws.cell(r, 1).border = box; ws.cell(r, 1).alignment = wrap
    ws.cell(r, 2, v).border = box; ws.cell(r, 2).alignment = wrap
    r += 1

# ---------------- Sheet 2: Take-off ----------------
ws = sheet("Window & Door Take-Off")
widths = [10, 26, 30, 16, 16, 6, 6, 10, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws["A1"] = "TAKE-OFF - BPG SCHEDULES T02 vs FENSTER PRICING 23/07/2026"; ws["A1"].font = sub_font
r = put(ws, 3, ["Type", "Marks (schedule)", "Schedule size WxH (mm)", "Priced size WxH", "Qty sched / priced", "Obs", "PAS24*", "Code", "Notes"], header=True)

takeoff = [
 ("Win 1", "W1F/W2F/WGF 21-35 (36)", "920 x 1350", "920 x 1350", "36 / 36", "Y", "GF:Y", "MAW", "All 36 obscure glazed per schedule - confirm obscure in Aplus price."),
 ("Win 1A", "W1F/W2F 01,06,11,16 (8)", "920 x 1350", "920 x 1350", "8 / 8", "Y", "-", "MAW", "All obscure."),
 ("Win 2", "W1F/W2F/WGF (14)", "1145 x 1135", "1146 x 1135", "14 / 14", "Y", "GF:Y", "MAW", "All obscure."),
 ("Win 3", "W3F 03,07,11,15 (4)", "1145 x 1135", "1146 x 960", "4 / 4", "Y", "-", "MAW", "HEIGHT DEVIATION: priced 960 vs schedule 1135 - re-check at survey / with Aplus."),
 ("Win 4", "W1F/W2F/WGF 02,07,12,17 (12)", "460 x 1390", "460 x 1390", "12 / 12", "N", "GF:Y", "SAW", ""),
 ("Win 5", "W1F/W2F 05,10,15,20 (8 pairs)", "1500 & 1525 x 1400", "1500 & 1525 x 1375", "16 / 16", "1", "-", "MAW", "8 marks, two frames each. Priced 1375 vs 1400 high. 2nr obscure split priced."),
 ("Win 6", "W1F/W2F/WGF 03,08,13,18 (10 pairs)", "720 & 260 x 1160", "720 & 260 x 1375", "20 / 20", "Y*", "GF:Y", "SAW", "HEIGHT DEVIATION: priced 1375 vs schedule 1160. 6 of 10 pairs obscure."),
 ("Win 7", "W3F 02,06,10,14 (4 pairs)", "1825 & 440 x 1185", "1825 & 440 x 1160", "8 / 8", "3/4", "-", "MAW/SAW", "3 of 4 obscure."),
 ("Win 8", "W3F x12", "1825 x 1185", "1812 x 1185", "12 / 12", "N", "-", "MAW", ""),
 ("Win 9", "W3F 04,08,12 (3 pairs)", "1825 & 1365 x 1185", "1825 & 1365 x 1160", "6 / 6", "N", "-", "MAW", ""),
 ("Win 10", "W3F 16 (1 pair)", "1825 & 1830 x 1185", "1825 & 1830 x 1160", "2 / 2", "Y", "-", "MAW", "Obscure per schedule."),
 ("Win 11", "W1F/W2F/W3F/WGF (16 pairs)", "825 x 1380 (x2 per mark)", "825 x 1355", "32 / 32", "part", "GF:Y", "MAW", "4 of 16 marks obscure (W3F 20,24,28)."),
 ("Win 12", "WLF 01,05,09 (3)", "470 x 1145", "470 x 1145", "3 / 3", "N", "-", "SAW", ""),
 ("Win 13", "WLF 03,07,11 (3)", "1150 x 1150", "1150 x 1155", "3 / 3", "N", "-", "MAW", ""),
 ("Win 14", "WLF 14-22 (7)", "910 x 625", "910 x 625", "7 / 7", "N", "-", "SAW", ""),
 ("Win 15", "W3F 01,05,09,13 (4)", "top 2280 x 1068 + btm 920 x 1375", "2280 x 1068 + 880 x 1375", "8 / 8", "N", "-", "LAW/MAW", "Coupled top+bottom."),
 ("Louvre 01", "WLF 13,16,19 (3)", "DIMENSIONS TBC ON SITE", "NOT PRICED", "3 / 0", "-", "n/a", "-", "MISSING: not in Aplus quote or pricing doc. Proposal covers louvres only 'where quoted'."),
 ("Door 1", "DGF 22,23 (2)", "1070 x 2348 + side screen 2280 x 1068", "1069 x 2348 only", "2+2 / 2+0", "N", "No*", "SAD", "MISSING: 2nr 2280x1068 fixed side screens not priced. Door has fixed louvred panel (737+332) - Aplus: louvred/panelled items NOT PAS24 tested."),
 ("Door 2", "DLF 02-12 (6)", "1260 x 2275 + 600 x 2275 light", "1806 x 2275 assembly", "6 / 6", "N", "Y", "DAD", "Side light within priced assembly."),
 ("Door 3", "DGF 13 (1)", "1000 + 530 + 570 x 2275", "2101 x 2160 assembly", "1 / 1", "N", "Y", "DAD", ""),
 ("Door 4", "DGF 18 (1)", "860 + 473 + 815 x 2160 + 460 light", "2148 x 2135 + 460 x 2135", "1 / 1", "N", "Y", "DAD/MAW", ""),
 ("Door 5", "DGF 20 (1)", "1525 x 3000 screen (962+538)", "1500 x 2135 + 1525 x 1375 top", "1 / 1", "Y(SC)", "Y", "DAD/LAW", "Self-closer required. Priced composition differs slightly - confirm at survey."),
 ("Door 6", "DGF 01 (1)", "1550/1595 x 2540 + 1270 x 2800 + 660 x 3000", "as schedule", "1 / 1", "N", "Y", "ELAW/DAD/MAW", "Technal STII screen - matches Aplus Logikal exactly."),
 ("Door 7", "DGF 06,11 (2)", "2700 x 2540 + 1270/1309 x 2800 + 604/623 x 3000", "as schedule", "2 / 2", "N", "Y", "ELAW/DAD/MAW", "Technal STII. Fenster added glass 274.80 + 1,000 additional."),
 ("Door 8", "WLG 15 (1)", "1120 x 2540 + 1270 x 2540 + 665 x 3000", "as schedule", "1 / 1", "N", "Y", "LAW/DAD/MAW", "Technal STII."),
]
for row in takeoff:
    fill = RED if "MISSING" in row[8] else (AMBER if "DEVIATION" in row[8] else None)
    r = put(ws, r, list(row), fill=fill)
r = put(ws, r, ["TOTALS", "", "", "", "191 windows / 191 priced (exact); doors reconcile; 3 louvres + 2 side screens unpriced", "76 obs", "", "", ""], fill=LIGHT)
ws.cell(r, 1, "* PAS24 'GF:Y' = ground-floor / easily accessible marks (WGF, 28nr) must be PAS24 for Part Q. Obscure count = 76 of 191 schedule entries flagged obscure.").font = Font(italic=True, size=9)

# ---------------- Sheet 3: Spec compliance ----------------
ws = sheet("Spec Compliance")
for i, w in enumerate([46, 22, 12, 70], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws["A1"] = "SPEC COMPLIANCE - CLIENT BILL C0234 + RBKC EMPLOYERS REQUIREMENTS"; ws["A1"].font = sub_font
r = put(ws, 3, ["Requirement", "Source", "Status", "Comment / action"], header=True)
spec = [
 ("All windows and doors 'fully in accordance with the Technal UK specification'", "Bill C0234 item A", "DEVIATION", "Windows + door types 1-5 quoted in Modeal (Aplus), Technal STII only on door screens 6-8. ERs deem 'or other approved' to follow proprietary references, and the proposal openly names Modeal - but the ITT requires qualifications to be declared. Action: declare Modeal as an alternative with equivalency evidence (profiles, U-values, PAS24 certs) and obtain CA approval in writing; Employer reserves the right to reject."),
 ("Part Q compliance + Secured by Design Silver; consult SBD officer", "ERs 9.x", "AT RISK", "Aplus note: items glazed with panels or louvres NOT PAS24 tested (affects Door Type 1 louvred door). Proposal says 'PAS24 hardware where quoted' - Part Q needs tested/certified doorsets and accessible windows, not just hardware. Evidence Modeal PAS24 certification for ground-floor (WGF, 28nr) and easily accessible units."),
 ("Toughened glass heat-soaked + BS/Kitemark on all specialist glass", "ERs Glazing", "PARTIAL", "Doors: SCA LAM/TGH HS - compliant. Windows: proposal hedges 'heat-soaked where quoted' - confirm with Aplus that ALL toughened panes are heat-soaked and Kitemarked, and remove the hedge."),
 ("Obscure glazing to 76 schedule entries (bathrooms etc.)", "Schedules 5100/5101", "CONFIRM", "Pricing splits obscure only on Types 5/6/7. Types 1/1A/2/3/10 etc. are wholly obscure per schedule - confirm obscure glass is in the Aplus glazed price for all 76 entries."),
 ("Weather seal external perimeter (Tremco one-part low-modulus) + internal white acrylic seal", "Bill C0234 item A", "CONFLICT", "Bill deems sealing included; Fenster lists external mastic GBP 5,356.22 as OPTIONAL. Roll into the base price or qualify explicitly. Confirm Tremco brand will be used."),
 ("Removal of existing windows/doors to Guildmore skip; recycling certificates; make good brickwork/plaster; clean all glazing on completion", "Bill C0234 prep + item C", "CONFLICT", "Proposal excludes making good, waste removal 'generally', internal finishing and final clean. Bill deems these included in item rates. Reconcile - either allow or list as declared qualifications Guildmore signs off."),
 ("Additional scaffolding to be allowed by subcontractor", "Bill C0234 item H", "EXCLUDED", "Proposal excludes access/lifting equipment - declared qualification; ensure Guildmore acknowledges."),
 ("Comprehensive guarantee (manufacturer + installer); insurance-backed; NHBC/Defects Insurer standards; durability assessment for windows", "Bill item B + ERs", "PARTIAL", "Proposal offers 10-year company warranty. Confirm insurance-backed guarantee availability and allow for durability/life-expectancy statement for windows."),
 ("Fully-fitted window sample + mock-ups; 6 weeks Employer review; benchmarking units", "ERs samples", "NOT ALLOWED", "No sample allowance visible in pricing. Small cost + programme impact - allow one fitted sample window."),
 ("Fabrication/shop drawings for CA approval before manufacture", "Bill + drawings note", "OK", "Proposal includes GA and sectional drawings within cost."),
 ("Site dimensions to be obtained by contractor before fabrication; drawing dims approximate", "Drawings note / bill E", "OK", "Proposal: survey after openings formed; deviations (Type 3 960 vs 1135, Type 6 1375 vs 1160, Type 5 1375 vs 1400, widths -13/-25mm) trued up at survey - watch for variation cost movement."),
 ("uPVC price variance + lead-in times for aluminium AND uPVC", "Bill item AA + collection", "NOT ANSWERED", "No uPVC alternative or lead-in times in the submission. Note: uPVC sits awkwardly with the Technal spec and Part Q - answer or formally decline."),
 ("Tender inclusive of 2.5% main contractor discount", "ITT conditions", "CONFIRM", "No MCD line in the pricing doc. If not built into markup, 2.5% (GBP 6,819) will come off at account."),
 ("Drawings status PRELIMINARY / WORK IN PROGRESS, subject to consultant + fire engineer review", "Drawings T02", "NOTE", "Price against a moving spec - keep the 30-day validity and re-confirm at order."),
]
for row in spec:
    fill = RED if row[2] in ("DEVIATION", "CONFLICT") else (AMBER if row[2] in ("AT RISK", "PARTIAL", "CONFIRM", "NOT ANSWERED", "NOT ALLOWED", "EXCLUDED") else LIGHT)
    r = put(ws, r, list(row), fill=fill)

# ---------------- Sheet 4: Pricing check ----------------
ws = sheet("Pricing Check")
for i, w in enumerate([50, 20, 20, 60], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws["A1"] = "PRICING VERIFICATION - FENSTER QUOTE vs APLUS SUPPLIER COSTS"; ws["A1"].font = sub_font
r = put(ws, 3, ["Check", "Quoted", "Verified", "Result"], header=True)
pricing = [
 ("Items subtotal (rows 9-56)", "233,091.68", "233,091.68", "Exact - line maths recompute correctly."),
 ("Installation", "39,680.00", "39,680.00", "Exact - recomputes from code labour (4 ELAW x250, 13 DAD x500... 146 MAW x160, 46 SAW x160, 6 LAW, 2 SAD)."),
 ("TOTAL ex VAT", "272,771.68", "272,771.68", "Exact (items + installation)."),
 ("Technal STII door screens vs Aplus Logikal offer", "17,499.74", "17,499.74", "Every frame price matches the Aplus Logikal offer penny-for-penny; Fenster added 274.80 glass + 1,000 additional on Door 7."),
 ("Modeal windows + coupled doors vs Aplus Crystal QP70171", "111,185.75", "110,517.34 carried", "GBP 668.41 of the Aplus quote is NOT carried into the priced rows (likely welded cill / bay layout lines). Small margin leak (~1.1k sell at markup) - reallocate."),
 ("Aplus basis", "-", "Glazed, supply only, delivered; 21-Jul-2026", "Current quote; no end-discount block (Crystal format) - nett figures used correctly."),
 ("Blended materials markup (items sell / supplier cost)", "-", "~1.80", "In line with house markups for MAW/SAW-dominated mix."),
 ("Missing scope value", "-", "est. 3,500-5,500 sell", "3nr louvres (dims TBC) + 2nr 2280x1068 side screens - not in Aplus, not priced."),
 ("2.5% main contractor discount", "not visible", "-6,819.29 if applied", "Confirm whether markup absorbs the ITT-required MCD."),
 ("Optional items", "mastic 5,356.22 + EPDM 8,276.91", "-", "Bill deems sealing included - risk of non-compliant tender while optional."),
]
for row in pricing:
    fill = AMBER if ("668" in row[3] or "Missing" in row[0] or "Confirm" in row[3] or "risk" in row[3]) else LIGHT
    r = put(ws, r, list(row), fill=fill)

# ---------------- Sheet 5: Issues ----------------
ws = sheet("Issues & Actions")
for i, w in enumerate([6, 90, 14, 40], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws["A1"] = "ISSUES & RECOMMENDED ACTIONS (ranked)"; ws["A1"].font = sub_font
r = put(ws, 3, ["#", "Issue", "Severity", "Action"], header=True)
issues = [
 (1, "Technal spec: bill demands Technal UK spec for ALL windows and doors; windows and door types 1-5 are Modeal. Disclosed in proposal but not declared as a formal alternative/qualification with equivalency evidence.", "HIGH", "Declare qualification + equivalency pack; get CA/Employer approval in writing before order."),
 (2, "Missing scope: 3nr Louvre Type 01 (WLF 13/16/19, dims TBC) and 2nr 2280x1068 fixed side screens to Door Type 1 (DGF 22/23) - in neither the Aplus quotes nor the pricing.", "HIGH", "Price (est. 3.5-5.5k sell) or exclude explicitly in the tender return."),
 (3, "2.5% main contractor discount required to be included; not visible in pricing (GBP 6,819 exposure).", "HIGH", "Confirm with Adam whether markup allows for MCD; adjust if not."),
 (4, "PAS24/Part Q/SBD Silver: louvred/panelled doors not PAS24 tested (Aplus note); 'PAS24 hardware where quoted' is weaker than certified doorsets; 28nr ground-floor windows need PAS24 evidence.", "HIGH", "Obtain Modeal PAS24 certificates; agree SBD position for louvred Door Type 1."),
 (5, "Mastic (5,356.22) and EPDM (8,276.91) optional while the bill deems perimeter sealing (Tremco) included; making good / waste / final clean also deemed included but excluded in proposal.", "MEDIUM", "Roll mastic into base or qualify; reconcile exclusions with bill 'deemed included' items."),
 (6, "GBP 668.41 of Aplus window quote not carried into priced rows.", "LOW", "Identify missed lines (welded cills/bays) and reallocate (~1.1k sell)."),
 (7, "Dimension deviations vs T02 schedule: Type 3 height 960 vs 1135; Type 6 height 1375 vs 1160; Type 5 1375 vs 1400; assorted widths -13/-25mm.", "MEDIUM", "True up at survey; check with Aplus which revision they priced; variation risk both ways."),
 (8, "Obscure glazing on 76 of 191 schedule entries; pricing only splits obscure on Types 5/6/7.", "MEDIUM", "Confirm Aplus glazed price covers obscure glass on all 76 (incl. all of Types 1/1A/2/3)."),
 (9, "Heat-soak hedge: 'heat-soaked where quoted' vs ER requirement for ALL toughened glass heat-soaked + Kitemark.", "MEDIUM", "Confirm with Aplus; remove hedge from proposal."),
 (10, "Bill item AA (uPVC price variance) and lead-in times (alu + uPVC) not answered.", "MEDIUM", "Answer or formally decline (uPVC conflicts with Technal spec/Part Q)."),
 (11, "Guarantee: proposal offers 10-yr company warranty; bill/ERs want comprehensive + insurance-backed + Defects Insurer standards + durability assessment; window sample not allowed for.", "MEDIUM", "Confirm IBG; allow one fitted sample window."),
 (12, "Commercial: tender return was 17/07/2026 (quote dated 23/07); proposal T&Cs (50% deposit, O&M after final payment) conflict with Guildmore subcontract (5% retention, JCT D&B step-down).", "LOW", "Confirm late submission accepted; expect Guildmore terms to prevail."),
]
for row in issues:
    fill = RED if row[2] == "HIGH" else (AMBER if row[2] == "MEDIUM" else LIGHT)
    r = put(ws, r, list(row), fill=fill)
r += 1
ws.cell(r, 2, "VERIFIED GOOD: 191/191 window quantities reconcile exactly against BPG T02; door schedule fully mapped; installation and totals recompute exactly; Technal door costs match Aplus penny-for-penny; both supplier quotes current (21-22 Jul 2026); blended markup in line with house rates.").font = Font(bold=True, color=GREEN)
ws.cell(r, 2).alignment = wrap

out = r"C:\Users\zacpl\Desktop\Glazing-Quote-Assistant\test-results\princess-beatrice\Princess Beatrice House - Take-Off & Tender Audit.xlsx"
wb.save(out)
print("saved", out)
