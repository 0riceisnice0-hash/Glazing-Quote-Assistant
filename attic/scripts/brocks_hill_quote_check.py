"""Brocks Hill Phase 2 Teaching Block (Spacemaker Developments, SMDT0173) - quote check.

Audits Gintare's outgoing tender (pricing xlsx + proposal docx, both dated 28/07/2026)
against the architect's schedules, the Employer's Requirements and the two supplier
quotes actually on file.

Sources, all read at source:
  - Window Schedule  23409-S2E-04-00-D-A-32 XX  P04  30/01/2026 TENDER ISSUE
  - Door Schedule    23409-S2E-04-00-D-A-31 XX  P06  30/01/2026 TENDER ISSUE
  - Brocks Hill - Employers Requirements 003.pdf, s.5.4.1 and s.7.5
  - Brocks Hill Phase 2 Building Works Pricing Schedule.xlsx (SMD's own quantities)
  - Brocks Hill - Phase 2 Building Works Preliminaries-2026-02-10.pdf
  - BSW Window Solutions QT253232, 22/07/2026  (Sheerline, windows)
  - Bellview Products 0000000503, 22/07/2026   (SMA Smart Wall Pocket, doors)

Output: outputs/Brocks Hill Phase 2 - Quote Check (schedules vs tender).xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = r"outputs\Brocks Hill Phase 2 - Quote Check (schedules vs tender).xlsx"

HDR = PatternFill("solid", fgColor="1F2A44")
HDRF = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
TITLE = Font(bold=True, size=13, color="1F2A44")
RED = PatternFill("solid", fgColor="FCE4E4")
AMB = PatternFill("solid", fgColor="FFF3D6")
GRN = PatternFill("solid", fgColor="E4F3E6")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# ------------------------------------------------------------------ as quoted
# Gintare's pricing document, 28/07/2026. Frame = supplier net; adder = house
# template code value x 75%; unit rate = frame + adder.
QUOTED = [
    # desc, code, size, qty, frame_each, adder, supplier_pos
    ("Door Type E.02",              "SAD",  "1010 x 2110", 1,  2509.0725,  900.00, "Bellview 001"),
    ("Door Type E.04",              "DAD",  "1810 x 2110", 2,  2723.4935, 1500.00, "Bellview 002 (1no only)"),
    ("Door/Window Type E.01, E.03", "SAD",  "1800 x 2110", 4,  2965.4885,  900.00, "Bellview 003+004"),
    ("Window Type E.02",            "ELAW", "1800 x 2100", 23, 1362.5700,  637.50, "BSW e.02"),
    ("Window Type E.04",            "ELAW", "1800 x 2400", 4,  1098.9000,  637.50, "BSW e.04"),
    ("Window Type E.05",            "LAW",  "1000 x 2100", 2,   620.4900,  487.50, "BSW e.05"),
    ("Window Type E.06",            "ELAW", "1800 x 2100", 1,   984.6400,  637.50, "BSW e.06"),
]
LABOUR = {"SAD": 250.0, "DAD": 500.0, "ELAW": 250.0, "LAW": 160.0}
INSTALL_AS_QUOTED = 9570.00
TOTAL_AS_QUOTED = 93673.3435
MASTIC_OPT = 1400.80
EPDM_OPT = 3375.755

# Supplier quotes as filed.
BSW_NET = 37960.33        # QT253232, "Total Nett Ex. VAT", net price includes discounts
BELLVIEW_NET = 17094.52   # 0000000503, net 20,111.20 less 15% end discount
BELLVIEW_GROSS = 20111.20

# Architect's schedules - external elements only.
SCHEDULE_WINDOWS = {"E.01": 4, "E.02": 23, "E.04": 4, "E.05": 3, "E.06": 1}
SCHEDULE_DOORS = {
    "E.01": ("Hall External Escape Door - STEEL, Strongdor Sportsdor or similar approved",
             ["ED.0.04", "ED.0.06", "ED.0.07", "ED.0.08", "ED.0.09"], 1810, 2110),
    "E.02": ("Classroom External Door - aluminium, glazed",
             ["ED.0.02", "ED.0.03", "ED.0.11", "ED.0.12", "ED.0.13"], 1200, 2100),
    "E.03": ("Plant Room Door - aluminium frame with ALUMINIUM LOUVRES",
             ["ED.0.01", "ED.0.05"], 1810, 2110),
    "E.04": ("Glazed Entrance Doors - assisted opening to ED.0.10",
             ["ED.0.10", "ED.0.14"], 1810, 2110),
}

# ------------------------------------------------------------- missing scope
# Every figure below is a BENCHMARK or a supplier's own rate reapplied - none is
# a firm price. Provenance is stated on every line.
SOLAR_RATE = 35.00        # engine solarControlExtra, GBP/m2
STEELDOR_DOUBLE = 2603.00  # register headline median, Strongdor Steeldor double

MISSING = [
    ("Door Type E.01 - 5no steel sports hall escape doors 1810x2110",
     5, STEELDOR_DOUBLE, 1500.00, 500.00,
     "Register headline median, Strongdor Steeldor double. Sportsdor is a specialist "
     "flush-face sports hall door with recessed panic furniture - treat as a FLOOR, not a price."),
    ("Door Type E.03 - 2no aluminium louvred plant room doors 1810x2110",
     2, 2723.4935, 1500.00, 500.00,
     "Bellview's own net rate for the identical 1810x2110 double (pos 002). Frame basis only - "
     "the aluminium louvre infill is NOT in this figure and has no rate on file."),
    ("Window Type E.05 - 3rd unit (schedule lists 3, quote covers 2)",
     1, 620.49, 487.50, 160.00,
     "BSW's own net rate for e.05 on QT253232."),
    ("Power-assisted operator to main entrance ED.0.10",
     1, 3000.00, 0.00, 0.00,
     "Grange Hill auto-operator allowance (judgement). Bellview quoted the AUTO SLIDE/SWING "
     "header profile but no operator and no price for one."),
]


def solar_area():
    """Every external window on the schedule is marked 'Solar Control Glazing'."""
    a = 23 * 1.80 * 2.10 + 4 * 1.80 * 2.40 + 3 * 1.00 * 2.10 + 1 * 1.80 * 2.10  # E.02/04/05/06
    a += 4 * 0.60 * 2.10                                                        # E.01 fixed fields
    return a


def build():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------- Summary
    ws = wb.active
    ws.title = "Summary"
    rows = []
    quoted_frames = sum(q[3] * q[4] for q in QUOTED)
    quoted_sell = sum(q[3] * (q[4] + q[5]) for q in QUOTED)
    labour_calc = sum(q[3] * LABOUR[q[1]] for q in QUOTED)

    missing_sell = sum(n * (s + a) + n * l for _, n, s, a, l in
                       [(m[0], m[1], m[2], m[3], m[4]) for m in MISSING])
    solar_sell = solar_area() * SOLAR_RATE
    corrected = TOTAL_AS_QUOTED + missing_sell + solar_sell
    with_ancillary = corrected + MASTIC_OPT + EPDM_OPT
    with_mcd = with_ancillary / 0.975

    rows = [
        ("Job", "Brocks Hill Phase 2 Teaching Block, Oadby, Leicestershire LE2 5WP"),
        ("Client / enquiry", "Spacemaker Developments Ltd (Martin Moore) - SMDT0173"),
        ("End client", "Lionheart Education Trust / Brocks Hill Primary School"),
        ("Tender return", "Friday 31 July 2026"),
        ("Contract", "JCT Design and Build 2024. Possession 24/08/2026, completion 12/03/2027."),
        ("", ""),
        ("TENDER AS DRAFTED (ex VAT)", TOTAL_AS_QUOTED),
        ("  of which supplier frames", quoted_frames),
        ("  of which house code adders", quoted_sell - quoted_frames),
        ("  of which installation", INSTALL_AS_QUOTED),
        ("Optional extras shown separately", MASTIC_OPT + EPDM_OPT),
        ("", ""),
        ("ARITHMETIC", "Reconciles exactly. Template adders land to the penny on all 7 rows; "
                       "installation recomputes from the labour codes to GBP 9,570.00; both "
                       "supplier quotes tie to the frame column."),
        ("", ""),
        ("SCOPE MISSING FROM THE TENDER (benchmark, not a price)", missing_sell),
        ("Solar control glass on %.2f m2 of window at GBP %.0f/m2" % (solar_area(), SOLAR_RATE),
         solar_sell),
        ("CORRECTED INDICATIVE TOTAL", corrected),
        ("  plus mastic + EPDM, both specification requirements", with_ancillary),
        ("  grossed for the 2.5% MCD the enquiry requires", with_mcd),
        ("", ""),
        ("STILL UNQUANTIFIED", "Triple glazing to the doors; trickle vents; restrictors, lockable "
                               "handles and keys; remote openers above 2.0m; manifestations; obscure "
                               "glazing; hold-open closers; PAS 24 / LPS 1175 SR2 certification; "
                               "delivery (both quotes are ex works); Part M threshold seals; flashings."),
        ("", ""),
        ("CROSS-CHECK", "SMD's own Building Works Pricing Schedule carries 48 m2 of External Doors. "
                        "The tender prices 24.96 m2 - about half. Adding the missing Type E.01 and "
                        "E.03 doors gives 51.70 m2, which brackets SMD's figure."),
        ("VERDICT", "DO NOT ISSUE. scripts/mary_checks.py returns 4 FAILED on this job."),
    ]
    ws["A1"] = "BROCKS HILL PHASE 2 - QUOTE CHECK"
    ws["A1"].font = TITLE
    ws["A2"] = "Gintare's tender dated 28/07/2026, checked against the tender schedules and the supplier quotes on file. Prepared 27/07/2026."
    ws["A2"].font = BASE
    r = 4
    for k, v in rows:
        ws.cell(r, 1, k).font = BOLD if k and not k.startswith("  ") else BASE
        c = ws.cell(r, 2, v)
        c.font = BASE
        c.alignment = WRAP
        if isinstance(v, float):
            c.number_format = '"GBP "#,##0.00'
        if k.startswith("TENDER AS DRAFTED"):
            c.fill = AMB
            c.font = BOLD
        if k.startswith("CORRECTED") or k.startswith("VERDICT") or k.startswith("SCOPE MISSING"):
            c.fill = RED
            c.font = BOLD
        if k == "ARITHMETIC":
            c.fill = GRN
        r += 1
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 96

    # ------------------------------------------------- Reconciliation sheet
    ws2 = wb.create_sheet("Reconciliation")
    ws2["A1"] = "TENDER vs SUPPLIER QUOTES - line by line"
    ws2["A1"].font = TITLE
    hdrs = ["Description", "Code", "Size", "Qty", "Frame each (net)", "House adder",
            "Unit rate", "Line total", "Labour", "Supplier position", "Check"]
    for i, h in enumerate(hdrs, 1):
        c = ws2.cell(3, i, h)
        c.fill = HDR
        c.font = HDRF
        c.alignment = WRAP
    r = 4
    for desc, code, size, qty, frame, adder, pos in QUOTED:
        check = "OK"
        fill = None
        if "1no only" in pos:
            check = ("Bellview quoted 1no. The tender sells 2no at the same rate. "
                     "GBP 2,723.49 of cost has no quote behind it.")
            fill = RED
        vals = [desc, code, size, qty, frame, adder, frame + adder, qty * (frame + adder),
                qty * LABOUR[code], pos, check]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(r, i, v)
            c.font = BASE
            c.alignment = WRAP if i in (1, 10, 11) else TOP
            c.border = THIN
            if isinstance(v, float):
                c.number_format = '#,##0.00'
            if fill and i == 11:
                c.fill = fill
        r += 1
    r += 1
    for label, val, note in [
        ("Frames sold", sum(q[3] * q[4] for q in QUOTED),
         "Column J of the pricing document, extended by quantity."),
        ("Supplier quotes on file", BSW_NET + BELLVIEW_NET,
         "BSW QT253232 GBP 37,960.33 + Bellview 0000000503 GBP 17,094.52 (net of the 15% end discount on GBP 20,111.20)."),
        ("Difference", sum(q[3] * q[4] for q in QUOTED) - (BSW_NET + BELLVIEW_NET),
         "Exactly one Door Type E.04. The pricing document's own 'Supplier used' tally (cell M5) "
         "understates the frame cost being sold by the same amount."),
        ("Installation as quoted", INSTALL_AS_QUOTED,
         "Recomputes exactly from the labour codes: SAD 250x5 + DAD 500x2 + ELAW 250x28 + LAW 160x2."),
        ("Total as quoted", TOTAL_AS_QUOTED, "Ties to cell I20."),
    ]:
        ws2.cell(r, 1, label).font = BOLD
        c = ws2.cell(r, 8, val)
        c.font = BOLD
        c.number_format = '"GBP "#,##0.00'
        c2 = ws2.cell(r, 10, note)
        c2.font = BASE
        c2.alignment = WRAP
        r += 1
    for c, w in zip("ABCDEFGHIJK", [30, 8, 14, 6, 15, 12, 12, 13, 10, 30, 46]):
        ws2.column_dimensions[c].width = w

    # ------------------------------------------------------ Missing scope
    ws3 = wb.create_sheet("Missing scope")
    ws3["A1"] = "SCOPE ON THE ARCHITECT'S SCHEDULES THAT IS NOT IN THE TENDER"
    ws3["A1"].font = TITLE
    ws3["A2"] = ("Every figure here is a benchmark or a supplier's own rate reapplied. "
                 "None is a firm price and none should go to SMD as one.")
    ws3["A2"].font = BASE
    hdrs = ["Item", "Qty", "Supply each", "House adder", "Labour each", "Sell", "Basis"]
    for i, h in enumerate(hdrs, 1):
        c = ws3.cell(4, i, h)
        c.fill = HDR
        c.font = HDRF
        c.alignment = WRAP
    r = 5
    for desc, qty, supply, adder, labour, basis in MISSING:
        sell = qty * (supply + adder + labour)
        for i, v in enumerate([desc, qty, supply, adder, labour, sell, basis], 1):
            c = ws3.cell(r, i, v)
            c.font = BASE
            c.alignment = WRAP if i in (1, 7) else TOP
            c.border = THIN
            if isinstance(v, float):
                c.number_format = '#,##0.00'
        r += 1
    sell = solar_area() * SOLAR_RATE
    for i, v in enumerate(["Solar control glass to all external windows", "%.2f m2" % solar_area(),
                           SOLAR_RATE, 0.0, 0.0, sell,
                           "Engine solarControlExtra GBP 35/m2. The window schedule marks every "
                           "external window 'Triple Glazed, Solar Control Glazing' and ER 7.5.17 "
                           "requires it; BSW quoted 'Clr' clear glass on every line."], 1):
        c = ws3.cell(r, i, v)
        c.font = BASE
        c.alignment = WRAP if i in (1, 7) else TOP
        c.border = THIN
        if isinstance(v, float):
            c.number_format = '#,##0.00'
    r += 2
    total_missing = sum(m[1] * (m[2] + m[3] + m[4]) for m in MISSING) + sell
    ws3.cell(r, 1, "TOTAL QUANTIFIED GAP").font = BOLD
    c = ws3.cell(r, 6, total_missing)
    c.font = BOLD
    c.fill = RED
    c.number_format = '"GBP "#,##0.00'
    r += 2
    ws3.cell(r, 1, "NOT QUANTIFIED - each needs a supplier return or a decision").font = BOLD
    r += 1
    for t in [
        "Triple glazing to the external doors. ER 7.5.1 and the door schedule both require it; "
        "Bellview quoted double (6.8 Lami / 4mm Tuff). BSW have already told Fenster in writing "
        "that Smart Wall is not available in triple glazing - so this is a system change or a "
        "formal non-compliance qualification, not a glass swap.",
        "Trickle vents, min 4000mm2 to the head of every window (ER 7.5.12). Not on QT253232 at all.",
        "Restricted opening hinges to 100mm with push-button release to BS 6375-2 (ER 7.5.9), and "
        "lockable espagnolette handles with 3no keys per openable window (ER 7.5.10). Neither quoted.",
        "Remotely operated mechanical openers to any openable window above 2.0m (ER 7.5.11).",
        "Manifestations - required on window types E.01, E.02, E.05, E.06 and door type E.04.",
        "Obscure glazing to W.1.19 (Student Toilets FF) and ED.0.13 (Student Toilets GF). Both quoted clear.",
        "Door closers with built-in hold-back and a 90 degree hold-open facility (ER 7.5.5 / 7.5.7). "
        "Bellview quoted NHO - Non Hold Open - closers on every door. That is the opposite of the requirement.",
        "Security certification: PAS 24:2007 / WCL 1 / LPS 1175 Issue 7 SR2 on doorsets, PAS 23-1:1999, "
        "SBD ironmongery, and BS 7950 / WCL 4 on ground floor windows. Neither quote references any of it.",
        "Delivery. Both quotes are ex works - QT253232 says so on its face. No delivery line in the tender.",
        "Low mobility Part M threshold seals (ER 7.5.6) and aluminium flashings / perimeter trims (ER 7.5.8).",
    ]:
        c = ws3.cell(r, 1, t)
        c.font = BASE
        c.alignment = WRAP
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws3.row_dimensions[r].height = 30
        r += 1
    for c, w in zip("ABCDEFG", [46, 10, 13, 12, 12, 13, 62]):
        ws3.column_dimensions[c].width = w

    # ---------------------------------------------------------- Findings
    ws4 = wb.create_sheet("Findings")
    ws4["A1"] = "FINDINGS"
    ws4["A1"].font = TITLE
    hdrs = ["#", "Severity", "Finding", "Source", "What it costs / what to do"]
    for i, h in enumerate(hdrs, 1):
        c = ws4.cell(3, i, h)
        c.fill = HDR
        c.font = HDRF
    findings = [
        ("HIGH", "SEVEN EXTERNAL DOORS ARE NOT IN THE TENDER. Door Type E.01 (5no steel sports "
                 "hall escape doors, ED.0.04/06/07/08/09) and Door Type E.03 (2no aluminium louvred "
                 "plant room doors, ED.0.01/05) appear on the door schedule and in neither supplier "
                 "quote nor the pricing document, and are not excluded.",
         "Door Schedule 23409-...-31 XX P06, 30/01/2026",
         "About GBP 32,462 of sell at benchmark, and the louvre infill on the E.03s is on top of that. "
         "Independently corroborated: SMD's own pricing schedule carries 48 m2 of external doors, "
         "the tender prices 24.96 m2."),
        ("HIGH", "FIVE FIRE ESCAPE DOORS WITH NO PANIC HARDWARE PRICED. All five Type E.01 doors are "
                 "marked Fire Escape on the schedule and require panic furniture to both leaves as a "
                 "specialist design. They are not priced at all.",
         "Door Schedule P06; mary_checks.py check_panic_hardware",
         "Falls away once the E.01 doors are priced - but they must be priced as certified escape "
         "doorsets, not plain doors."),
        ("HIGH", "THE PROPOSAL PROMISES TRIPLE GLAZING ON THE DOORS AND THE SUPPLIER HAS QUOTED "
                 "DOUBLE. Bellview 0000000503 quotes 6.8 Lami / 4mm Tuff on every door position - a "
                 "double glazed unit. The proposal tells SMD 'triple glazing throughout'.",
         "Bellview 0000000503; proposal p1 Executive Summary; ER 7.5.1; door schedule general spec",
         "BSW have already told Fenster in writing that Smart Wall products are not available in "
         "triple glazing. This needs a different door system or a formal qualification in the "
         "tender - it cannot be closed by changing the glass."),
        ("HIGH", "SOLAR CONTROL GLASS IS NOT PRICED ON ANY WINDOW. The window schedule marks every "
                 "external window 'Triple Glazed, Solar Control Glazing' and ER 7.5.17 requires it. "
                 "BSW quoted 'Clr' - clear - on every line.",
         "Window Schedule P04; ER 7.5.17; BSW QT253232",
         "About GBP 4,177 at the GBP 35/m2 benchmark. Same omission as Filwood and Crestwood this month."),
        ("HIGH", "THE 2.5% MCD THE ENQUIRY ASKS FOR IS NOT IN THE TENDER. Martin Moore's enquiry says "
                 "'Please allow for a 2.5% MCD'. There is no MCD line in the pricing document.",
         "Martin Moore, 14/07/2026 enquiry",
         "Grossed neutral the subtotal needs to rise by GBP 2,401.88. Taken off the bottom as on "
         "Princess Beatrice it costs GBP 2,341.84 of margin. Adam's call which."),
        ("MED", "ONE WINDOW SHORT. The window schedule lists 3no Type E.05 (W.0.14, W.1.19, W.1.20); "
                "BSW quoted 2 and the tender sells 2.",
         "Window Schedule P04",
         "GBP 1,267.99 at BSW's own rate."),
        ("MED", "ONE DOOR SOLD WITH NO SUPPLIER QUOTE BEHIND IT. The tender sells 2no Door Type E.04; "
                "Bellview quoted 1no. The second is sold at the same rate.",
         "Bellview 0000000503 pos 002; pricing document row 10",
         "GBP 2,723.49 of cost uncovered. The rate will very likely hold - same product, same size - "
         "but the quote has to be extended before order. Note only the quoted one carries the "
         "AUTO SLIDE/SWING header."),
        ("MED", "THE POWER-ASSISTED OPERATOR TO ED.0.10 IS NOT PRICED. The door schedule requires "
                "assisted opening to the main entrance. Bellview supplied the deep AUTO SLIDE/SWING "
                "header profile but no operator and no price for one.",
         "Door Schedule P06 Type E.04 note; Bellview pos 002",
         "About GBP 3,000 on the Grange Hill allowance. This was already RFI'd on 15/07 and is "
         "still open."),
        ("MED", "MASTIC AND EPDM ARE SPECIFICATION REQUIREMENTS BEING OFFERED AS OPTIONAL EXTRAS. "
                "The door schedule requires all openings sealed with non-setting mastic all round; "
                "ER 7.5.1 requires EPDM gaskets and weatherseals to BS 4255. Both sit in the "
                "OPTIONAL block at GBP 1,400.80 and GBP 3,375.76.",
         "Door Schedule P06 general spec; ER 7.5.1",
         "GBP 4,776.56 of specified work outside the tender sum. Note this is the opposite of what "
         "Adam ordered on Princess Beatrice - worth confirming which way he wants it now."),
        ("MED", "WINDOW IRONMONGERY IS SHORT OF THE ER. No trickle vents (ER 7.5.12, min 4000mm2 "
                "each), no restrictors or push-button release (ER 7.5.9, and the window schedule "
                "says 'Fitted with restrictors'), no lockable espagnolette handles or the 3no keys "
                "per window (ER 7.5.10), no remote openers above 2.0m (ER 7.5.11).",
         "ER 7.5.9-7.5.12; BSW QT253232",
         "Not quantified - needs a BSW requote. None of it is on the quote."),
        ("MED", "THE CLOSERS QUOTED ARE THE OPPOSITE OF THE ONES SPECIFIED. ER 7.5.5 requires "
                "concealed closers with built-in hold-back and 7.5.7 a 90 degree hold-open facility. "
                "Bellview quoted 'NHO CLOSER' and 'DOOR CLOSER NON HOLD OPEN' on every door.",
         "ER 7.5.5 / 7.5.7; Bellview 0000000503",
         "Requote. Low value, but it is a stated requirement quoted against."),
        ("MED", "NO SECURITY CERTIFICATION ANYWHERE. The door schedule requires doorsets certified to "
                "PAS 24:2007 / WCL 1 / LPS 1175 Issue 7 SR2, also PAS 23-1:1999, ironmongery to SBD "
                "standards, and ground floor windows to BS 7950:1997 / WCL 4. Neither quote "
                "references any of it. The proposal defers it: 'PAS 24 / SBD requirements and final "
                "ironmongery to be coordinated prior to manufacture'.",
         "Door Schedule P06 general spec; both quotes; proposal p1",
         "Fourth time this month after Princess Beatrice, Filwood and Vesuvius. Either get it "
         "certified or qualify it - a deferral is not a compliance statement."),
        ("MED", "BOTH SUPPLIER QUOTES EXPIRE BEFORE THE SITE STARTS. QT253232 and Bellview 0000000503 "
                "are both dated 22/07/2026 and valid 30 days - about 21/08/2026. Date of possession "
                "is 24/08/2026.",
         "Both quotes; Preliminaries clause 2.3",
         "Order has to be placed by ~21/08 or both prices are re-opened. Also: both quotes are ex "
         "works and no delivery is priced."),
        ("MED", "THE TENDER IS SILENT ON TWO THINGS THE ENQUIRY EXPRESSLY ASKED FOR. Martin Moore "
                "asked for VE opportunities to be highlighted and for lead-in times. The proposal "
                "offers neither, and neither supplier quote states a lead time.",
         "Martin Moore, 14/07/2026 enquiry",
         "Free marks on a competitive tender. Ask BSW and Bellview for lead times - they are needed "
         "against a 24/08 possession anyway."),
        ("MED", "FENSTER'S STANDARD T&Cs ARE UNQUALIFIED AGAINST JCT D&B 2024. The proposal offers "
                "50% deposit / 50% on completion, 30-day validity and a 10-year warranty. The "
                "preliminaries carry LADs of GBP 12,500 per calendar week, 3% retention, periodic "
                "payment under Alternative B, 12 years of PI insurance, collateral warranties and a "
                "12-month defects period.",
         "Preliminaries 2026-02-10, clauses 1.1/2.3/2.29.2/4.7/4.18.1/6.15/7.4",
         "Fenster is a sub-contractor to SMD so SMD's own sub-contract will flow down, but the "
         "proposal should say what it is and is not accepting. LADs of GBP 12,500/week are eight "
         "times Filwood's."),
        ("LOW", "NEITHER QUOTE STATES AN INTERNAL FINISH. ER 7.5.13 specifies RAL 7016 powder coated. "
                "BSW state 'Ext Colour: 7016M' and Bellview 'Profiles: Anthracite Grey' - the "
                "internal face is not mentioned on either.",
         "ER 7.5.13; both quotes; mary_checks check_finish_substitution returns ASK",
         "Same silence that cost Georgie's the white internal face. Confirm 7016 both sides is in "
         "the price."),
        ("LOW", "SCOPE BOUNDARIES ARE NOT WRITTEN DOWN. The 7no rooflights R.2.01-07 belong to the "
                "roofing manufacturer under ER 7.6.6, and the 8no internal timber screens and all "
                "internal doors are outside Fenster's package. None of them is excluded in the "
                "proposal - and the rooflights sit on the window schedule.",
         "Window Schedule P04; ER 7.6.6",
         "Correctly out of scope, wrongly left unsaid. A silent gap reads as included."),
        ("LOW", "WORKBOOK HYGIENE. Cells O12 and O17 read #VALUE!. Cells K3/L3/M3:M5 hold "
                "'Supplier used: BSW' and the supplier totals, and columns J and P hold frame cost - "
                "all outside the print area C1:I28.",
         "Pricing xlsx",
         "Print area is clean, so SEND PDF ONLY. Do not send the .xlsx - it leaks the supplier and "
         "the cost."),
        ("LOW", "THE U-VALUE BASIS IS AMBIGUOUS. The door schedule says 1.1 W/m2K is an "
                "area-weighted average across all external glazing; ER 5.4.1 tabulates per-element "
                "maxima of Window 1.1, Doors 1.2, Rooflight 2.2. QT253232 states no U-value at all.",
         "Door Schedule P06; ER 5.4.1",
         "RFI. No element is being rejected on this - but Fenster should say which basis it has "
         "priced to, and BSW should state a figure."),
    ]
    r = 4
    for i, (sev, txt, src, act) in enumerate(findings, 1):
        vals = [i, sev, txt, src, act]
        for j, v in enumerate(vals, 1):
            c = ws4.cell(r, j, v)
            c.font = BOLD if j == 2 else BASE
            c.alignment = WRAP
            c.border = THIN
            if j == 2:
                c.fill = {"HIGH": RED, "MED": AMB, "LOW": GRN}[sev]
        ws4.row_dimensions[r].height = 62
        r += 1
    for c, w in zip("ABCDE", [5, 10, 72, 38, 66]):
        ws4.column_dimensions[c].width = w

    # -------------------------------------------------------------- RFIs
    ws5 = wb.create_sheet("RFIs")
    ws5["A1"] = "RFIs / QUESTIONS TO PUT BEFORE THE TENDER GOES"
    ws5["A1"].font = TITLE
    ws5["A2"] = ("Mary cannot email SMD or the suppliers. These need a human to send them. "
                 "Martin Moore, Senior Quantity Surveyor, martin.moore@smd-ltd.com, 07564 581082.")
    ws5["A2"].font = BASE
    for i, h in enumerate(["#", "To", "Question"], 1):
        c = ws5.cell(4, i, h)
        c.fill = HDR
        c.font = HDRF
    rfis = [
        ("SMD", "Are the 5no Type E.01 steel sports hall escape doors (ED.0.04/06/07/08/09) and the "
                "2no Type E.03 louvred plant room doors (ED.0.01/05) in our package? They are on the "
                "door schedule but not in the BoQ we were issued. Our price excludes them as it stands."),
        ("SMD", "The door schedule and ER 7.5.1 require triple glazing to the external doors. Our door "
                "system is not available in triple glazing. Will a double glazed door meeting the "
                "1.2 W/m2K door U-value be accepted, or does the door package need a different system?"),
        ("SMD", "Is the 1.1 W/m2K an area-weighted average across all external glazing, as the door "
                "schedule states, or the per-element maximum tabulated at ER 5.4.1?"),
        ("SMD", "Confirm the solar control glass specification - the window schedule marks every "
                "window 'Solar Control Glazing' but names no product. We need a g-value to price to."),
        ("SMD", "Confirm the rooflights R.2.01-07 sit with the roofing package under ER 7.6.6 and are "
                "not in our scope."),
        ("SMD", "Which doors require access control? The door schedule leaves this to the Client to confirm."),
        ("SMD", "Confirm the manifestation pattern, or confirm it is a provisional allowance at this stage."),
        ("BSW", "Requote QT253232 with: solar control glass; trickle vents min 4000mm2 to every head; "
                "restricted opening hinges to 100mm with push-button release to BS 6375-2; lockable "
                "espagnolette handles with 3no keys per window; obscure glazing to W.1.19; the 3rd "
                "Type E.05; BS 7950 / WCL 4 certification on the ground floor windows; a stated "
                "U-value; internal finish confirmed RAL 7016; a lead time; and a delivered price."),
        ("Bellview", "Requote 0000000503 with: 2no Door Type E.04 not 1no; hold-open closers per ER "
                     "7.5.5 and 7.5.7 in place of the NHO closers quoted; the power operator to "
                     "ED.0.10; obscure glazing to ED.0.13; PAS 24 / LPS 1175 SR2 certification; Part M "
                     "threshold seals; internal finish confirmed; a lead time; and a delivered price. "
                     "Confirm in writing whether Smart Wall can be supplied triple glazed."),
        ("Strongdor or equal", "Quote 5no Type E.01 steel sports hall escape doorsets 1810x2110, PPC "
                               "anthracite, flush faces, recessed panic furniture to both leaves, fire "
                               "escape, no external access. Sportsdor or similar approved."),
        ("Bellview or equal", "Quote 2no Type E.03 aluminium louvred plant room doorsets 1810x2110, "
                              "PPC anthracite, 3 pairs hinges, suited mortice lock, hold stay bolt to "
                              "slave leaf."),
        ("Adam", "MCD: gross the subtotal up by 2.5% so the deduction is neutral, or take it off the "
                 "bottom as we did on Princess Beatrice?"),
        ("Adam", "Mastic and EPDM are specification requirements here. Into the tender sum as on "
                 "Princess Beatrice, or left as optional extras as on Crestwood?"),
    ]
    r = 5
    for i, (to, q) in enumerate(rfis, 1):
        for j, v in enumerate([i, to, q], 1):
            c = ws5.cell(r, j, v)
            c.font = BASE
            c.alignment = WRAP
            c.border = THIN
        ws5.row_dimensions[r].height = 46
        r += 1
    for c, w in zip("ABC", [5, 20, 120]):
        ws5.column_dimensions[c].width = w

    for s in wb.worksheets:
        s.freeze_panes = {"Summary": "A4", "Reconciliation": "A4",
                          "Missing scope": "A5", "Findings": "A4", "RFIs": "A5"}[s.title]
    wb.save(OUT)
    print("wrote", OUT)
    print("quoted total      GBP %10.2f" % TOTAL_AS_QUOTED)
    print("quantified gap    GBP %10.2f" % total_missing)
    print("corrected         GBP %10.2f" % (TOTAL_AS_QUOTED + total_missing))
    print("+ mastic/EPDM     GBP %10.2f" % (TOTAL_AS_QUOTED + total_missing + MASTIC_OPT + EPDM_OPT))
    print("+ 2.5%% MCD gross  GBP %10.2f" % ((TOTAL_AS_QUOTED + total_missing + MASTIC_OPT + EPDM_OPT) / 0.975))


if __name__ == "__main__":
    build()
