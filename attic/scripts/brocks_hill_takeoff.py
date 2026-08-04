# -*- coding: utf-8 -*-
"""Brocks Hill Phase 2 - take-off document (Adam, REQ-2 dashboard 28/07 20:30).

Greenfields 3-sheet format: Project Information / Window & Door Schedule / RFIs.
Every element on the architect's schedules, whether or not it is in the current
tender, with what is quoted against it.

Sources: Window Schedule 23409-S2E-04-00-D-A-32 XX P04 (30/01/26), Door Schedule
...-31 XX P06 (30/01/26), Employer's Requirements 003 s.5.4.1 / 7.5, Preliminaries
2026-02-10, BSW QT253232 and Bellview 0000000503 (both 22/07/2026).

Output: outputs/Brocks Hill Phase 2 - Take-Off.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = r"outputs\Brocks Hill Phase 2 - Take-Off.xlsx"

HDR = PatternFill("solid", fgColor="1F2A44")
HDRF = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
TITLE = Font(bold=True, size=13, color="1F2A44")
RED = PatternFill("solid", fgColor="FCE4E4")
AMB = PatternFill("solid", fgColor="FFF3D6")
GRN = PatternFill("solid", fgColor="E4F3E6")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

Q = "QUOTED"
N = "NOT QUOTED"

# ref, type, w, h, room/location, opening, glazing, supplier line, status, note
ELEMENTS = [
    # ---- ground floor windows
    ("W.0.01", "E.01", 600, 2100, "Classroom 03", "Fixed field", "Triple, solar control",
     "Bellview 003/004 (in 1800 element)", Q, "Coupled with door ED.0.02 into one 1800x2110 Smart Wall element"),
    ("W.0.02", "E.02", 1800, 2100, "Classroom 03", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.03", "E.02", 1800, 2100, "Classroom 03", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.04", "E.02", 1800, 2100, "Classroom 02", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.05", "E.02", 1800, 2100, "Classroom 02", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.06", "E.01", 600, 2100, "Classroom 02", "Fixed field", "Triple, solar control",
     "Bellview 003/004 (in 1800 element)", Q, "Coupled with door ED.0.03"),
    ("W.0.07", "E.02", 1800, 2100, "Meeting Room", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.08", "E.02", 1800, 2100, "Office", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.09", "E.02", 1800, 2100, "Library", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.10", "E.01", 600, 2100, "Library", "Fixed field", "Triple, solar control",
     "Bellview 003/004 (in 1800 element)", Q, "Coupled with door ED.0.11"),
    ("W.0.11", "E.02", 1800, 2100, "Classroom 01", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.12", "E.02", 1800, 2100, "Classroom 01", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.0.13", "E.01", 600, 2100, "Classroom 01", "Fixed field", "Triple, solar control",
     "Bellview 003/004 (in 1800 element)", Q, "Coupled with door ED.0.12"),
    ("W.0.14", "E.05", 1000, 2100, "Stair 02", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.05", Q, ""),
    # ---- first floor windows
    ("W.1.01", "E.02", 1800, 2100, "Office", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.02", "E.02", 1800, 2100, "Classroom 6", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.03", "E.02", 1800, 2100, "Classroom 6", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.04", "E.02", 1800, 2100, "Classroom 6", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.05", "E.02", 1800, 2100, "Classroom 5", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.06", "E.02", 1800, 2100, "Classroom 5", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.07", "E.02", 1800, 2100, "Classroom 5", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.08", "E.02", 1800, 2100, "Staff Room", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.09", "E.04", 1800, 2400, "Hall (high level)", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.04", Q, ""),
    ("W.1.10", "E.04", 1800, 2400, "Hall (high level)", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.04", Q, ""),
    ("W.1.11", "E.04", 1800, 2400, "Hall (high level)", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.04", Q, ""),
    ("W.1.12", "E.04", 1800, 2400, "Hall (high level)", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.04", Q, ""),
    ("W.1.13", "E.02", 1800, 2100, "General Purpose", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.14", "E.02", 1800, 2100, "General Purpose", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.15", "E.02", 1800, 2100, "General Purpose", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.16", "E.02", 1800, 2100, "Classroom 4", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.17", "E.02", 1800, 2100, "Classroom 4", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.18", "E.02", 1800, 2100, "Classroom 4", "Tilt/turn side panel + fixed", "Triple, solar control",
     "BSW e.02", Q, ""),
    ("W.1.19", "E.05", 1000, 2100, "Student Toilets FF", "Fixed", "Triple, solar control, TOUGHENED, OBSCURE",
     "BSW e.05", Q, "Obscure required by the window schedule note - BSW quoted CLEAR"),
    ("W.1.20", "E.05", 1000, 2100, "Stair 2", "Fixed", "Triple, solar control, TOUGHENED",
     "-", N, "THIRD E.05. Schedule lists 3; BSW quoted 2. GBP 620.49 supply at BSW's own rate."),
    ("W.1.21", "E.06", 1800, 2100, "Circulation FF", "Fixed", "Triple, solar control, TOUGHENED",
     "BSW e.06", Q, ""),
    # ---- external doors
    ("ED.0.01", "E.03", 1810, 2110, "Plant Room", "Double, aluminium LOUVRED leaf", "n/a - louvred",
     "-", N, "Aluminium frame with aluminium louvres. 3 pairs hinges, suited mortice lock, hold stay bolt to slave leaf. NO SUPPLIER."),
    ("ED.0.02", "E.02", 1200, 2100, "Classroom 03", "Single pivoted AFT, FIRE ESCAPE", "Double glazed as quoted",
     "Bellview 003/004", Q, "In 1800x2110 element with W.0.01. Concealed panic bar quoted."),
    ("ED.0.03", "E.02", 1200, 2100, "Classroom 02", "Single pivoted AFT, FIRE ESCAPE", "Double glazed as quoted",
     "Bellview 003/004", Q, "In 1800x2110 element with W.0.06. Concealed panic bar quoted."),
    ("ED.0.04", "E.01", 1810, 2110, "Hall", "Double, STEEL, FIRE ESCAPE", "Solid - no glazing",
     "-", N, "Specialist sports hall door, flush faces, recessed panic furniture to BOTH leaves. Strongdor Sportsdor or similar approved. NO SUPPLIER."),
    ("ED.0.05", "E.03", 1810, 2110, "Plant Room", "Double, aluminium LOUVRED leaf", "n/a - louvred",
     "-", N, "As ED.0.01. NO SUPPLIER."),
    ("ED.0.06", "E.01", 1810, 2110, "Hall", "Double, STEEL, FIRE ESCAPE", "Solid - no glazing",
     "-", N, "As ED.0.04. NO SUPPLIER."),
    ("ED.0.07", "E.01", 1810, 2110, "Hall", "Double, STEEL, FIRE ESCAPE", "Solid - no glazing",
     "-", N, "As ED.0.04. NO SUPPLIER."),
    ("ED.0.08", "E.01", 1810, 2110, "Servery", "Double, STEEL, FIRE ESCAPE", "Solid - no glazing",
     "-", N, "As ED.0.04. NO SUPPLIER."),
    ("ED.0.09", "E.01", 1810, 2110, "Equipment Store", "Double, STEEL, FIRE ESCAPE", "Solid - no glazing",
     "-", N, "As ED.0.04. NO SUPPLIER."),
    ("ED.0.10", "E.04", 1810, 2110, "Circulation - MAIN ENTRANCE", "Double pivoted AFT, assisted opening", "Double glazed as quoted",
     "Bellview 002", Q, "ASSISTED OPENING required. Bellview quoted the AUTO SLIDE/SWING header but NO operator - allow ~GBP 3,000."),
    ("ED.0.11", "E.02", 1200, 2100, "Library", "Single pivoted AFT, FIRE ESCAPE", "Double glazed as quoted",
     "Bellview 003/004", Q, "In 1800x2110 element with W.0.10. Concealed panic bar quoted."),
    ("ED.0.12", "E.02", 1200, 2100, "Classroom 01", "Single pivoted AFT, FIRE ESCAPE", "Double glazed as quoted",
     "Bellview 003/004", Q, "In 1800x2110 element with W.0.13. Concealed panic bar quoted."),
    ("ED.0.13", "E.02", 1010, 2100, "Student Toilets GF", "Single pivoted AFT, FIRE ESCAPE", "Double glazed, OBSCURE required",
     "Bellview 001", Q, "Obscure required by the door schedule note - Bellview quoted CLEAR."),
    ("ED.0.14", "E.04", 1810, 2110, "Circulation", "Double pivoted AFT", "Double glazed as quoted",
     "-", N, "SOLD but NOT QUOTED. Bellview 0000000503 covers 1no Type E.04, the tender sells 2. GBP 2,723.49 of cost uncovered."),
]

OUT_OF_SCOPE = [
    ("R.2.01 - R.2.07", "Rooflight", "7no", "Main Roof",
     "uPVC framed, triple glazed with solar coating, integrated automatic blinds, U 1.1",
     "ROOFING PACKAGE - ER 7.6.6 requires all roof system components including rooflights from one "
     "manufacturer for a single-point guarantee. Exclude in writing."),
    ("S.0.01 - S.0.04, S.1.01 - S.1.04", "I.01 / I.02", "8no", "Ground + First Floor",
     "Timber framed internal screens, single glazed, 30/60 min FR",
     "TIMBER - not Fenster's package. Exclude in writing."),
    ("ID.0.01 - ID.1.17", "I.01 - I.06", "39no", "Throughout",
     "Timber internal doors, 44mm solid core oak veneer, FD30S",
     "TIMBER - priced separately on SMD's own schedule (Internal Doors, 39nr). Exclude in writing."),
    ("ID.0.20, ID.0.21", "Shutter", "2no", "Store/Queue, Servery",
     "Roller shutters 2400x2600 and 1000x2500, FD30S",
     "SHUTTER SPECIALIST - not Fenster's package. Exclude in writing."),
    ("Access Hatch", "-", "1no", "Main Roof", "535 x 1195",
     "ROOF ACCESS - ER 7.8.2. Not Fenster's package. Exclude in writing."),
]


def build():
    wb = openpyxl.Workbook()

    # ------------------------------------------------- 1. Project Information
    ws = wb.active
    ws.title = "Project Information"
    ws["A1"] = "BROCKS HILL PHASE 2 TEACHING BLOCK - TAKE-OFF"
    ws["A1"].font = TITLE
    ws["A2"] = "Prepared for Adam Butcher, 28/07/2026. Every element on the architect's schedules, with what is quoted against it."
    ws["A2"].font = BASE

    info = [
        ("PROJECT", ""),
        ("Enquiry ref", "SMDT0173"),
        ("Enquirer", "Spacemaker Developments Ltd - Martin Moore, Senior QS, martin.moore@smd-ltd.com, 07564 581082"),
        ("End client", "Lionheart Education Trust / Brocks Hill Primary School"),
        ("Architect", "Surveyors to Education Ltd (S2E), Enderby, Leicester - enquire@s2e.org.uk"),
        ("Site", "Brocks Hill Teaching Block, Howdon Road, Oadby, Leicestershire LE2 5WP"),
        ("Scheme", "New teaching building - classrooms, hall, toilets, ancillary - plus external works"),
        ("", ""),
        ("DATES", ""),
        ("Tender return", "FRIDAY 31 JULY 2026"),
        ("Date of possession", "24 August 2026"),
        ("Date for completion", "12 March 2027"),
        ("Site constraint", "Occupied primary school. Phase 1 works complete 21/08/2026. Term dates issued with the pack."),
        ("", ""),
        ("CONTRACT", ""),
        ("Form", "JCT Design and Build Contract (DB), 2024 Edition"),
        ("Liquidated damages", "GBP 12,500 per calendar week"),
        ("Retention", "3 per cent"),
        ("Payment", "Periodically, Alternative B (NOT Fenster's standard 50/50)"),
        ("Defects period", "Twelve months from practical completion"),
        ("PI insurance", "Required for 12 years"),
        ("Bond", "Not required"),
        ("Collateral warranties", "Clause 7.4 applies - third party rights / collateral warranties"),
        ("MCD", "2.5% requested in the enquiry"),
        ("Also requested", "VE opportunities to be highlighted; lead-in times to be stated"),
        ("", ""),
        ("SOURCE DOCUMENTS", ""),
        ("Window schedule", "23409-S2E-04-00-D-A-32 XX, revision P04, 30/01/2026 TENDER ISSUE"),
        ("Door schedule", "23409-S2E-04-00-D-A-31 XX, revision P06, 30/01/2026 TENDER ISSUE"),
        ("Specification", "Brocks Hill - Employers Requirements 003.pdf, s.5.4.1 and s.7.5"),
        ("Preliminaries", "Brocks Hill - Phase 2 Building Works Preliminaries-2026-02-10.pdf"),
        ("Client quantities", "Brocks Hill Phase 2 Building Works Pricing Schedule.xlsx - Windows 136 m2, External Doors 48 m2"),
        ("", ""),
        ("MATERIALS AND FINISH", ""),
        ("Frames", "Thermally broken aluminium throughout (ER 7.5.1), EXCEPT door Type E.01 which the door schedule specifies as STEEL"),
        ("Finish", "Powder coated RAL 7016 (TBC) - ER 7.5.13. Neither supplier quote states an INTERNAL finish."),
        ("Windows quoted as", "Sheerline Prestige casement, RAL 7016M Anthracite, 158mm cill, Kenrick shootbolt, black Signature handles, hinge protectors"),
        ("Doors quoted as", "SMA Smart Wall Pocket pivoted anti-fingertrap, Anthracite Grey"),
        ("", ""),
        ("GLAZING", ""),
        ("Specified", "Factory sealed TRIPLE glazed, argon filled, to windows AND doors (ER 7.5.1, door schedule general note)"),
        ("Windows quoted", "4T/4T/6.8 Clr Tgh Lam triple, argon, 36mm black warm edge - COMPLIANT on triple, but CLEAR (no solar coating)"),
        ("Doors quoted", "6.8 Lami / 4mm Tuff - DOUBLE glazed. Smart Wall is not available in triple. Adam is calling SMD on this (REQ-2)."),
        ("Solar control", "Required on every external window (schedule + ER 7.5.17). NOT PRICED. No product named - g-value needed."),
        ("Safety glass", "Toughened or laminated to BS 6206 in critical zones (Part K); toughened to BR ADK on types E.02/E.04/E.05/E.06"),
        ("Obscure", "Toilets - W.1.19 and ED.0.13. Both quoted clear."),
        ("", ""),
        ("U-VALUES (ER 5.4.1)", ""),
        ("Window", "1.1 W/m2K"),
        ("Doors", "1.2 W/m2K"),
        ("Rooflight", "2.2 W/m2K"),
        ("Basis", "Door schedule calls 1.1 an AREA-WEIGHTED AVERAGE across all external glazing; ER 5.4.1 tabulates per-element maxima. Ambiguity - RFI 3."),
        ("Stated by supplier", "Bellview state Ug 1.0 on the doors. BSW state NO U-value at all on QT253232."),
        ("", ""),
        ("SECURITY AND STANDARDS", ""),
        ("Doorsets", "Certified to one of: PAS 24:2007, WCL 1, LPS 1175 Issue 7 Security Rating 2. Also PAS 23-1:1999. Ironmongery to SBD standards."),
        ("Windows", "Ground floor and easily accessible windows to BS 7950:1997 or WCL 4"),
        ("Status", "NEITHER SUPPLIER QUOTE REFERENCES ANY OF IT"),
        ("Other standards", "BS 4873 windows; BS 6262 glazing; BS 6375-1 weathering, exposure/design wind load 2000; BS 6375-2 restrictors; BS 4255 gaskets"),
        ("DfE", "DfE Specification GDB, Technical Annexes 2C, 2E, 2F; BB93 acoustics; glazing to Annex 2F; fingerguards throughout to Annex 2C"),
        ("", ""),
        ("IRONMONGERY / ANCILLARIES REQUIRED BY ER 7.5", ""),
        ("7.5.5", "External doors: fingersafe pivot hinges, concealed low-weight closers WITH BUILT-IN HOLD BACK, brush perimeter seals. Bellview quoted NON HOLD OPEN closers."),
        ("7.5.6", "Low mobility Part M threshold seal to every external door. Not priced."),
        ("7.5.7", "Part M opening force: max 30N at 0-30 deg, 22.5N at 30-60 deg, 90 deg hold-open facility."),
        ("7.5.8", "Aluminium flashings and perimeter trims. Not priced."),
        ("7.5.9", "Restricted opening hinges to 100mm, shootbolt locking friction hinges, push-button release to BS 6375-2. Not quoted."),
        ("7.5.10", "Lockable espagnolette handles, minimum 3no keys per openable window. Not quoted."),
        ("7.5.11", "Remotely operated mechanical openers to any openable window above 2.0m. Not quoted."),
        ("7.5.12", "Closable trickle vents to head of EVERY window, min 4000 mm2, concealed fixings, to Approved Document F2. NOT ON QT253232."),
        ("Manifestations", "Required on window types E.01, E.02, E.05, E.06 and door type E.04. Pattern to be agreed with the Client. Not priced."),
        ("Access control", "Door access control to the School's security provider, tied into existing security and fire alarm. Client to confirm which doors."),
        ("EPDM", "EPDM extruded rubber gaskets and weatherseals to BS 4255 - ER 7.5.1. Currently an optional extra at GBP 3,375.76."),
        ("Mastic", "All openings sealed with non-setting mastic all round - door schedule. Currently an optional extra at GBP 1,400.80. Per Adam's 28/07 ruling this goes IN and the exclusion line comes out."),
        ("", ""),
        ("WARRANTY", ""),
        ("Fenster standard", "10 years, glass and frames, supply and install"),
        ("Design life required", "Windows and external doors 25 years; rooflights 25 years (ER)"),
        ("Note", "We hold an IBG with the CPA and are FENSA registered. Both are built around replacement work in dwellings; this is new build and we are a sub-contractor, so eligibility here is unconfirmed."),
    ]
    r = 4
    for k, v in info:
        ck = ws.cell(r, 1, k)
        ck.font = BOLD
        ck.alignment = TOP
        if k and not v:
            ck.fill = HDR
            ck.font = HDRF
            ws.cell(r, 2, "").fill = HDR
        cv = ws.cell(r, 2, v)
        cv.font = BASE
        cv.alignment = WRAP
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 132

    # --------------------------------------------- 2. Window & Door Schedule
    ws2 = wb.create_sheet("Window & Door Schedule")
    ws2["A1"] = "WINDOW & DOOR SCHEDULE - EVERY EXTERNAL ELEMENT"
    ws2["A1"].font = TITLE
    ws2["A2"] = ("Taken from the architect's schedules, not from the bill. The BoQ SMD issued omits the "
                 "Type E.01 and E.03 doors and one Type E.05 window - they are listed here because they are "
                 "on the drawings.")
    ws2["A2"].font = BASE
    hdrs = ["Ref", "Type", "Width", "Height", "m2", "Location", "Configuration",
            "Glazing (as specified)", "Supplier line", "Status", "Notes"]
    for i, h in enumerate(hdrs, 1):
        c = ws2.cell(4, i, h)
        c.fill = HDR
        c.font = HDRF
        c.alignment = WRAP
    r = 5
    for ref, typ, w, h, loc, cfg, glz, sup, st, note in ELEMENTS:
        vals = [ref, typ, w, h, round(w * h / 1e6, 4), loc, cfg, glz, sup, st, note]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(r, i, v)
            c.font = BASE
            c.alignment = WRAP if i in (6, 7, 8, 9, 11) else TOP
            c.border = THIN
            if i == 5:
                c.number_format = "0.0000"
            if i == 10:
                c.fill = GRN if v == Q else RED
                c.font = BOLD
        r += 1

    r += 1
    ws2.cell(r, 1, "TOTALS").font = BOLD
    r += 1
    win = [e for e in ELEMENTS if e[0].startswith("W")]
    dr = [e for e in ELEMENTS if e[0].startswith("ED")]
    for label, group in (("External windows", win), ("External doors", dr)):
        qn = [e for e in group if e[8] == Q]
        nn = [e for e in group if e[8] == N]
        ws2.cell(r, 1, label).font = BOLD
        ws2.cell(r, 2, "%d on schedule" % len(group)).font = BASE
        ws2.cell(r, 4, "quoted %d" % len(qn)).font = BASE
        ws2.cell(r, 5, round(sum(e[2] * e[3] for e in qn) / 1e6, 2)).number_format = "0.00"
        ws2.cell(r, 6, "NOT quoted %d" % len(nn)).font = BOLD
        ws2.cell(r, 7, round(sum(e[2] * e[3] for e in nn) / 1e6, 2)).number_format = "0.00"
        ws2.cell(r, 8, "m2 not quoted").font = BASE
        r += 1
    r += 1
    for t in [
        "Door area quoted 24.96 m2 against SMD's own Building Works Pricing Schedule figure of 48 m2 for "
        "External Doors. Adding the 7 unquoted doors gives 51.70 m2, which brackets their number.",
        "Window area quoted 112.20 m2 (plus 5.04 m2 of E.01 fixed fields inside the door elements) against "
        "SMD's 136 m2 - the balance is broadly the 7 rooflights, which are the roofing package's.",
    ]:
        c = ws2.cell(r, 1, t)
        c.font = BASE
        c.alignment = WRAP
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws2.row_dimensions[r].height = 28
        r += 1

    r += 2
    ws2.cell(r, 1, "OUTSIDE FENSTER'S PACKAGE - exclude in writing, none of these is excluded today").font = BOLD
    r += 1
    for i, h in enumerate(["Ref", "Type", "Qty", "Location", "Description", "Why it is not ours"], 1):
        c = ws2.cell(r, i, h)
        c.fill = HDR
        c.font = HDRF
        c.alignment = WRAP
    r += 1
    for ref, typ, qty, loc, desc, why in OUT_OF_SCOPE:
        for i, v in enumerate([ref, typ, qty, loc, desc, why], 1):
            c = ws2.cell(r, i, v)
            c.font = BASE
            c.alignment = WRAP
            c.border = THIN
            c.fill = AMB
        r += 1
    for c, w in zip("ABCDEFGHIJK", [11, 7, 8, 8, 9, 22, 26, 27, 24, 12, 60]):
        ws2.column_dimensions[c].width = w

    # ------------------------------------------------------ 3. RFIs & Queries
    ws3 = wb.create_sheet("RFIs & Queries")
    ws3["A1"] = "RFIs & QUERIES"
    ws3["A1"].font = TITLE
    ws3["A2"] = "Mary cannot email SMD or the suppliers. These need a human to send them."
    ws3["A2"].font = BASE
    for i, h in enumerate(["#", "To", "Query", "Why it matters", "Status"], 1):
        c = ws3.cell(4, i, h)
        c.fill = HDR
        c.font = HDRF
    rfis = [
        ("SMD", "Are the 5no Type E.01 steel sports hall escape doors (ED.0.04/06/07/08/09) and the 2no "
                "Type E.03 louvred plant room doors (ED.0.01/05) in our package? They are on door schedule "
                "P06 but not in the BoQ issued to us.",
         "About GBP 32,462 of sell plus louvre infill. Neither type has a supplier. Your bill omitted them, "
         "so this is a package boundary question, not a pricing one.", "OPEN - REQ-13"),
        ("SMD", "Will a double glazed external door meeting the 1.2 W/m2K door U-value be accepted against "
                "the triple glazed specification at ER 7.5.1?",
         "SMA Smart Wall is not manufactured in triple glazing. Either the door package changes system or "
         "the deviation is qualified.", "Adam calling SMD - REQ-2"),
        ("SMD", "Is the 1.1 W/m2K an area-weighted average across all external glazing, as the door schedule "
                "states, or the per-element maximum tabulated at ER 5.4.1?",
         "Changes whether individual elements can be traded off against each other.", "OPEN"),
        ("SMD", "Confirm the solar control glass specification - a product or a g-value. The window schedule "
                "marks every window 'Solar Control Glazing' but names nothing.",
         "Cannot be priced without it. Currently absent from the tender entirely - about GBP 4,177 at "
         "benchmark.", "OPEN"),
        ("SMD", "Confirm rooflights R.2.01-07 sit with the roofing package under ER 7.6.6 and are not ours.",
         "ER 7.6.6 requires all roof components from one manufacturer for a single-point guarantee, which "
         "reads as excluding us. Needs confirming before we exclude it in writing.", "OPEN"),
        ("SMD", "Which doors require access control? The door schedule leaves this to the Client.",
         "Unpriced and unbounded as it stands.", "OPEN"),
        ("SMD", "Confirm the manifestation pattern, or confirm a provisional allowance is acceptable.",
         "Required on 4 window types and door type E.04. Not priced, not excluded.", "OPEN"),
        ("BSW", "Requote QT253232 with: solar control glass; trickle vents min 4000mm2 to every head; "
                "restricted opening hinges to 100mm with push-button release to BS 6375-2; lockable "
                "espagnolette handles with 3no keys per window; obscure to W.1.19; the 3rd Type E.05 "
                "(W.1.20); BS 7950 / WCL 4 certification; a stated U-value; internal finish confirmed "
                "RAL 7016; lead time; delivered price.",
         "Nine separate specification requirements absent from the current quote. Quote expires ~21/08.",
         "OPEN"),
        ("Bellview", "Requote 0000000503 with: 2no Type E.04 not 1no; hold-open closers per ER 7.5.5/7.5.7 "
                     "in place of the NHO closers quoted; the power operator to ED.0.10; obscure to ED.0.13; "
                     "PAS 24 / LPS 1175 SR2 certification; Part M threshold seals; internal finish; lead "
                     "time; delivered price. And confirm in writing whether Smart Wall can be supplied "
                     "triple glazed.",
         "One door is sold with no quote behind it, and the closers quoted are the opposite of those "
         "specified. Quote expires ~21/08.", "OPEN"),
        ("Steel doorset supplier", "Quote 5no Type E.01 steel sports hall escape doorsets 1810x2110, PPC "
                                   "anthracite, flush faces, recessed panic furniture to BOTH leaves, fire "
                                   "escape, no external access, fire signage. Strongdor Sportsdor or similar "
                                   "approved.",
         "No fabricator engaged. Tender closes Friday.", "OPEN - needs raising"),
        ("Louvred door supplier", "Quote 2no Type E.03 aluminium louvred plant room doorsets 1810x2110, PPC "
                                  "anthracite, 3 pairs hinges, suited mortice lock, hold stay bolt to slave leaf.",
         "No fabricator engaged. Tender closes Friday.", "OPEN - needs raising"),
        ("Adam", "MCD - gross the subtotal up 2.5% so the deduction is neutral, or take it off the bottom "
                 "as on Princess Beatrice?",
         "GBP 2,401.88 either onto the price or off the margin.", "OPEN"),
        ("Adam", "Mastic and EPDM.",
         "ANSWERED by your 28/07 ruling - the mastic exclusion comes out when we send pricing. Both are "
         "specification requirements here, so both go into the sum and the 'optional extra' line comes out "
         "of the proposal. GBP 4,776.56.", "CLOSED 28/07"),
    ]
    r = 5
    for i, (to, q, why, st) in enumerate(rfis, 1):
        for j, v in enumerate([i, to, q, why, st], 1):
            c = ws3.cell(r, j, v)
            c.font = BASE
            c.alignment = WRAP
            c.border = THIN
            if j == 5:
                c.fill = GRN if str(v).startswith("CLOSED") else AMB
        ws3.row_dimensions[r].height = 50
        r += 1
    for c, w in zip("ABCDE", [5, 22, 74, 58, 20]):
        ws3.column_dimensions[c].width = w

    for s in wb.worksheets:
        s.freeze_panes = {"Project Information": "A4", "Window & Door Schedule": "A5",
                          "RFIs & Queries": "A5"}[s.title]
    wb.save(OUT)

    qn = [e for e in ELEMENTS if e[8] == Q]
    nn = [e for e in ELEMENTS if e[8] == N]
    print("wrote", OUT)
    print("elements on schedule %d   quoted %d   NOT quoted %d" % (len(ELEMENTS), len(qn), len(nn)))
    print("  not quoted:", ", ".join(e[0] for e in nn))


if __name__ == "__main__":
    build()
