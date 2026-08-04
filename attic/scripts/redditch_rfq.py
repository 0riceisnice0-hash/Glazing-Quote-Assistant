# -*- coding: utf-8 -*-
"""Redditch Library - a supplier-facing RFQ schedule, with no prices in it.

A SEPARATE FILE ON PURPOSE. The take-off workbook carries our buy rates, our
margin and a competitor's tendered price, and it must never be the thing that
reaches a supplier. Gordon Court sent a client five supplier quotations under
the filename "Elevations" (REQ-28); the cheapest guard against that is for the
sendable document to be a different file that has never contained a price.

Nothing here is confidential to Fenster: references, sizes, configurations and
the client's own performance specification. All of it is in the tender pack.

  python scripts/redditch_rfq.py
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import redditch_takeoff as tk
from redditch_takeoff_doc import CONFIG, ELEV

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "outputs", "Redditch Library - RFQ Schedule for Supplier.xlsx")

HDR = PatternFill("solid", fgColor="1F2A44")
SUB = PatternFill("solid", fgColor="E8ECF3")
WHITE = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFC7D5")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

SPEC = [
    ("Project", "Redditch Library, 15 Market Place, Redditch, B98 8AR - external window and door replacement"),
    ("Quantity", "43 items across 41 references, 136.54 m2 total"),
    ("Windows", "Polyester powder coated THERMALLY BROKEN aluminium casement. The tender names Joedan's EL75mm Squareline as the minimum standard and expressly permits an alternative; please quote your nearest equivalent and state the system."),
    ("Doors", "Polyester powder coated commercial aluminium doorsets, refs 32.2, 34.2, 37, 39 (single) and 41 (double). The tender names AC100 Commercial non-thermal as the minimum standard."),
    ("PLEASE NOTE - refs 32 and 34", "Each is ONE coupled assembly, window frame joined directly to a door frame. Please quote each as a single coupled unit IN ONE SYSTEM AND ONE FRAME DEPTH throughout. Do not quote the window in a 75mm thermally broken system and the door in a 100mm non-thermal one - they cannot be joined."),
    ("U-values", "Windows 1.4 W/m2K area-weighted across the package. Doors 3.0 W/m2K area-weighted. Infill panels 1.2 W/m2K. Please state the whole-window Uw you are quoting, not a centre-pane Ug."),
    ("Glass", "Argon filled double glazed units, warm edge spacer. INNER 4mm clear soft coat low-E toughened. OUTER 4mm BRONZE ANTI-SUN toughened. Safety glass to all Part K critical locations."),
    ("Panels", "Solid phenolic core, faced both sides with 2.0mm polyester powder coated aluminium, MATT finish, 1.2 W/m2K."),
    ("Window hardware", "Shootbolt espagnolette locking. Securistyle Defender friction hinges. Separate concealed restrictor limiting initial opening to 100mm where required. TRICKLE VENTILATORS TO ALL CASEMENTS - the specification says these must be included."),
    ("Door hardware - refs 32.2, 34.2, 37 and 41 ONLY", "Anti-finger-trap stiles; 100mm bottom rail; 100mm midrail; Axim 8800 concealed non-hold-open transom closer; PR7100 exit panic device without external access and no external hardware; mill finished low threshold. Ref 41 is a double door and additionally needs flush bolts to head and threshold of the slave leaf."),
    ("Door hardware - ref 39", "NOT SPECIFIED in the tender. Please quote as a plain single commercial doorset with standard ironmongery and no panic set, and price any panic set separately so it can be added if the CA confirms it."),
    ("Finish", "Single standard RAL, the same colour internally and externally. The RAL is not yet stated by the client - please quote on a standard RAL and tell us which RALs are standard and what a non-standard or metallic colour would add."),
    ("Ancillaries", "45mm white cellular uPVC cloaking profile to the internal perimeter where required."),
    ("Supply basis", "Supply only, factory glazed, delivered to site. Installation is ours. Please state your delivery terms and any carriage threshold ON THE FACE OF THE QUOTATION."),
    ("Validity", "PLEASE STATE HOW LONG THE PRICE IS FIRM. The main contract tender must stay open 10 weeks from submission and the preliminaries say not less than 3 months, so a 30-day quotation leaves us exposed. If you can hold longer, please say so and for how long."),
    ("Warranty", "Please state the guarantee period, what it covers, when it starts, and anything capped by cycles rather than time."),
    ("SIZES ARE INDICATIVE", "Taken from the client's tender schedule. The specification requires a measured survey of every opening before ordering and states that the client's dimensions must not be relied on. Quote on these for tender purposes; sizes will be confirmed by survey."),
    ("RAKED UNITS - refs 16, 17 and 18", "These are drawn on the elevations as PARALLELOGRAMS on a raking stair wall - sloping head and sloping cill. The 2250 x 2304 below is a bounding box, not the frame. Please price them as raked units and tell us what the rake adds over a rectangle of the same area."),
    ("Ref 38", "The schedule says 6 fixed lights; the elevation drawing shows a 5-wide by 3-high grid of 15 panes. Please quote the 15-pane arrangement and note the difference."),
    ("Refs 29, 30 and 31", "The tender gives NO configuration for these three. Please quote as fixed lights."),
]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Enquiry")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 104
    c = ws.cell(row=1, column=1, value="Redditch Library - request for quotation, windows and doors")
    c.fill, c.font = HDR, WHITE
    ws.cell(row=1, column=2).fill = HDR
    r = 3
    for k, v in SPEC:
        for col, val in ((1, k), (2, v)):
            cc = ws.cell(row=r, column=col, value=val)
            cc.border, cc.alignment = THIN, WRAP
            if col == 1:
                cc.font = BOLD
        r += 1

    ws = wb.create_sheet("Schedule")
    for col, w in zip("ABCDEFGH", (9, 7, 10, 10, 9, 9, 40, 22)):
        ws.column_dimensions[col].width = w
    head = ["Ref", "Elev", "Width mm", "Height mm", "Area m2", "Type", "Configuration", "Your rate GBP"]
    for i, v in enumerate(head, start=1):
        cc = ws.cell(row=1, column=i, value=v)
        cc.fill, cc.font, cc.border, cc.alignment = HDR, WHITE, THIN, WRAP
    ws.freeze_panes = "A2"
    r = 2
    total = 0.0
    for ref, wd, ht, th, fl, fp, sd, dd in tk.SCHEDULE:
        area = round(wd / 1000.0 * ht / 1000.0, 3)
        total += wd / 1000.0 * ht / 1000.0        # sum unrounded, or the total reads 136.53
        vals = [ref, ELEV.get(ref, ""), wd, ht, area,
                "door" if (sd or dd) else "window", CONFIG.get(ref, ""), None]
        for i, v in enumerate(vals, start=1):
            cc = ws.cell(row=r, column=i, value=v)
            cc.border, cc.alignment = THIN, WRAP
        r += 1
    for i, v in enumerate(["TOTAL", "", "", "", round(total, 2), "43 items", "", None], start=1):
        cc = ws.cell(row=r, column=i, value=v)
        cc.fill, cc.font, cc.border = SUB, BOLD, THIN

    wb.properties.creator = "Fenster Glazing & Locks Ltd"
    wb.properties.lastModifiedBy = "Fenster Glazing & Locks Ltd"
    wb.properties.title = "Redditch Library - RFQ Schedule"
    wb.properties.company = "Fenster Glazing & Locks Ltd"
    wb.save(OUT)
    print("written %s" % OUT)


if __name__ == "__main__":
    main()
